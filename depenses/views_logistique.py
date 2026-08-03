from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F, DecimalField, IntegerField, Value
from django.db.models.functions import Coalesce
from django.db import models, transaction
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models_logistique import (
    CategorieArticle, Article, BienEtablissement, 
    MouvementStock, Inventaire, LigneInventaire, ContributionPapierRame
)
from .forms import (
    CategorieArticleForm, ArticleForm, BienEtablissementForm, MouvementStockForm,
    VenteFournitureForm,
    InventaireForm, LigneInventaireForm, ContributionPapierRameForm
)


def _biens_pour_utilisateur(user):
    from utilisateurs.utils import user_is_superadmin, user_school

    biens = BienEtablissement.objects.filter(actif=True)
    if user_is_superadmin(user):
        return biens
    ecole = user_school(user)
    if not ecole:
        return biens.none()
    return biens.filter(cree_par__profil__ecole=ecole)


def _articles_fournitures_pour_utilisateur(user, actifs_seulement=True):
    """Fournitures visibles par l'utilisateur, strictement limitees a son ecole."""
    from utilisateurs.utils import filter_by_user_school

    articles = Article.objects.select_related('categorie').filter(
        categorie__type_categorie='FOURNITURE',
    )
    if actifs_seulement:
        articles = articles.filter(actif=True)
    return filter_by_user_school(articles, user, 'cree_par__profil__ecole')


def _categorie_fournitures_defaut():
    categorie, _ = CategorieArticle.objects.get_or_create(
        code='FOURNITURES',
        defaults={
            'nom': 'Fournitures scolaires',
            'type_categorie': 'FOURNITURE',
            'description': 'Produits scolaires destines a la vente.',
            'actif': True,
        },
    )
    champs = []
    if categorie.type_categorie != 'FOURNITURE':
        categorie.type_categorie = 'FOURNITURE'
        champs.append('type_categorie')
    if not categorie.actif:
        categorie.actif = True
        champs.append('actif')
    if champs:
        categorie.save(update_fields=champs)
    return categorie


def _contributions_pour_utilisateur(user, annee_scolaire=None):
    from utilisateurs.utils import filter_by_user_school

    contributions = ContributionPapierRame.objects.select_related(
        'eleve', 'eleve__classe', 'eleve__classe__ecole', 'cree_par'
    )
    contributions = filter_by_user_school(contributions, user, 'eleve__classe__ecole')
    if annee_scolaire:
        contributions = contributions.filter(eleve__classe__annee_scolaire=annee_scolaire)
    return contributions


@login_required
def dashboard_logistique(request):
    """Tableau de bord simplifié : biens de l'école et papier RAM."""
    from utilisateurs.utils import user_school, filter_by_user_school
    from eleves.models import Eleve
    from eleves.utils_annee import get_annee_active

    ecole = user_school(request.user)
    annee_active = get_annee_active(request, ecole) if ecole else None
    biens = _biens_pour_utilisateur(request.user)
    contributions = _contributions_pour_utilisateur(request.user, annee_active)

    stats_biens = biens.aggregate(
        total_biens=Count('id'),
        somme_quantite_achetee=Sum('quantite_achetee'),
        somme_quantite_utilisee=Sum('quantite_utilisee'),
        somme_quantite_gatee=Sum('quantite_gatee'),
        valeur_totale=Sum(
            F('quantite_achetee') * F('prix_achat_unitaire'),
            output_field=DecimalField(max_digits=20, decimal_places=0),
        ),
    )
    quantite_achetee = stats_biens.pop('somme_quantite_achetee') or 0
    quantite_utilisee = stats_biens.pop('somme_quantite_utilisee') or 0
    quantite_gatee = stats_biens.pop('somme_quantite_gatee') or 0
    stats_biens['quantite_achetee'] = quantite_achetee
    stats_biens['quantite_utilisee'] = quantite_utilisee
    stats_biens['quantite_gatee'] = quantite_gatee
    stats_biens['quantite_disponible'] = max(
        quantite_achetee - quantite_utilisee - quantite_gatee,
        0,
    )
    stats_biens['valeur_totale'] = stats_biens['valeur_totale'] or 0

    stats_papier = contributions.aggregate(
        total_paquets=Sum('nombre_paquets'),
        total_argent=Sum('montant_paye'),
    )
    stats_papier['total_paquets'] = stats_papier['total_paquets'] or 0
    stats_papier['total_argent'] = stats_papier['total_argent'] or 0
    stats_papier['eleves_enregistres'] = contributions.values('eleve_id').distinct().count()

    eleves = Eleve.objects.filter(statut='ACTIF')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    if annee_active:
        eleves = eleves.filter(classe__annee_scolaire=annee_active)
    stats_papier['total_eleves'] = eleves.count()
    stats_papier['eleves_non_enregistres'] = max(
        stats_papier['total_eleves'] - stats_papier['eleves_enregistres'],
        0,
    )

    context = {
        'titre_page': 'Gestion logistique',
        'annee_active': annee_active,
        'stats_biens': stats_biens,
        'stats_papier': stats_papier,
        'derniers_biens': biens.order_by('-date_modification')[:6],
        'dernieres_contributions': contributions[:8],
    }
    return render(request, 'depenses/logistique/dashboard.html', context)


@login_required
def liste_articles(request):
    """Tableau de bord du stock et des ventes de fournitures scolaires."""
    q = request.GET.get('q', '').strip()
    categorie_id = request.GET.get('categorie', '')
    alerte = request.GET.get('alerte', '')

    articles = _articles_fournitures_pour_utilisateur(request.user)
    if q:
        articles = articles.filter(
            Q(code_article__icontains=q)
            | Q(nom__icontains=q)
            | Q(marque__icontains=q)
            | Q(reference__icontains=q)
        )
    if categorie_id:
        articles = articles.filter(categorie_id=categorie_id)
    if alerte == 'oui':
        articles = articles.filter(stock_actuel__lte=F('stock_minimum'))

    articles = articles.annotate(
        quantite_approvisionnee=Coalesce(
            Sum('mouvements__quantite', filter=Q(mouvements__type_mouvement='ENTREE')),
            Value(0),
            output_field=IntegerField(),
        ),
        quantite_vendue=Coalesce(
            Sum(
                'mouvements__quantite',
                filter=Q(mouvements__type_mouvement='SORTIE', mouvements__motif='VENTE'),
            ),
            Value(0),
            output_field=IntegerField(),
        ),
        chiffre_affaires=Coalesce(
            Sum(
                'mouvements__montant_total',
                filter=Q(mouvements__type_mouvement='SORTIE', mouvements__motif='VENTE'),
            ),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=20, decimal_places=0),
        ),
    ).order_by('nom')

    articles = list(articles)
    stats = {
        'produits': len(articles),
        'quantite_approvisionnee': 0,
        'quantite_vendue': 0,
        'quantite_restante': 0,
        'valeur_stock': Decimal('0'),
        'chiffre_affaires': Decimal('0'),
        'solde': Decimal('0'),
    }
    for article in articles:
        article.cout_produits_vendus = article.quantite_vendue * article.prix_unitaire
        article.solde_ventes = article.chiffre_affaires - article.cout_produits_vendus
        stats['quantite_approvisionnee'] += article.quantite_approvisionnee
        stats['quantite_vendue'] += article.quantite_vendue
        stats['quantite_restante'] += article.stock_actuel
        stats['valeur_stock'] += article.valeur_stock
        stats['chiffre_affaires'] += article.chiffre_affaires
        stats['solde'] += article.solde_ventes

    categories = CategorieArticle.objects.filter(actif=True, type_categorie='FOURNITURE')
    articles_ids = [article.pk for article in articles]
    dernieres_ventes = MouvementStock.objects.select_related('article', 'cree_par').filter(
        article_id__in=articles_ids,
        type_mouvement='SORTIE',
        motif='VENTE',
    ).order_by('-date_mouvement')[:8]

    return render(request, 'depenses/logistique/liste_articles.html', {
        'titre_page': 'Gestion et vente des fournitures scolaires',
        'articles': articles,
        'categories': categories,
        'stats': stats,
        'dernieres_ventes': dernieres_ventes,
        'q': q,
        'categorie_id': categorie_id,
        'alerte': alerte,
    })


@login_required
def liste_articles_ancienne(request):
    """Liste des articles en stock"""
    from utilisateurs.utils import user_school

    # Filtres
    q = request.GET.get('q', '')
    categorie_id = request.GET.get('categorie', '')
    etat = request.GET.get('etat', '')
    alerte = request.GET.get('alerte', '')

    articles = Article.objects.select_related('categorie').filter(actif=True)
    # Sécurité : filtrer par école
    ecole = user_school(request.user)
    if ecole:
        articles = articles.filter(cree_par__profil__ecole=ecole)
    
    if q:
        articles = articles.filter(
            Q(code_article__icontains=q) |
            Q(nom__icontains=q) |
            Q(marque__icontains=q) |
            Q(reference__icontains=q)
        )
    
    if categorie_id:
        articles = articles.filter(categorie_id=categorie_id)
    
    if etat:
        articles = articles.filter(etat=etat)
    
    if alerte == 'oui':
        articles = articles.filter(stock_actuel__lte=models.F('stock_minimum'))
    
    categories = CategorieArticle.objects.filter(actif=True)
    
    context = {
        'titre_page': 'Stock & Articles',
        'articles': articles,
        'categories': categories,
        'q': q,
        'categorie_id': categorie_id,
        'etat': etat,
        'alerte': alerte,
    }
    
    return render(request, 'depenses/logistique/liste_articles.html', context)


@login_required
def liste_biens(request):
    """Liste simplifiée des biens de l'établissement."""
    q = request.GET.get('q', '').strip()
    type_bien = request.GET.get('type_bien', '')
    etat = request.GET.get('etat', '')

    biens = _biens_pour_utilisateur(request.user)
    
    if q:
        biens = biens.filter(
            Q(code_bien__icontains=q) |
            Q(nom__icontains=q) |
            Q(localisation__icontains=q)
        )
    
    if type_bien:
        biens = biens.filter(type_bien=type_bien)
    
    if etat:
        biens = biens.filter(etat=etat)

    stats = biens.aggregate(
        total=Count('id'),
        achetee=Sum('quantite_achetee'),
        utilisee=Sum('quantite_utilisee'),
        gatee=Sum('quantite_gatee'),
    )
    stats = {key: value or 0 for key, value in stats.items()}
    stats['disponible'] = max(stats['achetee'] - stats['utilisee'] - stats['gatee'], 0)

    context = {
        'titre_page': "Biens de l'établissement",
        'biens': biens.order_by('nom'),
        'stats': stats,
        'q': q,
        'type_bien': type_bien,
        'etat': etat,
        'types_bien': BienEtablissement.TYPE_CHOICES,
        'etats_bien': BienEtablissement.ETAT_CHOICES,
    }
    return render(request, 'depenses/logistique/liste_biens.html', context)


@login_required
def creer_bien(request):
    """Ajouter un bien de l'établissement."""
    if request.method == 'POST':
        form = BienEtablissementForm(request.POST, request.FILES)
        if form.is_valid():
            bien = form.save(commit=False)
            bien.cree_par = request.user
            
            if not bien.code_bien:
                today = date.today()
                prefix = f"BIEN-{today.strftime('%Y%m%d')}"
                last_bien = BienEtablissement.objects.filter(
                    code_bien__startswith=prefix
                ).order_by('-code_bien').first()
                
                if last_bien:
                    last_num = int(last_bien.code_bien.split('-')[-1])
                    bien.code_bien = f"{prefix}-{last_num + 1:04d}"
                else:
                    bien.code_bien = f"{prefix}-0001"
            
            bien.save()
            messages.success(request, f'Bien "{bien.nom}" créé avec succès.')
            return redirect('depenses:liste_biens')
    else:
        form = BienEtablissementForm()
    
    context = {
        'titre_page': 'Nouveau Bien',
        'form': form,
    }
    
    return render(request, 'depenses/logistique/form_bien.html', context)


@login_required
def modifier_bien(request, bien_id):
    """Modifier un bien, avec isolation stricte par école."""
    bien = get_object_or_404(_biens_pour_utilisateur(request.user), pk=bien_id)
    
    if request.method == 'POST':
        form = BienEtablissementForm(request.POST, request.FILES, instance=bien)
        if form.is_valid():
            form.save()
            messages.success(request, f'Bien "{bien.nom}" modifié avec succès.')
            return redirect('depenses:liste_biens')
    else:
        form = BienEtablissementForm(instance=bien)
    
    context = {
        'titre_page': 'Modifier Bien',
        'form': form,
        'bien': bien,
    }
    
    return render(request, 'depenses/logistique/form_bien.html', context)


@login_required
def liste_contributions_papier(request):
    """Historique consultable et filtrable des contributions papier RAM."""
    from utilisateurs.utils import user_school
    from eleves.utils_annee import get_annee_active

    ecole = user_school(request.user)
    annee_active = get_annee_active(request, ecole) if ecole else None
    q = request.GET.get('q', '').strip()
    mode = request.GET.get('mode', '').strip()
    contributions = _contributions_pour_utilisateur(request.user, annee_active)

    if q:
        contributions = contributions.filter(
            Q(eleve__matricule__icontains=q)
            | Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q)
            | Q(eleve__classe__nom__icontains=q)
        )
    if mode in {'PAPIER', 'ARGENT'}:
        contributions = contributions.filter(type_contribution=mode)

    stats = contributions.aggregate(
        total_paquets=Sum('nombre_paquets'),
        total_argent=Sum('montant_paye'),
    )
    stats['total_paquets'] = stats['total_paquets'] or 0
    stats['total_argent'] = stats['total_argent'] or 0
    stats['total_enregistrements'] = contributions.count()
    stats['total_eleves'] = contributions.values('eleve_id').distinct().count()
    page_obj = Paginator(contributions, 50).get_page(request.GET.get('page'))

    return render(request, 'depenses/logistique/liste_contributions_papier.html', {
        'titre_page': 'Gestion du papier RAM',
        'annee_active': annee_active,
        'page_obj': page_obj,
        'q': q,
        'mode': mode,
        'stats': stats,
    })


@login_required
def creer_contribution_papier(request):
    """Enregistrer des paquets de papier ou un paiement à la place."""
    from utilisateurs.utils import user_school
    from eleves.utils_annee import get_annee_active

    ecole = user_school(request.user)
    annee_active = get_annee_active(request, ecole) if ecole else None
    kwargs = {'user': request.user, 'annee_scolaire': annee_active}
    if request.method == 'POST':
        form = ContributionPapierRameForm(request.POST, **kwargs)
        if form.is_valid():
            contribution = form.save(commit=False)
            contribution.cree_par = request.user
            contribution.save()
            messages.success(
                request,
                f'Contribution papier RAM de {contribution.eleve.nom_complet} enregistrée.',
            )
            return redirect('depenses:liste_contributions_papier')
    else:
        initial = {}
        if request.GET.get('eleve'):
            initial['eleve'] = request.GET['eleve']
        form = ContributionPapierRameForm(initial=initial, **kwargs)

    return render(request, 'depenses/logistique/form_contribution_papier.html', {
        'titre_page': 'Nouvelle contribution papier RAM',
        'annee_active': annee_active,
        'form': form,
    })


@login_required
def modifier_contribution_papier(request, contribution_id):
    """Corriger un enregistrement papier RAM de la même école."""
    from utilisateurs.utils import user_school
    from eleves.utils_annee import get_annee_active

    ecole = user_school(request.user)
    annee_active = get_annee_active(request, ecole) if ecole else None
    contribution = get_object_or_404(
        _contributions_pour_utilisateur(request.user),
        pk=contribution_id,
    )
    kwargs = {
        'instance': contribution,
        'user': request.user,
        'annee_scolaire': contribution.eleve.classe.annee_scolaire,
    }
    if request.method == 'POST':
        form = ContributionPapierRameForm(request.POST, **kwargs)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contribution papier RAM modifiée avec succès.')
            return redirect('depenses:liste_contributions_papier')
    else:
        form = ContributionPapierRameForm(**kwargs)

    return render(request, 'depenses/logistique/form_contribution_papier.html', {
        'titre_page': 'Modifier la contribution papier RAM',
        'annee_active': annee_active,
        'form': form,
        'contribution': contribution,
    })


@login_required
def ancienne_logistique_redirection(request, *args, **kwargs):
    """Conserver les anciens liens sans exposer les écrans de stock retirés."""
    messages.info(
        request,
        "La gestion Stock, Articles, Inventaires et Mouvements a été remplacée par la gestion simplifiée des biens.",
    )
    return redirect('depenses:dashboard_logistique')


@login_required
def liste_mouvements(request):
    """Historique des entrees, ventes et autres sorties de fournitures."""
    article_id = request.GET.get('article', '')
    type_mouvement = request.GET.get('type', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    articles = _articles_fournitures_pour_utilisateur(request.user).order_by('nom')
    mouvements = MouvementStock.objects.select_related('article', 'cree_par').filter(
        article__in=articles,
    )

    if article_id:
        mouvements = mouvements.filter(article_id=article_id)
    if type_mouvement:
        mouvements = mouvements.filter(type_mouvement=type_mouvement)
    if date_debut:
        mouvements = mouvements.filter(date_mouvement__date__gte=date_debut)
    if date_fin:
        mouvements = mouvements.filter(date_mouvement__date__lte=date_fin)

    return render(request, 'depenses/logistique/liste_mouvements.html', {
        'titre_page': 'Historique des fournitures scolaires',
        'mouvements': mouvements.order_by('-date_mouvement'),
        'articles': articles,
        'article_id': article_id,
        'type_mouvement': type_mouvement,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


@login_required
def liste_mouvements_ancienne(request):
    """Liste des mouvements de stock"""
    from utilisateurs.utils import user_school

    # Filtres
    article_id = request.GET.get('article', '')
    type_mouvement = request.GET.get('type', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    mouvements = MouvementStock.objects.select_related(
        'article', 'cree_par'
    ).all()
    # Sécurité : filtrer par école
    ecole = user_school(request.user)
    if ecole:
        mouvements = mouvements.filter(cree_par__profil__ecole=ecole)
    
    if article_id:
        mouvements = mouvements.filter(article_id=article_id)
    
    if type_mouvement:
        mouvements = mouvements.filter(type_mouvement=type_mouvement)
    
    if date_debut:
        mouvements = mouvements.filter(date_mouvement__gte=date_debut)
    
    if date_fin:
        mouvements = mouvements.filter(date_mouvement__lte=date_fin)
    
    articles = Article.objects.filter(actif=True)
    
    context = {
        'titre_page': 'Mouvements de Stock',
        'mouvements': mouvements,
        'articles': articles,
        'article_id': article_id,
        'type_mouvement': type_mouvement,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'depenses/logistique/liste_mouvements.html', context)


@login_required
def liste_inventaires(request):
    """Liste des inventaires"""
    
    inventaires = Inventaire.objects.select_related(
        'cree_par', 'valide_par'
    ).order_by('-date_inventaire')
    
    context = {
        'titre_page': 'Inventaires',
        'inventaires': inventaires,
    }
    
    return render(request, 'depenses/logistique/liste_inventaires.html', context)


@login_required
def creer_mouvement(request):
    """Enregistrer un approvisionnement ou une sortie de stock."""
    kwargs = {'user': request.user}
    mouvement_enregistre = False
    if request.method == 'POST':
        form = MouvementStockForm(request.POST, **kwargs)
        if form.is_valid():
            with transaction.atomic():
                article = get_object_or_404(
                    _articles_fournitures_pour_utilisateur(request.user).select_for_update(),
                    pk=form.cleaned_data['article'].pk,
                )
                quantite = form.cleaned_data['quantite']
                type_mouvement = form.cleaned_data['type_mouvement']
                if type_mouvement == 'SORTIE' and quantite > article.stock_actuel:
                    form.add_error(
                        'quantite',
                        f"Stock insuffisant : {article.stock_actuel} unite(s) disponible(s).",
                    )
                else:
                    mouvement = form.save(commit=False)
                    mouvement.article = article
                    mouvement.cree_par = request.user
                    mouvement.numero_mouvement = _generer_code(
                        MouvementStock, 'numero_mouvement', 'MVT'
                    )
                    if mouvement.prix_unitaire is None:
                        mouvement.prix_unitaire = (
                            article.prix_vente_unitaire
                            if mouvement.motif == 'VENTE'
                            else article.prix_unitaire
                        )
                    mouvement.save()
                    mouvement_enregistre = True
            if mouvement_enregistre:
                messages.success(request, 'Mouvement de stock enregistre avec succes.')
                return redirect('depenses:liste_articles')
    else:
        initial = {
            'article': request.GET.get('article', ''),
            'type_mouvement': 'ENTREE',
            'motif': 'ACHAT',
            'date_mouvement': timezone.localtime().strftime('%Y-%m-%dT%H:%M'),
        }
        article = _articles_fournitures_pour_utilisateur(request.user).filter(
            pk=request.GET.get('article') or None,
        ).first()
        if article:
            initial['prix_unitaire'] = article.prix_unitaire
        form = MouvementStockForm(initial=initial, **kwargs)

    return render(request, 'depenses/logistique/form_mouvement.html', {
        'titre_page': 'Approvisionner le stock',
        'form': form,
    })


@login_required
def creer_vente_fourniture(request):
    """Enregistrer une vente et diminuer le stock de facon atomique."""
    kwargs = {
        'user': request.user,
        'article_id': request.GET.get('article'),
    }
    vente_enregistree = False
    if request.method == 'POST':
        form = VenteFournitureForm(request.POST, **kwargs)
        if form.is_valid():
            with transaction.atomic():
                article = get_object_or_404(
                    _articles_fournitures_pour_utilisateur(request.user).select_for_update(),
                    pk=form.cleaned_data['article'].pk,
                )
                quantite = form.cleaned_data['quantite']
                if quantite > article.stock_actuel:
                    form.add_error(
                        'quantite',
                        f"Stock insuffisant : {article.stock_actuel} unite(s) disponible(s).",
                    )
                else:
                    MouvementStock.objects.create(
                        numero_mouvement=_generer_code(
                            MouvementStock, 'numero_mouvement', 'MVT'
                        ),
                        article=article,
                        type_mouvement='SORTIE',
                        motif='VENTE',
                        quantite=quantite,
                        prix_unitaire=form.cleaned_data['prix_vente_unitaire'],
                        destinataire=form.cleaned_data.get('acheteur', ''),
                        document_reference=form.cleaned_data.get('document_reference', ''),
                        observations=form.cleaned_data.get('observations', ''),
                        cree_par=request.user,
                    )
                    vente_enregistree = True
            if vente_enregistree:
                messages.success(
                    request,
                    f"Vente de {quantite} unite(s) de {article.nom} enregistree avec succes.",
                )
                return redirect('depenses:liste_articles')
    else:
        form = VenteFournitureForm(**kwargs)

    return render(request, 'depenses/logistique/form_vente_fourniture.html', {
        'titre_page': 'Enregistrer une vente de fournitures',
        'form': form,
    })


@login_required
def creer_mouvement_ancien(request):
    """Créer un mouvement de stock"""
    
    if request.method == 'POST':
        form = MouvementStockForm(request.POST)
        if form.is_valid():
            mouvement = form.save(commit=False)
            mouvement.cree_par = request.user
            
            # Générer le numéro de mouvement
            today = date.today()
            prefix = f"MVT-{today.strftime('%Y%m%d')}"
            last_mvt = MouvementStock.objects.filter(
                numero_mouvement__startswith=prefix
            ).order_by('-numero_mouvement').first()
            
            if last_mvt:
                last_num = int(last_mvt.numero_mouvement.split('-')[-1])
                mouvement.numero_mouvement = f"{prefix}-{last_num + 1:04d}"
            else:
                mouvement.numero_mouvement = f"{prefix}-0001"
            
            mouvement.save()
            messages.success(request, 'Mouvement de stock créé avec succès.')
            return redirect('depenses:liste_mouvements')
    else:
        form = MouvementStockForm()
    
    context = {
        'titre_page': 'Nouveau Mouvement',
        'form': form,
    }
    
    return render(request, 'depenses/logistique/form_mouvement.html', context)


def _generer_code(modele, champ, prefixe):
    """Génère un code séquentiel du type PREFIXE-YYYYMMDD-0001."""
    today = date.today()
    base = f"{prefixe}-{today.strftime('%Y%m%d')}"
    dernier = modele.objects.filter(
        **{f"{champ}__startswith": base}
    ).order_by(f"-{champ}").first()
    if dernier:
        try:
            num = int(getattr(dernier, champ).split('-')[-1]) + 1
        except (ValueError, IndexError):
            num = 1
    else:
        num = 1
    return f"{base}-{num:04d}"


# ===== CRUD ARTICLES =====

@login_required
def detail_article(request, article_id):
    """Detail d'une fourniture avec ses ventes et ses mouvements."""
    article = get_object_or_404(
        _articles_fournitures_pour_utilisateur(request.user, actifs_seulement=False),
        pk=article_id,
    )
    mouvements = article.mouvements.select_related('cree_par').order_by('-date_mouvement')[:50]
    ventes = article.mouvements.filter(type_mouvement='SORTIE', motif='VENTE').aggregate(
        quantite=Coalesce(Sum('quantite'), Value(0), output_field=IntegerField()),
        chiffre_affaires=Coalesce(
            Sum('montant_total'),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=20, decimal_places=0),
        ),
    )
    ventes['solde'] = ventes['chiffre_affaires'] - ventes['quantite'] * article.prix_unitaire
    return render(request, 'depenses/logistique/detail_article.html', {
        'titre_page': f"Produit : {article.nom}",
        'article': article,
        'mouvements': mouvements,
        'ventes': ventes,
    })


@login_required
def detail_article_ancien(request, article_id):
    """Détail d'un article avec historique des mouvements"""
    article = get_object_or_404(Article, pk=article_id)
    mouvements = article.mouvements.select_related('cree_par').order_by('-date_mouvement')[:50]

    context = {
        'titre_page': f"Article : {article.nom}",
        'article': article,
        'mouvements': mouvements,
    }
    return render(request, 'depenses/logistique/detail_article.html', context)


@login_required
def creer_article(request):
    """Creer une fourniture et, si necessaire, son stock initial."""
    categorie = _categorie_fournitures_defaut()
    kwargs = {'user': request.user}
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, **kwargs)
        if form.is_valid():
            with transaction.atomic():
                article = form.save(commit=False)
                article.cree_par = request.user
                if not article.code_article:
                    article.code_article = _generer_code(Article, 'code_article', 'FOUR')
                article.stock_actuel = 0
                article.save()
                stock_initial = form.cleaned_data.get('stock_initial') or 0
                if stock_initial:
                    MouvementStock.objects.create(
                        numero_mouvement=_generer_code(
                            MouvementStock, 'numero_mouvement', 'MVT'
                        ),
                        article=article,
                        type_mouvement='ENTREE',
                        motif='ACHAT',
                        quantite=stock_initial,
                        prix_unitaire=article.prix_unitaire,
                        observations='Stock initial du produit',
                        cree_par=request.user,
                    )
            messages.success(request, f'Produit "{article.nom}" ajoute avec succes.')
            return redirect('depenses:liste_articles')
    else:
        form = ArticleForm(initial={'categorie': categorie}, **kwargs)

    return render(request, 'depenses/logistique/form_article.html', {
        'titre_page': 'Nouveau produit scolaire',
        'form': form,
    })


@login_required
def creer_article_ancien(request):
    """Créer un article en stock"""
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES)
        if form.is_valid():
            article = form.save(commit=False)
            article.cree_par = request.user
            if not article.code_article:
                article.code_article = _generer_code(Article, 'code_article', 'ART')
            article.save()
            messages.success(request, f'Article "{article.nom}" créé avec succès.')
            return redirect('depenses:detail_article', article_id=article.pk)
    else:
        form = ArticleForm()

    context = {
        'titre_page': 'Nouvel Article',
        'form': form,
    }
    return render(request, 'depenses/logistique/form_article.html', context)


@login_required
def modifier_article(request, article_id):
    """Modifier un produit de la meme ecole."""
    article = get_object_or_404(
        _articles_fournitures_pour_utilisateur(request.user),
        pk=article_id,
    )
    kwargs = {'instance': article, 'user': request.user}
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, **kwargs)
        if form.is_valid():
            form.save()
            messages.success(request, f'Produit "{article.nom}" modifie avec succes.')
            return redirect('depenses:liste_articles')
    else:
        form = ArticleForm(**kwargs)
    return render(request, 'depenses/logistique/form_article.html', {
        'titre_page': f"Modifier : {article.nom}",
        'form': form,
        'article': article,
    })


@login_required
def modifier_article_ancien(request, article_id):
    """Modifier un article en stock"""
    article = get_object_or_404(Article, pk=article_id)

    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            form.save()
            messages.success(request, f'Article "{article.nom}" modifié avec succès.')
            return redirect('depenses:detail_article', article_id=article.pk)
    else:
        form = ArticleForm(instance=article)

    context = {
        'titre_page': f"Modifier : {article.nom}",
        'form': form,
        'article': article,
    }
    return render(request, 'depenses/logistique/form_article.html', context)


@login_required
def supprimer_article(request, article_id):
    """Archiver un produit de la meme ecole."""
    article = get_object_or_404(
        _articles_fournitures_pour_utilisateur(request.user),
        pk=article_id,
    )
    if request.method == 'POST':
        article.actif = False
        article.save(update_fields=['actif'])
        messages.success(request, f'Produit "{article.nom}" archive.')
        return redirect('depenses:liste_articles')
    return render(request, 'depenses/logistique/confirmer_suppression_article.html', {
        'titre_page': 'Archiver un produit',
        'article': article,
    })


@login_required
def supprimer_article_ancien(request, article_id):
    """Désactiver (suppression logique) un article"""
    article = get_object_or_404(Article, pk=article_id)
    if request.method == 'POST':
        article.actif = False
        article.save(update_fields=['actif'])
        messages.success(request, f'Article "{article.nom}" archivé.')
        return redirect('depenses:liste_articles')

    context = {
        'titre_page': 'Archiver un article',
        'article': article,
    }
    return render(request, 'depenses/logistique/confirmer_suppression_article.html', context)


# ===== CATÉGORIES D'ARTICLES =====

@login_required
def gestion_categories_articles(request):
    """Liste et création des catégories d'articles"""
    if request.method == 'POST':
        form = CategorieArticleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie créée avec succès.')
            return redirect('depenses:gestion_categories_articles')
    else:
        form = CategorieArticleForm()

    categories = CategorieArticle.objects.annotate(
        nb_articles=Count('articles')
    ).order_by('type_categorie', 'nom')

    context = {
        'titre_page': "Catégories d'articles",
        'form': form,
        'categories': categories,
    }
    return render(request, 'depenses/logistique/categories_articles.html', context)


@login_required
def modifier_categorie_article(request, categorie_id):
    """Modifier une catégorie d'article"""
    categorie = get_object_or_404(CategorieArticle, pk=categorie_id)
    if request.method == 'POST':
        form = CategorieArticleForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie modifiée avec succès.')
            return redirect('depenses:gestion_categories_articles')
    else:
        form = CategorieArticleForm(instance=categorie)

    categories = CategorieArticle.objects.annotate(
        nb_articles=Count('articles')
    ).order_by('type_categorie', 'nom')

    context = {
        'titre_page': "Modifier la catégorie",
        'form': form,
        'categories': categories,
        'categorie_edition': categorie,
    }
    return render(request, 'depenses/logistique/categories_articles.html', context)


# ===== INVENTAIRES =====

@login_required
def creer_inventaire(request):
    """Démarrer un inventaire : crée l'inventaire et une ligne par article actif"""
    if request.method == 'POST':
        form = InventaireForm(request.POST)
        if form.is_valid():
            inventaire = form.save(commit=False)
            inventaire.cree_par = request.user
            inventaire.numero_inventaire = _generer_code(Inventaire, 'numero_inventaire', 'INV')
            inventaire.save()

            # Générer les lignes pour tous les articles actifs
            articles = Article.objects.filter(actif=True)
            for article in articles:
                LigneInventaire.objects.create(
                    inventaire=inventaire,
                    article=article,
                    stock_theorique=article.stock_actuel,
                    stock_physique=article.stock_actuel,
                    prix_unitaire=article.prix_unitaire,
                    valeur_theorique=article.stock_actuel * article.prix_unitaire,
                    valeur_physique=article.stock_actuel * article.prix_unitaire,
                )
            inventaire.nombre_articles = articles.count()
            inventaire.save(update_fields=['nombre_articles'])

            messages.success(request, f'Inventaire {inventaire.numero_inventaire} démarré.')
            return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)
    else:
        form = InventaireForm(initial={'date_inventaire': date.today()})

    context = {
        'titre_page': 'Nouvel Inventaire',
        'form': form,
    }
    return render(request, 'depenses/logistique/form_inventaire.html', context)


@login_required
def detail_inventaire(request, inventaire_id):
    """Saisie du stock physique et synthèse d'un inventaire"""
    inventaire = get_object_or_404(Inventaire, pk=inventaire_id)
    lignes = inventaire.lignes.select_related('article').order_by('article__nom')

    if request.method == 'POST' and inventaire.statut == 'EN_COURS':
        for ligne in lignes:
            valeur = request.POST.get(f'stock_physique_{ligne.pk}')
            if valeur is not None and valeur != '':
                try:
                    ligne.stock_physique = int(valeur)
                    ligne.save()  # recalcule écart et valeurs (cf. LigneInventaire.save)
                except (ValueError, TypeError):
                    continue
        # Recalcul de la synthèse
        agg = lignes.aggregate(valeur=Sum('valeur_physique'))
        inventaire.valeur_totale = agg['valeur'] or 0
        inventaire.ecarts_detectes = lignes.exclude(ecart=0).count()
        inventaire.statut = 'TERMINE'
        inventaire.save(update_fields=['valeur_totale', 'ecarts_detectes', 'statut'])
        messages.success(request, 'Saisie enregistrée. Inventaire prêt à être validé.')
        return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)

    context = {
        'titre_page': f"Inventaire {inventaire.numero_inventaire}",
        'inventaire': inventaire,
        'lignes': lignes,
    }
    return render(request, 'depenses/logistique/detail_inventaire.html', context)


@login_required
def valider_inventaire(request, inventaire_id):
    """Valider l'inventaire et ajuster le stock réel via des mouvements"""
    inventaire = get_object_or_404(Inventaire, pk=inventaire_id)

    if request.method == 'POST':
        if inventaire.statut in ('VALIDE', 'ANNULE'):
            messages.warning(request, 'Cet inventaire est déjà clôturé.')
            return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)

        lignes_ecart = inventaire.lignes.exclude(ecart=0).select_related('article')
        for ligne in lignes_ecart:
            # Ajuster le stock de l'article à la valeur physique constatée
            mvt = MouvementStock(
                article=ligne.article,
                type_mouvement='AJUSTEMENT',
                motif='INVENTAIRE',
                quantite=ligne.stock_physique,
                prix_unitaire=ligne.prix_unitaire,
                destinataire='',
                document_reference=inventaire.numero_inventaire,
                observations=f"Ajustement inventaire (écart {ligne.ecart})",
                cree_par=request.user,
            )
            mvt.numero_mouvement = _generer_code(MouvementStock, 'numero_mouvement', 'MVT')
            mvt.save()  # met à jour le stock de l'article (type AJUSTEMENT)

        inventaire.statut = 'VALIDE'
        inventaire.valide_par = request.user
        inventaire.date_validation = timezone.now()
        inventaire.save(update_fields=['statut', 'valide_par', 'date_validation'])
        messages.success(request, f'Inventaire {inventaire.numero_inventaire} validé. Stock ajusté.')

    return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)


@login_required
def export_stock_excel(request):
    """Exporter le tableau des fournitures de l'ecole au format Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fournitures scolaires'
    headers = [
        'Code', 'Produit', 'Categorie', 'Quantite en stock', 'Quantite vendue',
        'Reste', "Prix d'achat unitaire", 'Prix de vente unitaire',
        "Chiffre d'affaires", 'Solde (marge realisee)',
    ]
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    articles = _articles_fournitures_pour_utilisateur(request.user).annotate(
        quantite_vendue=Coalesce(
            Sum(
                'mouvements__quantite',
                filter=Q(mouvements__type_mouvement='SORTIE', mouvements__motif='VENTE'),
            ),
            Value(0),
            output_field=IntegerField(),
        ),
        chiffre_affaires=Coalesce(
            Sum(
                'mouvements__montant_total',
                filter=Q(mouvements__type_mouvement='SORTIE', mouvements__motif='VENTE'),
            ),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=20, decimal_places=0),
        ),
    ).order_by('nom')
    for row, article in enumerate(articles, 2):
        solde = article.chiffre_affaires - article.quantite_vendue * article.prix_unitaire
        valeurs = [
            article.code_article,
            article.nom,
            article.categorie.nom,
            article.stock_actuel,
            article.quantite_vendue,
            article.stock_actuel,
            float(article.prix_unitaire),
            float(article.prix_vente_unitaire),
            float(article.chiffre_affaires),
            float(solde),
        ]
        for col, valeur in enumerate(valeurs, 1):
            ws.cell(row=row, column=col, value=valeur)
    for col in ws.columns:
        largeur = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(largeur + 2, 40)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fournitures_{date.today()}.xlsx'
    wb.save(response)
    return response


@login_required
def export_stock_excel_ancien(request):
    """Exporter le stock en Excel"""
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Articles"
    
    # En-têtes
    headers = [
        'Code Article', 'Nom', 'Catégorie', 'Stock Actuel', 
        'Stock Minimum', 'Prix Unitaire', 'Valeur Stock', 'État', 'Emplacement'
    ]
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    articles = Article.objects.select_related('categorie').filter(actif=True)
    
    for row, article in enumerate(articles, 2):
        ws.cell(row=row, column=1, value=article.code_article)
        ws.cell(row=row, column=2, value=article.nom)
        ws.cell(row=row, column=3, value=article.categorie.nom)
        ws.cell(row=row, column=4, value=article.stock_actuel)
        ws.cell(row=row, column=5, value=article.stock_minimum)
        ws.cell(row=row, column=6, value=float(article.prix_unitaire))
        ws.cell(row=row, column=7, value=float(article.valeur_stock))
        ws.cell(row=row, column=8, value=article.get_etat_display())
        ws.cell(row=row, column=9, value=article.emplacement)
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=stock_articles_{date.today()}.xlsx'
    
    wb.save(response)
    return response
