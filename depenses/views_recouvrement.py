"""Vues des nouveaux modules du menu Recouvrement (ex-Dépenses) :
Cuisine, Document, Versement et Informatique (abonnements élèves).

Les modules Cuisine / Document / Versement partagent la même forme
(date automatique, un champ principal, montant, observation) : ils sont
donc pilotés par un petit tableau de configuration (MODULES_SIMPLES) au
lieu de dupliquer 3 fois les mêmes vues CRUD/dashboard/export.
"""
import io
from datetime import timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from eleves.models import Eleve
from ecole_moderne.security_decorators import require_school_object
from utilisateurs.permissions import can_add_expenses, can_modify_expenses, can_delete_expenses, can_delete_subscriptions
from utilisateurs.utils import user_is_superadmin, user_school, filter_by_user_school

from .models_recouvrement import (
    DepenseCuisine, DepenseDocument, Versement, AbonnementInformatique
)
from .forms import (
    DepenseCuisineForm, DepenseDocumentForm, VersementForm, AbonnementInformatiqueForm
)


# =====================================================================
# Tableau de bord général du menu Recouvrement (vue d'ensemble en cartes)
# =====================================================================

@login_required
def tableau_bord_general(request):
    """Vue d'ensemble générale du menu Recouvrement : une carte par sous-module,
    plus des totaux agrégés en haut de page."""
    from .models import Depense
    from .models_logistique import Article, BienEtablissement
    from .models_bibliotheque import Livre, Emprunt

    user = request.user
    ecole = None if user_is_superadmin(user) else user_school(user)

    def _depenses_qs():
        qs = Depense.objects.all()
        if not user_is_superadmin(user):
            qs = qs.none() if ecole is None else qs.filter(cree_par__profil__ecole=ecole)
        return qs

    depenses_qs = _depenses_qs()
    cuisine_qs = _base_qs_module('cuisine', user)
    document_qs = _base_qs_module('document', user)
    versement_qs = _base_qs_module('versement', user)

    informatique_qs = AbonnementInformatique.objects.all()
    if not user_is_superadmin(user):
        informatique_qs = filter_by_user_school(informatique_qs, user, 'eleve__classe__ecole')

    logistique_qs = BienEtablissement.objects.filter(actif=True)
    if not user_is_superadmin(user):
        logistique_qs = logistique_qs.none() if ecole is None else logistique_qs.filter(cree_par__profil__ecole=ecole)

    bibliotheque_qs = Livre.objects.all()
    if not user_is_superadmin(user):
        bibliotheque_qs = bibliotheque_qs.none() if ecole is None else bibliotheque_qs.filter(cree_par__profil__ecole=ecole)

    montant_total_general = (
        (depenses_qs.aggregate(t=Sum('montant_ttc'))['t'] or Decimal('0')) +
        (cuisine_qs.aggregate(t=Sum('montant'))['t'] or Decimal('0')) +
        (document_qs.aggregate(t=Sum('montant'))['t'] or Decimal('0')) +
        (versement_qs.aggregate(t=Sum('montant'))['t'] or Decimal('0'))
    )

    cartes = [
        {
            'titre': 'Dépenses courantes', 'icone': 'fa-receipt', 'couleur': 'primary',
            'description': 'Factures, fournisseurs, catégories et budgets.',
            'href': reverse('depenses:dashboard_courantes'),
            'total': depenses_qs.count(),
            'montant': depenses_qs.aggregate(t=Sum('montant_ttc'))['t'] or Decimal('0'),
        },
        {
            'titre': 'Logistique & Fournitures', 'icone': 'fa-boxes', 'couleur': 'info',
            'description': 'Biens, stock et vente de fournitures scolaires.',
            'href': reverse('depenses:dashboard_logistique'),
            'total': logistique_qs.count(),
            'montant': None,
        },
        {
            'titre': 'Bibliothèque', 'icone': 'fa-book-reader', 'couleur': 'success',
            'description': 'Catalogue, emprunts et réservations de livres.',
            'href': reverse('depenses:dashboard_bibliotheque'),
            'total': bibliotheque_qs.count(),
            'montant': None,
        },
        {
            'titre': 'Cuisine', 'icone': 'fa-utensils', 'couleur': 'orange',
            'description': 'Dépenses de la cuisine (denrées, matériel...).',
            'href': reverse('depenses:dashboard_module_simple', args=['cuisine']),
            'total': cuisine_qs.count(),
            'montant': cuisine_qs.aggregate(t=Sum('montant'))['t'] or Decimal('0'),
        },
        {
            'titre': 'Document', 'icone': 'fa-file-alt', 'couleur': 'purple',
            'description': 'Dépenses liées aux documents administratifs.',
            'href': reverse('depenses:dashboard_module_simple', args=['document']),
            'total': document_qs.count(),
            'montant': document_qs.aggregate(t=Sum('montant'))['t'] or Decimal('0'),
        },
        {
            'titre': 'Versement', 'icone': 'fa-hand-holding-usd', 'couleur': 'rose',
            'description': 'Versements effectués par l\'établissement.',
            'href': reverse('depenses:dashboard_module_simple', args=['versement']),
            'total': versement_qs.count(),
            'montant': versement_qs.aggregate(t=Sum('montant'))['t'] or Decimal('0'),
        },
        {
            'titre': 'Informatique', 'icone': 'fa-laptop-code', 'couleur': 'magic',
            'description': 'Abonnements informatique des élèves et alertes.',
            'href': reverse('depenses:dashboard_informatique'),
            'total': informatique_qs.count(),
            'montant': informatique_qs.filter(statut='ACTIF').aggregate(t=Sum('montant'))['t'] or Decimal('0'),
        },
        {
            'titre': 'Salaires enseignants', 'icone': 'fa-chalkboard-teacher', 'couleur': 'success',
            'description': 'Suivi des salaires payés par enseignant et par mois.',
            'href': reverse('depenses:dashboard_salaires'),
            'total': _salaires_payes_qs(user).count(),
            'montant': _salaires_payes_qs(user).aggregate(t=Sum('salaire_net'))['t'] or Decimal('0'),
        },
    ]

    context = {
        'cartes': cartes,
        'montant_total_general': montant_total_general,
    }
    return render(request, 'depenses/tableau_bord.html', context)


# =====================================================================
# Modules simples : Cuisine / Document / Versement
# =====================================================================

MODULES_SIMPLES = {
    'cuisine': {
        'model': DepenseCuisine,
        'form': DepenseCuisineForm,
        'titre': 'Dépenses Cuisine',
        'champ_principal': 'designation',
        'champ_principal_label': 'Désignation',
        'icone': 'fa-utensils',
    },
    'document': {
        'model': DepenseDocument,
        'form': DepenseDocumentForm,
        'titre': 'Dépenses Document',
        'champ_principal': 'designation',
        'champ_principal_label': 'Désignation',
        'icone': 'fa-file-alt',
    },
    'versement': {
        'model': Versement,
        'form': VersementForm,
        'titre': 'Versements',
        'champ_principal': 'lieu_versement',
        'champ_principal_label': 'Lieu de versement',
        'icone': 'fa-hand-holding-usd',
    },
}


def _base_qs_module(cle, user):
    cfg = MODULES_SIMPLES[cle]
    qs = cfg['model'].objects.all()
    if not user_is_superadmin(user):
        ecole = user_school(user)
        if ecole is None:
            return qs.none()
        qs = qs.filter(cree_par__profil__ecole=ecole)
    return qs


@login_required
def liste_module_simple(request, cle):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    qs = _base_qs_module(cle, request.user).order_by('-date', '-date_creation')

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(**{f"{cfg['champ_principal']}__icontains": q}) | Q(observation__icontains=q)
        )

    date_debut = request.GET.get('date_debut') or ''
    date_fin = request.GET.get('date_fin') or ''
    if date_debut:
        qs = qs.filter(date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date__lte=date_fin)

    total_montant = qs.aggregate(total=Sum('montant'))['total'] or Decimal('0')

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'cle': cle,
        'cfg': cfg,
        'page_obj': page_obj,
        'total_montant': total_montant,
        'q': q,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'depenses/recouvrement/liste.html', context)


@login_required
@can_add_expenses
def ajouter_module_simple(request, cle):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    if request.method == 'POST':
        form = cfg['form'](request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.cree_par = request.user
            instance.save()
            messages.success(request, f"{cfg['titre']} : enregistrement ajouté avec succès.")
            return redirect('depenses:liste_module_simple', cle=cle)
    else:
        form = cfg['form']()
    return render(request, 'depenses/recouvrement/form.html', {'cle': cle, 'cfg': cfg, 'form': form})


@login_required
@can_modify_expenses
def modifier_module_simple(request, cle, pk):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    instance = get_object_or_404(_base_qs_module(cle, request.user), pk=pk)
    if request.method == 'POST':
        form = cfg['form'](request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, f"{cfg['titre']} : enregistrement modifié avec succès.")
            return redirect('depenses:liste_module_simple', cle=cle)
    else:
        form = cfg['form'](instance=instance)
    return render(request, 'depenses/recouvrement/form.html', {
        'cle': cle, 'cfg': cfg, 'form': form, 'instance': instance,
    })


@login_required
@can_delete_expenses
def supprimer_module_simple(request, cle, pk):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    instance = get_object_or_404(_base_qs_module(cle, request.user), pk=pk)
    if request.method == 'POST':
        instance.delete()
        messages.success(request, f"{cfg['titre']} : enregistrement supprimé.")
        return redirect('depenses:liste_module_simple', cle=cle)
    return render(request, 'depenses/recouvrement/confirmer_suppression.html', {
        'cle': cle, 'cfg': cfg, 'instance': instance,
    })


@login_required
def dashboard_module_simple(request, cle):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    qs = _base_qs_module(cle, request.user)

    total_count = qs.count()
    total_montant = qs.aggregate(total=Sum('montant'))['total'] or Decimal('0')

    today = timezone.localdate()
    debut_mois = today.replace(day=1)
    montant_mois = qs.filter(date__gte=debut_mois).aggregate(total=Sum('montant'))['total'] or Decimal('0')
    count_mois = qs.filter(date__gte=debut_mois).count()

    recents = qs.order_by('-date', '-date_creation')[:10]

    context = {
        'cle': cle,
        'cfg': cfg,
        'total_count': total_count,
        'total_montant': total_montant,
        'montant_mois': montant_mois,
        'count_mois': count_mois,
        'recents': recents,
    }
    return render(request, 'depenses/recouvrement/dashboard.html', context)


@login_required
def export_module_simple_excel(request, cle):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    qs = _base_qs_module(cle, request.user).order_by('-date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg['titre'][:31]

    headers = ['Date', cfg['champ_principal_label'], 'Montant (GNF)', 'Observation']
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, item in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=item.date.strftime('%d/%m/%Y') if item.date else '')
        ws.cell(row=row, column=2, value=getattr(item, cfg['champ_principal']))
        ws.cell(row=row, column=3, value=float(item.montant))
        ws.cell(row=row, column=4, value=item.observation)

    for col in ws.columns:
        largeur = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(largeur + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename={cle}_{timezone.localdate()}.xlsx"
    wb.save(response)
    return response


@login_required
def export_module_simple_pdf(request, cle):
    if cle not in MODULES_SIMPLES:
        raise Http404()
    cfg = MODULES_SIMPLES[cle]
    qs = _base_qs_module(cle, request.user).order_by('-date')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    titre_style = ParagraphStyle('Titre', parent=styles['Heading1'], fontSize=16, alignment=1)
    story.append(Paragraph(cfg['titre'], titre_style))
    story.append(Paragraph(f"Édité le {timezone.localdate().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 16))

    data = [['Date', cfg['champ_principal_label'], 'Montant (GNF)', 'Observation']]
    total = Decimal('0')
    for item in qs:
        total += item.montant or Decimal('0')
        data.append([
            item.date.strftime('%d/%m/%Y') if item.date else '',
            getattr(item, cfg['champ_principal']),
            f"{item.montant:,.0f}".replace(',', ' '),
            (item.observation or '')[:60],
        ])
    data.append(['', 'TOTAL', f"{total:,.0f}".replace(',', ' '), ''])

    table = Table(data, colWidths=[70, 190, 90, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f1f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    story.append(table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename={cle}_{timezone.localdate()}.pdf"
    return response


# =====================================================================
# Module Informatique (abonnements élèves)
# =====================================================================

@login_required
def liste_abonnements_informatique(request):
    qs = AbonnementInformatique.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__matricule__icontains=q)
        )

    filtre = (request.GET.get('filtre') or '').strip().lower()
    today = timezone.localdate()
    if filtre == 'actif':
        qs = qs.filter(statut='ACTIF')
    elif filtre == 'expire':
        qs = qs.filter(statut='EXPIRE')
    elif filtre == 'proche_expiration':
        qs = qs.filter(statut='ACTIF', date_fin__gte=today, date_fin__lte=today + timedelta(days=7))

    qs = qs.order_by('-updated_at')
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'q': q,
        'filtre': filtre,
    }
    return render(request, 'depenses/recouvrement/informatique_liste.html', context)


@login_required
def recherche_eleve_informatique(request):
    """Recherche d'élèves par matricule/nom pour créer un abonnement (API JSON)."""
    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return HttpResponse('[]', content_type='application/json')

    qs = Eleve.objects.select_related('classe')
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')

    qs = qs.filter(
        Q(matricule__icontains=q) | Q(nom__icontains=q) | Q(prenom__icontains=q)
    )[:15]

    import json
    resultats = [
        {
            'id': e.id,
            'matricule': e.matricule or '',
            'nom_complet': f"{e.prenom} {e.nom}",
            'classe': e.classe.nom if e.classe else '',
        }
        for e in qs
    ]
    return HttpResponse(json.dumps(resultats), content_type='application/json')


@login_required
def ajouter_abonnement_informatique(request):
    if request.method == 'POST':
        form = AbonnementInformatiqueForm(request.POST)
        if form.is_valid():
            abonnement = form.save(commit=False)
            abonnement.cree_par = request.user
            abonnement.save()
            messages.success(request, f"Abonnement informatique créé pour {abonnement.eleve}")
            return redirect('depenses:liste_abonnements_informatique')
    else:
        form = AbonnementInformatiqueForm()
        eleve_id = request.GET.get('eleve')
        if eleve_id:
            try:
                form.initial['eleve'] = Eleve.objects.get(pk=eleve_id)
            except Eleve.DoesNotExist:
                pass

    if not user_is_superadmin(request.user):
        form.fields['eleve'].queryset = filter_by_user_school(Eleve.objects.all(), request.user, 'classe__ecole')

    return render(request, 'depenses/recouvrement/informatique_form.html', {'form': form})


@login_required
@require_school_object(model=AbonnementInformatique, pk_kwarg='pk', field_path='eleve__classe__ecole')
def modifier_abonnement_informatique(request, pk):
    abonnement = get_object_or_404(AbonnementInformatique, pk=pk)
    if request.method == 'POST':
        form = AbonnementInformatiqueForm(request.POST, instance=abonnement)
        if form.is_valid():
            form.save()
            messages.success(request, f"Abonnement informatique modifié pour {abonnement.eleve}")
            return redirect('depenses:liste_abonnements_informatique')
    else:
        form = AbonnementInformatiqueForm(instance=abonnement)

    if not user_is_superadmin(request.user):
        form.fields['eleve'].queryset = filter_by_user_school(Eleve.objects.all(), request.user, 'classe__ecole')

    return render(request, 'depenses/recouvrement/informatique_form.html', {'form': form, 'abonnement': abonnement})


@login_required
@can_delete_subscriptions
@require_school_object(model=AbonnementInformatique, pk_kwarg='pk', field_path='eleve__classe__ecole')
def supprimer_abonnement_informatique(request, pk):
    abonnement = get_object_or_404(AbonnementInformatique, pk=pk)
    if request.method == 'POST':
        eleve_nom = str(abonnement.eleve)
        abonnement.delete()
        messages.success(request, f"Abonnement informatique supprimé pour {eleve_nom}")
        return redirect('depenses:liste_abonnements_informatique')
    return render(request, 'depenses/recouvrement/informatique_confirmer_suppression.html', {'abonnement': abonnement})


@login_required
def dashboard_informatique(request):
    qs = AbonnementInformatique.objects.select_related('eleve', 'eleve__classe')
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    today = timezone.localdate()
    total = qs.count()
    actifs = qs.filter(statut='ACTIF').count()
    expires = qs.filter(statut='EXPIRE').count()
    suspendus = qs.filter(statut='SUSPENDU').count()

    abonnements_expires = qs.filter(statut='ACTIF', date_fin__lt=today)
    abonnements_proche_expiration = qs.filter(
        statut='ACTIF', date_fin__gte=today, date_fin__lte=today + timedelta(days=7)
    )

    montant_total = qs.filter(statut='ACTIF').aggregate(total=Sum('montant'))['total'] or Decimal('0')

    context = {
        'total': total,
        'actifs': actifs,
        'expires': expires,
        'suspendus': suspendus,
        'abonnements_expires': abonnements_expires,
        'abonnements_proche_expiration': abonnements_proche_expiration,
        'nb_expires': abonnements_expires.count(),
        'nb_proche_expiration': abonnements_proche_expiration.count(),
        'montant_total': montant_total,
    }
    return render(request, 'depenses/recouvrement/informatique_dashboard.html', context)


@login_required
@require_school_object(model=AbonnementInformatique, pk_kwarg='pk', field_path='eleve__classe__ecole')
def carte_abonnement_informatique_pdf(request, pk):
    """Génère la carte d'abonnement informatique PDF d'un élève."""
    abonnement = get_object_or_404(
        AbonnementInformatique.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole'),
        pk=pk
    )

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from django.contrib.staticfiles import finders
    import os

    buffer = io.BytesIO()
    # Format carte (85.6mm x 54mm, format carte bancaire standard) sur fond A4 centré
    carte_w, carte_h = 85.6 * mm, 54 * mm
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    x0 = (width - carte_w) / 2
    y0 = (height - carte_h) / 2

    c.setFillColorRGB(0.05, 0.15, 0.35)
    c.roundRect(x0, y0, carte_w, carte_h, 8, fill=1, stroke=0)

    ecole_obj = getattr(getattr(abonnement.eleve, 'classe', None), 'ecole', None)
    logo_path = None
    try:
        if ecole_obj and getattr(ecole_obj, 'logo', None) and hasattr(ecole_obj.logo, 'path'):
            if os.path.exists(ecole_obj.logo.path):
                logo_path = ecole_obj.logo.path
    except Exception:
        logo_path = None
    if not logo_path:
        logo_path = finders.find('logos/logo.png')
    if logo_path:
        try:
            c.drawImage(ImageReader(logo_path), x0 + 6, y0 + carte_h - 22, width=16, height=16,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    c.setFillColorRGB(1, 1, 1)
    c.setFont('Helvetica-Bold', 9)
    nom_ecole = ecole_obj.nom if ecole_obj and getattr(ecole_obj, 'nom', None) else 'MySchoolGN'
    c.drawString(x0 + 26, y0 + carte_h - 16, nom_ecole[:32])

    c.setFont('Helvetica-Bold', 8)
    c.drawString(x0 + 6, y0 + carte_h - 30, "CARTE D'ABONNEMENT INFORMATIQUE")

    c.setFont('Helvetica', 8)
    c.drawString(x0 + 6, y0 + carte_h - 42, f"Élève : {abonnement.eleve.prenom} {abonnement.eleve.nom}")
    c.drawString(x0 + 6, y0 + carte_h - 52, f"Matricule : {abonnement.eleve.matricule or '-'}")
    classe_nom = abonnement.eleve.classe.nom if abonnement.eleve.classe else '-'
    c.drawString(x0 + 6, y0 + carte_h - 62, f"Classe : {classe_nom}")
    c.drawString(x0 + 6, y0 + carte_h - 74, f"Validité : {abonnement.date_debut.strftime('%d/%m/%Y')} au {abonnement.date_fin.strftime('%d/%m/%Y')}")
    c.drawString(x0 + 6, y0 + carte_h - 84, f"Statut : {abonnement.get_statut_display()}")

    c.showPage()
    c.save()
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename=carte_informatique_{abonnement.eleve.matricule or abonnement.pk}.pdf"
    return response


@login_required
def export_informatique_excel(request):
    qs = AbonnementInformatique.objects.select_related('eleve', 'eleve__classe').order_by('-updated_at')
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Abonnements Informatique'
    headers = ['Matricule', 'Élève', 'Classe', 'Montant (GNF)', 'Début', 'Fin', 'Statut']
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, abo in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=abo.eleve.matricule or '')
        ws.cell(row=row, column=2, value=f"{abo.eleve.prenom} {abo.eleve.nom}")
        ws.cell(row=row, column=3, value=abo.eleve.classe.nom if abo.eleve.classe else '')
        ws.cell(row=row, column=4, value=float(abo.montant))
        ws.cell(row=row, column=5, value=abo.date_debut.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=6, value=abo.date_fin.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=7, value=abo.get_statut_display())

    for col in ws.columns:
        largeur = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(largeur + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename=abonnements_informatique_{timezone.localdate()}.xlsx"
    wb.save(response)
    return response


@login_required
def export_informatique_pdf(request):
    qs = AbonnementInformatique.objects.select_related('eleve', 'eleve__classe').order_by('-updated_at')
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    titre_style = ParagraphStyle('Titre', parent=styles['Heading1'], fontSize=16, alignment=1)
    story.append(Paragraph('Abonnements Informatique', titre_style))
    story.append(Paragraph(f"Édité le {timezone.localdate().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 16))

    data = [['Matricule', 'Élève', 'Classe', 'Montant', 'Début', 'Fin', 'Statut']]
    for abo in qs:
        data.append([
            abo.eleve.matricule or '',
            f"{abo.eleve.prenom} {abo.eleve.nom}"[:22],
            abo.eleve.classe.nom if abo.eleve.classe else '',
            f"{abo.montant:,.0f}".replace(',', ' '),
            abo.date_debut.strftime('%d/%m/%Y'),
            abo.date_fin.strftime('%d/%m/%Y'),
            abo.get_statut_display(),
        ])

    table = Table(data, colWidths=[55, 110, 60, 60, 55, 55, 50])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename=abonnements_informatique_{timezone.localdate()}.pdf"
    return response


# =====================================================================
# Module Salaires enseignants (suivi des montants payés par mois)
# =====================================================================

MOIS_NOMS_FR = [
    '', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
]

NIVEAUX_ENSEIGNANT_ORDRE = ['Maternelle', 'Primaire', 'Collège', 'Lycée', 'Autre']


def _niveau_depuis_classe(niveau_classe):
    """Regroupe les niveaux détaillés de Classe.NIVEAUX_CHOICES en 4 grandes familles."""
    if not niveau_classe:
        return 'Autre'
    if niveau_classe in ('GARDERIE', 'MATERNELLE'):
        return 'Maternelle'
    if niveau_classe.startswith('PRIMAIRE'):
        return 'Primaire'
    if niveau_classe.startswith('COLLEGE'):
        return 'Collège'
    if niveau_classe in ('LYCEE_11', 'LYCEE_12', 'TERMINALE'):
        return 'Lycée'
    return 'Autre'


def _niveau_enseignant(enseignant, affectations_par_enseignant):
    """Détermine le niveau (Maternelle/Primaire/Collège/Lycée) d'un enseignant
    à partir de sa classe réellement affectée (la plus récente), avec un
    repli sur son type de rémunération si aucune affectation n'existe."""
    affectations = affectations_par_enseignant.get(enseignant.id) or []
    if affectations:
        derniere = max(affectations, key=lambda a: (a.actif, a.date_debut))
        if derniere.classe_id:
            return _niveau_depuis_classe(derniere.classe.niveau)

    mapping_type = {
        'GARDERIE': 'Maternelle',
        'MATERNELLE': 'Maternelle',
        'PRIMAIRE': 'Primaire',
        'SECONDAIRE': 'Collège',
        'ADMINISTRATEUR': 'Autre',
    }
    return mapping_type.get(enseignant.type_enseignant, 'Autre')


def _salaires_payes_qs(user):
    """États de salaire marqués payés, filtrés par école pour les non-superadmins."""
    from salaires.models import EtatSalaire
    qs = EtatSalaire.objects.filter(paye=True).select_related('enseignant', 'periode', 'enseignant__ecole')
    if not user_is_superadmin(user):
        ecole = user_school(user)
        qs = qs.none() if ecole is None else qs.filter(enseignant__ecole=ecole)
    return qs


def _enseignants_scope_qs(user):
    """Liste des enseignants visibles par l'utilisateur (scoping par école).

    Les démissionnaires sont exclus: ils ne sont plus en poste et ne doivent
    pas continuer à accumuler un salaire dû mois après mois. Les enseignants
    en congé ou suspendus restent visibles (toujours en poste, la question de
    leur paie pendant cette période relève d'une décision de gestion, pas
    d'un filtrage automatique).
    """
    from salaires.models import Enseignant, StatutEnseignant
    qs = Enseignant.objects.select_related('ecole').exclude(statut=StatutEnseignant.DEMISSIONNAIRE)
    if not user_is_superadmin(user):
        ecole = user_school(user)
        qs = qs.none() if ecole is None else qs.filter(ecole=ecole)
    return qs


def _mois_disponibles(user):
    """Liste des mois (annee, mois) ayant une période de salaire créée pour
    l'école de l'utilisateur, complétée par le mois en cours s'il est absent,
    triée du plus récent au plus ancien."""
    from salaires.models import PeriodeSalaire

    qs = PeriodeSalaire.objects.all()
    if not user_is_superadmin(user):
        ecole = user_school(user)
        qs = qs.none() if ecole is None else qs.filter(ecole=ecole)

    cles = set(qs.values_list('annee', 'mois').distinct())
    today = timezone.localdate()
    cles.add((today.year, today.month))

    mois_tries = sorted(cles, reverse=True)
    return [
        {'annee': annee, 'mois': mois, 'libelle': f"{MOIS_NOMS_FR[mois]} {annee}"}
        for annee, mois in mois_tries
    ]


def _construire_lignes_salaires_mois(user, annee, mois, niveau=None):
    """Construit, pour un mois donné, une ligne par enseignant : montant dû
    (état de salaire s'il existe, sinon estimation via calculer_salaire_mensuel)
    et montant reçu (montant net si l'état est marqué payé, sinon 0).

    Un enseignant nouvellement ajouté dans le module Salaires apparaît donc
    immédiatement ici, même sans aucun état de salaire calculé pour ce mois.
    """
    from salaires.models import AffectationClasse, EtatSalaire

    enseignants = list(_enseignants_scope_qs(user).order_by('nom', 'prenoms'))
    enseignant_ids = [e.id for e in enseignants]

    affectations_par_enseignant = {}
    if enseignant_ids:
        for aff in AffectationClasse.objects.filter(enseignant_id__in=enseignant_ids).select_related('classe'):
            affectations_par_enseignant.setdefault(aff.enseignant_id, []).append(aff)

    etats_par_enseignant = {}
    if enseignant_ids:
        etats = EtatSalaire.objects.filter(
            enseignant_id__in=enseignant_ids, periode__annee=annee, periode__mois=mois
        )
        etats_par_enseignant = {e.enseignant_id: e for e in etats}

    lignes = []
    for ens in enseignants:
        niveau_ens = _niveau_enseignant(ens, affectations_par_enseignant)
        if niveau and niveau_ens != niveau:
            continue

        etat = etats_par_enseignant.get(ens.id)
        if etat is not None:
            montant = etat.salaire_net
            estime = False
        else:
            montant = ens.calculer_salaire_mensuel()
            estime = True

        montant_recu = etat.salaire_net if (etat and etat.paye) else Decimal('0')

        lignes.append({
            'enseignant_id': ens.id,
            'nom': ens.nom,
            'prenoms': ens.prenoms,
            'niveau': niveau_ens,
            'montant': montant,
            'montant_recu': montant_recu,
            'paye': bool(etat and etat.paye),
            'estime': estime,
        })

    return lignes


def _construire_pivot_salaires(user):
    """Construit le tableau croisé enseignant x mois des salaires payés.

    Retourne (periodes, lignes, totaux_par_periode, totaux_par_niveau) où :
    - periodes: liste ordonnée de dicts {annee, mois, libelle} présentes dans les données
    - lignes: liste de dicts par enseignant {nom, prenoms, niveau, par_periode: {(annee,mois): montant}, total}
    - totaux_par_periode: {(annee,mois): montant total tous enseignants}
    - totaux_par_niveau: {niveau: montant total}
    """
    from salaires.models import Enseignant, AffectationClasse

    qs = _salaires_payes_qs(user).order_by('periode__annee', 'periode__mois')

    enseignant_ids = set(qs.values_list('enseignant_id', flat=True))
    affectations_par_enseignant = {}
    if enseignant_ids:
        for aff in AffectationClasse.objects.filter(enseignant_id__in=enseignant_ids).select_related('classe'):
            affectations_par_enseignant.setdefault(aff.enseignant_id, []).append(aff)

    periodes_vues = {}
    lignes_par_enseignant = {}
    totaux_par_periode = {}
    totaux_par_niveau = {}

    for etat in qs:
        cle_periode = (etat.periode.annee, etat.periode.mois)
        if cle_periode not in periodes_vues:
            periodes_vues[cle_periode] = {
                'annee': etat.periode.annee,
                'mois': etat.periode.mois,
                'libelle': f"{MOIS_NOMS_FR[etat.periode.mois]} {etat.periode.annee}",
            }

        ens = etat.enseignant
        if ens.id not in lignes_par_enseignant:
            lignes_par_enseignant[ens.id] = {
                'nom': ens.nom,
                'prenoms': ens.prenoms,
                'niveau': _niveau_enseignant(ens, affectations_par_enseignant),
                'par_periode': {},
                'total': Decimal('0'),
            }
        ligne = lignes_par_enseignant[ens.id]
        montant = etat.salaire_net or Decimal('0')
        ligne['par_periode'][cle_periode] = ligne['par_periode'].get(cle_periode, Decimal('0')) + montant
        ligne['total'] += montant

        totaux_par_periode[cle_periode] = totaux_par_periode.get(cle_periode, Decimal('0')) + montant
        totaux_par_niveau[ligne['niveau']] = totaux_par_niveau.get(ligne['niveau'], Decimal('0')) + montant

    periodes = sorted(periodes_vues.values(), key=lambda p: (p['annee'], p['mois']))
    lignes = sorted(lignes_par_enseignant.values(), key=lambda l: (l['nom'], l['prenoms']))

    return periodes, lignes, totaux_par_periode, totaux_par_niveau


@login_required
def dashboard_salaires(request):
    """Tableau de bord Recouvrement > Salaires enseignants : liste, pour le mois
    sélectionné (le mois courant par défaut), chaque enseignant avec le montant
    dû et le montant reçu affichés en face de son nom. Le tableau s'ajuste
    automatiquement au mois choisi et peut être filtré par niveau
    (Maternelle/Primaire/Collège/Lycée). Un enseignant nouvellement créé dans
    le module Salaires apparaît immédiatement ici, même sans paiement encore
    enregistré."""
    today = timezone.localdate()

    mois_disponibles = _mois_disponibles(request.user)
    try:
        annee_sel = int(request.GET.get('annee') or today.year)
        mois_sel = int(request.GET.get('mois') or today.month)
    except (TypeError, ValueError):
        annee_sel, mois_sel = today.year, today.month
    niveau_sel = request.GET.get('niveau') or ''

    lignes = _construire_lignes_salaires_mois(request.user, annee_sel, mois_sel, niveau_sel or None)

    total_du = sum((l['montant'] or Decimal('0') for l in lignes), Decimal('0'))
    total_recu = sum((l['montant_recu'] for l in lignes), Decimal('0'))
    reste_a_payer = total_du - total_recu

    # Répartition par niveau pour le mois sélectionné (indépendante du filtre niveau,
    # pour toujours afficher la vue d'ensemble des 4 familles).
    lignes_tous_niveaux = (
        lignes if not niveau_sel else _construire_lignes_salaires_mois(request.user, annee_sel, mois_sel)
    )
    totaux_par_niveau = {}
    for l in lignes_tous_niveaux:
        totaux_par_niveau[l['niveau']] = totaux_par_niveau.get(l['niveau'], Decimal('0')) + l['montant_recu']
    totaux_par_niveau_ordonnes = [
        {'niveau': niveau, 'montant': totaux_par_niveau.get(niveau, Decimal('0'))}
        for niveau in NIVEAUX_ENSEIGNANT_ORDRE
        if totaux_par_niveau.get(niveau)
    ]

    # Tendance sur les mois payés (historique global, non filtré par mois sélectionné).
    periodes, _lignes_pivot, totaux_par_periode, _t = _construire_pivot_salaires(request.user)
    evolution_mensuelle = [
        {'libelle': p['libelle'], 'montant': totaux_par_periode.get((p['annee'], p['mois']), Decimal('0'))}
        for p in periodes
    ]
    max_mensuel = max((e['montant'] for e in evolution_mensuelle), default=Decimal('0'))
    for e in evolution_mensuelle:
        e['part'] = int((e['montant'] / max_mensuel) * 100) if max_mensuel else 0
    total_general = sum(totaux_par_periode.values(), Decimal('0'))

    context = {
        'mois_disponibles': mois_disponibles,
        'niveau_choices': NIVEAUX_ENSEIGNANT_ORDRE,
        'annee_sel': annee_sel,
        'mois_sel': mois_sel,
        'niveau_sel': niveau_sel,
        'mois_sel_libelle': f"{MOIS_NOMS_FR[mois_sel]} {annee_sel}",
        'lignes': lignes,
        'total_du': total_du,
        'total_recu': total_recu,
        'reste_a_payer': reste_a_payer,
        'total_general': total_general,
        'totaux_par_niveau': totaux_par_niveau_ordonnes,
        'evolution_mensuelle': evolution_mensuelle,
    }
    return render(request, 'depenses/recouvrement/salaires_dashboard.html', context)


@login_required
def export_salaires_excel(request):
    """Export du mois sélectionné (mêmes filtres mois/niveau que le tableau de bord) :
    une ligne par enseignant avec le montant dû et le montant reçu."""
    today = timezone.localdate()
    try:
        annee_sel = int(request.GET.get('annee') or today.year)
        mois_sel = int(request.GET.get('mois') or today.month)
    except (TypeError, ValueError):
        annee_sel, mois_sel = today.year, today.month
    niveau_sel = request.GET.get('niveau') or ''

    lignes = _construire_lignes_salaires_mois(request.user, annee_sel, mois_sel, niveau_sel or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Salaires enseignants'

    mois_libelle = f"{MOIS_NOMS_FR[mois_sel]} {annee_sel}"
    headers = ['Mois', 'Prénoms', 'Nom', 'Niveau', 'Montant (GNF)', 'Montant reçu (GNF)', 'Statut']
    header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row = 2
    total_du = Decimal('0')
    total_recu = Decimal('0')
    for ligne in lignes:
        montant = ligne['montant'] or Decimal('0')
        total_du += montant
        total_recu += ligne['montant_recu']
        ws.cell(row=row, column=1, value=mois_libelle)
        ws.cell(row=row, column=2, value=ligne['prenoms'])
        ws.cell(row=row, column=3, value=ligne['nom'])
        ws.cell(row=row, column=4, value=ligne['niveau'])
        ws.cell(row=row, column=5, value=float(montant))
        ws.cell(row=row, column=6, value=float(ligne['montant_recu']))
        ws.cell(row=row, column=7, value='Payé' if ligne['paye'] else 'Non payé')
        row += 1

    ws.cell(row=row, column=4, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(total_du)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_recu)).font = Font(bold=True)

    for col_cells in ws.columns:
        largeur = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(largeur + 2, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename=salaires_{mois_sel:02d}_{annee_sel}.xlsx"
    wb.save(response)
    return response
