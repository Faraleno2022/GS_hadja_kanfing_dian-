"""Exports Excel et PDF partagés par les modules de recouvrement."""

import io
import os

from django.contrib.staticfiles import finders
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

BLEU = colors.HexColor('#0d6efd')
GRIS = colors.HexColor('#f1f3f5')


def formater_montant(valeur):
    """1234567 -> « 1 234 567 » (séparateur insécable évité pour Excel/PDF)."""
    try:
        return f"{int(valeur):,}".replace(',', ' ')
    except (TypeError, ValueError):
        return '0'


def export_excel(nom_fichier, titre, colonnes, lignes, ligne_total=None):
    """Classeur d'une feuille: en-têtes colorés, colonnes ajustées, total en gras."""
    wb = Workbook()
    ws = wb.active
    ws.title = titre[:31] or 'Export'

    ws.append(colonnes)
    entete_fond = PatternFill('solid', fgColor='0D6EFD')
    for cellule in ws[1]:
        cellule.font = Font(bold=True, color='FFFFFF')
        cellule.fill = entete_fond
        cellule.alignment = Alignment(horizontal='center', vertical='center')

    for ligne in lignes:
        ws.append(list(ligne))

    if ligne_total:
        ws.append(list(ligne_total))
        for cellule in ws[ws.max_row]:
            cellule.font = Font(bold=True)

    for index, colonne in enumerate(colonnes, start=1):
        largeur = max(
            [len(str(colonne))]
            + [len(str(ligne[index - 1])) for ligne in lignes if len(ligne) >= index]
            or [len(str(colonne))]
        )
        ws.column_dimensions[get_column_letter(index)].width = min(max(largeur + 4, 12), 45)
    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response


def export_pdf(nom_fichier, titre, sous_titre, colonnes, lignes, ligne_total=None,
               ecole=None, paysage=False):
    """Tableau PDF prêt à imprimer, avec en-tête établissement et pied de page."""
    buffer = io.BytesIO()
    format_page = landscape(A4) if paysage else A4
    doc = SimpleDocTemplate(
        buffer, pagesize=format_page,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=18 * mm,
        title=titre,
    )

    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle(
        'TitreRecouvrement', parent=styles['Title'], fontSize=15, spaceAfter=2,
        textColor=colors.HexColor('#1a1a2e'),
    )
    style_sous_titre = ParagraphStyle(
        'SousTitreRecouvrement', parent=styles['Normal'], fontSize=9,
        alignment=1, textColor=colors.HexColor('#6c757d'), spaceAfter=8,
    )
    style_cellule = ParagraphStyle(
        'CelluleRecouvrement', parent=styles['Normal'], fontSize=8, leading=10,
    )

    elements = []
    nom_ecole = getattr(ecole, 'nom', '') or ''
    if nom_ecole:
        elements.append(Paragraph(nom_ecole.upper(), style_sous_titre))
    elements.append(Paragraph(titre, style_titre))
    if sous_titre:
        elements.append(Paragraph(sous_titre, style_sous_titre))
    elements.append(Spacer(1, 4))

    donnees = [[Paragraph(f"<b>{colonne}</b>", style_cellule) for colonne in colonnes]]
    for ligne in lignes:
        donnees.append([Paragraph(str(valeur), style_cellule) for valeur in ligne])
    if ligne_total:
        donnees.append([
            Paragraph(f"<b>{valeur}</b>", style_cellule) for valeur in ligne_total
        ])

    largeur_utile = format_page[0] - 30 * mm
    table = Table(donnees, repeatRows=1, colWidths=[largeur_utile / len(colonnes)] * len(colonnes))
    style_table = [
        ('BACKGROUND', (0, 0), (-1, 0), BLEU),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#adb5bd')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, GRIS]),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    if ligne_total:
        style_table.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e7f1ff')))
    table.setStyle(TableStyle(style_table))
    elements.append(table)

    edite_le = timezone.localtime().strftime('%d/%m/%Y à %H:%M')

    def _pied_de_page(canvas_obj, document):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 7.5)
        canvas_obj.setFillColor(colors.HexColor('#6c757d'))
        canvas_obj.drawString(15 * mm, 10 * mm, f"Édité le {edite_le}")
        canvas_obj.drawRightString(
            format_page[0] - 15 * mm, 10 * mm, f"Page {document.page}"
        )
        canvas_obj.restoreState()

    doc.build(elements, onFirstPage=_pied_de_page, onLaterPages=_pied_de_page)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nom_fichier}"'
    return response


def _chemin_logo(ecole):
    try:
        # `.path` lève une ValueError quand aucun fichier n'est associé
        chemin = getattr(getattr(ecole, 'logo', None), 'path', None)
    except ValueError:
        chemin = None
    if chemin and os.path.exists(chemin):
        return chemin
    return finders.find('logos/logo.jpeg')


def carte_abonnement_pdf(abonnement):
    """Carte d'abonnement informatique au format carte bancaire (85,6 × 54 mm)."""
    from reportlab.lib.utils import ImageReader

    eleve = abonnement.eleve
    ecole = getattr(getattr(eleve, 'classe', None), 'ecole', None)

    largeur_carte, hauteur_carte = 85.6 * mm, 54 * mm
    buffer = io.BytesIO()
    c = pdf_canvas.Canvas(buffer, pagesize=A4)
    largeur_page, hauteur_page = A4
    x0 = (largeur_page - largeur_carte) / 2
    y0 = hauteur_page - 60 * mm - hauteur_carte

    # Fond et bandeau
    c.setFillColor(colors.white)
    c.roundRect(x0, y0, largeur_carte, hauteur_carte, 4 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#0d3b66'))
    c.roundRect(x0, y0 + hauteur_carte - 14 * mm, largeur_carte, 14 * mm, 4 * mm, stroke=0, fill=1)
    c.rect(x0, y0 + hauteur_carte - 14 * mm, largeur_carte, 4 * mm, stroke=0, fill=1)

    # Logo + titre
    chemin_logo = _chemin_logo(ecole)
    if chemin_logo:
        try:
            c.drawImage(
                ImageReader(chemin_logo), x0 + 3 * mm, y0 + hauteur_carte - 12.5 * mm,
                width=11 * mm, height=11 * mm, preserveAspectRatio=True, mask='auto',
            )
        except Exception:
            pass
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 7.5)
    c.drawString(x0 + 16 * mm, y0 + hauteur_carte - 6 * mm, "CARTE D'ABONNEMENT INFORMATIQUE")
    c.setFont('Helvetica', 6)
    nom_ecole = (getattr(ecole, 'nom', '') or '')[:46]
    c.drawString(x0 + 16 * mm, y0 + hauteur_carte - 10 * mm, nom_ecole)

    # Photo de l'élève
    x_photo = x0 + largeur_carte - 24 * mm
    y_photo = y0 + 16 * mm
    try:
        photo = getattr(getattr(eleve, 'photo', None), 'path', None)
    except ValueError:
        photo = None
    photo_dessinee = False
    if photo and os.path.exists(photo):
        try:
            c.drawImage(
                ImageReader(photo), x_photo, y_photo, width=20 * mm, height=24 * mm,
                preserveAspectRatio=True, mask='auto',
            )
            photo_dessinee = True
        except Exception:
            photo_dessinee = False
    if not photo_dessinee:
        c.setStrokeColor(colors.HexColor('#adb5bd'))
        c.setFillColor(colors.HexColor('#f1f3f5'))
        c.rect(x_photo, y_photo, 20 * mm, 24 * mm, stroke=1, fill=1)
        initiales = ''.join(
            partie[0].upper()
            for partie in f"{eleve.prenom} {eleve.nom}".split()[:2]
        ) or 'E'
        c.setFillColor(colors.HexColor('#6c757d'))
        c.setFont('Helvetica-Bold', 14)
        c.drawCentredString(x_photo + 10 * mm, y_photo + 10 * mm, initiales)

    # Identité et abonnement
    y = y0 + hauteur_carte - 20 * mm

    def ligne(libelle, valeur):
        nonlocal y
        c.setFillColor(colors.HexColor('#6c757d'))
        c.setFont('Helvetica', 5.5)
        c.drawString(x0 + 4 * mm, y, libelle.upper())
        c.setFillColor(colors.HexColor('#1a1a2e'))
        c.setFont('Helvetica-Bold', 7.5)
        c.drawString(x0 + 4 * mm, y - 3.6 * mm, str(valeur)[:34])
        y -= 8 * mm

    ligne('Élève', f"{eleve.prenom} {eleve.nom}")
    ligne('Matricule', eleve.matricule or '-')
    ligne('Classe', getattr(getattr(eleve, 'classe', None), 'nom', '-') or '-')
    ligne(
        'Période',
        f"{abonnement.date_debut:%d/%m/%Y} → {abonnement.date_fin:%d/%m/%Y}",
    )

    # Bandeau bas: montant et statut
    c.setFillColor(colors.HexColor('#f1f3f5'))
    c.rect(x0, y0, largeur_carte, 9 * mm, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#1a1a2e'))
    c.setFont('Helvetica-Bold', 7)
    c.drawString(x0 + 4 * mm, y0 + 3.4 * mm, f"{formater_montant(abonnement.montant)} GNF")
    couleur_statut = {
        'ACTIF': colors.HexColor('#198754'),
        'EXPIRE': colors.HexColor('#dc3545'),
    }.get(abonnement.statut_effectif, colors.HexColor('#6c757d'))
    c.setFillColor(couleur_statut)
    c.drawRightString(
        x0 + largeur_carte - 4 * mm, y0 + 3.4 * mm, abonnement.libelle_statut.upper()
    )

    # Contour
    c.setStrokeColor(colors.HexColor('#0d3b66'))
    c.setLineWidth(0.8)
    c.roundRect(x0, y0, largeur_carte, hauteur_carte, 4 * mm, stroke=1, fill=0)

    # Mention sous la carte
    c.setFillColor(colors.HexColor('#6c757d'))
    c.setFont('Helvetica', 8)
    c.drawCentredString(
        largeur_page / 2, y0 - 10 * mm,
        "Carte à présenter à l'entrée de la salle informatique — découper suivant le contour.",
    )

    c.showPage()
    c.save()

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="carte_informatique_{abonnement.id}.pdf"'
    )
    return response
