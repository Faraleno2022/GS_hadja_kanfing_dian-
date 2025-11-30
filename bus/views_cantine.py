"""
Vues pour la gestion des abonnements à la cantine scolaire
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io

from eleves.models import Eleve
from .models import AbonnementCantine
from .forms import AbonnementCantineForm
from utilisateurs.utils import user_is_admin, user_is_superadmin, filter_by_user_school
from utilisateurs.permissions import can_delete_subscriptions
from ecole_moderne.security_decorators import require_school_object


@login_required
def tableau_bord_cantine(request):
    """Tableau de bord des abonnements cantine avec alertes"""
    qs = AbonnementCantine.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    
    # Statistiques générales
    total = qs.count()
    actifs = qs.filter(statut='ACTIF').count()
    expires = qs.filter(statut='EXPIRE').count()
    suspendus = qs.filter(statut='SUSPENDU').count()
    
    # Alertes
    today = timezone.localdate()
    
    # Abonnements expirés (non marqués comme EXPIRE)
    abonnements_expires = [a for a in qs.filter(statut='ACTIF') if a.est_expire]
    
    # Abonnements proches de l'expiration (dans les 7 jours)
    abonnements_proche_expiration = [a for a in qs.filter(statut='ACTIF') if a.est_proche_expiration and not a.est_expire]
    
    # Abonnements critiques (expire dans 3 jours ou moins)
    abonnements_critiques = [a for a in abonnements_proche_expiration if a.jours_restants <= 3]
    
    # Statistiques par type de repas
    stats_type_repas = {
        'dejeuner': qs.filter(type_repas='DEJEUNER', statut='ACTIF').count(),
        'gouter': qs.filter(type_repas='GOUTER', statut='ACTIF').count(),
        'complet': qs.filter(type_repas='COMPLET', statut='ACTIF').count(),
    }
    
    # Revenus mensuels estimés
    revenus_mensuel = qs.filter(
        statut='ACTIF',
        periodicite='MENSUEL'
    ).aggregate(total=Sum('montant'))['total'] or 0
    
    context = {
        'titre_page': 'Gestion Cantine Scolaire',
        'total': total,
        'actifs': actifs,
        'expires': expires,
        'suspendus': suspendus,
        'abonnements_expires': abonnements_expires,
        'abonnements_proche_expiration': abonnements_proche_expiration,
        'abonnements_critiques': abonnements_critiques,
        'nb_expires': len(abonnements_expires),
        'nb_proche_expiration': len(abonnements_proche_expiration),
        'nb_critiques': len(abonnements_critiques),
        'stats_type_repas': stats_type_repas,
        'revenus_mensuel': revenus_mensuel,
    }
    return render(request, 'bus/cantine/tableau_bord.html', context)


@login_required
def liste_abonnements_cantine(request):
    """Liste des abonnements cantine avec filtres et recherche"""
    qs = AbonnementCantine.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    
    # Recherche
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__matricule__icontains=q) |
            Q(contact_parent__icontains=q)
        )
    
    # Filtres
    filtre = (request.GET.get('filtre') or '').strip().lower()
    if filtre == 'actif':
        qs = qs.filter(statut='ACTIF')
    elif filtre == 'expire':
        qs = qs.filter(statut='EXPIRE')
    elif filtre == 'suspendu':
        qs = qs.filter(statut='SUSPENDU')
    elif filtre == 'proche_expiration':
        today = timezone.localdate()
        date_limite = today + timedelta(days=7)
        qs = qs.filter(statut='ACTIF', date_expiration__lte=date_limite, date_expiration__gte=today)
    elif filtre == 'critique':
        today = timezone.localdate()
        date_limite = today + timedelta(days=3)
        qs = qs.filter(statut='ACTIF', date_expiration__lte=date_limite, date_expiration__gte=today)
    
    # Filtre par type de repas
    type_repas = request.GET.get('type_repas')
    if type_repas:
        qs = qs.filter(type_repas=type_repas)
    
    # Filtre par classe
    classe_id = request.GET.get('classe')
    if classe_id:
        qs = qs.filter(eleve__classe_id=classe_id)
    
    # Tri
    qs = qs.order_by('-date_expiration', 'eleve__nom')
    
    context = {
        'titre_page': 'Liste des Abonnements Cantine',
        'abonnements': qs,
        'q': q,
        'filtre': filtre,
        'type_repas': type_repas,
    }
    return render(request, 'bus/cantine/liste.html', context)


@login_required
def creer_abonnement_cantine(request):
    """Créer un nouvel abonnement cantine"""
    if request.method == 'POST':
        form = AbonnementCantineForm(request.POST)
        if form.is_valid():
            abonnement = form.save()
            messages.success(request, f"Abonnement cantine créé pour {abonnement.eleve}")
            return redirect('bus:liste_abonnements_cantine')
    else:
        form = AbonnementCantineForm()
        
        # Pré-remplir l'élève si fourni dans l'URL
        eleve_id = request.GET.get('eleve')
        if eleve_id:
            try:
                eleve = Eleve.objects.get(pk=eleve_id)
                form.initial['eleve'] = eleve
                if eleve.responsable_principal:
                    form.initial['contact_parent'] = eleve.responsable_principal.telephone
            except Eleve.DoesNotExist:
                pass
    
    # Filtrer les élèves par école de l'utilisateur
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        form.fields['eleve'].queryset = filter_by_user_school(
            Eleve.objects.all(), 
            request.user, 
            'classe__ecole'
        )
    
    context = {
        'titre_page': 'Nouvel Abonnement Cantine',
        'form': form,
    }
    return render(request, 'bus/cantine/form.html', context)


@login_required
@require_school_object(model=AbonnementCantine, pk_kwarg='pk', field_path='eleve__classe__ecole')
def modifier_abonnement_cantine(request, pk):
    """Modifier un abonnement cantine existant"""
    abonnement = get_object_or_404(AbonnementCantine, pk=pk)
    
    if request.method == 'POST':
        form = AbonnementCantineForm(request.POST, instance=abonnement)
        if form.is_valid():
            form.save()
            messages.success(request, f"Abonnement cantine modifié pour {abonnement.eleve}")
            return redirect('bus:liste_abonnements_cantine')
    else:
        form = AbonnementCantineForm(instance=abonnement)
    
    # Filtrer les élèves par école de l'utilisateur
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        form.fields['eleve'].queryset = filter_by_user_school(
            Eleve.objects.all(), 
            request.user, 
            'classe__ecole'
        )
    
    context = {
        'titre_page': 'Modifier Abonnement Cantine',
        'form': form,
        'abonnement': abonnement,
    }
    return render(request, 'bus/cantine/form.html', context)


@login_required
@can_delete_subscriptions
@require_school_object(model=AbonnementCantine, pk_kwarg='pk', field_path='eleve__classe__ecole')
def supprimer_abonnement_cantine(request, pk):
    """Supprimer définitivement un abonnement cantine"""
    abonnement = get_object_or_404(AbonnementCantine, pk=pk)
    
    if request.method == 'POST':
        eleve_nom = str(abonnement.eleve)
        abonnement.delete()
        messages.success(request, f"Abonnement cantine supprimé définitivement pour {eleve_nom}")
        return redirect('bus:liste_abonnements_cantine')
    
    context = {
        'titre_page': 'Supprimer Abonnement Cantine',
        'abonnement': abonnement,
    }
    return render(request, 'bus/cantine/confirmer_suppression.html', context)


@login_required
def export_cantine_excel(request):
    """Exporter la liste des abonnements cantine en Excel"""
    qs = AbonnementCantine.objects.select_related('eleve', 'eleve__classe')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    
    # Appliquer les mêmes filtres que la liste
    filtre = request.GET.get('filtre', '').strip().lower()
    if filtre == 'actif':
        qs = qs.filter(statut='ACTIF')
    elif filtre == 'expire':
        qs = qs.filter(statut='EXPIRE')
    elif filtre == 'proche_expiration':
        today = timezone.localdate()
        date_limite = today + timedelta(days=7)
        qs = qs.filter(statut='ACTIF', date_expiration__lte=date_limite, date_expiration__gte=today)
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Abonnements Cantine"
    
    # En-têtes
    headers = [
        'Matricule', 'Nom', 'Prénom', 'Classe', 'Type Repas', 'Périodicité', 
        'Montant (GNF)', 'Date Début', 'Date Expiration', 'Jours Restants', 
        'Statut', 'Régime Alimentaire', 'Allergies', 'Contact Parent'
    ]
    ws.append(headers)
    
    # Données
    for abo in qs:
        ws.append([
            abo.eleve.matricule or '',
            abo.eleve.nom,
            abo.eleve.prenom,
            abo.eleve.classe.nom if abo.eleve.classe else '',
            abo.get_type_repas_display(),
            abo.get_periodicite_display(),
            float(abo.montant),
            abo.date_debut.strftime('%d/%m/%Y'),
            abo.date_expiration.strftime('%d/%m/%Y'),
            abo.jours_restants,
            abo.get_statut_display(),
            abo.regime_alimentaire or '',
            abo.allergies or '',
            abo.contact_parent or '',
        ])
    
    # Ajuster la largeur des colonnes
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Réponse HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="abonnements_cantine_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response


@login_required
def alertes_cantine_json(request):
    """API JSON pour récupérer les alertes cantine (pour dashboard)"""
    qs = AbonnementCantine.objects.select_related('eleve', 'eleve__classe')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    
    # Abonnements expirés
    expires = [a for a in qs.filter(statut='ACTIF') if a.est_expire]
    
    # Abonnements proches de l'expiration
    proche_expiration = [a for a in qs.filter(statut='ACTIF') if a.est_proche_expiration and not a.est_expire]
    
    # Abonnements critiques (3 jours ou moins)
    critiques = [a for a in proche_expiration if a.jours_restants <= 3]
    
    data = {
        'total': qs.count(),
        'actifs': qs.filter(statut='ACTIF').count(),
        'expires': {
            'count': len(expires),
            'abonnements': [
                {
                    'id': a.id,
                    'eleve': str(a.eleve),
                    'classe': a.eleve.classe.nom if a.eleve.classe else '',
                    'date_expiration': a.date_expiration.strftime('%d/%m/%Y'),
                    'jours_restants': a.jours_restants,
                }
                for a in expires[:10]  # Limiter à 10
            ]
        },
        'proche_expiration': {
            'count': len(proche_expiration),
            'abonnements': [
                {
                    'id': a.id,
                    'eleve': str(a.eleve),
                    'classe': a.eleve.classe.nom if a.eleve.classe else '',
                    'date_expiration': a.date_expiration.strftime('%d/%m/%Y'),
                    'jours_restants': a.jours_restants,
                }
                for a in proche_expiration[:10]
            ]
        },
        'critiques': {
            'count': len(critiques),
            'abonnements': [
                {
                    'id': a.id,
                    'eleve': str(a.eleve),
                    'classe': a.eleve.classe.nom if a.eleve.classe else '',
                    'date_expiration': a.date_expiration.strftime('%d/%m/%Y'),
                    'jours_restants': a.jours_restants,
                }
                for a in critiques
            ]
        }
    }
    
    return JsonResponse(data)


@login_required
def get_eleve_info_json(request, eleve_id):
    """API JSON pour récupérer les informations d'un élève"""
    import os
    
    try:
        eleve = Eleve.objects.select_related('classe', 'responsable_principal').get(pk=eleve_id)
        
        # Vérifier les permissions - seul le superuser peut voir toutes les écoles
        if not user_is_superadmin(request.user):
            eleves_qs = filter_by_user_school(Eleve.objects.all(), request.user, 'classe__ecole')
            if not eleves_qs.filter(pk=eleve_id).exists():
                return JsonResponse({'error': 'Permission refusée'}, status=403)
        
        # Gérer la photo de manière sécurisée
        photo_url = None
        if eleve.photo:
            try:
                # Vérifier que le fichier existe physiquement
                if hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
                    photo_url = eleve.photo.url
                elif hasattr(eleve.photo, 'url'):
                    # Si pas de path (stockage distant), utiliser l'URL directement
                    photo_url = eleve.photo.url
            except Exception as e:
                # En cas d'erreur, on ignore simplement la photo
                print(f"Erreur lors de la récupération de la photo pour l'élève {eleve_id}: {e}")
                photo_url = None
        
        data = {
            'success': True,
            'nom_complet': f"{eleve.nom} {eleve.prenom}",
            'nom': eleve.nom,
            'prenom': eleve.prenom,
            'matricule': eleve.matricule or '',
            'classe': eleve.classe.nom if eleve.classe else '',
            'classe_id': eleve.classe.id if eleve.classe else None,
            'telephone_parent': eleve.responsable_principal.telephone if eleve.responsable_principal else '',
            'email_parent': eleve.responsable_principal.email if eleve.responsable_principal else '',
            'photo_url': photo_url,
        }
        return JsonResponse(data)
    except Eleve.DoesNotExist:
        return JsonResponse({'error': 'Élève non trouvé'}, status=404)


@login_required
@require_school_object(model=AbonnementCantine, pk_kwarg='abo_id', field_path='eleve__classe__ecole')
def generer_recu_cantine_pdf(request, abo_id):
    """Génère un reçu PDF pour un abonnement cantine"""
    abo = get_object_or_404(
        AbonnementCantine.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole'), 
        id=abo_id
    )
    
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from django.contrib.staticfiles import finders
        import os
    except Exception:
        messages.error(request, "ReportLab requis (pip install reportlab)")
        return redirect('bus:cantine_liste')
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        ecole_obj = getattr(getattr(abo.eleve, 'classe', None), 'ecole', None)
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=ecole_obj)
    except Exception:
        pass
    
    # Logo de l'école en haut à gauche
    try:
        logo_path = None
        ecole_obj = getattr(getattr(abo.eleve, 'classe', None), 'ecole', None)
        
        if ecole_obj and hasattr(ecole_obj, 'logo'):
            school_logo_path = getattr(getattr(ecole_obj, 'logo', None), 'path', None)
            if school_logo_path and os.path.exists(school_logo_path):
                logo_path = school_logo_path
        
        if not logo_path:
            logo_path = finders.find('logos/logo.png')
        
        if logo_path:
            try:
                logo_img = ImageReader(logo_path)
                logo_w, logo_h = 60, 60
                c.drawImage(logo_img, 40, height - 100, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
    except Exception:
        pass
    
    # Titre
    c.setFont('Helvetica-Bold', 18)
    title = 'REÇU ABONNEMENT CANTINE SCOLAIRE'
    tw = c.stringWidth(title, 'Helvetica-Bold', 18)
    c.drawString((width - tw)/2, height - 50, title)
    
    # Nom de l'école sous le titre
    try:
        ecole_nom = getattr(ecole_obj, 'nom', '')
        if ecole_nom:
            c.setFont('Helvetica-Bold', 12)
            tw_ecole = c.stringWidth(ecole_nom, 'Helvetica-Bold', 12)
            c.drawString((width - tw_ecole)/2, height - 70, ecole_nom)
    except Exception:
        pass
    
    # Photo de l'élève en haut à droite
    try:
        img_drawn = False
        img_w, img_h = 100, 100
        x_img = width - 40 - img_w
        y_img = height - 40 - img_h
        
        el = abo.eleve
        photo_path = getattr(getattr(el, 'photo', None), 'path', None)
        
        if photo_path and os.path.exists(photo_path):
            try:
                img = ImageReader(photo_path)
                c.drawImage(img, x_img, y_img, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
                img_drawn = True
            except Exception:
                img_drawn = False
        
        if not img_drawn:
            # Dessiner un placeholder avec initiales
            nom_complet = f"{getattr(el, 'prenom', '')} {getattr(el, 'nom', '')}".strip()
            initiales = ''.join([p[0].upper() for p in nom_complet.split()[:2]]) or 'E'
            c.setLineWidth(1)
            try:
                c.roundRect(x_img, y_img, img_w, img_h, 8)
            except Exception:
                c.rect(x_img, y_img, img_w, img_h)
            c.setFont('Helvetica-Bold', 24)
            c.drawCentredString(x_img + img_w/2, y_img + img_h/2 - 8, initiales)
            c.setFont('Helvetica', 8)
            c.drawCentredString(x_img + img_w/2, y_img + 6, "Pas de photo")
        
        # Afficher le nom de l'élève sous l'image/placeholder
        try:
            nom_aff = f"{getattr(el, 'prenom', '')} {getattr(el, 'nom', '')}".strip()
            if nom_aff:
                c.setFont('Helvetica-Bold', 10)
                c.drawCentredString(x_img + img_w/2, y_img - 12, nom_aff)
        except Exception:
            pass
    except Exception:
        pass
    
    # Corps
    y = height - 110
    c.setFont('Helvetica', 12)
    
    def line(lbl, val):
        nonlocal y
        c.setFont('Helvetica-Bold', 12)
        c.drawString(40, y, f"{lbl} :")
        c.setFont('Helvetica', 12)
        c.drawString(200, y, str(val))
        y -= 20
    
    el = abo.eleve
    line('Élève', f"{el.prenom} {el.nom} ({el.matricule})")
    line('Classe', getattr(el.classe, 'nom', ''))
    line('École', getattr(getattr(el.classe, 'ecole', None), 'nom', ''))
    line('Type de repas', abo.get_type_repas_display())
    line('Périodicité', abo.get_periodicite_display())
    line('Montant', f"{int(abo.montant):,}".replace(',', ' ') + ' GNF')
    line('Début', abo.date_debut.strftime('%d/%m/%Y') if abo.date_debut else '')
    line('Expiration', abo.date_expiration.strftime('%d/%m/%Y') if abo.date_expiration else '')
    
    # Calcul et affichage de la durée en jours
    if abo.date_debut and abo.date_expiration:
        duree_jours = (abo.date_expiration - abo.date_debut).days
        line('Durée', f"{duree_jours} jours")
    
    # Informations spécifiques à la cantine
    if abo.regime_alimentaire:
        line('Régime alimentaire', abo.regime_alimentaire)
    if abo.allergies:
        line('Allergies', abo.allergies)
    if abo.contact_parent:
        line('Contact parent', abo.contact_parent)
    
    # Pied de page avec date d'émission
    y -= 30
    c.setFont('Helvetica', 9)
    c.setFillGray(0.5)
    c.drawString(40, y, f"Reçu émis le {timezone.localdate().strftime('%d/%m/%Y')}")
    c.setFillGray(0)
    
    # Signature
    y -= 40
    c.setFont('Helvetica-Bold', 10)
    c.drawString(width - 200, y, "Signature et cachet")
    c.line(width - 200, y - 5, width - 40, y - 5)
    
    c.showPage()
    c.save()
    
    pdf = buffer.getvalue()
    buffer.close()
    
    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename=recu_cantine_{abo.id}.pdf'
    resp.write(pdf)
    return resp
