"""Exports PDF et Excel des encaissements regroupes par mode de paiement."""

import io
from decimal import Decimal

from django.db.models import Count, DecimalField, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date

from eleves.models import Ecole
from rapports.utils import _get_logo_path
from utilisateurs.permissions import can_view_reports
from utilisateurs.utils import filter_by_user_school, user_is_superadmin, user_school

from .models import Paiement
from .rapports_professionnels import (
    BLUE,
    GREEN,
    GREY,
    _build,
    _document,
    _money,
    _pdf_primitives,
    _title_elements,
)


ZERO = Decimal('0')


def _date_param(request, name, default):
    raw_value = (request.GET.get(name) or '').strip()
    if not raw_value:
        return default
    value = parse_date(raw_value)
    if value is None:
        raise ValueError(f"La date « {name} » doit être au format AAAA-MM-JJ.")
    return value


def _display_user(user):
    full_name = (user.get_full_name() or '').strip()
    return full_name or user.get_username() or 'Système'


def collect_modes_encaissement_data(request):
    """Retourne les paiements valides, regroupes par mode, sur une periode."""
    today = timezone.localdate()
    start = _date_param(request, 'du', today.replace(day=1))
    end = _date_param(request, 'au', today)
    if start > end:
        raise ValueError("La date de début doit précéder la date de fin.")

    payments = Paiement.objects.filter(
        statut='VALIDE',
        date_paiement__gte=start,
        date_paiement__lte=end,
    )
    payments = filter_by_user_school(
        payments, request.user, 'eleve__classe__ecole'
    )

    school = user_school(request.user)
    if user_is_superadmin(request.user):
        school_ids = list(
            payments.values_list('eleve__classe__ecole_id', flat=True)
            .distinct()[:2]
        )
        school = (
            Ecole.objects.filter(pk=school_ids[0]).first()
            if len(school_ids) == 1
            else None
        )

    amount_field = DecimalField(max_digits=18, decimal_places=0)
    grouped = (
        payments
        .values('mode_paiement__nom')
        .annotate(
            count=Count('id'),
            amount=Coalesce(
                Sum('montant'),
                Value(ZERO, output_field=amount_field),
                output_field=amount_field,
            ),
        )
        .order_by('-amount', 'mode_paiement__nom')
    )

    raw_rows = list(grouped)
    total_amount = sum((item['amount'] or ZERO for item in raw_rows), ZERO)
    rows = []
    for item in raw_rows:
        amount = item['amount'] or ZERO
        rows.append({
            'mode': item['mode_paiement__nom'] or 'Non précisé',
            'count': int(item['count'] or 0),
            'amount': amount,
            'share': (amount / total_amount * Decimal('100')) if total_amount else ZERO,
        })

    generated_at = timezone.localtime()
    return {
        'school': school,
        'school_name': school.nom if school else 'ÉTABLISSEMENTS AUTORISÉS',
        'school_year': '',
        'scope_label': f"École : {school.nom}" if school else 'Tous les établissements autorisés',
        'start': start,
        'end': end,
        'cutoff': end,
        'period_label': f"Du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}",
        'reference': f"MDE-{generated_at.strftime('%Y%m%d-%H%M')}",
        'generated_at': generated_at,
        'generated_by': _display_user(request.user),
        'rows': rows,
        'mode_count': len(rows),
        'payment_count': sum(item['count'] for item in rows),
        'total_amount': total_amount,
    }


def build_modes_encaissement_pdf(data):
    """Construit le rapport PDF, avec le logo d'ecole et son filigrane."""
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer

    title = "MONTANTS PAR MODE D'ENCAISSEMENT"
    styles, _paragraph, table, kpis, on_page = _pdf_primitives(data, title)
    buffer, document = _document(data, title)
    elements = _title_elements(data, styles, title, data['period_label'])
    elements.append(kpis([
        ('Modes utilisés', str(data['mode_count']), BLUE),
        ('Paiements validés', str(data['payment_count']), GREEN),
        ('Total encaissé', f"{_money(data['total_amount'])} GNF", GREEN),
    ]))
    elements.append(Spacer(1, 0.25 * cm))

    rows = [['Mode d’encaissement', 'Opérations', 'Montant encaissé (GNF)', 'Part du total']]
    for item in data['rows']:
        rows.append([
            item['mode'],
            item['count'],
            _money(item['amount']),
            f"{item['share']:.1f} %",
        ])
    rows.append([
        'TOTAL',
        data['payment_count'],
        _money(data['total_amount']),
        '100,0 %' if data['total_amount'] else '0,0 %',
    ])
    elements.append(table(
        rows,
        widths=[10 * cm, 4 * cm, 6 * cm, 4 * cm],
        numeric_columns=(1, 2, 3),
        total_row=True,
    ))
    elements.append(Spacer(1, 0.25 * cm))
    if data['rows']:
        note = (
            "Les montants correspondent uniquement aux paiements validés dans la période. "
            "La part de chaque mode est calculée sur le total effectivement encaissé."
        )
    else:
        note = "Aucun paiement validé n'a été trouvé dans la période sélectionnée."
    elements.append(Paragraph(note, styles['Note']))
    return _build(document, buffer, elements, data, title, on_page)


def build_modes_encaissement_workbook(data):
    """Construit un classeur Excel exploitable et controle comptablement."""
    import openpyxl
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Modes d'encaissement"
    sheet.sheet_view.showGridLines = False

    dark_blue = BLUE.replace('#', '')
    light_blue = 'DCEAF3'
    border_side = Side(style='thin', color='AEBBC4')
    border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side,
    )

    sheet.merge_cells('A1:D1')
    title_cell = sheet['A1']
    title_cell.value = f"{data['school_name']} — MONTANTS PAR MODE D'ENCAISSEMENT"
    title_cell.font = Font(bold=True, size=14, color=dark_blue)
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.row_dimensions[1].height = 36

    sheet.merge_cells('A2:D2')
    sheet['A2'] = f"{data['scope_label']} | {data['period_label']}"
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A3:D3')
    sheet['A3'] = (
        f"Réf. {data['reference']} | Généré par {data['generated_by']} le "
        f"{data['generated_at'].strftime('%d/%m/%Y à %H:%M')}"
    )
    sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A3'].font = Font(size=9, color=GREY.replace('#', ''))

    logo_path = _get_logo_path(data.get('school'))
    if logo_path:
        try:
            logo = ExcelImage(logo_path)
            ratio = (logo.width / logo.height) if logo.height else 1
            logo.height = 30
            logo.width = 30 * ratio
            sheet.add_image(logo, 'A1')
        except Exception:
            pass

    headers = [
        "Mode d'encaissement",
        "Nombre d'opérations",
        'Montant encaissé (GNF)',
        'Part du total',
    ]
    header_row = 5
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, label)
        cell.fill = PatternFill('solid', fgColor=dark_blue)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    sheet.row_dimensions[header_row].height = 28

    first_data_row = header_row + 1
    for item in data['rows']:
        row = sheet.max_row + 1
        sheet.cell(row, 1, item['mode'])
        sheet.cell(row, 2, item['count'])
        sheet.cell(row, 3, int(item['amount']))
        sheet.cell(row, 4, float(item['share'] / Decimal('100')))
        for cell in sheet[row]:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
        sheet.cell(row, 2).number_format = '#,##0'
        sheet.cell(row, 3).number_format = '#,##0 "GNF"'
        sheet.cell(row, 4).number_format = '0.0%'

    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, 'TOTAL')
    sheet.cell(total_row, 2, data['payment_count'])
    sheet.cell(total_row, 3, int(data['total_amount']))
    sheet.cell(total_row, 4, 1.0 if data['total_amount'] else 0.0)
    for cell in sheet[total_row]:
        cell.fill = PatternFill('solid', fgColor=light_blue)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(vertical='center')
    sheet.cell(total_row, 2).number_format = '#,##0'
    sheet.cell(total_row, 3).number_format = '#,##0 "GNF"'
    sheet.cell(total_row, 4).number_format = '0.0%'

    sheet.freeze_panes = 'A6'
    last_data_row = max(first_data_row, total_row - 1)
    sheet.auto_filter.ref = f'A{header_row}:D{last_data_row}'
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 18
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = '1:5'
    sheet.oddFooter.center.text = f"{data['reference']} — Page &P / &N"
    return workbook


def _bad_request(exc):
    return HttpResponse(str(exc), status=400, content_type='text/plain; charset=utf-8')


def _response_filename(data, extension):
    return f"modes_encaissement_{data['start'].isoformat()}_{data['end'].isoformat()}.{extension}"


@can_view_reports
def export_modes_encaissement_pdf(request):
    try:
        data = collect_modes_encaissement_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    response = HttpResponse(
        build_modes_encaissement_pdf(data).getvalue(),
        content_type='application/pdf',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{_response_filename(data, "pdf")}"'
    )
    return response


@can_view_reports
def export_modes_encaissement_excel(request):
    try:
        data = collect_modes_encaissement_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    buffer = io.BytesIO()
    build_modes_encaissement_workbook(data).save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{_response_filename(data, "xlsx")}"'
    )
    return response
