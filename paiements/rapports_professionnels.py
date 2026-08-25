"""Rapports professionnels des encaissements et du recouvrement scolaire."""

import io
import re
import uuid
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from xml.sax.saxutils import escape

from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.dateparse import parse_date

from eleves.models import Classe
from eleves.utils_annee import get_annee_active
from rapports.utils import _get_logo_path
from utilisateurs.permissions import can_view_reports
from utilisateurs.utils import filter_by_user_school, user_school

from .allocation import (
    ALLOCATION_COMPONENTS,
    allocate_amount_sequentially,
    allocate_discounts,
    registration_kind_for_type,
)
from .models import (
    EcheancierPaiement,
    ModePaiement,
    Paiement,
    PaiementRemise,
    Relance,
)


ZERO = Decimal('0')
BLUE = '#174A6E'
BLUE_LIGHT = '#DCEAF3'
GREEN = '#207A54'
ORANGE = '#C2761C'
RED = '#B53A3A'
GREY = '#5D6973'

ACCOUNTING_COMPONENTS = (
    ('inscription', 'Inscription'),
    ('reinscription', 'Réinscription'),
    ('tranche_1', 'Tranche 1'),
    ('tranche_2', 'Tranche 2'),
    ('tranche_3', 'Tranche 3'),
    ('non_affecte', 'Non affecté / à contrôler'),
)


def _money(value):
    return f"{int(value or 0):,}".replace(',', ' ')


def _percentage(amount, base):
    amount = Decimal(str(amount or 0))
    base = Decimal(str(base or 0))
    return amount / base * Decimal('100') if base > 0 else ZERO


def _discount_precision(discount, balance):
    if Decimal(str(discount or 0)) <= 0:
        return '-'
    if Decimal(str(balance or 0)) <= 0:
        return 'Soldé avec remise appliquée au paiement'
    return 'Remise appliquée au paiement'


def _safe_filename(value):
    return re.sub(r'[^\w-]+', '_', value or '').strip('_') or 'etablissement'


def _display_user(user):
    full_name = (user.get_full_name() or '').strip() if user else ''
    if full_name:
        words = full_name.split()
        if len(words) == 2 and words[0].casefold() == words[1].casefold():
            full_name = words[0]
    return full_name or getattr(user, 'username', '') or 'Système'


def _make_report_reference(prefix, generated_at):
    return (
        f"{prefix}-{generated_at.strftime('%Y%m%d-%H%M%S')}-"
        f"{uuid.uuid4().hex[:6].upper()}"
    )


def _requires_external_reference(mode_name):
    normalized = (mode_name or '').casefold()
    return not any(token in normalized for token in ('espèce', 'espece', 'cash', 'caisse'))


def _empty_accounting_allocation():
    return {key: ZERO for key, _label in ACCOUNTING_COMPONENTS}


def _parse_filters(request):
    def read_date(name):
        raw = (request.GET.get(name) or '').strip()
        if not raw:
            return None
        parsed = parse_date(raw)
        if parsed is None:
            raise ValueError(f"La date « {name} » doit être au format AAAA-MM-JJ.")
        return parsed

    start = read_date('du')
    end = read_date('au')
    if start and end and start > end:
        raise ValueError("La date de début doit précéder la date de fin.")

    classe_id = (request.GET.get('classe_id') or '').strip()
    classes = filter_by_user_school(
        Classe.objects.select_related('ecole').order_by('ecole__nom', 'niveau', 'nom'),
        request.user,
        'ecole',
    )
    if classe_id:
        if not classe_id.isdigit():
            raise ValueError("La classe sélectionnée est invalide.")
        classes = classes.filter(pk=int(classe_id))

    classes = list(classes)
    if classe_id and not classes:
        raise ValueError("La classe sélectionnée est introuvable ou non autorisée.")
    school = classes[0].ecole if classes else user_school(request.user)
    requested_year = (request.GET.get('annee_scolaire') or '').strip()
    if requested_year and not re.fullmatch(r'\d{4}-\d{4}', requested_year):
        raise ValueError("L'année scolaire doit être au format AAAA-AAAA.")
    if requested_year:
        school_year = requested_year
    elif len(classes) == 1:
        school_year = classes[0].annee_scolaire
    elif school:
        school_year = get_annee_active(request, school) or ''
    else:
        school_year = ''

    if school_year:
        classes = [item for item in classes if item.annee_scolaire == school_year]

    mode_id = (request.GET.get('mode_id') or '').strip()
    selected_mode = None
    if mode_id:
        if not mode_id.isdigit():
            raise ValueError("Le mode d'encaissement sélectionné est invalide.")
        selected_mode = ModePaiement.objects.filter(pk=int(mode_id)).first()
        if selected_mode is None:
            raise ValueError("Le mode d'encaissement sélectionné est introuvable.")

    search = (request.GET.get('q') or '').strip()[:100]
    situation = (request.GET.get('situation') or '').strip()
    if situation not in {'', 'solde', 'reste', 'sans_echeancier'}:
        raise ValueError("La situation de paiement est invalide.")

    class_ids = [item.pk for item in classes]
    school_ids = sorted({item.ecole_id for item in classes})
    if len(classes) == 1:
        scope_label = f"Classe : {classes[0].nom}"
    elif len(school_ids) == 1 and school:
        scope_label = "Tout l'établissement"
    elif classes:
        scope_label = "Périmètre multi-établissements autorisé"
    else:
        scope_label = "Aucune classe dans le périmètre"

    today = timezone.localdate()
    if start and start > today:
        raise ValueError("La date de début ne peut pas être postérieure à la date du jour.")
    requested_end = end
    period_adjusted = bool(requested_end and requested_end > today)
    end = min(requested_end or today, today)
    cutoff = end
    generated_at = timezone.localtime()
    return {
        'classes': classes,
        'class_ids': class_ids,
        'school': school if len(school_ids) <= 1 else None,
        'school_name': (
            school.nom if school and len(school_ids) <= 1 else 'ÉTABLISSEMENTS AUTORISÉS'
        ),
        'school_year': school_year,
        'selected_mode': selected_mode,
        'search': search,
        'situation': situation,
        'scope_label': scope_label,
        'start': start,
        'end': end,
        'requested_end': requested_end,
        'period_adjusted': period_adjusted,
        'cutoff': cutoff,
        'historical_cutoff': cutoff < today,
        'generated_at': generated_at,
        'generated_by': _display_user(request.user),
    }


def _period_label(data):
    if data['start'] and data['end']:
        return f"Du {data['start'].strftime('%d/%m/%Y')} au {data['end'].strftime('%d/%m/%Y')}"
    if data['start']:
        return f"Depuis le {data['start'].strftime('%d/%m/%Y')}"
    if data['end']:
        return f"Jusqu'au {data['end'].strftime('%d/%m/%Y')}"
    return "Toutes les opérations disponibles"


def _payments_queryset(scope):
    queryset = (
        Paiement.objects
        .filter(eleve__classe_id__in=scope['class_ids'])
        .select_related(
            'eleve', 'eleve__classe', 'type_paiement', 'mode_paiement',
            'cree_par', 'valide_par',
        )
        .order_by('date_paiement', 'numero_recu')
    )
    if scope['school_year']:
        queryset = queryset.filter(annee_scolaire=scope['school_year'])
    if scope['start']:
        queryset = queryset.filter(date_paiement__gte=scope['start'])
    if scope['end']:
        queryset = queryset.filter(date_paiement__lte=scope['end'])
    if scope['selected_mode']:
        queryset = queryset.filter(mode_paiement=scope['selected_mode'])
    if scope['search']:
        queryset = queryset.filter(
            Q(eleve__nom__icontains=scope['search'])
            | Q(eleve__prenom__icontains=scope['search'])
            | Q(eleve__matricule__icontains=scope['search'])
            | Q(numero_recu__icontains=scope['search'])
            | Q(reference_externe__icontains=scope['search'])
        )
    return queryset


def _validated_payment_allocations(scope, selected_payments):
    """Reconstruit la ventilation réelle des encaissements sélectionnés.

    Les paiements antérieurs à la période sont rejoués pour positionner le
    solde initial, sans être ajoutés aux totaux du rapport. Cette approche
    évite de répartir un paiement combiné d'après son seul libellé.
    """
    selected_ids = {item.pk for item in selected_payments}
    if not selected_ids:
        return {}

    student_ids = {item.eleve_id for item in selected_payments}
    schedules = EcheancierPaiement.objects.filter(eleve_id__in=student_ids)
    if scope['school_year']:
        schedules = schedules.filter(annee_scolaire=scope['school_year'])
    schedules_by_key = {
        (item.eleve_id, item.annee_scolaire): item
        for item in schedules
    }

    history = (
        Paiement.objects
        .filter(
            eleve_id__in=student_ids,
            statut='VALIDE',
            date_paiement__lte=scope['end'],
        )
        .select_related('type_paiement')
        .order_by(
            'eleve_id', 'annee_scolaire', 'date_paiement', 'date_creation', 'pk',
        )
    )
    if scope['school_year']:
        history = history.filter(annee_scolaire=scope['school_year'])

    running_paid = {}
    results = {}
    for payment in history:
        key = (payment.eleve_id, payment.annee_scolaire)
        schedule = schedules_by_key.get(key)
        allocation = _empty_accounting_allocation()
        if schedule is None:
            allocation['non_affecte'] = payment.montant or ZERO
        else:
            initial_paid = running_paid.setdefault(
                key,
                {
                    component: ZERO
                    for component, _due_field, _paid_field in ALLOCATION_COMPONENTS
                },
            )
            raw_allocation, paid_after, unapplied = allocate_amount_sequentially(
                schedule, payment.montant, initial_paid=initial_paid,
            )
            running_paid[key] = paid_after
            registration_kind = registration_kind_for_type(payment.type_paiement)
            if registration_kind not in {'inscription', 'reinscription'}:
                registration_kind = (
                    'reinscription' if schedule.est_reinscription else 'inscription'
                )
            allocation[registration_kind] = raw_allocation['inscription']
            allocation['tranche_1'] = raw_allocation['tranche_1']
            allocation['tranche_2'] = raw_allocation['tranche_2']
            allocation['tranche_3'] = raw_allocation['tranche_3']
            allocation['non_affecte'] = unapplied
        if payment.pk in selected_ids:
            results[payment.pk] = allocation
    return results


def collect_accounting_data(request):
    data = _parse_filters(request)
    payments = list(_payments_queryset(data))
    validated = [item for item in payments if item.statut == 'VALIDE']
    payment_allocations = _validated_payment_allocations(data, validated)
    discounts = list(
        PaiementRemise.objects
        .filter(paiement__in=validated)
        .select_related('paiement', 'remise')
    )
    discount_by_payment = defaultdict(lambda: ZERO)
    discount_by_reason = defaultdict(lambda: ZERO)
    for discount in discounts:
        amount = discount.montant_remise or ZERO
        discount_by_payment[discount.paiement_id] += amount
        discount_by_reason[discount.remise.get_motif_display()] += amount

    status_labels = dict(Paiement.STATUT_CHOICES)
    by_status = {}
    for code, label in Paiement.STATUT_CHOICES:
        rows = [item for item in payments if item.statut == code]
        by_status[code] = {
            'label': label,
            'count': len(rows),
            'amount': sum((item.montant or ZERO) for item in rows),
        }

    by_mode = defaultdict(lambda: {
        'count': 0,
        'amount': ZERO,
        'reference_required': 0,
        'reference_present': 0,
        'reference_missing': 0,
        'reference_missing_amount': ZERO,
    })
    by_type = defaultdict(lambda: {'count': 0, 'amount': ZERO})
    by_class = defaultdict(lambda: {'count': 0, 'amount': ZERO, 'discount': ZERO})
    by_component = {
        key: {'label': label, 'count': 0, 'amount': ZERO}
        for key, label in ACCOUNTING_COMPONENTS
    }
    payment_rows = []
    for payment in validated:
        amount = payment.montant or ZERO
        discount = discount_by_payment[payment.pk]
        mode = payment.mode_paiement.nom if payment.mode_paiement_id else 'Non précisé'
        payment_type = payment.type_paiement.nom if payment.type_paiement_id else 'Non précisé'
        class_name = payment.eleve.classe.nom
        by_mode[mode]['count'] += 1
        by_mode[mode]['amount'] += amount
        requires_reference = _requires_external_reference(mode)
        has_reference = bool((payment.reference_externe or '').strip())
        if requires_reference:
            by_mode[mode]['reference_required'] += 1
            if has_reference:
                by_mode[mode]['reference_present'] += 1
            else:
                by_mode[mode]['reference_missing'] += 1
                by_mode[mode]['reference_missing_amount'] += amount
        by_type[payment_type]['count'] += 1
        by_type[payment_type]['amount'] += amount
        by_class[class_name]['count'] += 1
        by_class[class_name]['amount'] += amount
        by_class[class_name]['discount'] += discount
        allocation = payment_allocations.get(payment.pk, _empty_accounting_allocation())
        for component, component_amount in allocation.items():
            if component_amount > 0:
                by_component[component]['count'] += 1
                by_component[component]['amount'] += component_amount
        cashier = _display_user(payment.cree_par)
        validator = _display_user(payment.valide_par)
        payment_rows.append({
            'student_id': payment.eleve_id,
            'school_year': payment.annee_scolaire,
            'date': payment.date_paiement,
            'receipt': payment.numero_recu,
            'student': payment.eleve.nom_complet,
            'matricule': payment.eleve.matricule,
            'class': class_name,
            'type': payment_type,
            'mode': mode,
            'amount': amount,
            'discount': discount,
            'reference': payment.reference_externe or '-',
            'reference_status': (
                'Non requise' if not requires_reference
                else ('Complète' if has_reference else 'À compléter')
            ),
            'cashier': cashier,
            'validator': validator,
            'approval': cashier if cashier == validator else f"{cashier} / {validator}",
            'allocation': allocation,
        })

    total_validated = sum((item.montant or ZERO) for item in validated)
    total_discounts = sum(discount_by_payment.values(), ZERO)
    data.update({
        'period_label': _period_label(data),
        'report_reference': _make_report_reference('RC', data['generated_at']),
        'payments': payments,
        'payment_rows': payment_rows,
        'by_status': by_status,
        'by_mode': dict(sorted(by_mode.items())),
        'by_type': dict(sorted(by_type.items())),
        'by_class': dict(sorted(by_class.items())),
        'by_component': by_component,
        'discount_by_reason': dict(sorted(discount_by_reason.items())),
        'payment_count': len(payments),
        'validated_count': len(validated),
        'total_validated': total_validated,
        'total_discounts': total_discounts,
        'total_coverage': total_validated + total_discounts,
        'reference_missing_count': sum(
            item['reference_missing'] for item in by_mode.values()
        ),
        'reference_missing_amount': sum(
            (item['reference_missing_amount'] for item in by_mode.values()), ZERO
        ),
        'unallocated_total': by_component['non_affecte']['amount'],
    })
    return data


def collect_payment_modes_data(request):
    """Prépare le rapport dédié aux encaissements validés par mode."""
    data = collect_accounting_data(request)
    student_keys = {
        (item['student_id'], item['school_year'])
        for item in data['payment_rows']
    }
    student_ids = {key[0] for key in student_keys}

    schedules = EcheancierPaiement.objects.filter(eleve_id__in=student_ids)
    if data['school_year']:
        schedules = schedules.filter(annee_scolaire=data['school_year'])
    schedules_by_key = {
        (item.eleve_id, item.annee_scolaire): item
        for item in schedules
    }

    payment_history = Paiement.objects.filter(
        eleve_id__in=student_ids,
        statut='VALIDE',
        date_paiement__lte=data['cutoff'],
    )
    if data['school_year']:
        payment_history = payment_history.filter(annee_scolaire=data['school_year'])
    paid_by_key = {
        (item['eleve_id'], item['annee_scolaire']): item['total'] or ZERO
        for item in payment_history.values('eleve_id', 'annee_scolaire').annotate(
            total=Sum('montant')
        )
    }

    discount_history = PaiementRemise.objects.filter(
        paiement__eleve_id__in=student_ids,
        paiement__statut='VALIDE',
        paiement__date_paiement__lte=data['cutoff'],
    )
    if data['school_year']:
        discount_history = discount_history.filter(
            paiement__annee_scolaire=data['school_year']
        )
    discount_by_key = {
        (
            item['paiement__eleve_id'],
            item['paiement__annee_scolaire'],
        ): item['total'] or ZERO
        for item in discount_history.values(
            'paiement__eleve_id', 'paiement__annee_scolaire'
        ).annotate(total=Sum('montant_remise'))
    }

    balances = {}
    for key in student_keys:
        schedule = schedules_by_key.get(key)
        amount_due = schedule.total_du if schedule else ZERO
        amount_paid = paid_by_key.get(key, ZERO)
        discount = discount_by_key.get(key, ZERO)
        balance = max(ZERO, amount_due - amount_paid - discount)
        balances[key] = {
            'amount_due': amount_due,
            'amount_paid': amount_paid,
            'discount': discount,
            'balance': balance,
            'status': (
                'Échéancier absent' if schedule is None
                else ('Soldé' if balance <= ZERO else 'Reste à payer')
            ),
        }

    if data['situation']:
        expected_status = {
            'solde': 'Soldé',
            'reste': 'Reste à payer',
            'sans_echeancier': 'Échéancier absent',
        }[data['situation']]
        allowed_keys = {
            key for key, values in balances.items()
            if values['status'] == expected_status
        }
        data['payment_rows'] = [
            item for item in data['payment_rows']
            if (item['student_id'], item['school_year']) in allowed_keys
        ]
        balances = {
            key: values for key, values in balances.items()
            if key in allowed_keys
        }
        student_keys = allowed_keys

    by_mode = defaultdict(lambda: {
        'count': 0,
        'amount': ZERO,
        'reference_required': 0,
        'reference_present': 0,
        'reference_missing': 0,
        'reference_missing_amount': ZERO,
    })
    daily = defaultdict(lambda: {'count': 0, 'amount': ZERO})
    for item in data['payment_rows']:
        mode_values = by_mode[item['mode']]
        mode_values['count'] += 1
        mode_values['amount'] += item['amount']
        if item['reference_status'] != 'Non requise':
            mode_values['reference_required'] += 1
            if item['reference_status'] == 'Complète':
                mode_values['reference_present'] += 1
            else:
                mode_values['reference_missing'] += 1
                mode_values['reference_missing_amount'] += item['amount']
        daily_key = (item['date'], item['mode'])
        daily[daily_key]['count'] += 1
        daily[daily_key]['amount'] += item['amount']

    grouped_students = {}
    for item in data['payment_rows']:
        key = (item['mode'], item['student_id'], item['school_year'])
        row = grouped_students.setdefault(key, {
            'mode': item['mode'],
            'student_id': item['student_id'],
            'school_year': item['school_year'],
            'student': item['student'],
            'matricule': item['matricule'],
            'class': item['class'],
            'operation_count': 0,
            'collected': ZERO,
            'last_date': item['date'],
            'last_receipt': item['receipt'],
        })
        row['operation_count'] += 1
        row['collected'] += item['amount']
        if item['date'] >= row['last_date']:
            row['last_date'] = item['date']
            row['last_receipt'] = item['receipt']
        row.update(balances[(item['student_id'], item['school_year'])])

    student_mode_rows = sorted(
        grouped_students.values(),
        key=lambda item: (
            item['mode'].casefold(), item['class'].casefold(),
            item['student'].casefold(),
        ),
    )
    unique_balances = list(balances.values())
    total_validated = sum(
        (item['amount'] for item in data['payment_rows']), ZERO
    )
    total_discounts = sum(
        (item['discount'] for item in data['payment_rows']), ZERO
    )
    by_mode = dict(sorted(by_mode.items()))

    data.update({
        'report_reference': _make_report_reference('EM', data['generated_at']),
        'by_mode': by_mode,
        'mode_count': len(by_mode),
        'validated_count': len(data['payment_rows']),
        'payment_count': len(data['payment_rows']),
        'total_validated': total_validated,
        'total_discounts': total_discounts,
        'total_coverage': total_validated + total_discounts,
        'reference_missing_count': sum(
            item['reference_missing'] for item in by_mode.values()
        ),
        'reference_missing_amount': sum(
            (item['reference_missing_amount'] for item in by_mode.values()), ZERO
        ),
        'student_count': len(student_keys),
        'student_mode_rows': student_mode_rows,
        'student_total_due': sum(
            (item['amount_due'] for item in unique_balances), ZERO
        ),
        'student_total_paid': sum(
            (item['amount_paid'] for item in unique_balances), ZERO
        ),
        'student_total_discount': sum(
            (item['discount'] for item in unique_balances), ZERO
        ),
        'student_total_balance': sum(
            (item['balance'] for item in unique_balances), ZERO
        ),
        'settled_student_count': sum(
            1 for item in unique_balances if item['status'] == 'Soldé'
        ),
        'daily_modes': [
            {
                'date': key[0],
                'mode': key[1],
                'count': values['count'],
                'amount': values['amount'],
            }
            for key, values in sorted(
                daily.items(), key=lambda item: (item[0][0], item[0][1])
            )
        ],
    })
    return data


def _schedule_components(schedule):
    return (
        ('Admission', 'frais_inscription_du', 'frais_inscription_paye', 'date_echeance_inscription'),
        ('Tranche 1', 'tranche_1_due', 'tranche_1_payee', 'date_echeance_tranche_1'),
        ('Tranche 2', 'tranche_2_due', 'tranche_2_payee', 'date_echeance_tranche_2'),
        ('Tranche 3', 'tranche_3_due', 'tranche_3_payee', 'date_echeance_tranche_3'),
    )


def collect_recovery_data(request):
    data = _parse_filters(request)
    schedules_qs = (
        EcheancierPaiement.objects
        .filter(eleve__classe_id__in=data['class_ids'], eleve__statut='ACTIF')
        .select_related(
            'eleve', 'eleve__classe', 'eleve__responsable_principal',
            'eleve__classe__ecole',
        )
        .order_by('eleve__classe__nom', 'eleve__nom', 'eleve__prenom')
    )
    if data['school_year']:
        schedules_qs = schedules_qs.filter(annee_scolaire=data['school_year'])
    schedules = list(schedules_qs)
    schedule_students = [item.eleve_id for item in schedules]

    cash_rows = (
        Paiement.objects
        .filter(
            eleve_id__in=schedule_students,
            statut='VALIDE',
            date_paiement__lte=data['cutoff'],
        )
        .values('eleve_id', 'annee_scolaire')
        .annotate(total=Sum('montant'))
    )
    cash_by_schedule = {
        (item['eleve_id'], item['annee_scolaire']): item['total'] or ZERO
        for item in cash_rows
    }

    validated_discounts = list(
        PaiementRemise.objects
        .filter(
            paiement__eleve_id__in=schedule_students,
            paiement__statut='VALIDE',
            paiement__date_paiement__lte=data['cutoff'],
        )
        .select_related('paiement', 'remise')
        .order_by('paiement__date_paiement', 'paiement_id', 'id')
    )
    discounts_by_schedule = defaultdict(list)
    for item in validated_discounts:
        discounts_by_schedule[(item.paiement.eleve_id, item.paiement.annee_scolaire)].append(item)

    relances_qs = (
        Relance.objects
        .filter(eleve__classe_id__in=data['class_ids'])
        .select_related('eleve', 'eleve__classe', 'cree_par')
        .order_by('eleve_id', '-date_creation')
    )
    all_relances = [
        item for item in relances_qs
        if timezone.localdate(item.date_creation) <= data['cutoff']
    ]
    relances_by_student = defaultdict(list)
    for reminder in all_relances:
        relances_by_student[reminder.eleve_id].append(reminder)

    period_relances = all_relances
    if data['start']:
        period_relances = [item for item in period_relances if item.date_creation.date() >= data['start']]
    if data['end']:
        period_relances = [item for item in period_relances if item.date_creation.date() <= data['end']]

    class_summary = defaultdict(lambda: {
        'students': 0, 'due': ZERO, 'cash': ZERO, 'discount': ZERO,
        'discount_applied': ZERO,
        'balance': ZERO, 'overdue': ZERO, 'upcoming': ZERO, 'reminders': 0,
    })
    aging = {
        '1-30 jours': {'count': 0, 'amount': ZERO},
        '31-60 jours': {'count': 0, 'amount': ZERO},
        '61-90 jours': {'count': 0, 'amount': ZERO},
        'Plus de 90 jours': {'count': 0, 'amount': ZERO},
    }
    priority_rows = []
    student_rows = []
    settled_count = partial_count = unpaid_count = overdue_count = 0
    total_due = total_cash = total_discount = total_discount_applied = ZERO
    total_balance = total_overdue = total_upcoming = ZERO

    for schedule in schedules:
        discounts = discounts_by_schedule[(schedule.eleve_id, schedule.annee_scolaire)]
        recorded_cash = schedule.total_paye or ZERO
        payment_cash = cash_by_schedule.get(
            (schedule.eleve_id, schedule.annee_scolaire), ZERO,
        )
        cash_source = (
            max(recorded_cash, payment_cash)
            if data['cutoff'] >= timezone.localdate()
            else payment_cash
        )
        _cash_allocation, cash_paid, _unallocated = allocate_amount_sequentially(
            schedule,
            cash_source,
            initial_paid={
                key: ZERO for key, _due_field, _paid_field in ALLOCATION_COMPONENTS
            },
        )
        balances_after_cash = {
            key: max(
                ZERO,
                Decimal(str(getattr(schedule, due_field, 0) or 0))
                - cash_paid[key],
            )
            for key, due_field, _paid_field in ALLOCATION_COMPONENTS
        }
        discount_allocation, net_balances = allocate_discounts(
            schedule, discounts, balances=balances_after_cash,
        )
        discount_applied = sum(discount_allocation.values(), ZERO)
        discount_total = sum(
            (max(ZERO, Decimal(str(item.montant_remise or 0))) for item in discounts),
            ZERO,
        )
        due = schedule.total_du or ZERO
        cash = sum(cash_paid.values(), ZERO)
        balance = sum(net_balances.values(), ZERO)
        discount_rate = _percentage(discount_total, due)
        discount_details = []
        for discount in discounts:
            rule = discount.remise
            if rule.type_remise == 'POURCENTAGE':
                value = f"{rule.valeur:.2f}".rstrip('0').rstrip('.')
                rule_value = f"{value} %"
            else:
                rule_value = f"{_money(discount.montant_remise)} GNF"
            discount_details.append(f"{rule.nom} ({rule_value})")
        discount_detail = ', '.join(discount_details) if discount_details else '-'
        discount_precision = _discount_precision(discount_total, balance)
        overdue = upcoming = ZERO
        oldest_overdue = None
        for index, (_label, _due_field, _paid_field, date_field) in enumerate(_schedule_components(schedule)):
            key = ALLOCATION_COMPONENTS[index][0]
            due_date = getattr(schedule, date_field)
            remaining = net_balances[key]
            if remaining <= 0:
                continue
            if due_date < data['cutoff']:
                overdue += remaining
                if oldest_overdue is None or due_date < oldest_overdue:
                    oldest_overdue = due_date
                component_days = (data['cutoff'] - due_date).days
                if component_days <= 30:
                    component_bucket = '1-30 jours'
                elif component_days <= 60:
                    component_bucket = '31-60 jours'
                elif component_days <= 90:
                    component_bucket = '61-90 jours'
                else:
                    component_bucket = 'Plus de 90 jours'
                aging[component_bucket]['count'] += 1
                aging[component_bucket]['amount'] += remaining
            elif data['cutoff'] <= due_date <= data['cutoff'] + timedelta(days=30):
                upcoming += remaining

        reminders = relances_by_student[schedule.eleve_id]
        latest = reminders[0] if reminders else None
        class_name = schedule.eleve.classe.nom
        summary = class_summary[class_name]
        summary['students'] += 1
        summary['due'] += due
        summary['cash'] += cash
        summary['discount'] += discount_total
        summary['discount_applied'] += discount_applied
        summary['balance'] += balance
        summary['overdue'] += overdue
        summary['upcoming'] += upcoming
        summary['reminders'] += len(reminders)

        total_due += due
        total_cash += cash
        total_discount += discount_total
        total_discount_applied += discount_applied
        total_balance += balance
        total_overdue += overdue
        total_upcoming += upcoming
        if balance <= 0:
            settled_count += 1
            recovery_status = (
                'Soldé (remise appliquée)' if discount_total > 0 else 'Soldé'
            )
        elif cash + discount_applied > 0:
            partial_count += 1
            recovery_status = 'En retard' if overdue > 0 else 'Paiement partiel'
        else:
            unpaid_count += 1
            recovery_status = 'En retard' if overdue > 0 else 'À payer'

        responsible = schedule.eleve.responsable_principal
        student_rows.append({
            'matricule': schedule.eleve.matricule,
            'student': schedule.eleve.nom_complet,
            'class': class_name,
            'responsible': responsible.nom_complet if responsible else '-',
            'phone': responsible.telephone if responsible else '-',
            'due': due,
            'cash': cash,
            'discount': discount_total,
            'discount_applied': discount_applied,
            'discount_rate': discount_rate,
            'discount_detail': discount_detail,
            'discount_precision': discount_precision,
            'balance': balance,
            'overdue': overdue,
            'upcoming': upcoming,
            'status': recovery_status,
            'reminder_count': len(reminders),
            'last_reminder': latest.date_creation if latest else None,
            'last_status': latest.get_statut_display() if latest else 'Jamais relancé',
        })

        if overdue > 0:
            overdue_count += 1
            days = (data['cutoff'] - oldest_overdue).days if oldest_overdue else 0
            priority_rows.append({
                'matricule': schedule.eleve.matricule,
                'student': schedule.eleve.nom_complet,
                'class': class_name,
                'responsible': responsible.nom_complet if responsible else '-',
                'phone': responsible.telephone if responsible else '-',
                'due': due,
                'cash': cash,
                'discount': discount_total,
                'discount_applied': discount_applied,
                'discount_rate': discount_rate,
                'discount_precision': discount_precision,
                'coverage': cash + discount_applied,
                'balance': balance,
                'overdue': overdue,
                'days': days,
                'reminder_count': len(reminders),
                'last_reminder': latest.date_creation if latest else None,
                'last_status': latest.get_statut_display() if latest else 'Jamais relancé',
            })

    priority_rows.sort(key=lambda item: (-item['overdue'], -item['days'], item['student']))
    reminder_by_channel = defaultdict(lambda: {'count': 0, 'sent': 0, 'failed': 0})
    reminder_by_status = defaultdict(int)
    for reminder in period_relances:
        label = reminder.get_canal_display()
        reminder_by_channel[label]['count'] += 1
        if reminder.statut == 'ENVOYEE':
            reminder_by_channel[label]['sent'] += 1
        if reminder.statut == 'ECHEC':
            reminder_by_channel[label]['failed'] += 1
        reminder_by_status[reminder.get_statut_display()] += 1

    validated_period = [
        item for item in _payments_queryset(data)
        if item.statut == 'VALIDE' and item.date_paiement <= data['cutoff']
    ]
    period_cash = sum((item.montant or ZERO) for item in validated_period)
    recovery_rate = _percentage(total_cash + total_discount_applied, total_due)
    discount_rate = _percentage(total_discount, total_due)
    discount_rows = [item for item in student_rows if item['discount'] > 0]
    data.update({
        'period_label': _period_label(data),
        'report_reference': _make_report_reference('RR', data['generated_at']),
        'schedules': schedules,
        'schedule_count': len(schedules),
        'class_summary': dict(sorted(class_summary.items())),
        'aging': aging,
        'priority_rows': priority_rows,
        'student_rows': student_rows,
        'discount_rows': discount_rows,
        'total_due': total_due,
        'total_cash': total_cash,
        'total_discount': total_discount,
        'total_discount_applied': total_discount_applied,
        'discount_rate': discount_rate,
        'total_coverage': total_cash + total_discount_applied,
        'total_balance': total_balance,
        'total_overdue': total_overdue,
        'total_upcoming': total_upcoming,
        'recovery_rate': recovery_rate,
        'settled_count': settled_count,
        'partial_count': partial_count,
        'unpaid_count': unpaid_count,
        'overdue_count': overdue_count,
        'period_payment_count': len(validated_period),
        'period_cash': period_cash,
        'period_relances': period_relances,
        'reminder_by_channel': dict(sorted(reminder_by_channel.items())),
        'reminder_by_status': dict(sorted(reminder_by_status.items())),
    })
    return data


def _draw_logo_watermark(canvas, logo_path, page_width, page_height):
    """Dessine le logo de l'école en filigrane discret au centre de la page."""
    if not logo_path:
        return

    canvas.saveState()
    try:
        try:
            canvas.setFillAlpha(0.06)
        except Exception:
            pass
        watermark_width = page_width * 0.46
        watermark_height = page_height * 0.46
        canvas.drawImage(
            logo_path,
            (page_width - watermark_width) / 2,
            (page_height - watermark_height) / 2,
            width=watermark_width,
            height=watermark_height,
            preserveAspectRatio=True,
            mask='auto',
            anchor='c',
        )
    except Exception:
        # Un fichier de logo illisible ne doit jamais bloquer l'export.
        pass
    finally:
        canvas.restoreState()


def _pdf_primitives(data, title):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import Paragraph, Table, TableStyle

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='ReportTitle', parent=styles['Title'], fontName='Helvetica-Bold',
        fontSize=18, leading=21, textColor=colors.HexColor(BLUE), alignment=TA_LEFT,
        spaceAfter=5,
    ))
    styles.add(ParagraphStyle(
        name='ReportSubTitle', parent=styles['Normal'], fontSize=9, leading=12,
        textColor=colors.HexColor(GREY), spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name='SectionTitle', parent=styles['Heading2'], fontName='Helvetica-Bold',
        fontSize=12, leading=15, textColor=colors.HexColor(BLUE),
        spaceBefore=9, spaceAfter=5,
    ))
    styles.add(ParagraphStyle(
        name='SmallCell', parent=styles['Normal'], fontSize=6.8, leading=8.2,
        textColor=colors.HexColor('#202B33'),
    ))
    styles.add(ParagraphStyle(
        name='SmallCellCenter', parent=styles['SmallCell'], alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name='HeaderCell', parent=styles['SmallCellCenter'],
        textColor=colors.white, fontName='Helvetica-Bold',
    ))
    styles.add(ParagraphStyle(
        name='SmallCellRight', parent=styles['SmallCell'], alignment=TA_RIGHT,
    ))
    styles.add(ParagraphStyle(
        name='Note', parent=styles['Normal'], fontSize=7.2, leading=9,
        textColor=colors.HexColor(GREY),
    ))

    def paragraph(value, style='SmallCell'):
        safe_value = escape(str(value if value is not None else '')).replace('\n', '<br/>')
        return Paragraph(safe_value, styles[style])

    def table(
        rows, widths=None, numeric_from=None, numeric_columns=None,
        total_row=False, font_size=6.8,
    ):
        rendered = []
        for row_index, row in enumerate(rows):
            rendered.append([
                paragraph(value, 'HeaderCell' if row_index == 0 else 'SmallCell')
                for value in row
            ])
        commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(BLUE)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#AEBBC4')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F7FA')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        if numeric_from is not None:
            commands.append(('ALIGN', (numeric_from, 1), (-1, -1), 'RIGHT'))
        for column in numeric_columns or ():
            commands.append(('ALIGN', (column, 1), (column, -1), 'RIGHT'))
        if total_row and len(rows) > 1:
            commands.extend([
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor(BLUE_LIGHT)),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ])
        return Table(rendered, colWidths=widths, repeatRows=1, style=TableStyle(commands))

    def kpis(items):
        cells = []
        for label, value, color in items:
            cells.append([
                [paragraph(label, 'SmallCellCenter')],
                [Paragraph(f"<b>{escape(str(value))}</b>", ParagraphStyle(
                    f'Kpi{len(cells)}', parent=styles['Normal'], fontSize=13,
                    leading=15, alignment=TA_CENTER, textColor=colors.HexColor(color),
                ))],
            ])
        outer = Table(
            [[Table(cell, colWidths=[4.15 * cm], style=TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F7FAFC')),
                ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#C7D4DD')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ])) for cell in cells]],
            colWidths=[4.35 * cm] * len(cells),
        )
        return outer

    page_width, page_height = landscape(A4)
    logo_path = _get_logo_path(data['school']) if data.get('school') else ''

    def draw_page_chrome(canvas, page_number, page_count):
        _draw_logo_watermark(canvas, logo_path, page_width, page_height)
        canvas.saveState()
        canvas.setTitle(title)
        canvas.setAuthor(data['generated_by'])
        canvas.setSubject(
            f"{data['scope_label']} - {data.get('school_year') or 'année non précisée'} - "
            f"{data['report_reference']}"
        )
        if logo_path:
            try:
                canvas.drawImage(
                    logo_path, 0.8 * cm, page_height - 1.25 * cm,
                    width=1.4 * cm, height=0.75 * cm,
                    preserveAspectRatio=True, mask='auto',
                )
            except Exception:
                pass
        canvas.setFillColor(colors.HexColor(BLUE))
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(2.4 * cm, page_height - 0.9 * cm, data['school_name'])
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor(GREY))
        canvas.drawRightString(page_width - 0.8 * cm, page_height - 0.9 * cm, title)
        canvas.setStrokeColor(colors.HexColor('#9DB4C5'))
        canvas.line(0.8 * cm, page_height - 1.35 * cm, page_width - 0.8 * cm, page_height - 1.35 * cm)
        canvas.line(0.8 * cm, 0.8 * cm, page_width - 0.8 * cm, 0.8 * cm)
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor(GREY))
        canvas.drawString(
            0.8 * cm, 0.48 * cm,
            f"Confidentiel - Réf. {data['report_reference']} - Généré par {data['generated_by']} "
            f"le {data['generated_at'].strftime('%d/%m/%Y à %H:%M')}",
        )
        canvas.drawRightString(
            page_width - 0.8 * cm,
            0.48 * cm,
            f"Page {page_number}/{page_count}",
        )
        canvas.restoreState()

    class NumberedCanvas(pdf_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_page_states)
            for page_number, state in enumerate(self._saved_page_states, 1):
                self.__dict__.update(state)
                draw_page_chrome(self, page_number, page_count)
                pdf_canvas.Canvas.showPage(self)
            pdf_canvas.Canvas.save(self)

    def on_page(canvas, doc):
        # L'en-tête et le pied de page sont ajoutés après le contenu par le
        # canvas numéroté. Ils ne peuvent ainsi pas être masqués par la suite
        # d'un tableau réparti sur plusieurs pages.
        return None

    return styles, paragraph, table, kpis, on_page, NumberedCanvas


def _title_elements(data, styles, title, subtitle):
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.units import cm

    details = [data['scope_label']]
    if data.get('school_year'):
        details.append(f"Année scolaire : {data['school_year']}")
    details.append(subtitle)
    details.append(f"Réf. {data['report_reference']}")
    return [
        Spacer(1, 0.1 * cm),
        Paragraph(escape(title), styles['ReportTitle']),
        Paragraph(escape(' | '.join(details)), styles['ReportSubTitle']),
    ]


def build_accounting_pdf(data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table

    title = 'RAPPORT COMPTABLE DES ENCAISSEMENTS'
    styles, p, table, kpis, on_page, numbered_canvas = _pdf_primitives(data, title)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.65 * cm, bottomMargin=1.05 * cm,
        leftMargin=0.8 * cm, rightMargin=0.8 * cm,
        title=title, author=data['generated_by'],
    )
    elements = _title_elements(data, styles, title, data['period_label'])
    if data['period_adjusted']:
        elements.append(Paragraph(
            escape(
                "La date de fin demandée "
                f"({data['requested_end'].strftime('%d/%m/%Y')}) était future : "
                f"le rapport est arrêté au {data['end'].strftime('%d/%m/%Y')}."
            ),
            styles['Note'],
        ))
    elements.append(kpis([
        ('Encaissements validés', f"{_money(data['total_validated'])} GNF", GREEN),
        ('Paiements validés', str(data['validated_count']), BLUE),
        ('Remises accordées', f"{_money(data['total_discounts'])} GNF", ORANGE),
        ('Couverture totale', f"{_money(data['total_coverage'])} GNF", BLUE),
        ('En attente', str(data['by_status']['EN_ATTENTE']['count']), ORANGE),
        ('Remboursés', str(data['by_status']['REMBOURSE']['count']), RED),
    ]))

    elements.append(Paragraph('1. SYNTHÈSE PAR STATUT', styles['SectionTitle']))
    status_rows = [['Statut', 'Nombre', 'Montant (GNF)', 'Part des opérations']]
    for code, _label in Paiement.STATUT_CHOICES:
        item = data['by_status'][code]
        share = (item['count'] / data['payment_count'] * 100) if data['payment_count'] else 0
        status_rows.append([item['label'], item['count'], _money(item['amount']), f"{share:.1f} %"])
    status_rows.append(['TOTAL', data['payment_count'], _money(sum((p.montant or ZERO) for p in data['payments'])), '100 %'])
    elements.append(table(status_rows, widths=[7 * cm, 4 * cm, 6 * cm, 5 * cm], numeric_from=1, total_row=True))

    elements.append(Paragraph('2. VENTILATION ET CONTRÔLE DES JUSTIFICATIFS', styles['SectionTitle']))
    mode_rows = [[
        'Mode', 'Opérations', 'Montant (GNF)', '% validé',
        'Réf. manquantes', 'Montant à justifier',
    ]]
    for label, item in data['by_mode'].items():
        pct = item['amount'] / data['total_validated'] * 100 if data['total_validated'] else ZERO
        mode_rows.append([
            label, item['count'], _money(item['amount']), f"{pct:.1f} %",
            item['reference_missing'], _money(item['reference_missing_amount']),
        ])
    if len(mode_rows) == 1:
        mode_rows.append(['Aucun encaissement validé', 0, '0', '0 %', 0, '0'])
    elements.append(table(
        mode_rows,
        widths=[5.2*cm, 2.6*cm, 4.2*cm, 2.4*cm, 3.6*cm, 4.4*cm],
        numeric_from=1,
    ))
    elements.append(Paragraph(
        "Les écarts de caisse, banque ou Mobile Money nécessitent les bordereaux "
        "externes. Le tableau ci-dessus contrôle déjà la présence des références.",
        styles['Note'],
    ))

    type_rows = [['Type de paiement', 'Opérations', 'Montant (GNF)']]
    for label, item in data['by_type'].items():
        type_rows.append([label, item['count'], _money(item['amount'])])
    if len(type_rows) == 1:
        type_rows.append(['Aucun encaissement validé', 0, '0'])
    component_rows = [['Affectation réelle', 'Opérations', 'Montant (GNF)']]
    for item in data['by_component'].values():
        if item['amount'] > 0 or item['label'] != 'Non affecté / à contrôler':
            component_rows.append([
                item['label'], item['count'], _money(item['amount']),
            ])
    component_rows.append([
        'TOTAL', '-', _money(sum((item['amount'] for item in data['by_component'].values()), ZERO)),
    ])
    elements.append(Table([[table(type_rows, widths=[6.2*cm, 2.5*cm, 4*cm], numeric_from=1),
                            table(component_rows, widths=[6.2*cm, 2.5*cm, 4*cm], numeric_from=1, total_row=True)]],
                          colWidths=[14*cm, 13*cm]))
    if data['unallocated_total'] > 0:
        elements.append(Paragraph(
            f"Alerte : {_money(data['unallocated_total'])} GNF ne peuvent pas être "
            "rattachés avec certitude à un échéancier et doivent être contrôlés.",
            styles['Note'],
        ))

    class_rows = [['Classe', 'Paiements', 'Encaissé (GNF)', 'Remises (GNF)', 'Couverture (GNF)']]
    for label, item in data['by_class'].items():
        class_rows.append([label, item['count'], _money(item['amount']), _money(item['discount']), _money(item['amount'] + item['discount'])])
    if len(class_rows) == 1:
        class_rows.append(['Aucune classe avec encaissement', 0, '0', '0', '0'])
    class_rows.append(['TOTAL', data['validated_count'], _money(data['total_validated']), _money(data['total_discounts']), _money(data['total_coverage'])])
    elements.append(Paragraph('3. SYNTHÈSE PAR CLASSE', styles['SectionTitle']))
    elements.append(table(class_rows, widths=[8*cm, 3*cm, 5*cm, 5*cm, 5*cm], numeric_from=1, total_row=True))

    discount_rows = [['Motif de remise', 'Montant (GNF)']]
    for label, amount in data['discount_by_reason'].items():
        discount_rows.append([label, _money(amount)])
    if len(discount_rows) == 1:
        discount_rows.append(['Aucune remise sur les paiements validés', '0'])
    discount_rows.append(['TOTAL', _money(data['total_discounts'])])
    elements.append(Paragraph('4. REMISES ET RÉDUCTIONS', styles['SectionTitle']))
    elements.append(table(discount_rows, widths=[13*cm, 6*cm], numeric_from=1, total_row=True))

    if data['payment_rows']:
        elements.append(Paragraph('5. JOURNAL DÉTAILLÉ DES ENCAISSEMENTS VALIDÉS', styles['SectionTitle']))
        elements.append(Paragraph(
            escape(f"{data['scope_label']} | {data['period_label']}"),
            styles['ReportSubTitle'],
        ))
        detail_rows = [['Date', 'Reçu', 'Matricule / Élève', 'Classe', 'Type', 'Mode', 'Montant', 'Remise', 'Référence', 'Validation']]
        for item in data['payment_rows']:
            detail_rows.append([
                item['date'].strftime('%d/%m/%Y'), item['receipt'],
                f"{item['matricule']}\n{item['student']}", item['class'], item['type'], item['mode'],
                _money(item['amount']), _money(item['discount']), item['reference'],
                item['approval'],
            ])
        elements.append(table(
            detail_rows,
            widths=[1.8*cm, 2.2*cm, 4.2*cm, 2.8*cm, 3.5*cm, 2.5*cm, 2.6*cm, 2.2*cm, 2.8*cm, 3.2*cm],
            numeric_columns=(6, 7),
        ))
    else:
        elements.append(Paragraph(
            'Aucune opération validée pour les filtres sélectionnés. Le rapport est volontairement limité à une page de synthèse.',
            styles['Note'],
        ))

    elements.extend([
        Spacer(1, 0.35 * cm),
        Paragraph(
            "Contrôle recommandé : rapprocher les montants par mode avec les bordereaux de caisse, relevés Mobile Money, chèques et relevés bancaires.",
            styles['Note'],
        ),
        Spacer(1, 0.35 * cm),
        table([['Établi par', 'Contrôlé par', 'Validé par'], [data['generated_by'], 'Nom / Signature', 'Direction / Signature']],
              widths=[8.8*cm, 8.8*cm, 8.8*cm]),
    ])
    doc.build(
        elements,
        onFirstPage=on_page,
        onLaterPages=on_page,
        canvasmaker=numbered_canvas,
    )
    buffer.seek(0)
    return buffer


def build_payment_modes_pdf(data):
    """Construit un PDF autonome de rapprochement par mode d'encaissement."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    title = "RAPPORT DES ENCAISSEMENTS PAR MODE"
    styles, _p, table, kpis, on_page, numbered_canvas = _pdf_primitives(
        data, title
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=1.65 * cm,
        bottomMargin=1.05 * cm,
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        title=title,
        author=data['generated_by'],
    )
    elements = _title_elements(data, styles, title, data['period_label'])
    if data['period_adjusted']:
        elements.append(Paragraph(
            escape(
                "La date de fin demandée "
                f"({data['requested_end'].strftime('%d/%m/%Y')}) était future : "
                f"le rapport est arrêté au {data['end'].strftime('%d/%m/%Y')}."
            ),
            styles['Note'],
        ))

    elements.append(kpis([
        ('Total encaissé', f"{_money(data['total_validated'])} GNF", GREEN),
        ('Opérations validées', str(data['validated_count']), BLUE),
        ('Modes utilisés', str(data['mode_count']), BLUE),
        ('Montant à justifier', f"{_money(data['reference_missing_amount'])} GNF", ORANGE),
    ]))

    elements.append(Paragraph(
        "1. SYNTHÈSE PAR MODE D'ENCAISSEMENT", styles['SectionTitle']
    ))
    mode_rows = [[
        'Mode', 'Opérations', 'Montant encaissé (GNF)', 'Part du total',
        'Montant moyen (GNF)', 'Réf. manquantes', 'Montant à justifier (GNF)',
    ]]
    for label, item in data['by_mode'].items():
        share = _percentage(item['amount'], data['total_validated'])
        average = item['amount'] / item['count'] if item['count'] else ZERO
        mode_rows.append([
            label, item['count'], _money(item['amount']), f"{share:.1f} %",
            _money(average), item['reference_missing'],
            _money(item['reference_missing_amount']),
        ])
    if len(mode_rows) == 1:
        mode_rows.append([
            'Aucun encaissement validé', 0, '0', '0 %', '0', 0, '0'
        ])
    mode_rows.append([
        'TOTAL', data['validated_count'], _money(data['total_validated']),
        '100 %' if data['total_validated'] else '0 %',
        _money(
            data['total_validated'] / data['validated_count']
            if data['validated_count'] else ZERO
        ),
        data['reference_missing_count'],
        _money(data['reference_missing_amount']),
    ])
    elements.append(table(
        mode_rows,
        widths=[4.8*cm, 2.4*cm, 4.2*cm, 2.6*cm, 3.8*cm, 3.2*cm, 4.2*cm],
        numeric_from=1,
        total_row=True,
    ))
    elements.append(Paragraph(
        "Seuls les paiements au statut « Validé » sont comptabilisés. Les références "
        "externes sont contrôlées pour les modes hors espèces.",
        styles['Note'],
    ))

    elements.append(Paragraph(
        "2. RAPPROCHEMENT JOURNALIER PAR MODE", styles['SectionTitle']
    ))
    daily_rows = [['Date', 'Mode', 'Opérations', 'Montant encaissé (GNF)']]
    for item in data['daily_modes']:
        daily_rows.append([
            item['date'].strftime('%d/%m/%Y'), item['mode'], item['count'],
            _money(item['amount']),
        ])
    if len(daily_rows) == 1:
        daily_rows.append(['-', 'Aucun encaissement validé', 0, '0'])
    daily_rows.append([
        'TOTAL', '', data['validated_count'], _money(data['total_validated'])
    ])
    elements.append(table(
        daily_rows,
        widths=[4*cm, 8*cm, 4*cm, 6*cm],
        numeric_from=2,
        total_row=True,
    ))

    elements.append(Paragraph(
        "3. JOURNAL DÉTAILLÉ DES ENCAISSEMENTS", styles['SectionTitle']
    ))
    detail_rows = [[
        'Mode', 'Date', 'Reçu', 'Matricule / Élève', 'Classe', 'Type',
        'Montant (GNF)', 'Référence', 'Contrôle',
    ]]
    for item in sorted(
        data['payment_rows'], key=lambda row: (row['mode'], row['date'], row['receipt'])
    ):
        detail_rows.append([
            item['mode'], item['date'].strftime('%d/%m/%Y'), item['receipt'],
            f"{item['matricule']}\n{item['student']}", item['class'], item['type'],
            _money(item['amount']), item['reference'], item['reference_status'],
        ])
    if len(detail_rows) == 1:
        detail_rows.append([
            '-', '-', '-', 'Aucun encaissement validé', '-', '-', '0', '-', '-'
        ])
    elements.append(table(
        detail_rows,
        widths=[3*cm, 2.2*cm, 2.5*cm, 4.8*cm, 3.2*cm, 3.6*cm, 3*cm, 3.2*cm, 3.2*cm],
        numeric_columns=(6,),
    ))
    elements.extend([
        Spacer(1, 0.3 * cm),
        Paragraph(
            "Rapprochement recommandé : comparer chaque total avec le journal de "
            "caisse, les relevés Mobile Money et les relevés bancaires correspondants.",
            styles['Note'],
        ),
    ])
    doc.build(
        elements,
        onFirstPage=on_page,
        onLaterPages=on_page,
        canvasmaker=numbered_canvas,
    )
    buffer.seek(0)
    return buffer


def build_recovery_pdf(data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table

    title = 'RAPPORT PROFESSIONNEL DE RECOUVREMENT'
    styles, p, table, kpis, on_page, numbered_canvas = _pdf_primitives(data, title)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.65 * cm, bottomMargin=1.05 * cm,
        leftMargin=0.8 * cm, rightMargin=0.8 * cm,
        title=title, author=data['generated_by'],
    )
    subtitle = f"Situation arrêtée au {data['cutoff'].strftime('%d/%m/%Y')} | Activité : {data['period_label']}"
    elements = _title_elements(data, styles, title, subtitle)
    if data['period_adjusted']:
        elements.append(Paragraph(
            escape(
                "La date de fin demandée "
                f"({data['requested_end'].strftime('%d/%m/%Y')}) était future : "
                f"le rapport est arrêté au {data['end'].strftime('%d/%m/%Y')}."
            ),
            styles['Note'],
        ))
    elements.append(kpis([
        ('Créances totales', f"{_money(data['total_due'])} GNF", BLUE),
        ('Couverture', f"{_money(data['total_coverage'])} GNF", GREEN),
        ('Solde à recouvrer', f"{_money(data['total_balance'])} GNF", ORANGE),
        ('Retard exigible', f"{_money(data['total_overdue'])} GNF", RED),
        ('Taux de recouvrement', f"{data['recovery_rate']:.1f} %", GREEN),
        ('Élèves en retard', str(data['overdue_count']), RED),
    ]))

    elements.append(Paragraph('1. PORTEFEUILLE DE RECOUVREMENT', styles['SectionTitle']))
    portfolio_rows = [
        ['Indicateur', 'Élèves', 'Montant (GNF)', 'Observation'],
        ['Créances brutes', data['schedule_count'], _money(data['total_due']), 'Échéanciers actifs suivis'],
        ['Encaissements cumulés', '-', _money(data['total_cash']), 'Paiements affectés'],
        ['Remises enregistrées', '-', _money(data['total_discount']), f"{data['discount_rate']:.1f} % des créances brutes"],
        ['Solde à recouvrer', data['schedule_count'] - data['settled_count'], _money(data['total_balance']), 'Après paiements et remises'],
        ['Élèves soldés', data['settled_count'], '-', 'Aucun solde restant'],
        ['Paiement partiel', data['partial_count'], '-', 'Couverture incomplète'],
        ['Sans paiement', data['unpaid_count'], '-', 'Aucune couverture enregistrée'],
        ['En retard', data['overdue_count'], _money(data['total_overdue']), f"À la date du {data['cutoff'].strftime('%d/%m/%Y')}"] ,
        ['Échéances sous 30 jours', '-', _money(data['total_upcoming']), 'Prévention à organiser'],
        ['Encaissements sur la période', data['period_payment_count'], _money(data['period_cash']), data['period_label']],
    ]
    elements.append(table(portfolio_rows, widths=[6.5*cm, 3*cm, 5.5*cm, 11*cm], numeric_from=1))

    class_rows = [['Classe', 'Élèves', 'Dû', 'Encaissé', 'Remises', 'Rem. %', 'Solde', 'Retard', 'À 30 jours', 'Taux']]
    for label, item in data['class_summary'].items():
        rate = _percentage(item['cash'] + item['discount_applied'], item['due'])
        discount_rate = _percentage(item['discount'], item['due'])
        class_rows.append([
            label, item['students'], _money(item['due']), _money(item['cash']), _money(item['discount']),
            f"{discount_rate:.1f} %", _money(item['balance']), _money(item['overdue']),
            _money(item['upcoming']), f"{rate:.1f} %",
        ])
    if len(class_rows) == 1:
        class_rows.append(['Aucun échéancier actif', 0, '0', '0', '0', '0 %', '0', '0', '0', '0 %'])
    class_rows.append(['TOTAL', data['schedule_count'], _money(data['total_due']), _money(data['total_cash']),
                       _money(data['total_discount']), f"{data['discount_rate']:.1f} %",
                       _money(data['total_balance']), _money(data['total_overdue']),
                       _money(data['total_upcoming']), f"{data['recovery_rate']:.1f} %"])
    elements.append(Paragraph('2. PERFORMANCE PAR CLASSE', styles['SectionTitle']))
    elements.append(table(
        class_rows,
        widths=[4.1*cm, 1.5*cm, 2.7*cm, 2.7*cm, 2.4*cm, 2.1*cm, 2.7*cm, 2.7*cm, 2.7*cm, 2.4*cm],
        numeric_from=1,
        total_row=True,
    ))

    aging_rows = [['Ancienneté du retard', 'Échéances', 'Montant en retard (GNF)', 'Part du retard']]
    for label, item in data['aging'].items():
        share = item['amount'] / data['total_overdue'] * 100 if data['total_overdue'] else ZERO
        aging_rows.append([label, item['count'], _money(item['amount']), f"{share:.1f} %"])
    aging_rows.append([
        'TOTAL', sum(item['count'] for item in data['aging'].values()),
        _money(data['total_overdue']), '100 %' if data['total_overdue'] else '0 %',
    ])
    elements.append(Paragraph('3. BALANCE ÂGÉE DES IMPAYÉS', styles['SectionTitle']))
    elements.append(table(aging_rows, widths=[8*cm, 3*cm, 6*cm, 4*cm], numeric_from=1, total_row=True))

    elements.append(PageBreak())
    elements.append(Paragraph('4. REMISES APPLIQUÉES PAR ÉLÈVE', styles['SectionTitle']))
    elements.append(Paragraph(
        "Le taux de remise correspond au montant enregistré rapporté au total dû. "
        "La précision signale les dossiers soldés avec l'aide d'une remise.",
        styles['ReportSubTitle'],
    ))
    if data['discount_rows']:
        discount_rows = [[
            'Matricule / Élève', 'Classe', 'Dû', 'Payé', 'Remise', 'Remise %',
            'Solde', 'Précision',
        ]]
        for item in data['discount_rows']:
            discount_rows.append([
                f"{item['matricule']}\n{item['student']}", item['class'],
                _money(item['due']), _money(item['cash']), _money(item['discount']),
                f"{item['discount_rate']:.1f} %", _money(item['balance']),
                item['discount_precision'],
            ])
        elements.append(table(
            discount_rows,
            widths=[4.5*cm, 2.5*cm, 3*cm, 3*cm, 3*cm, 2.1*cm, 3*cm, 5.4*cm],
            numeric_columns=(2, 3, 4, 5, 6),
        ))
    else:
        elements.append(Paragraph('Aucune remise validée à la date d’arrêt.', styles['Note']))

    if data['priority_rows']:
        elements.append(Paragraph('5. DOSSIERS PRIORITAIRES DE RECOUVREMENT', styles['SectionTitle']))
        elements.append(Paragraph(
            f"Situation arrêtée au {data['cutoff'].strftime('%d/%m/%Y')} | Classement par montant en retard décroissant",
            styles['ReportSubTitle'],
        ))
        priority = [['Matricule / Élève', 'Classe', 'Responsable / Téléphone', 'Solde', 'Retard', 'Jours', 'Relances', 'Dernière action']]
        for item in data['priority_rows']:
            last = item['last_reminder'].strftime('%d/%m/%Y') if item['last_reminder'] else '-'
            priority.append([
                f"{item['matricule']}\n{item['student']}", item['class'],
                f"{item['responsible']}\n{item['phone']}", _money(item['balance']), _money(item['overdue']),
                item['days'], item['reminder_count'], f"{last}\n{item['last_status']}",
            ])
        elements.append(table(
            priority,
            widths=[4.3*cm, 3*cm, 5*cm, 3*cm, 3*cm, 1.7*cm, 1.7*cm, 4*cm],
            numeric_columns=(3, 4, 5, 6),
        ))
    else:
        elements.append(Paragraph('Aucun dossier en retard à la date d’arrêt.', styles['Note']))

    elements.append(Paragraph('6. PILOTAGE DES RELANCES', styles['SectionTitle']))
    channel_rows = [['Canal', 'Actions', 'Envoyées', 'Échecs', 'Taux de succès']]
    for label, item in data['reminder_by_channel'].items():
        success = item['sent'] / item['count'] * 100 if item['count'] else 0
        channel_rows.append([label, item['count'], item['sent'], item['failed'], f"{success:.1f} %"])
    if len(channel_rows) == 1:
        channel_rows.append(['Aucune relance sur la période', 0, 0, 0, '0 %'])
    status_rows = [['Statut de relance', 'Nombre']]
    for label, count in data['reminder_by_status'].items():
        status_rows.append([label, count])
    if len(status_rows) == 1:
        status_rows.append(['Aucune relance', 0])
    elements.append(Table([[table(channel_rows, widths=[4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm], numeric_from=1),
                            table(status_rows, widths=[7*cm, 3*cm], numeric_from=1)]],
                          colWidths=[15*cm, 11*cm]))
    elements.extend([
        Spacer(1, 0.35 * cm),
        Paragraph(
            (
                "Méthode : la situation historique est reconstruite à partir des paiements "
                "et remises validés jusqu'à la date d'arrêt."
                if data['historical_cutoff']
                else "Méthode : la situation courante rapproche les paiements validés et les cumuls de l'échéancier."
            ),
            styles['Note'],
        ),
        Spacer(1, 0.15 * cm),
        Paragraph(
            "Plan d'action recommandé : traiter d'abord les dossiers les plus anciens et les montants les plus élevés, puis prévenir les échéances attendues dans les 30 prochains jours.",
            styles['Note'],
        ),
        Spacer(1, 0.35 * cm),
        table([['Responsable recouvrement', 'Comptabilité', 'Direction'],
               [data['generated_by'], 'Nom / Signature', 'Visa / Signature']],
              widths=[8.8*cm, 8.8*cm, 8.8*cm]),
    ])
    doc.build(
        elements,
        onFirstPage=on_page,
        onLaterPages=on_page,
        canvasmaker=numbered_canvas,
    )
    buffer.seek(0)
    return buffer


def _excel_workbook(data, report_kind):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    dark_fill = PatternFill('solid', fgColor=BLUE.replace('#', ''))
    light_fill = PatternFill('solid', fgColor=BLUE_LIGHT.replace('#', ''))
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)
    border_side = Side(style='thin', color='B7C4CC')
    border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side,
    )

    def sheet(name, title, headers):
        ws = wb.create_sheet(name[:31])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = ws.cell(1, 1, f"{data['school_name']} - {title}")
        cell.font = Font(bold=True, size=14, color=BLUE.replace('#', ''))
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(
            2, 1,
            f"{data['scope_label']} | Année {data.get('school_year') or '-'} | {data.get('period_label', '')}",
        ).alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        ws.cell(
            3, 1,
            f"Réf. {data['report_reference']} | Généré par {data['generated_by']} "
            f"le {data['generated_at'].strftime('%d/%m/%Y à %H:%M')}",
        ).alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 22
        for col, label in enumerate(headers, 1):
            item = ws.cell(5, col, label)
            item.fill = dark_fill
            item.font = white_font
            item.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            item.border = border
        ws.freeze_panes = 'A6'
        ws.auto_filter.ref = f"A5:{openpyxl.utils.get_column_letter(len(headers))}5"
        return ws

    def append(ws, values, total=False):
        ws.append(values)
        row = ws.max_row
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if total:
                cell.fill = light_fill
                cell.font = bold_font
            if isinstance(cell.value, int) and cell.column > 1:
                cell.number_format = '#,##0'
            elif isinstance(cell.value, float):
                cell.number_format = '0.0'

    if report_kind == 'modes':
        from openpyxl.chart import BarChart, Reference

        ws = sheet(
            'Synthèse par mode',
            "ENCAISSEMENTS PAR MODE",
            [
                'Mode', 'Opérations', 'Montant encaissé (GNF)', 'Part du total',
                'Montant moyen (GNF)', 'Réf. manquantes',
                'Montant à justifier (GNF)',
            ],
        )
        first_mode_row = 6
        for label, item in data['by_mode'].items():
            append(ws, [
                label, item['count'], int(item['amount']), None, None,
                item['reference_missing'], int(item['reference_missing_amount']),
            ])
        if not data['by_mode']:
            append(ws, ['Aucun encaissement validé', 0, 0, 0, 0, 0, 0])
        last_mode_row = ws.max_row
        total_row = last_mode_row + 1
        append(ws, [
            'TOTAL', f'=SUM(B{first_mode_row}:B{last_mode_row})',
            f'=SUM(C{first_mode_row}:C{last_mode_row})',
            f'=IF(C{total_row}>0,1,0)',
            f'=IFERROR(C{total_row}/B{total_row},0)',
            f'=SUM(F{first_mode_row}:F{last_mode_row})',
            f'=SUM(G{first_mode_row}:G{last_mode_row})',
        ], total=True)
        for row in range(first_mode_row, last_mode_row + 1):
            ws.cell(row, 4, f'=IFERROR(C{row}/$C${total_row},0)')
            ws.cell(row, 5, f'=IFERROR(C{row}/B{row},0)')
        for row in range(first_mode_row, total_row + 1):
            ws.cell(row, 3).number_format = '#,##0 "GNF"'
            ws.cell(row, 4).number_format = '0.0%'
            ws.cell(row, 5).number_format = '#,##0 "GNF"'
            ws.cell(row, 7).number_format = '#,##0 "GNF"'

        if data['by_mode']:
            chart = BarChart()
            chart.type = 'bar'
            chart.style = 10
            chart.title = "Montants encaissés par mode"
            chart.y_axis.title = "Mode"
            chart.x_axis.title = "Montant (GNF)"
            chart.height = 7
            chart.width = 13
            chart.add_data(
                Reference(ws, min_col=3, min_row=5, max_row=last_mode_row),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(ws, min_col=1, min_row=first_mode_row, max_row=last_mode_row)
            )
            chart.legend = None
            ws.add_chart(chart, 'I5')

        ws = sheet(
            'Rapprochement journalier',
            "ENCAISSEMENTS JOURNALIERS PAR MODE",
            ['Date', 'Mode', 'Opérations', 'Montant encaissé (GNF)'],
        )
        for item in data['daily_modes']:
            append(ws, [
                item['date'], item['mode'], item['count'], int(item['amount'])
            ])
        if not data['daily_modes']:
            append(ws, [None, 'Aucun encaissement validé', 0, 0])
        for row in range(6, ws.max_row + 1):
            ws.cell(row, 1).number_format = 'dd/mm/yyyy'
            ws.cell(row, 4).number_format = '#,##0 "GNF"'

        ws = sheet(
            'Détail encaissements',
            "JOURNAL DÉTAILLÉ PAR MODE",
            [
                'Mode', 'Date', 'Reçu', 'Matricule', 'Élève', 'Classe', 'Type',
                'Montant (GNF)', 'Référence', 'Contrôle référence',
                'Caissier', 'Validateur',
            ],
        )
        for item in sorted(
            data['payment_rows'],
            key=lambda row: (row['mode'], row['date'], row['receipt']),
        ):
            append(ws, [
                item['mode'], item['date'], item['receipt'], item['matricule'],
                item['student'], item['class'], item['type'], int(item['amount']),
                item['reference'], item['reference_status'], item['cashier'],
                item['validator'],
            ])
        for row in range(6, ws.max_row + 1):
            ws.cell(row, 2).number_format = 'dd/mm/yyyy'
            ws.cell(row, 8).number_format = '#,##0 "GNF"'
    elif report_kind == 'accounting':
        ws = sheet('Synthèse', 'RAPPORT COMPTABLE DES ENCAISSEMENTS', ['Indicateur', 'Nombre', 'Montant (GNF)'])
        append(ws, ['Paiements validés', data['validated_count'], int(data['total_validated'])])
        append(ws, ['Remises accordées', len(data['discount_by_reason']), int(data['total_discounts'])])
        append(ws, ['Couverture totale', data['validated_count'], int(data['total_coverage'])], total=True)
        ws = sheet('Journal validé', 'JOURNAL DES ENCAISSEMENTS VALIDÉS', [
            'Date', 'Reçu', 'Matricule', 'Élève', 'Classe', 'Type', 'Mode',
            'Montant', 'Remise', 'Référence', 'Contrôle référence',
            'Inscription', 'Réinscription', 'Tranche 1', 'Tranche 2', 'Tranche 3',
            'Non affecté', 'Caissier', 'Validateur',
        ])
        for item in data['payment_rows']:
            append(ws, [
                item['date'], item['receipt'], item['matricule'], item['student'],
                item['class'], item['type'], item['mode'], int(item['amount']),
                int(item['discount']), item['reference'], item['reference_status'],
                int(item['allocation']['inscription']),
                int(item['allocation']['reinscription']),
                int(item['allocation']['tranche_1']),
                int(item['allocation']['tranche_2']),
                int(item['allocation']['tranche_3']),
                int(item['allocation']['non_affecte']),
                item['cashier'], item['validator'],
            ])
        ws = sheet('Affectations', 'VENTILATION RÉELLE DES ENCAISSEMENTS', ['Composante', 'Opérations', 'Montant (GNF)'])
        for item in data['by_component'].values():
            append(ws, [item['label'], item['count'], int(item['amount'])])
        append(ws, [
            'TOTAL', '-',
            int(sum((item['amount'] for item in data['by_component'].values()), ZERO)),
        ], total=True)
        ws = sheet('Statuts', 'PAIEMENTS PAR STATUT', ['Statut', 'Nombre', 'Montant (GNF)'])
        for code, _label in Paiement.STATUT_CHOICES:
            item = data['by_status'][code]
            append(ws, [item['label'], item['count'], int(item['amount'])])
        append(
            ws,
            ['TOTAL', data['payment_count'], int(sum((item.montant or ZERO) for item in data['payments']))],
            total=True,
        )
        ws = sheet('Ventilations', 'VENTILATIONS ET CONTRÔLE DES JUSTIFICATIFS', [
            'Dimension', 'Libellé', 'Nombre', 'Montant (GNF)', 'Remises (GNF)',
            'Références manquantes', 'Montant à justifier (GNF)',
        ])
        for label, item in data['by_mode'].items():
            append(ws, [
                'Mode', label, item['count'], int(item['amount']), 0,
                item['reference_missing'], int(item['reference_missing_amount']),
            ])
        for label, item in data['by_type'].items():
            append(ws, ['Type', label, item['count'], int(item['amount']), 0, 0, 0])
        for label, item in data['by_class'].items():
            append(ws, ['Classe', label, item['count'], int(item['amount']), int(item['discount']), 0, 0])
        ws = sheet('Remises', 'REMISES ET RÉDUCTIONS', ['Motif', 'Montant (GNF)'])
        for label, amount in data['discount_by_reason'].items():
            append(ws, [label, int(amount)])
        append(ws, ['TOTAL', int(data['total_discounts'])], total=True)
    else:
        ws = sheet(
            'Synthèse', 'RAPPORT PROFESSIONNEL DE RECOUVREMENT',
            ['Indicateur', 'Nombre', 'Montant (GNF)', 'Valeur / Observation'],
        )
        for label, count, amount, observation in [
            ('Créances brutes', data['schedule_count'], data['total_due'], 'Échéanciers actifs suivis'),
            ('Encaissements cumulés', None, data['total_cash'], 'Paiements affectés'),
            ('Remises enregistrées', None, data['total_discount'], f"{data['discount_rate']:.1f} % des créances brutes"),
            ('Solde à recouvrer', data['schedule_count'] - data['settled_count'], data['total_balance'], 'Après paiements et remises'),
            ('Élèves soldés', data['settled_count'], None, 'Aucun solde restant'),
            ('Paiement partiel', data['partial_count'], None, 'Couverture incomplète'),
            ('Sans paiement', data['unpaid_count'], None, 'Aucune couverture enregistrée'),
            ('Élèves en retard', data['overdue_count'], data['total_overdue'], f"Situation au {data['cutoff'].strftime('%d/%m/%Y')}"),
            ('Échéances sous 30 jours', None, data['total_upcoming'], 'Prévention à organiser'),
            ('Encaissements sur la période', data['period_payment_count'], data['period_cash'], data['period_label']),
            ('Taux de recouvrement', None, None, f"{data['recovery_rate']:.1f} %"),
        ]:
            append(ws, [label, count, int(amount) if amount is not None else None, observation])
        ws = sheet(
            'Portefeuille élèves', 'PORTEFEUILLE DÉTAILLÉ DU RECOUVREMENT',
            ['Matricule', 'Élève', 'Classe', 'Responsable', 'Téléphone', 'Dû', 'Encaissé', 'Remises', 'Remise (%)', 'Détail remise', 'Précision remise', 'Solde', 'Retard', 'À 30 jours', 'Situation', 'Relances', 'Dernière relance', 'Statut relance'],
        )
        for item in data['student_rows']:
            append(ws, [
                item['matricule'], item['student'], item['class'], item['responsible'],
                item['phone'], int(item['due']), int(item['cash']), int(item['discount']),
                float(round(item['discount_rate'], 1)), item['discount_detail'],
                item['discount_precision'], int(item['balance']), int(item['overdue']),
                int(item['upcoming']), item['status'], item['reminder_count'],
                timezone.localtime(item['last_reminder']).replace(tzinfo=None)
                if item['last_reminder'] else None,
                item['last_status'],
            ])
        ws = sheet(
            'Remises élèves', 'DÉTAIL DES REMISES PAR ÉLÈVE',
            ['Matricule', 'Élève', 'Classe', 'Dû', 'Encaissé', 'Remise (GNF)', 'Remise (%)', 'Solde', 'Situation', 'Détail remise', 'Précision'],
        )
        for item in data['discount_rows']:
            append(ws, [
                item['matricule'], item['student'], item['class'], int(item['due']),
                int(item['cash']), int(item['discount']),
                float(round(item['discount_rate'], 1)), int(item['balance']),
                item['status'], item['discount_detail'], item['discount_precision'],
            ])
        ws = sheet('Classes', 'PERFORMANCE PAR CLASSE', ['Classe', 'Élèves', 'Dû', 'Encaissé', 'Remises', 'Remise (%)', 'Solde', 'Retard', 'À 30 jours', 'Taux (%)'])
        for label, item in data['class_summary'].items():
            rate = _percentage(item['cash'] + item['discount_applied'], item['due'])
            discount_rate = _percentage(item['discount'], item['due'])
            append(ws, [label, item['students'], int(item['due']), int(item['cash']), int(item['discount']), float(round(discount_rate, 1)), int(item['balance']), int(item['overdue']), int(item['upcoming']), float(round(rate, 1))])
        append(ws, [
            'TOTAL', data['schedule_count'], int(data['total_due']), int(data['total_cash']),
            int(data['total_discount']), float(round(data['discount_rate'], 1)),
            int(data['total_balance']), int(data['total_overdue']),
            int(data['total_upcoming']), float(round(data['recovery_rate'], 1)),
        ], total=True)
        ws = sheet('Balance âgée', 'BALANCE ÂGÉE DES IMPAYÉS', ['Ancienneté', 'Échéances', 'Montant (GNF)'])
        for label, item in data['aging'].items():
            append(ws, [label, item['count'], int(item['amount'])])
        append(ws, [
            'TOTAL', sum(item['count'] for item in data['aging'].values()),
            int(data['total_overdue']),
        ], total=True)
        ws = sheet('Priorités', 'DOSSIERS PRIORITAIRES', ['Matricule', 'Élève', 'Classe', 'Responsable', 'Téléphone', 'Encaissé', 'Remise', 'Remise (%)', 'Précision remise', 'Solde', 'Retard', 'Jours', 'Relances', 'Dernière relance', 'Statut'])
        for item in data['priority_rows']:
            append(ws, [
                item['matricule'], item['student'], item['class'], item['responsible'],
                item['phone'], int(item['cash']), int(item['discount']),
                float(round(item['discount_rate'], 1)), item['discount_precision'],
                int(item['balance']), int(item['overdue']), item['days'], item['reminder_count'],
                timezone.localtime(item['last_reminder']).replace(tzinfo=None)
                if item['last_reminder'] else None,
                item['last_status'],
            ])
        ws = sheet('Relances', 'PILOTAGE DES RELANCES', ['Dimension', 'Libellé', 'Actions', 'Envoyées', 'Échecs', 'Taux succès (%)'])
        for label, item in data['reminder_by_channel'].items():
            success = item['sent'] / item['count'] * 100 if item['count'] else 0
            append(ws, ['Canal', label, item['count'], item['sent'], item['failed'], round(success, 1)])
        for label, count in data['reminder_by_status'].items():
            append(ws, ['Statut', label, count, None, None, None])
        ws = sheet(
            'Journal relances', 'JOURNAL DÉTAILLÉ DES RELANCES',
            ['Date', 'Matricule', 'Élève', 'Classe', 'Canal', 'Statut', 'Solde estimé', 'Message', 'Créé par'],
        )
        for item in data['period_relances']:
            append(ws, [
                timezone.localtime(item.date_creation).replace(tzinfo=None),
                item.eleve.matricule, item.eleve.nom_complet,
                item.eleve.classe.nom, item.get_canal_display(), item.get_statut_display(),
                int(item.solde_estime or 0), item.message, _display_user(item.cree_par),
            ])

    for ws in wb.worksheets:
        for column in ws.columns:
            letter = openpyxl.utils.get_column_letter(column[0].column)
            max_len = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = '1:5'
    return wb


def _bad_request(exc):
    return HttpResponse(str(exc), status=400, content_type='text/plain; charset=utf-8')


def _pdf_response(buffer, prefix, data):
    suffix = _safe_filename(data['classes'][0].nom) if len(data['classes']) == 1 else 'etablissement'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{prefix}_{suffix}_{date.today().isoformat()}.pdf"'
    return response


def _excel_response(workbook, prefix, data):
    buffer = io.BytesIO()
    workbook.save(buffer)
    suffix = _safe_filename(data['classes'][0].nom) if len(data['classes']) == 1 else 'etablissement'
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{prefix}_{suffix}_{date.today().isoformat()}.xlsx"'
    return response


@can_view_reports
def modes_encaissement_tableau(request):
    """Affiche les encaissements et le solde des élèves dans un tableau filtrable."""
    try:
        data = collect_payment_modes_data(request)
    except ValueError as exc:
        return _bad_request(exc)

    situation = data['situation']

    rows = data['student_mode_rows']
    unique_students = {
        (item['student_id'], item['school_year']): item
        for item in rows
    }
    mode_summary = defaultdict(lambda: {
        'students': set(), 'operations': 0, 'collected': ZERO, 'balance': ZERO,
    })
    for item in rows:
        summary = mode_summary[item['mode']]
        summary['students'].add((item['student_id'], item['school_year']))
        summary['operations'] += item['operation_count']
        summary['collected'] += item['collected']
        summary['balance'] += item['balance']
    mode_rows = [
        {
            'mode': mode,
            'student_count': len(values['students']),
            'operations': values['operations'],
            'collected': values['collected'],
            'balance': values['balance'],
            'share': _percentage(
                values['collected'],
                sum((item['collected'] for item in rows), ZERO),
            ),
        }
        for mode, values in sorted(mode_summary.items())
    ]

    all_classes = filter_by_user_school(
        Classe.objects.select_related('ecole').order_by('ecole__nom', 'niveau', 'nom'),
        request.user,
        'ecole',
    )
    year_options = list(
        all_classes.exclude(annee_scolaire='')
        .values_list('annee_scolaire', flat=True)
        .distinct()
        .order_by('-annee_scolaire')
    )
    paginator = Paginator(rows, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {
        'data': data,
        'page_obj': page_obj,
        'mode_rows': mode_rows,
        'classes_filtre': all_classes,
        'modes_filtre': ModePaiement.objects.order_by('nom'),
        'annees_filtre': year_options,
        'filtre_annee': data['school_year'],
        'filtre_classe_id': (request.GET.get('classe_id') or '').strip(),
        'filtre_mode_id': (
            str(data['selected_mode'].pk) if data['selected_mode'] else ''
        ),
        'filtre_du': data['start'],
        'filtre_au': data['requested_end'],
        'filtre_q': data['search'],
        'filtre_situation': situation,
        'filtered_student_count': len(unique_students),
        'filtered_total_collected': sum(
            (item['collected'] for item in rows), ZERO
        ),
        'filtered_total_due': sum(
            (item['amount_due'] for item in unique_students.values()), ZERO
        ),
        'filtered_total_paid': sum(
            (item['amount_paid'] for item in unique_students.values()), ZERO
        ),
        'filtered_total_discount': sum(
            (item['discount'] for item in unique_students.values()), ZERO
        ),
        'filtered_total_balance': sum(
            (item['balance'] for item in unique_students.values()), ZERO
        ),
    }
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    template = (
        'paiements/_modes_encaissement_resultats.html'
        if is_ajax else 'paiements/modes_encaissement.html'
    )
    return render(request, template, context)


@can_view_reports
def apercu_rapport_comptable(request):
    """Affiche exactement les données utilisées par les exports comptables."""
    try:
        data = collect_accounting_data(request)
    except ValueError as exc:
        return _bad_request(exc)

    status_rows = []
    for code, _label in Paiement.STATUT_CHOICES:
        item = data['by_status'][code]
        status_rows.append({
            **item,
            'code': code,
            'share': _percentage(item['count'], data['payment_count']),
        })

    mode_rows = [
        {
            **item,
            'label': label,
            'share': _percentage(item['amount'], data['total_validated']),
        }
        for label, item in data['by_mode'].items()
    ]
    component_rows = [
        item
        for item in data['by_component'].values()
        if item['amount'] > 0 or item['label'] != 'Non affecté / à contrôler'
    ]
    context = {
        'data': data,
        'status_rows': status_rows,
        'mode_rows': mode_rows,
        'type_rows': [
            {**item, 'label': label}
            for label, item in data['by_type'].items()
        ],
        'component_rows': component_rows,
        'component_total': sum(
            (item['amount'] for item in data['by_component'].values()), ZERO
        ),
        'class_rows': [
            {**item, 'label': label, 'coverage': item['amount'] + item['discount']}
            for label, item in data['by_class'].items()
        ],
        'discount_rows': [
            {'label': label, 'amount': amount}
            for label, amount in data['discount_by_reason'].items()
        ],
        'all_payments_total': sum(
            (payment.montant or ZERO for payment in data['payments']), ZERO
        ),
    }
    return render(request, 'paiements/apercu_rapport_comptable.html', context)


@can_view_reports
def export_comptabilite_pdf(request):
    try:
        data = collect_accounting_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    return _pdf_response(build_accounting_pdf(data), 'rapport_comptable', data)


@can_view_reports
def export_comptabilite_excel(request):
    try:
        data = collect_accounting_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    return _excel_response(_excel_workbook(data, 'accounting'), 'rapport_comptable', data)


@can_view_reports
def export_modes_encaissement_pdf(request):
    try:
        data = collect_payment_modes_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    return _pdf_response(
        build_payment_modes_pdf(data), 'encaissements_par_mode', data
    )


@can_view_reports
def export_modes_encaissement_excel(request):
    try:
        data = collect_payment_modes_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    return _excel_response(
        _excel_workbook(data, 'modes'), 'encaissements_par_mode', data
    )


@can_view_reports
def export_recouvrement_pdf(request):
    try:
        data = collect_recovery_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    return _pdf_response(build_recovery_pdf(data), 'rapport_recouvrement', data)


@can_view_reports
def export_recouvrement_excel(request):
    try:
        data = collect_recovery_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    return _excel_response(_excel_workbook(data, 'recovery'), 'rapport_recouvrement', data)
