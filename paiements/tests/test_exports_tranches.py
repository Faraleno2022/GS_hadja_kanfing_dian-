from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from eleves.models import Classe, Ecole, Eleve, Responsable
from paiements.models import (
    EcheancierPaiement,
    ModePaiement,
    Paiement,
    PaiementRemise,
    RemiseReduction,
    TypePaiement,
)
from utilisateurs.models import Profil


class ExportTranchesParClasseTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username='admin-export-tranches',
            email='admin-export@example.com',
            password='mot-de-passe-test',
        )
        self.client.force_login(self.user)
        self.ecole = Ecole.objects.create(
            nom='École export',
            adresse='Conakry',
            telephone='+224620000001',
            directeur='Direction export',
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole,
            nom='Classe export',
            niveau='COLLEGE_8',
            annee_scolaire='2025-2026',
        )
        responsable = Responsable.objects.create(
            prenom='Parent',
            nom='Export',
            relation='PERE',
            telephone='+224620000002',
        )
        self.eleve = Eleve.objects.create(
            matricule='EXP-001',
            prenom='Élève',
            nom='Export',
            sexe='F',
            date_naissance=date(2014, 1, 1),
            classe=self.classe,
            date_inscription=date(2025, 9, 1),
            responsable_principal=responsable,
        )
        EcheancierPaiement.objects.create(
            eleve=self.eleve,
            annee_scolaire='2025-2026',
            nature_frais=EcheancierPaiement.NATURE_REINSCRIPTION,
            frais_inscription_du=Decimal('30000'),
            frais_inscription_paye=Decimal('30000'),
            tranche_1_due=Decimal('500000'),
            tranche_1_payee=Decimal('100000'),
            tranche_2_due=Decimal('500000'),
            tranche_3_due=Decimal('500000'),
            date_echeance_inscription=date(2025, 9, 30),
            date_echeance_tranche_1=date(2026, 1, 10),
            date_echeance_tranche_2=date(2026, 3, 5),
            date_echeance_tranche_3=date(2026, 5, 5),
        )

    def _creer_utilisateur(self, username, role, ecole=..., **permissions):
        user = get_user_model().objects.create_user(
            username=username, password='mot-de-passe-test'
        )
        # Le signal utilisateurs.signals cree deja un profil (role COMPTABLE).
        profil, _ = Profil.objects.get_or_create(user=user)
        profil.role = role
        profil.ecole = self.ecole if ecole is ... else ecole
        for champ, valeur in permissions.items():
            setattr(profil, champ, valeur)
        profil.save()
        # Ce profil reste en cache sur `user`: on relit l'utilisateur pour ne
        # pas reecrire l'ancien role par megarde au prochain save().
        return get_user_model().objects.get(pk=user.pk)

    def _tables_pdf(self):
        """Capture les tables passees a ReportLab pendant la generation."""
        from reportlab.platypus import Table as ReportLabTable

        tables = []

        def capture_table(data, *args, **kwargs):
            tables.append(data)
            return ReportLabTable(data, *args, **kwargs)

        return tables, patch('reportlab.platypus.Table', side_effect=capture_table)

    def _appliquer_remise_et_solder(self):
        schedule = EcheancierPaiement.objects.get(eleve=self.eleve)
        schedule.tranche_1_payee = Decimal('500000')
        schedule.tranche_2_payee = Decimal('500000')
        schedule.tranche_3_payee = Decimal('400000')
        schedule.save()
        payment_type = TypePaiement.objects.create(nom='Scolarité avec remise export')
        payment_mode = ModePaiement.objects.create(nom='Espèces export')
        payment = Paiement.objects.create(
            eleve=self.eleve,
            type_paiement=payment_type,
            mode_paiement=payment_mode,
            numero_recu='EXP-REM-001',
            montant=Decimal('1430000'),
            annee_scolaire='2025-2026',
            date_paiement=date(2026, 2, 15),
            statut='VALIDE',
            cree_par=self.user,
            valide_par=self.user,
        )
        discount = RemiseReduction.objects.create(
            nom='Remise fratrie export',
            type_remise='POURCENTAGE',
            valeur=Decimal('5.00'),
            motif='FRATRIE',
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 8, 31),
        )
        PaiementRemise.objects.create(
            paiement=payment,
            remise=discount,
            montant_remise=Decimal('100000'),
        )

    def _creer_cas_remise_t1_avec_ancien_report_t2(self):
        """Reproduit le cas FARA LENO: la remise T1 ne devient pas du cash T2."""
        schedule = EcheancierPaiement.objects.get(eleve=self.eleve)
        schedule.nature_frais = EcheancierPaiement.NATURE_INSCRIPTION
        schedule.frais_inscription_du = Decimal('50000')
        schedule.frais_inscription_paye = Decimal('50000')
        schedule.tranche_1_payee = Decimal('475000')
        schedule.tranche_2_payee = Decimal('25000')
        schedule.tranche_3_payee = Decimal('0')
        schedule.save()

        payment_type = TypePaiement.objects.create(nom='Inscription + Tranche 1')
        payment_mode = ModePaiement.objects.create(nom='Cash')
        payment = Paiement.objects.create(
            eleve=self.eleve,
            type_paiement=payment_type,
            mode_paiement=payment_mode,
            numero_recu='REC20260001',
            montant=Decimal('550000'),
            annee_scolaire='2025-2026',
            date_paiement=date(2026, 8, 23),
            statut='VALIDE',
            cree_par=self.user,
            valide_par=self.user,
        )
        discount = RemiseReduction.objects.create(
            nom='Remise scolarité 5% (T1)',
            type_remise='POURCENTAGE',
            valeur=Decimal('5.00'),
            motif='AUTRE',
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 8, 31),
        )
        PaiementRemise.objects.create(
            paiement=payment,
            remise=discount,
            montant_remise=Decimal('25000'),
        )
        return payment

    @property
    def filtres(self):
        return {'classe': self.classe.pk, 'annee_scolaire': '2025-2026'}

    def test_excel_separe_inscription_et_reinscription(self):
        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            self.filtres,
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertGreaterEqual(len(workbook.worksheets), 2)
        sheet = workbook.worksheets[1]
        self.assertEqual(sheet.cell(2, 2).value, 'Inscription payée')
        self.assertEqual(sheet.cell(2, 3).value, 'Réinscription payée')
        self.assertEqual(sheet.cell(3, 2).value, 0)
        self.assertEqual(sheet.cell(3, 3).value, 30000)
        self.assertEqual(sheet.cell(3, 8).value, 130000)

    def test_excel_affiche_montant_taux_remise_et_statut_solde(self):
        self._appliquer_remise_et_solder()

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            self.filtres,
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook.worksheets[1]
        self.assertEqual(sheet.cell(2, 9).value, 'Remise')
        self.assertEqual(sheet.cell(2, 10).value, 'Remise (%)')
        self.assertEqual(sheet.cell(3, 6).value, 400000)
        self.assertEqual(sheet.cell(3, 8).value, 1430000)
        self.assertEqual(sheet.cell(3, 9).value, 100000)
        self.assertEqual(sheet.cell(3, 10).value, 5)
        self.assertEqual(sheet.cell(3, 11).value, 0)
        self.assertEqual(sheet.cell(2, 12).value, '% payé')
        self.assertEqual(sheet.cell(3, 12).value, 100)
        self.assertEqual(sheet.cell(3, 13).value, 'Soldé')

    def test_pdf_contient_la_colonne_reinscription(self):
        from reportlab.platypus import Table as ReportLabTable

        tables = []

        def capture_table(data, *args, **kwargs):
            tables.append(data)
            return ReportLabTable(data, *args, **kwargs)

        with patch('reportlab.platypus.Table', side_effect=capture_table):
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self.filtres,
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        header = [cell.getPlainText() for cell in tables[0][0]]
        self.assertEqual(header[1:3], ['Inscription payée', 'Réinscription payée'])
        self.assertEqual(tables[0][1][1:3], ['0', '30 000'])

    def test_pdf_tranches_utilise_le_logo_de_ecole_filtree(self):
        with patch('paiements.views_tranches._draw_header_and_watermark') as entete:
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self.filtres,
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(entete.called)
        self.assertTrue(all(
            appel.kwargs['ecole'] == self.ecole
            for appel in entete.call_args_list
        ))

    def test_pdf_paiements_filtres_utilise_le_logo_de_ecole_filtree(self):
        with patch('paiements.export_paiements_filtres._draw_header_and_watermark') as entete:
            response = self.client.get(
                reverse('paiements:export_paiements_filtres_pdf'),
                {'classe_id': self.classe.pk, 'annee': '2025-2026'},
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(entete.called)
        self.assertTrue(all(
            appel.kwargs['ecole'] == self.ecole
            for appel in entete.call_args_list
        ))

    def test_pdf_affiche_montant_taux_remise_et_statut_solde(self):
        from reportlab.platypus import Table as ReportLabTable

        self._appliquer_remise_et_solder()
        tables = []

        def capture_table(data, *args, **kwargs):
            tables.append(data)
            return ReportLabTable(data, *args, **kwargs)

        with patch('reportlab.platypus.Table', side_effect=capture_table):
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self.filtres,
            )

        self.assertEqual(response.status_code, 200)
        header = [cell.getPlainText() for cell in tables[0][0]]
        row = [cell.getPlainText() if hasattr(cell, 'getPlainText') else cell for cell in tables[0][1]]
        self.assertEqual(
            header[8:13],
            ['Remise', 'Remise (%)', 'Reste', '% payé', 'Situation / précision'],
        )
        self.assertEqual(row[5:11], ['400 000', '1 530 000', '1 430 000', '100 000', '5 %', '0'])
        self.assertEqual(row[11], '100 %')
        self.assertEqual(row[12], 'Soldé')

    def test_excel_ne_reporte_pas_la_remise_t1_dans_la_tranche_2(self):
        self._creer_cas_remise_t1_avec_ancien_report_t2()

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            self.filtres,
        )

        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.content), data_only=True).worksheets[1]
        self.assertEqual(sheet.cell(3, 2).value, 50000)
        self.assertEqual(sheet.cell(3, 4).value, 500000)
        self.assertEqual(sheet.cell(3, 5).value, 0)
        self.assertEqual(sheet.cell(3, 6).value, 0)
        self.assertEqual(sheet.cell(3, 8).value, 550000)
        self.assertEqual(sheet.cell(3, 9).value, 25000)
        self.assertEqual(sheet.cell(3, 10).value, 5)
        self.assertEqual(sheet.cell(3, 11).value, 975000)

    def test_pdf_ne_reporte_pas_la_remise_t1_dans_la_tranche_2(self):
        self._creer_cas_remise_t1_avec_ancien_report_t2()
        tables, capture = self._tables_pdf()

        with capture:
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self.filtres,
            )

        self.assertEqual(response.status_code, 200)
        row = [
            cell.getPlainText() if hasattr(cell, 'getPlainText') else cell
            for cell in tables[0][1]
        ]
        self.assertEqual(
            row[1:11],
            ['50 000', '0', '500 000', '0', '0', '1 550 000', '550 000', '25 000', '5 %', '975 000'],
        )

    def test_recu_pdf_ne_reporte_pas_la_remise_t1_dans_la_tranche_2(self):
        from pypdf import PdfReader

        payment = self._creer_cas_remise_t1_avec_ancien_report_t2()

        response = self.client.get(
            reverse('paiements:generer_recu_pdf', kwargs={'paiement_id': payment.pk})
        )

        self.assertEqual(response.status_code, 200)
        text = '\n'.join(
            page.extract_text() or ''
            for page in PdfReader(BytesIO(response.content)).pages
        )
        allocation = text.split('Affectation du paiement', 1)[1].split(
            "Informations de l'élève", 1
        )[0]
        self.assertIn('Inscription: 50 000 GNF', allocation)
        self.assertIn('1ère tranche: 500 000 GNF', allocation)
        self.assertIn('2ème tranche: 0 GNF', allocation)
        self.assertIn('3ème tranche: 0 GNF', allocation)
        self.assertNotIn('Remise accordée', text)
        self.assertNotIn('Dette couverte par ce reçu', text)
        self.assertIn('Montant payé : 525 000 GNF', text)
        self.assertNotIn('Montant payé : 550 000 GNF', text)
        self.assertIn('Remise de : 25 000 GNF', text)
        self.assertIn('Remise scolarité 5% (T1) : 25 000 GNF', text)
        self.assertLess(text.index('Montant payé'), text.index('Remise de'))

    def test_recu_public_pdf_garde_la_remise_uniquement_dans_son_bloc(self):
        from pypdf import PdfReader

        from paiements.recu_public import generer_token_recu

        payment = self._creer_cas_remise_t1_avec_ancien_report_t2()
        response = self.client.get(
            reverse('paiements:recu_public_pdf', kwargs={'paiement_id': payment.pk}),
            {'token': generer_token_recu(payment.pk)},
        )

        self.assertEqual(response.status_code, 200)
        text = '\n'.join(
            page.extract_text() or ''
            for page in PdfReader(BytesIO(response.content)).pages
        )
        allocation = text.split('Affectation du paiement', 1)[1].split(
            'SITUATION FINANCIÈRE', 1
        )[0]
        self.assertIn('1ère tranche: 500 000 GNF', allocation)
        self.assertIn('2ème tranche: 0 GNF', allocation)
        self.assertIn('3ème tranche: 0 GNF', allocation)
        self.assertNotIn('Remise accordée', text)
        self.assertNotIn('Dette couverte', text)
        self.assertIn('Montant payé : 525 000 GNF', text)
        self.assertNotIn('Montant payé : 550 000 GNF', text)
        self.assertIn('Remise de : 25 000 GNF', text)
        self.assertIn('Remise scolarité 5% (T1) : 25 000 GNF', text)
        self.assertLess(text.index('Montant payé'), text.index('Remise de'))

    # ── Accès ─────────────────────────────────────────────────────────
    def test_directeur_peut_telecharger_le_pdf(self):
        self.client.force_login(self._creer_utilisateur('directeur-export', 'DIRECTEUR'))

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_pdf'), self.filtres
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('tranches_par_classe_', response['Content-Disposition'])

    def test_secretaire_peut_telecharger_le_excel(self):
        self.client.force_login(self._creer_utilisateur('secretaire-export', 'SECRETAIRE'))

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'), self.filtres
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml.sheet', response['Content-Type'])
        self.assertIn('.xlsx', response['Content-Disposition'])

    def test_enseignant_prive_de_rapports_est_refuse(self):
        self.client.force_login(self._creer_utilisateur(
            'prof-export', 'ENSEIGNANT',
            peut_consulter_rapports=False, peut_generer_rapports=False,
        ))

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_pdf'), self.filtres
        )

        self.assertEqual(response.status_code, 403)

    def test_compte_sans_ecole_n_exporte_aucune_classe(self):
        self.client.force_login(
            self._creer_utilisateur('comptable-sans-ecole', 'COMPTABLE', ecole=None)
        )

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'), self.filtres
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(len(workbook.worksheets), 1)

    # ── Filtres et totaux ─────────────────────────────────────────────
    def test_filtre_annee_de_la_page_paiements_est_pris_en_compte(self):
        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            {'classe': self.classe.pk, 'annee': '2025-2026'},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(len(workbook.worksheets), 2)
        self.assertEqual(workbook.worksheets[1].cell(3, 3).value, 30000)

    def test_excel_termine_par_une_ligne_de_total_hors_filtre(self):
        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'), self.filtres
        )

        sheet = load_workbook(BytesIO(response.content), data_only=True).worksheets[1]
        self.assertEqual(sheet.cell(sheet.max_row, 1).value, 'TOTAL CLASSE')
        self.assertEqual(sheet.cell(sheet.max_row, 7).value, 1530000)
        self.assertEqual(sheet.cell(sheet.max_row, 8).value, 130000)
        self.assertTrue(sheet.auto_filter.ref.endswith(str(sheet.max_row - 1)))

    def test_pdf_termine_par_une_ligne_de_total(self):
        tables, capture = self._tables_pdf()

        with capture:
            self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'), self.filtres
            )

        derniere = tables[0][-1]
        self.assertEqual(derniere[0].getPlainText().strip(), 'TOTAL CLASSE')
        self.assertEqual(derniere[6], '1 530 000')
        self.assertEqual(derniere[7], '130 000')
