"""Tests des exports PDF/Excel regroupes par mode d'encaissement."""

import io
import importlib.util
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from eleves.models import Classe, Ecole, Eleve, Responsable
from paiements.export_modes_encaissement import collect_modes_encaissement_data
from paiements.models import EcheancierPaiement, ModePaiement, Paiement, TypePaiement
from paiements.tests.support import TEST_MIDDLEWARE
from paiements.views_modes_encaissement import collect_modes_students_data


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class ExportModesEncaissementTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='direction', password='pass1234',
            first_name='Aminata', last_name='Diallo',
        )
        self.ecole = self._ecole('École Principale', '+224600000001')
        self.autre_ecole = self._ecole('École Hors Périmètre', '+224600000002')
        self.user.profil.role = 'ADMIN'
        self.user.profil.telephone = '+224600000003'
        self.user.profil.ecole = self.ecole
        self.user.profil.save()
        self.client.force_login(self.user)
        self.today = timezone.localdate()
        self.month_start = self.today.replace(day=1)

        self.classe = self._classe(self.ecole, '7ème A')
        self.autre_classe = self._classe(self.autre_ecole, '8ème B')
        self.responsable = Responsable.objects.create(
            prenom='Mamadou', nom='Sow', relation='PERE',
            telephone='+224600000004', adresse='Conakry',
        )
        self.eleve = self._eleve('MDE-001', 'Fatou', self.classe)
        self.autre_eleve = self._eleve('MDE-002', 'Binta', self.autre_classe)
        self.type_paiement = TypePaiement.objects.create(nom='Scolarité export modes')
        self.especes = ModePaiement.objects.create(nom='Espèces')
        self.orange = ModePaiement.objects.create(nom='Orange Money')

    @staticmethod
    def _ecole(nom, telephone):
        return Ecole.objects.create(
            nom=nom, adresse='Conakry', telephone=telephone,
            directeur='Direction', email=f"{telephone[-3:]}@example.com",
        )

    @staticmethod
    def _classe(ecole, nom):
        return Classe.objects.create(
            ecole=ecole, nom=nom, niveau='COLLEGE_7',
            annee_scolaire='2026-2027', capacite_max=40,
        )

    def _eleve(self, matricule, prenom, classe):
        return Eleve.objects.create(
            matricule=matricule, prenom=prenom, nom='Camara', sexe='F',
            date_naissance=date(2012, 5, 4), lieu_naissance='Conakry',
            classe=classe, date_inscription=self.today,
            responsable_principal=self.responsable,
        )

    def _paiement(self, eleve, mode, montant, jour=None, statut='VALIDE'):
        return Paiement.objects.create(
            eleve=eleve,
            type_paiement=self.type_paiement,
            mode_paiement=mode,
            numero_recu=f"MDE-{Paiement.objects.count() + 1:04d}",
            montant=Decimal(str(montant)),
            date_paiement=jour or self.today,
            annee_scolaire='2026-2027',
            statut=statut,
            cree_par=self.user,
            valide_par=self.user if statut == 'VALIDE' else None,
        )

    def _echeancier(self, eleve, total=300000):
        return EcheancierPaiement.objects.create(
            eleve=eleve,
            annee_scolaire='2026-2027',
            frais_inscription_du=Decimal('0'),
            tranche_1_due=Decimal(str(total)),
            tranche_2_due=Decimal('0'),
            tranche_3_due=Decimal('0'),
            date_echeance_inscription=self.today - timedelta(days=30),
            date_echeance_tranche_1=self.today - timedelta(days=10),
            date_echeance_tranche_2=self.today + timedelta(days=30),
            date_echeance_tranche_3=self.today + timedelta(days=60),
        )

    def _request(self, **params):
        request = RequestFactory().get('/', params)
        request.user = self.user
        return request

    def _creer_encaissements(self):
        self._paiement(self.eleve, self.especes, 100000)
        self._paiement(self.eleve, self.especes, 50000)
        self._paiement(self.eleve, self.orange, 75000)
        # Ces trois lignes ne doivent jamais entrer dans le rapport.
        self._paiement(self.eleve, self.orange, 20000, statut='EN_ATTENTE')
        self._paiement(
            self.eleve, self.especes, 60000,
            jour=self.month_start - timedelta(days=1),
        )
        self._paiement(self.autre_eleve, self.orange, 999000)

    def test_collecte_regroupe_uniquement_les_valides_du_mois_et_de_l_ecole(self):
        self._creer_encaissements()

        data = collect_modes_encaissement_data(self._request())

        self.assertEqual(data['school'].pk, self.ecole.pk)
        self.assertNotIn('HORS PÉRIMÈTRE', data['scope_label'])
        self.assertEqual(data['payment_count'], 3)
        self.assertEqual(data['total_amount'], Decimal('225000'))
        self.assertEqual(
            [(row['mode'], row['count'], row['amount']) for row in data['rows']],
            [
                ('Espèces', 2, Decimal('150000')),
                ('Orange Money', 1, Decimal('75000')),
            ],
        )
        self.assertAlmostEqual(float(data['rows'][0]['share']), 66.666, places=2)

    def test_export_excel_contient_montants_parts_et_total(self):
        self._creer_encaissements()

        response = self.client.get(reverse('paiements:export_modes_encaissement_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])
        self.assertTrue(response.content.startswith(b'PK'))
        workbook = load_workbook(io.BytesIO(response.content), data_only=False)
        sheet = workbook["Modes d'encaissement"]
        self.assertEqual(
            [sheet.cell(6, column).value for column in range(1, 5)],
            ['Espèces', 2, 150000, 2 / 3],
        )
        self.assertEqual(
            [sheet.cell(7, column).value for column in range(1, 4)],
            ['Orange Money', 1, 75000],
        )
        self.assertAlmostEqual(sheet.cell(7, 4).value, 1 / 3, places=6)
        self.assertEqual(
            [sheet.cell(8, column).value for column in range(1, 5)],
            ['TOTAL', 3, 225000, 1],
        )
        self.assertEqual(sheet.cell(6, 4).number_format, '0.0%')

    def test_export_pdf_contient_les_deux_modes_et_le_total(self):
        self._creer_encaissements()

        response = self.client.get(reverse('paiements:export_modes_encaissement_pdf'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))
        # La lecture textuelle renforce le contrôle lorsque pypdf est présent,
        # sans en faire une dépendance de production de l'application.
        if importlib.util.find_spec('pypdf'):
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(response.content))
            text = '\n'.join(page.extract_text() or '' for page in reader.pages)
            self.assertIn("MONTANTS PAR MODE D'ENCAISSEMENT", text)
            self.assertIn('Espèces', text)
            self.assertIn('Orange Money', text)
            self.assertIn('225 000', text)
            self.assertIn('66.7 %', text)

    def test_periode_personnalisee_est_respectee(self):
        previous_day = self.month_start - timedelta(days=1)
        self._paiement(self.eleve, self.especes, 60000, jour=previous_day)
        self._paiement(self.eleve, self.orange, 75000)

        data = collect_modes_encaissement_data(self._request(
            du=previous_day.isoformat(), au=previous_day.isoformat(),
        ))

        self.assertEqual(data['payment_count'], 1)
        self.assertEqual(data['total_amount'], Decimal('60000'))

    def test_dates_invalides_retournent_400(self):
        response = self.client.get(
            reverse('paiements:export_modes_encaissement_pdf'),
            {'du': '16-08-2026'},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('AAAA-MM-JJ', response.content.decode('utf-8'))

    def test_tableau_de_bord_propose_les_deux_exports(self):
        response = self.client.get(reverse('paiements:tableau_bord'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response, reverse('paiements:export_modes_encaissement_pdf')
        )
        self.assertContains(
            response, reverse('paiements:export_modes_encaissement_excel')
        )
        self.assertContains(
            response, reverse('paiements:modes_encaissement_soldes')
        )

    def test_tableau_detaille_affiche_montant_par_mode_et_solde_global_unique(self):
        self._echeancier(self.eleve)
        self._paiement(self.eleve, self.especes, 100000)
        self._paiement(self.eleve, self.orange, 50000)

        data = collect_modes_students_data(self._request())

        self.assertEqual(len(data['rows']), 2)
        self.assertEqual(data['summary']['student_count'], 1)
        self.assertEqual(data['summary']['payment_count'], 2)
        self.assertEqual(data['summary']['period_amount'], Decimal('150000'))
        # Le même élève apparaît sous deux modes, mais son solde ne doit être
        # compté qu'une seule fois dans la synthèse.
        self.assertEqual(data['summary']['balance'], Decimal('150000'))
        self.assertEqual(data['summary']['remaining_count'], 1)
        self.assertTrue(all(row['situation']['balance'] == Decimal('150000') for row in data['rows']))

    def test_filtres_dynamiques_mode_et_situation_limitent_les_lignes(self):
        self._echeancier(self.eleve)
        self._paiement(self.eleve, self.especes, 100000)
        self._paiement(self.eleve, self.orange, 50000)

        data = collect_modes_students_data(self._request(
            mode_id=self.orange.pk, situation='reste',
        ))
        self.assertEqual(len(data['rows']), 1)
        self.assertEqual(data['rows'][0]['mode'], 'Orange Money')
        self.assertEqual(data['rows'][0]['period_amount'], Decimal('50000'))

        soldes = collect_modes_students_data(self._request(situation='solde'))
        self.assertEqual(len(soldes['rows']), 0)

    def test_reponse_ajax_ne_retourne_que_le_tableau_filtre(self):
        self._echeancier(self.eleve)
        self._paiement(self.eleve, self.especes, 100000)
        self._paiement(self.eleve, self.orange, 50000)

        response = self.client.get(
            reverse('paiements:modes_encaissement_soldes'),
            {'mode_id': self.orange.pk},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, 'paiements/_modes_encaissement_soldes_resultats.html'
        )
        self.assertContains(response, 'Orange Money')
        self.assertNotContains(response, '<form', html=False)

    def test_classe_hors_ecole_est_refusee_sur_le_tableau_detaille(self):
        response = self.client.get(
            reverse('paiements:modes_encaissement_soldes'),
            {'classe_id': self.autre_classe.pk},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('non autorisée', response.content.decode('utf-8'))
