"""Séparation stricte des frais d'inscription et de réinscription.

Un élève ne paie qu'un seul frais d'admission : soit une inscription (nouvel
élève), soit une réinscription (élève qui revient). Les deux natures partagent
le même poste dans l'échéancier (``frais_inscription_du``) et se distinguent
par ``nature_frais``. Les rapports ne doivent donc jamais compter le même
montant dans les deux colonnes.
"""

from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse

from eleves.models import Classe, Ecole, Eleve, Responsable
from paiements.models import (
    EcheancierPaiement,
    ModePaiement,
    Paiement,
    PaiementRemise,
    RemiseReduction,
    TypePaiement,
)
from paiements.tests.support import TEST_MIDDLEWARE
from utilisateurs.models import Profil


def _texte_cellule(cellule):
    """Texte d'une cellule de tableau PDF, Paragraph ou chaîne simple."""
    texte = cellule.getPlainText() if hasattr(cellule, 'getPlainText') else str(cellule)
    # Les montants sont mis en forme avec une espace insecable pour ne pas etre
    # coupes en fin de colonne : la ramener a une espace ordinaire.
    return texte.replace(' ', ' ')


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class RepartitionInscriptionReinscriptionTests(TestCase):
    def setUp(self):
        self.ecole = Ecole.objects.create(
            nom="École Répartition", adresse="Conakry",
            telephone="+224620100031", directeur="Direction",
        )
        self.classe = Classe.objects.create(
            nom="6ÈME ANNÉE", ecole=self.ecole, niveau="PRIMAIRE_6",
            annee_scolaire="2025-2026",
        )
        self.responsable = Responsable.objects.create(
            nom="Diallo", prenom="Mariama", telephone="+224620100032",
            relation="PERE",
        )

        # Nouvel élève : frais d'inscription de 30 000.
        self.nouvel_eleve = self._creer_eleve("MAT-INSC-1", "Camara", "Awa")
        self._creer_echeancier(self.nouvel_eleve, 30000, 'INSCRIPTION')

        # Élève qui revient : frais de réinscription de 20 000.
        self.ancien_eleve = self._creer_eleve("MAT-REINSC-1", "Bah", "Ibrahima")
        self._creer_echeancier(self.ancien_eleve, 20000, 'REINSCRIPTION')

        User = get_user_model()
        self.comptable = User.objects.create_user(
            username="comptable_repartition", password="pass12345",
        )
        Profil.objects.update_or_create(
            user=self.comptable,
            defaults={
                'role': 'COMPTABLE', 'ecole': self.ecole,
                'telephone': "+224620100033",
                'peut_consulter_rapports': True, 'is_validated': True,
            },
        )
        self.client.force_login(self.comptable)

    def _creer_eleve(self, matricule, nom, prenom):
        return Eleve.objects.create(
            nom=nom, prenom=prenom, matricule=matricule, classe=self.classe,
            sexe='M', date_naissance=date(2015, 1, 1),
            lieu_naissance="Conakry", date_inscription=date(2025, 9, 1),
            responsable_principal=self.responsable,
        )

    def _creer_echeancier(self, eleve, frais_admission, nature):
        return EcheancierPaiement.objects.create(
            eleve=eleve, annee_scolaire="2025-2026",
            frais_inscription_du=frais_admission,
            nature_frais=nature,
            tranche_1_due=100000, tranche_2_due=0, tranche_3_due=0,
            frais_inscription_paye=0,
            tranche_1_payee=0, tranche_2_payee=0, tranche_3_payee=0,
            date_echeance_inscription=date(2025, 9, 1),
            date_echeance_tranche_1=date(2025, 10, 1),
            date_echeance_tranche_2=date(2026, 1, 1),
            date_echeance_tranche_3=date(2026, 4, 1),
        )

    def _appliquer_remise_soldante(self, eleve):
        type_paiement = TypePaiement.objects.create(nom="Inscription + Tranche 1")
        mode = ModePaiement.objects.create(nom="Espèces test export")
        paiement = Paiement.objects.create(
            eleve=eleve, type_paiement=type_paiement, mode_paiement=mode,
            numero_recu="REMISE-EXPORT-001", montant=Decimal("110000"),
            date_paiement=date(2026, 8, 15), annee_scolaire="2025-2026",
            statut="VALIDE", cree_par=self.comptable, valide_par=self.comptable,
        )
        remise = RemiseReduction.objects.create(
            nom="Remise test export", type_remise="MONTANT_FIXE",
            valeur=Decimal("20000"), motif="SOCIALE",
            date_debut=date(2025, 9, 1), date_fin=date(2026, 8, 31), actif=True,
        )
        PaiementRemise.objects.create(
            paiement=paiement, remise=remise, montant_remise=Decimal("20000"),
            applique_tranche_1=True, montant_tranche_1=Decimal("20000"),
            base_calcul="TRANCHE", montant_base=Decimal("100000"),
            motif="GESTE_COMMERCIAL",
        )

    def test_colonnes_inscription_et_reinscription_ne_se_recouvrent_pas(self):
        response = self.client.get(reverse('paiements:liste_paiements'))
        self.assertEqual(response.status_code, 200)

        totaux = response.context['totaux_du']
        # La colonne Inscription ne doit contenir que les vraies inscriptions.
        self.assertEqual(totaux['frais_inscription_total'], 30000)
        self.assertEqual(totaux['frais_reinscription_total'], 20000)
        # Le total dû reste complet : scolarité + les deux natures d'admission.
        self.assertEqual(totaux['du_global_net'], 200000 + 30000 + 20000)
        # Part de la réinscription dans l'ensemble des frais d'admission.
        self.assertAlmostEqual(totaux['frais_reinscription_pct'], 40.0, places=2)

    def test_detail_par_classe_ne_double_compte_pas(self):
        response = self.client.get(reverse('paiements:liste_paiements'))
        self.assertEqual(response.status_code, 200)

        lignes = response.context['totaux_du_detail_classes']
        self.assertEqual(len(lignes), 1)
        ligne = lignes[0]
        self.assertEqual(ligne['frais_inscription_total'], 30000)
        self.assertEqual(ligne['frais_reinscription_total'], 20000)
        self.assertEqual(ligne['du_global_net'], 200000 + 30000 + 20000)
        self.assertAlmostEqual(ligne['frais_reinscription_pct'], 40.0, places=2)

    def test_export_excel_recap_ne_double_compte_pas(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(reverse('paiements:export_recap_par_classe_excel'))
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        entetes = [cell.value for cell in sheet[1]]
        ligne = dict(zip(entetes, [cell.value for cell in sheet[2]]))

        self.assertEqual(ligne['Inscription'], 30000)
        self.assertEqual(ligne['Réinscription'], 20000)
        self.assertEqual(ligne['Total dû net'], 200000 + 30000 + 20000)
        self.assertAlmostEqual(float(ligne['Réinscription %']), 40.0, places=2)

    def test_classe_entierement_en_reinscription_laisse_inscription_a_zero(self):
        EcheancierPaiement.objects.filter(eleve=self.nouvel_eleve).update(
            nature_frais='REINSCRIPTION', frais_inscription_du=20000,
        )

        response = self.client.get(reverse('paiements:liste_paiements'))
        totaux = response.context['totaux_du']

        self.assertEqual(totaux['frais_inscription_total'], 0)
        self.assertEqual(totaux['frais_reinscription_total'], 40000)
        self.assertAlmostEqual(totaux['frais_reinscription_pct'], 100.0, places=2)
        self.assertEqual(totaux['du_global_net'], 200000 + 40000)

    def test_export_tranches_excel_separe_inscription_et_reinscription(self):
        from io import BytesIO

        from openpyxl import load_workbook

        EcheancierPaiement.objects.filter(eleve=self.nouvel_eleve).update(
            frais_inscription_paye=30000,
        )
        EcheancierPaiement.objects.filter(eleve=self.ancien_eleve).update(
            frais_inscription_paye=20000,
        )

        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            {'annee_scolaire': '2025-2026'},
        )
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook[self.classe.nom[:25]]
        headers = [cell.value for cell in sheet[2]]
        rows = {
            row[0].value: dict(zip(headers, [cell.value for cell in row]))
            for row in sheet.iter_rows(min_row=3)
        }

        nouvel = rows[self.nouvel_eleve.nom_complet]
        ancien = rows[self.ancien_eleve.nom_complet]
        self.assertEqual(nouvel['Inscription payée'], 30000)
        self.assertEqual(nouvel['Réinscription payée'], 0)
        self.assertEqual(ancien['Inscription payée'], 0)
        self.assertEqual(ancien['Réinscription payée'], 20000)

    def test_export_tranches_pdf_contient_la_colonne_reinscription(self):
        from unittest.mock import patch

        from reportlab.platypus import Table

        EcheancierPaiement.objects.filter(eleve=self.ancien_eleve).update(
            frais_inscription_paye=20000,
        )

        with patch('reportlab.platypus.SimpleDocTemplate.build') as build_pdf:
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                {'annee_scolaire': '2025-2026'},
            )

        self.assertEqual(response.status_code, 200)
        elements = build_pdf.call_args.args[0]
        table = next(element for element in elements if isinstance(element, Table))
        headers = [cell.getPlainText() for cell in table._cellvalues[0]]
        self.assertIn('Réinscription payée', headers)

        rows = {
            row[0].getPlainText(): dict(
                zip(headers, [_texte_cellule(cellule) for cellule in row])
            )
            for row in table._cellvalues[1:]
        }
        ancien = rows[self.ancien_eleve.nom_complet]
        self.assertEqual(ancien['Inscription payée'], '0')
        self.assertEqual(ancien['Réinscription payée'], '20 000')

    def test_export_tranches_excel_affiche_remise_taux_et_solde(self):
        from io import BytesIO

        from openpyxl import load_workbook

        self._appliquer_remise_soldante(self.nouvel_eleve)
        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            {'annee_scolaire': '2025-2026'},
        )

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook[self.classe.nom[:25]]
        headers = [cell.value for cell in sheet[2]]
        rows = {
            row[0].value: dict(zip(headers, [cell.value for cell in row]))
            for row in sheet.iter_rows(min_row=3)
        }
        nouvel = rows[self.nouvel_eleve.nom_complet]

        self.assertEqual(nouvel['Encaissé'], 110000)
        self.assertEqual(nouvel['Remise'], 20000)
        # Taux calcule sur la base retenue a la saisie (montant_base), et non
        # sur le total du qui inclut l'inscription, jamais remisee.
        self.assertEqual(nouvel['Remise (%)'], 20)
        self.assertEqual(nouvel['Reste'], 0)
        self.assertIn('Soldé avec remise', nouvel['Situation / précision'])
        self.assertIn('20 000 GNF', nouvel['Situation / précision'])

    def test_export_tranches_pdf_affiche_remise_taux_et_precision(self):
        from unittest.mock import patch

        from reportlab.platypus import Table

        self._appliquer_remise_soldante(self.nouvel_eleve)
        with patch('reportlab.platypus.SimpleDocTemplate.build') as build_pdf:
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                {'annee_scolaire': '2025-2026'},
            )

        self.assertEqual(response.status_code, 200)
        elements = build_pdf.call_args.args[0]
        table = next(element for element in elements if isinstance(element, Table))
        headers = [cell.getPlainText() for cell in table._cellvalues[0]]
        rows = {
            row[0].getPlainText(): dict(
                zip(headers, [_texte_cellule(cellule) for cellule in row])
            )
            for row in table._cellvalues[1:]
        }
        nouvel = rows[self.nouvel_eleve.nom_complet]

        self.assertEqual(nouvel['Remise'], '20 000')
        self.assertEqual(nouvel['Remise (%)'], '20 %')
        self.assertIn('Soldé avec remise', nouvel['Situation / précision'])
        self.assertIn('20 000 GNF', nouvel['Situation / précision'])

    def test_logo_ecole_est_present_sur_exports_tranches_et_filigrane(self):
        from pathlib import Path
        from unittest.mock import MagicMock, patch

        from django.conf import settings
        from reportlab.lib.pagesizes import A4, landscape

        logo_path = str(Path(settings.BASE_DIR) / 'static' / 'logos' / 'logo.jpeg')
        with patch('rapports.utils._get_logo_path', return_value=logo_path):
            with patch('reportlab.platypus.SimpleDocTemplate.build') as build_pdf:
                response = self.client.get(
                    reverse('paiements:export_tranches_par_classe_pdf'),
                    {'annee_scolaire': '2025-2026'},
                )

            self.assertEqual(response.status_code, 200)
            canvas = MagicMock()
            canvas._pagesize = landscape(A4)
            callback = build_pdf.call_args.kwargs['onFirstPage']
            callback(canvas, MagicMock(pagesize=landscape(A4)))
            self.assertGreaterEqual(canvas.drawImage.call_count, 2)

        with patch('paiements.views_tranches._get_logo_path', return_value=logo_path):
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_excel'),
                {'annee_scolaire': '2025-2026'},
            )
        from io import BytesIO
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content))
        self.assertTrue(all(sheet._images for sheet in workbook.worksheets))

    def _appliquer_remise_sans_report(self):
        """Cas réel : 550 000 encaissés et 25 000 de remise sur T1."""
        echeancier = self.nouvel_eleve.echeancier
        echeancier.frais_inscription_du = Decimal('50000')
        echeancier.tranche_1_due = Decimal('500000')
        echeancier.tranche_2_due = Decimal('500000')
        echeancier.tranche_3_due = Decimal('500000')
        echeancier.save()

        paiement = Paiement.objects.create(
            eleve=self.nouvel_eleve,
            type_paiement=TypePaiement.objects.create(
                nom='Inscription + Tranche 1'
            ),
            mode_paiement=ModePaiement.objects.create(nom='Cash'),
            numero_recu='REC20260001',
            montant=Decimal('550000'),
            date_paiement=date(2026, 8, 15),
            annee_scolaire='2025-2026',
            statut='VALIDE',
            cree_par=self.comptable,
            valide_par=self.comptable,
        )
        remise = RemiseReduction.objects.create(
            nom='Remise scolarité 5%',
            type_remise='POURCENTAGE',
            valeur=Decimal('5'),
            motif='SOCIALE',
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 8, 31),
            actif=True,
        )
        PaiementRemise.objects.create(
            paiement=paiement,
            remise=remise,
            montant_remise=Decimal('25000'),
            applique_tranche_1=True,
            montant_tranche_1=Decimal('25000'),
            base_calcul='TRANCHE',
            montant_base=Decimal('500000'),
            motif='GESTE_COMMERCIAL',
            deduite_du_paiement=False,
        )
        return paiement

    def test_export_tranches_excel_ne_reporte_pas_la_remise_sur_t2(self):
        from io import BytesIO

        from openpyxl import load_workbook

        self._appliquer_remise_sans_report()
        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            {'annee_scolaire': '2025-2026'},
        )

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook[self.classe.nom[:25]]
        headers = [cell.value for cell in sheet[2]]
        rows = {
            row[0].value: dict(zip(headers, [cell.value for cell in row]))
            for row in sheet.iter_rows(min_row=3)
        }
        nouvel = rows[self.nouvel_eleve.nom_complet]

        self.assertEqual(nouvel['Inscription payée'], 50000)
        self.assertEqual(nouvel['Tranche 1 payée'], 500000)
        self.assertEqual(nouvel['Tranche 2 payée'], 0)
        self.assertEqual(nouvel['Tranche 3 payée'], 0)
        self.assertEqual(nouvel['Encaissé'], 550000)
        self.assertEqual(nouvel['Remise'], 25000)
        self.assertEqual(nouvel['Remise (%)'], 5)
        self.assertEqual(nouvel['Reste'], 975000)

    def test_export_tranches_pdf_ne_reporte_pas_la_remise_sur_t2(self):
        from unittest.mock import patch

        from reportlab.platypus import Table

        self._appliquer_remise_sans_report()
        with patch('reportlab.platypus.SimpleDocTemplate.build') as build_pdf:
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                {'annee_scolaire': '2025-2026'},
            )

        self.assertEqual(response.status_code, 200)
        elements = build_pdf.call_args.args[0]
        table = next(element for element in elements if isinstance(element, Table))
        headers = [cell.getPlainText() for cell in table._cellvalues[0]]
        rows = {
            row[0].getPlainText(): dict(
                zip(headers, [_texte_cellule(cellule) for cellule in row])
            )
            for row in table._cellvalues[1:]
        }
        nouvel = rows[self.nouvel_eleve.nom_complet]

        self.assertEqual(nouvel['Inscription payée'], '50 000')
        self.assertEqual(nouvel['Tranche 1 payée'], '500 000')
        self.assertEqual(nouvel['Tranche 2 payée'], '0')
        self.assertEqual(nouvel['Tranche 3 payée'], '0')
        self.assertEqual(nouvel['Encaissé'], '550 000')
        self.assertEqual(nouvel['Remise'], '25 000')
        self.assertEqual(nouvel['Remise (%)'], '5 %')
        self.assertEqual(nouvel['Reste'], '975 000')

    def test_recu_pdf_ne_reporte_pas_la_remise_sur_t2_ni_t3(self):
        from io import BytesIO

        from pypdf import PdfReader

        paiement = self._appliquer_remise_sans_report()
        response = self.client.get(reverse(
            'paiements:generer_recu_pdf',
            kwargs={'paiement_id': paiement.id},
        ))

        self.assertEqual(response.status_code, 200)
        texte = '\n'.join(
            page.extract_text() or ''
            for page in PdfReader(BytesIO(response.content)).pages
        )
        self.assertIn('Montant payé : 550 000 GNF', texte)
        self.assertIn('Total remises : -25 000 GNF', texte)
        self.assertIn('Montant net payé : 550 000 GNF', texte)
        self.assertIn('Inscription: 50 000 GNF', texte)
        self.assertIn('1ère tranche: 500 000 GNF', texte)
        self.assertIn('2ème tranche: 0 GNF', texte)
        self.assertIn('3ème tranche: 0 GNF', texte)
        self.assertIn('Solde global restant : 975 000 GNF', texte)
        self.assertIn('Remise scolarité 5% (T1) : -25 000 GNF', texte)
