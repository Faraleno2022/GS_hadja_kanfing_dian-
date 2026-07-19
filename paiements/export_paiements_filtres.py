"""
Export de la liste des paiements en respectant TOUS les filtres de la page
/paiements/ : recherche (q), statut, année, classe, mode, nature (type) et
niveau de paiement (retard / reste à payer / soldé).

Deux formats: Excel et PDF.
"""
import io
import re
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import (
    Q, F, Case, When, Value, DecimalField, ExpressionWrapper,
)
from django.http import HttpResponse

from eleves.models import Classe
from utilisateurs.utils import filter_by_user_school
from .models import Paiement, EcheancierPaiement, ModePaiement, TypePaiement


def _fmt_gnf(v):
    try:
        return f"{int(v):,}".replace(',', ' ')
    except (TypeError, ValueError):
        return '0'


def filtrer_paiements(request):
    """Applique les mêmes filtres que la vue liste_paiements et renvoie
    (queryset ordonné, liste de libellés de filtres actifs)."""
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()
    annee = (request.GET.get('annee') or '').strip()
    classe_id = (request.GET.get('classe_id') or '').strip()
    mode_id = (request.GET.get('mode_id') or '').strip()
    type_id = (request.GET.get('type_id') or '').strip()
    situation = (request.GET.get('situation') or '').strip()

    qs = (Paiement.objects
          .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole',
                          'type_paiement', 'mode_paiement')
          .exclude(statut='ANNULE')
          .order_by('eleve__classe__nom', 'eleve__nom', 'eleve__prenom', '-date_paiement'))
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    libelles = []

    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q) | Q(reference_externe__icontains=q)
            | Q(observations__icontains=q) | Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q) | Q(eleve__matricule__icontains=q)
        )
        libelles.append(f'Recherche « {q} »')
    if statut:
        qs = qs.filter(statut=statut)
        libelles.append(f'Statut : {statut}')
    if annee:
        qs = qs.filter(eleve__classe__annee_scolaire=annee)
        libelles.append(f'Année : {annee}')
    if classe_id.isdigit():
        qs = qs.filter(eleve__classe_id=int(classe_id))
        c = Classe.objects.filter(pk=int(classe_id)).first()
        libelles.append(f'Classe : {c.nom}' if c else 'Classe')
    if mode_id.isdigit():
        qs = qs.filter(mode_paiement_id=int(mode_id))
        m = ModePaiement.objects.filter(pk=int(mode_id)).first()
        libelles.append(f'Mode : {m.nom}' if m else 'Mode')
    if type_id.isdigit():
        qs = qs.filter(type_paiement_id=int(type_id))
        t = TypePaiement.objects.filter(pk=int(type_id)).first()
        libelles.append(f'Nature : {t.nom}' if t else 'Nature')

    if situation in ('retard', 'reste', 'solde'):
        today = date.today()
        exig = (
            Case(When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
                 default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
            + Case(When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
                   default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
            + Case(When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
                   default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
            + Case(When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
                   default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
        )
        paye = (F('frais_inscription_paye') + F('tranche_1_payee')
                + F('tranche_2_payee') + F('tranche_3_payee'))
        du_total = (F('frais_inscription_du') + F('tranche_1_due')
                    + F('tranche_2_due') + F('tranche_3_due'))
        eche = EcheancierPaiement.objects.annotate(
            _retard=ExpressionWrapper(exig - paye, output_field=DecimalField(max_digits=12, decimal_places=0)),
            _reste=ExpressionWrapper(du_total - paye, output_field=DecimalField(max_digits=12, decimal_places=0)),
        )
        if situation == 'retard':
            eche = eche.filter(_retard__gt=0)
            libelles.append('Niveau : en retard')
        elif situation == 'reste':
            eche = eche.filter(_reste__gt=0)
            libelles.append('Niveau : reste à payer')
        elif situation == 'solde':
            eche = eche.filter(_reste__lte=0)
            libelles.append('Niveau : soldé')
        qs = qs.filter(eleve_id__in=list(eche.values_list('eleve_id', flat=True)))

    return qs, libelles


@login_required
def export_paiements_filtres_excel(request):
    """Exporte en Excel la liste des paiements filtrée."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("Module openpyxl non installé", status=500)

    qs, libelles = filtrer_paiements(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Paiements filtrés'

    entete_font = Font(bold=True, color='FFFFFF')
    entete_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
    centre = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='DDDDDD')
    bordure = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:J1')
    ws['A1'] = 'LISTE DES PAIEMENTS' + (f" — {' | '.join(libelles)}" if libelles else '')
    ws['A1'].font = Font(bold=True, size=12, color='007BFF')
    ws['A1'].alignment = centre

    headers = ['N°', 'Matricule', 'Élève', 'Classe', 'Nature', 'Mode',
               'Date', 'N° Reçu', 'Statut', 'Montant (GNF)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = entete_font
        c.fill = entete_fill
        c.alignment = centre
        c.border = bordure

    ligne = 3
    total = Decimal('0')
    statut_labels = dict(Paiement.STATUT_CHOICES)
    for i, p in enumerate(qs.iterator(), 1):
        ligne += 1
        total += p.montant or Decimal('0')
        valeurs = [
            i, p.eleve.matricule or '', f"{p.eleve.prenom} {p.eleve.nom}",
            p.eleve.classe.nom if p.eleve.classe else '',
            p.type_paiement.nom if p.type_paiement else '',
            p.mode_paiement.nom if p.mode_paiement else '',
            p.date_paiement.strftime('%d/%m/%Y') if p.date_paiement else '',
            p.numero_recu or '', statut_labels.get(p.statut, p.statut),
            int(p.montant or 0),
        ]
        for col, val in enumerate(valeurs, 1):
            c = ws.cell(row=ligne, column=col, value=val)
            c.border = bordure
            if col == 10:
                c.number_format = '#,##0'

    ligne += 1
    c = ws.cell(row=ligne, column=9, value='TOTAL')
    c.font = Font(bold=True)
    m = ws.cell(row=ligne, column=10, value=int(total))
    m.font = Font(bold=True)
    m.number_format = '#,##0'

    for col, larg in zip('ABCDEFGHIJ', [5, 14, 28, 18, 16, 16, 12, 14, 12, 16]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="paiements_filtres_{date.today().isoformat()}.xlsx"')
    return response


@login_required
def export_paiements_filtres_pdf(request):
    """Exporte en PDF la liste des paiements filtrée."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    qs, libelles = filtrer_paiements(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
                            leftMargin=0.8 * cm, rightMargin=0.8 * cm)
    styles = getSampleStyleSheet()
    titre = ParagraphStyle('T', parent=styles['Heading1'], fontSize=13,
                           textColor=colors.HexColor('#007bff'), alignment=TA_CENTER, spaceAfter=2)
    sous = ParagraphStyle('S', parent=styles['Normal'], fontSize=9,
                          alignment=TA_CENTER, spaceAfter=6)

    elements = [Paragraph("<b>LISTE DES PAIEMENTS</b>", titre)]
    ligne_filtres = ' | '.join(libelles) if libelles else 'Aucun filtre'
    elements.append(Paragraph(
        f"Filtres : {ligne_filtres} — édité le {date.today().strftime('%d/%m/%Y')}", sous))

    data = [['N°', 'Matricule', 'Élève', 'Classe', 'Nature', 'Mode', 'Date', 'N° Reçu', 'Statut', 'Montant (GNF)']]
    total = Decimal('0')
    statut_labels = dict(Paiement.STATUT_CHOICES)
    for i, p in enumerate(qs.iterator(), 1):
        total += p.montant or Decimal('0')
        data.append([
            str(i), p.eleve.matricule or '', f"{p.eleve.prenom} {p.eleve.nom}",
            p.eleve.classe.nom if p.eleve.classe else '',
            p.type_paiement.nom if p.type_paiement else '',
            p.mode_paiement.nom if p.mode_paiement else '',
            p.date_paiement.strftime('%d/%m/%Y') if p.date_paiement else '',
            p.numero_recu or '', statut_labels.get(p.statut, p.statut),
            _fmt_gnf(p.montant),
        ])
    data.append(['', '', '', '', '', '', '', '', 'TOTAL', _fmt_gnf(total)])

    table = Table(
        data, repeatRows=1,
        colWidths=[1.0 * cm, 2.6 * cm, 6.0 * cm, 3.6 * cm, 3.0 * cm, 3.0 * cm,
                   2.2 * cm, 2.8 * cm, 2.2 * cm, 3.0 * cm])
    n = len(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f4f6f7')]),
        ('BACKGROUND', (0, n - 1), (-1, n - 1), colors.HexColor('#d6eaf8')),
        ('FONTNAME', (0, n - 1), (-1, n - 1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(
        f"Total : {len(data) - 2} paiement(s) — {_fmt_gnf(total)} GNF", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="paiements_filtres_{date.today().isoformat()}.pdf"')
    return response
