from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from datetime import date, datetime
from decimal import Decimal

from eleves.models import Classe
from eleves.utils_annee import get_annee_active
from paiements.models import Paiement
from paiements.allocation import (
    INSCRIPTION,
    TRANCHE_1,
    TRANCHE_2,
    TRANCHE_3,
    is_reinscription_payment,
    normalize_payment_type,
)
from paiements.payment_engine import situation_echeancier
from utilisateurs.utils import user_is_admin, user_school
from rapports.utils import _draw_header_and_watermark, _get_logo_path

# ReportLab
# ReportLab: fera l'objet d'un import différé dans la vue PDF


def _annee_vers_dates(annee_scolaire: str):
    try:
        deb, fin = annee_scolaire.split('-')
        an_deb = int(deb)
        an_fin = int(fin)
        return an_deb, an_fin
    except Exception:
        # Fallback année en cours, période des réinscriptions comprise.
        from .payment_engine import school_year_from_date
        deb, fin = school_year_from_date(timezone.now().date()).split('-')
        return int(deb), int(fin)


ZERO = Decimal('0')


def _taux_remise(remise, base):
    remise = Decimal(str(remise or 0))
    base = Decimal(str(base or 0))
    return (remise / base * Decimal('100')) if base > 0 else ZERO


def _format_taux(taux):
    """« 50 % » plutôt que « 50,0 % » quand le taux saisi est entier."""
    taux = Decimal(str(taux or 0))
    if taux == taux.to_integral_value():
        return f"{int(taux)} %"
    return f"{taux:.1f}".replace('.', ',') + ' %'


def _liens_remise_eleve(eleve, annee_scolaire):
    """Remises validées de l'élève sur l'année, avec la remise d'origine.

    Même périmètre que `remises_par_tranche` : on garde les paiements de
    l'année et les anciens paiements sans année tombant dans ses bornes.
    """
    from django.db.models import Q

    from paiements.models import PaiementRemise
    from paiements.payment_engine import school_year_bounds

    liens = PaiementRemise.objects.select_related('remise', 'paiement').filter(
        paiement__eleve=eleve,
        paiement__statut='VALIDE',
    )
    if annee_scolaire:
        debut, fin = school_year_bounds(annee_scolaire)
        legacy = Q(paiement__annee_scolaire='')
        if debut and fin:
            legacy &= Q(paiement__date_paiement__range=(debut, fin))
        liens = liens.filter(Q(paiement__annee_scolaire=annee_scolaire) | legacy)
    return list(liens)


def _taux_remise_affiche(liens, remise_totale, total_du):
    """(valeur, libellé, plusieurs_taux) du pourcentage de remise à afficher.

    On restitue le pourcentage réellement choisi par l'utilisateur au moment
    de la saisie. Le taux n'est recalculé que pour les remises exprimées en
    montant fixe, et alors sur la base retenue à la saisie plutôt que sur le
    total dû (qui inclut l'inscription, jamais remisée).
    """
    remise_totale = Decimal(str(remise_totale or 0))
    if remise_totale <= 0:
        return ZERO, '', False

    choisis = []
    base_fixe = ZERO
    for lien in liens:
        remise = getattr(lien, 'remise', None)
        if remise is not None and getattr(remise, 'type_remise', '') == 'POURCENTAGE':
            valeur = Decimal(str(getattr(remise, 'valeur', 0) or 0))
            if valeur > 0 and valeur not in choisis:
                choisis.append(valeur)
        else:
            base_fixe += Decimal(str(getattr(lien, 'montant_base', 0) or 0))

    if len(choisis) == 1:
        return choisis[0], _format_taux(choisis[0]), False
    if choisis:
        # Plusieurs taux saisis : on les affiche tous plutôt qu'une moyenne.
        return (
            max(choisis),
            ' + '.join(_format_taux(taux) for taux in choisis),
            True,
        )

    base = base_fixe if base_fixe > 0 else Decimal(str(total_du or 0))
    taux = _taux_remise(remise_totale, base)
    return taux, _format_taux(taux), False


def _situation_avec_remise(
    total_du, encaisse, remise, reste, *, avec_echeancier, taux_libelle=''
):
    """Libellé explicite, notamment lorsqu'une remise permet de solder l'élève."""
    if not avec_echeancier:
        situation = 'Sans échéancier'
    elif reste <= 0:
        situation = 'Soldé'
    elif encaisse + remise > 0:
        situation = 'Paiement partiel'
    else:
        situation = 'À payer'

    if remise > 0:
        montant = f"{remise:,.0f}".replace(',', ' ')
        detail = f"{montant} GNF ({taux_libelle})" if taux_libelle else f"{montant} GNF"
        if avec_echeancier and reste <= 0:
            return f"Soldé avec remise — {detail}"
        return f"{situation} — remise appliquée : {detail}"
    return situation


def _donnees_tranches_eleve(eleve, annee_scolaire):
    """Source unique des lignes PDF et Excel du rapport par classe."""
    valeurs = {
        'inscription': ZERO,
        'reinscription': ZERO,
        'tranche_1': ZERO,
        'tranche_2': ZERO,
        'tranche_3': ZERO,
        'total_du': ZERO,
        'encaisse': ZERO,
        'remise': ZERO,
        'reste': ZERO,
    }
    echeancier = getattr(eleve, 'echeancier', None)
    avec_echeancier = bool(
        echeancier is not None
        and (not annee_scolaire or echeancier.annee_scolaire == annee_scolaire)
    )

    if avec_echeancier:
        situation = situation_echeancier(echeancier)
        payes_affiches = situation.get('payes_affiches', situation['payes'])
        admission = payes_affiches[INSCRIPTION]
        if getattr(echeancier, 'nature_frais', 'INSCRIPTION') == 'REINSCRIPTION':
            valeurs['reinscription'] = admission
        else:
            valeurs['inscription'] = admission
        valeurs.update({
            'tranche_1': payes_affiches[TRANCHE_1],
            'tranche_2': payes_affiches[TRANCHE_2],
            'tranche_3': payes_affiches[TRANCHE_3],
            'total_du': situation['total_du'],
            'encaisse': situation['total_encaisse'],
            'remise': situation['total_remises'],
            'reste': situation['solde_restant'],
        })
    else:
        # Les montants restent visibles même pour un ancien dossier sans échéancier.
        paiements = Paiement.objects.filter(eleve=eleve, statut='VALIDE')
        if annee_scolaire:
            paiements = paiements.filter(annee_scolaire=annee_scolaire)
        paiements = list(paiements.select_related('type_paiement').prefetch_related('remises'))
        for paiement in paiements:
            montant = Decimal(str(paiement.montant or 0))
            valeurs['encaisse'] += montant
            type_nom = getattr(paiement.type_paiement, 'nom', '')
            type_normalise = normalize_payment_type(type_nom)
            if is_reinscription_payment(type_nom):
                valeurs['reinscription'] += montant
            elif 'inscription' in type_normalise:
                valeurs['inscription'] += montant
            elif 'tranche 1' in type_normalise:
                valeurs['tranche_1'] += montant
            elif 'tranche 2' in type_normalise:
                valeurs['tranche_2'] += montant
            elif 'tranche 3' in type_normalise:
                valeurs['tranche_3'] += montant
            valeurs['remise'] += sum(
                (Decimal(str(lien.montant_remise or 0)) for lien in paiement.remises.all()),
                ZERO,
            )

    annee_ref = annee_scolaire or (
        getattr(echeancier, 'annee_scolaire', '') if avec_echeancier else ''
    )
    liens_remise = _liens_remise_eleve(eleve, annee_ref) if valeurs['remise'] > 0 else []
    taux, taux_libelle, plusieurs_taux = _taux_remise_affiche(
        liens_remise, valeurs['remise'], valeurs['total_du']
    )
    valeurs['taux_remise'] = taux
    valeurs['taux_remise_libelle'] = taux_libelle
    valeurs['taux_remise_multiple'] = plusieurs_taux
    valeurs['situation'] = _situation_avec_remise(
        valeurs['total_du'], valeurs['encaisse'], valeurs['remise'], valeurs['reste'],
        avec_echeancier=avec_echeancier, taux_libelle=taux_libelle,
    )
    return valeurs


@login_required
def export_tranches_par_classe_pdf(request):
    """Export PDF des tranches par classe avec logo entête et filigrane.

    Filtres GET:
    - ecole: id de l'école
    - classe: id de la classe
    - annee_scolaire: ex '2024-2025'

    Respecte la séparation par école pour les non-admins.
    """
    # Contrôle d'accès: Admin ou Comptable uniquement
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Lecture et validation des paramètres
    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    # Scope classes (filtrées par année active)
    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    annee_active = get_annee_active(request, ecole_user) if ecole_user else None
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    if annee_active and not annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_active)
    elif annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_scolaire)

    # Anti-abus: limiter le nombre de classes exportées en une requête
    classes = list(classes.order_by('ecole__nom', 'niveau', 'nom')[:200])
    ecoles = {getattr(classe, 'ecole_id', None) for classe in classes}
    ecole_rapport = classes[0].ecole if classes and len(ecoles) == 1 else None

    # Préparer réponse PDF
    response = HttpResponse(content_type='application/pdf')
    suffix = datetime.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="tranches_par_classe_{suffix}.pdf"'

    # Import différé de ReportLab pour éviter les erreurs si non installé
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab n'est pas installé. Veuillez exécuter: pip install reportlab", status=500)

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=60, bottomMargin=30
    )
    elements = []
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=7, leading=8)
    header_cell = ParagraphStyle(
        'HeaderCell', parent=cell, alignment=1, fontName='Helvetica-Bold'
    )
    # Les montants passent par un Paragraph : ReportLab les confine alors à la
    # largeur de leur colonne au lieu de les laisser déborder sur la suivante.
    num_cell = ParagraphStyle('NumCell', parent=cell, alignment=2)

    titre = 'Tranches par classe'
    if annee_scolaire:
        titre += f" – Année {annee_scolaire}"
    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Spacer(1, 0.5*cm))

    header = [
        'Élève', 'Inscription payée', 'Réinscription payée', 'Tranche 1 payée', 'Tranche 2 payée', 'Tranche 3 payée',
        'Total dû', 'Encaissé', 'Remise', 'Remise (%)', 'Reste', 'Situation / précision'
    ]

    def P(x):
        return Paragraph(str(x or ''), cell)

    def N(montant):
        """Montant GNF cadré à droite, insécable pour rester sur une ligne."""
        texte = f"{montant:,.0f}".replace(',', ' ')
        return Paragraph(texte, num_cell)

    def Pct(libelle):
        """Taux saisi par l'utilisateur, tiret quand aucune remise n'existe."""
        return Paragraph((libelle or '—').replace(' %', ' %'), num_cell)

    # Parcours des classes
    for classe in classes:
        # Titre de la classe
        titre_classe = f"Classe: {classe.nom} – {getattr(classe.ecole, 'nom', '')}"
        elements.append(Paragraph(titre_classe, styles['Heading2']))
        elements.append(Spacer(1, 0.2*cm))

        data = [[Paragraph(label, header_cell) for label in header]]

        # Élèves de la classe
        # Utiliser le related_name défini sur Eleve.classe = 'eleves'
        eleves = getattr(classe, 'eleves', None)
        if eleves is not None:
            eleves = eleves.all().order_by('nom', 'prenom')
        else:
            eleves = []

        for e in eleves:
            valeurs = _donnees_tranches_eleve(e, annee_scolaire)

            # Construire le nom de l'élève sans déclencher d'erreur si un attribut manque
            nom_affiche = getattr(e, 'nom_complet', None) or f"{getattr(e, 'prenom', '')} {getattr(e, 'nom', '')}".strip()
            data.append([
                P(nom_affiche),
                N(valeurs['inscription']),
                N(valeurs['reinscription']),
                N(valeurs['tranche_1']),
                N(valeurs['tranche_2']),
                N(valeurs['tranche_3']),
                N(valeurs['total_du']),
                N(valeurs['encaisse']),
                N(valeurs['remise']),
                Pct(valeurs['taux_remise_libelle']),
                N(valeurs['reste']),
                P(valeurs['situation']),
            ])

        # Construire la table pour la classe
        # Somme = 28,2 cm, soit la largeur utile d'un A4 paysage (marges 20 pt).
        col_widths = (
            [3.8*cm] + [1.7*cm] * 5
            + [2.1*cm, 2.1*cm, 2.2*cm, 2.1*cm, 2.1*cm, 5.3*cm]
        )
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('FONTSIZE', (0,1), (-1,-1), 7),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('ALIGN', (-1,1), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.6*cm))

    # Construire le document avec le logo de l'école en en-tête et en filigrane.
    def _header_wrapper(canvas, doc_):
        _draw_header_and_watermark(
            canvas, doc_, ecole=ecole_rapport, titre_override='Tranches par classe',
        )

    doc.build(elements, onFirstPage=_header_wrapper, onLaterPages=_header_wrapper)
    return response

@login_required
def export_tranches_par_classe_excel(request):
    """Export Excel des tranches, encaissements, remises, soldes et situations.

    Filtres GET facultatifs: ecole, classe/classe_id, annee_scolaire.
    Respecte la séparation par école pour non-admin.
    """
    # Contrôle d'accès
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Import openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
    except Exception:
        return HttpResponse("OpenPyXL n'est pas installé. Veuillez exécuter: pip install openpyxl", status=500)

    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    annee_active_xl = get_annee_active(request, ecole_user) if ecole_user else None
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    if annee_active_xl and not annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_active_xl)
    elif annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_scolaire)
    classes = list(classes.order_by('ecole__nom', 'niveau', 'nom')[:200])
    ecoles = {getattr(classe, 'ecole_id', None) for classe in classes}
    ecole_rapport = classes[0].ecole if classes and len(ecoles) == 1 else None

    def ajouter_logo(ws, ecole):
        logo_path = _get_logo_path(ecole) if ecole else ''
        if not logo_path:
            return
        try:
            logo = ExcelImage(logo_path)
            ratio = (logo.width / logo.height) if logo.height else 1
            logo.height = 30
            logo.width = 30 * ratio
            ws.add_image(logo, 'A1')
            ws.row_dimensions[1].height = 34
        except Exception:
            pass

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = 'Index'
    ws_index.append(['', 'Tranches par classe', f"Année: {annee_scolaire}" if annee_scolaire else ''])
    ws_index.append(['Écoles / Classes listées:'])
    ajouter_logo(ws_index, ecole_rapport)

    headers = [
        'Élève', 'Inscription payée', 'Réinscription payée', 'Tranche 1 payée',
        'Tranche 2 payée', 'Tranche 3 payée', 'Total dû', 'Encaissé',
        'Remise', 'Remise (%)', 'Reste', 'Situation / précision',
    ]

    for idx, classe in enumerate(classes, start=1):
        sheet_name = f"{classe.nom[:25]}"  # Limite Excel <=31
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['', f"Classe: {classe.nom} – {getattr(classe.ecole, 'nom', '')}"])
        ws.append(headers)
        ajouter_logo(ws, classe.ecole)

        eleves_mgr = getattr(classe, 'eleves', None)
        eleves = eleves_mgr.all().order_by('nom', 'prenom') if eleves_mgr is not None else []

        for e in eleves:
            valeurs = _donnees_tranches_eleve(e, annee_scolaire)

            # Un seul taux saisi : on garde une vraie valeur numérique.
            # Plusieurs taux : le texte « 50 % + 25 % » évite d'inventer une moyenne.
            if valeurs['remise'] <= 0:
                taux_cellule = None
            elif valeurs['taux_remise_multiple']:
                taux_cellule = valeurs['taux_remise_libelle']
            else:
                taux_cellule = float(round(valeurs['taux_remise'], 1))

            ws.append([
                getattr(e, 'nom_complet', f"{e.prenom} {e.nom}"),
                int(valeurs['inscription']), int(valeurs['reinscription']),
                int(valeurs['tranche_1']), int(valeurs['tranche_2']),
                int(valeurs['tranche_3']), int(valeurs['total_du']),
                int(valeurs['encaisse']), int(valeurs['remise']),
                taux_cellule, int(valeurs['reste']),
                valeurs['situation'],
            ])

        # Ajuster largeur colonnes simple
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = (
                34 if col == 12 else (22 if col == 1 else 16)
            )
        for row in range(3, ws.max_row + 1):
            # Chaque montant garde sa colonne : format nombre + alignement à
            # droite, et le taux affiche son symbole sans déborder sur le reste.
            for col in range(2, 10):
                cellule = ws.cell(row, col)
                cellule.number_format = '#,##0'
                cellule.alignment = Alignment(horizontal='right')
            ws.cell(row, 11).number_format = '#,##0'
            ws.cell(row, 11).alignment = Alignment(horizontal='right')
            taux = ws.cell(row, 10)
            taux.alignment = Alignment(horizontal='right')
            if isinstance(taux.value, (int, float)):
                taux.number_format = '0.0" %"'
            ws.cell(row, 12).alignment = Alignment(horizontal='left', wrap_text=True)

        # Index line
        ws_index.append([getattr(classe.ecole, 'nom', ''), classe.nom, sheet_name])

    # Supprimer la feuille par défaut si vide
    if ws_index.max_row == 2:
        ws_index.append(['Aucune classe'])

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    suffix = datetime.now().strftime('%Y%m%d')
    filename = f'tranches_par_classe_{suffix}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
