"""
Rapport comptable complet en PDF et Excel.

Sections, groupées par classe pour tout l'établissement (ou une classe):
  1. Paiements validés (élève, reçu, date, type, mode, montant)
  2. Retards de paiement (exigible - payé effectif, remises incluses)
  3. Relances (canal, statut, solde estimé)
  4. Statistiques de rapprochement par mode (Espèces, Chèque, Virement, ...)
     - totaux globaux par mode + matrice classe x mode

Filtres GET: ?classe_id=  ?du=AAAA-MM-JJ  ?au=AAAA-MM-JJ
"""
import io
import re
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import (
    Case, When, Value, F, Q, Sum, Count, DecimalField, ExpressionWrapper,
)
from django.db.models.functions import Coalesce, Least
from django.http import HttpResponse

from eleves.models import Classe, Eleve
from utilisateurs.utils import filter_by_user_school, user_school
from .models import Paiement, EcheancierPaiement, Relance


def _fmt_gnf(v):
    """1234567 -> '1 234 567'"""
    try:
        return f"{int(v):,}".replace(',', ' ')
    except (TypeError, ValueError):
        return '0'


def _collecter_donnees(request):
    """Rassemble toutes les données du rapport, filtrées par école/classe/dates."""
    classe_id = (request.GET.get('classe_id') or '').strip()
    du = (request.GET.get('du') or '').strip()
    au = (request.GET.get('au') or '').strip()

    classes = Classe.objects.select_related('ecole').order_by('niveau', 'nom')
    classes = filter_by_user_school(classes, request.user, 'ecole')
    if classe_id.isdigit():
        classes = classes.filter(pk=int(classe_id))
    classes = list(classes)
    classes_ids = [c.id for c in classes]

    ecole = classes[0].ecole if classes else user_school(request.user)

    # ── 1. Paiements validés ──────────────────────────────────────────
    paiements = (Paiement.objects
                 .filter(statut='VALIDE', eleve__classe_id__in=classes_ids)
                 .select_related('eleve', 'eleve__classe', 'type_paiement', 'mode_paiement')
                 .order_by('eleve__classe__nom', 'date_paiement', 'numero_recu'))
    if du:
        paiements = paiements.filter(date_paiement__gte=du)
    if au:
        paiements = paiements.filter(date_paiement__lte=au)
    paiements = list(paiements)

    paiements_par_classe = {}
    for p in paiements:
        paiements_par_classe.setdefault(p.eleve.classe_id, []).append(p)

    # ── 2. Retards (même logique que rapport_retards) ─────────────────
    today = date.today()
    exigible_expr = (
        Case(When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
             default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
        + Case(When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
               default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
        + Case(When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
               default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
        + Case(When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
               default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0))
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee')
        + F('tranche_2_payee') + F('tranche_3_payee') + remises_applicables
    )
    retard_expr = ExpressionWrapper(
        exigible_expr - paye_effectif_expr,
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    retards = (EcheancierPaiement.objects
               .filter(eleve__classe_id__in=classes_ids, eleve__statut='ACTIF')
               .select_related('eleve', 'eleve__classe')
               .annotate(retard=retard_expr, exigible=exigible_expr)
               .filter(retard__gt=0)
               .order_by('eleve__classe__nom', '-retard'))
    retards = list(retards)

    retards_par_classe = {}
    for e in retards:
        retards_par_classe.setdefault(e.eleve.classe_id, []).append(e)

    # ── 3. Relances ───────────────────────────────────────────────────
    relances = (Relance.objects
                .filter(eleve__classe_id__in=classes_ids)
                .select_related('eleve', 'eleve__classe')
                .order_by('eleve__classe__nom', '-date_creation'))
    if du:
        relances = relances.filter(date_creation__date__gte=du)
    if au:
        relances = relances.filter(date_creation__date__lte=au)
    relances = list(relances)

    relances_par_classe = {}
    for r in relances:
        relances_par_classe.setdefault(r.eleve.classe_id, []).append(r)

    # ── 4. Statistiques de rapprochement par mode ─────────────────────
    stats_modes = {}   # mode -> {'nb': n, 'total': montant}
    matrice = {}       # (classe_id, mode) -> montant
    for p in paiements:
        mode = p.mode_paiement.nom if p.mode_paiement else 'Autre'
        s = stats_modes.setdefault(mode, {'nb': 0, 'total': Decimal('0')})
        s['nb'] += 1
        s['total'] += p.montant or Decimal('0')
        cle = (p.eleve.classe_id, mode)
        matrice[cle] = matrice.get(cle, Decimal('0')) + (p.montant or Decimal('0'))
    modes_ordonnes = sorted(stats_modes.keys())

    return {
        'ecole': ecole,
        'classes': classes,
        'du': du, 'au': au,
        'paiements_par_classe': paiements_par_classe,
        'retards_par_classe': retards_par_classe,
        'relances_par_classe': relances_par_classe,
        'stats_modes': stats_modes,
        'modes_ordonnes': modes_ordonnes,
        'matrice': matrice,
        'total_paiements': sum((p.montant or Decimal('0')) for p in paiements),
        'nb_paiements': len(paiements),
        'total_retards': sum((e.retard or Decimal('0')) for e in retards),
        'nb_retards': len(retards),
        'nb_relances': len(relances),
    }


def _periode_str(d):
    if d['du'] and d['au']:
        return f"Période: du {d['du']} au {d['au']}"
    if d['du']:
        return f"Période: depuis le {d['du']}"
    if d['au']:
        return f"Période: jusqu'au {d['au']}"
    return "Période: toute l'année"


@login_required
def export_comptabilite_pdf(request):
    """Rapport comptable complet en PDF (paysage)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    d = _collecter_donnees(request)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
                            leftMargin=0.8 * cm, rightMargin=0.8 * cm)
    styles = getSampleStyleSheet()
    titre = ParagraphStyle('T', parent=styles['Heading1'], fontSize=14,
                           textColor=colors.HexColor('#007bff'), alignment=TA_CENTER, spaceAfter=2)
    sous_titre = ParagraphStyle('ST', parent=styles['Normal'], fontSize=10,
                                alignment=TA_CENTER, spaceAfter=8)
    section = ParagraphStyle('S', parent=styles['Heading2'], fontSize=12,
                             textColor=colors.HexColor('#1a5276'), spaceBefore=8, spaceAfter=4)
    classe_style = ParagraphStyle('C', parent=styles['Heading3'], fontSize=10, spaceBefore=6, spaceAfter=2)

    def style_tableau(nb_lignes):
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f7')]),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('BACKGROUND', (0, nb_lignes - 1), (-1, nb_lignes - 1), colors.HexColor('#d6eaf8')),
            ('FONTNAME', (0, nb_lignes - 1), (-1, nb_lignes - 1), 'Helvetica-Bold'),
        ])

    elements = []
    nom_ecole = d['ecole'].nom.upper() if d['ecole'] else "ÉTABLISSEMENT"
    elements.append(Paragraph(f"<b>{nom_ecole}</b>", titre))
    elements.append(Paragraph(
        f"RAPPORT COMPTABLE — {_periode_str(d)} — édité le {date.today().strftime('%d/%m/%Y')}",
        sous_titre))

    # Synthèse générale
    elements.append(Table([
        ['Paiements validés', 'Montant encaissé (GNF)', 'Élèves en retard', 'Montant en retard (GNF)', 'Relances'],
        [str(d['nb_paiements']), _fmt_gnf(d['total_paiements']), str(d['nb_retards']),
         _fmt_gnf(d['total_retards']), str(d['nb_relances'])],
    ], colWidths=[5.4 * cm] * 5, style=TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])))

    # ── Section 1: Paiements par classe ───────────────────────────────
    elements.append(Paragraph("1. PAIEMENTS PAR CLASSE", section))
    for classe in d['classes']:
        lst = d['paiements_par_classe'].get(classe.id)
        if not lst:
            continue
        elements.append(Paragraph(f"{classe.nom} — {len(lst)} paiement(s)", classe_style))
        data = [['N° Reçu', 'Date', 'Matricule', 'Élève', 'Type', 'Mode', 'Montant (GNF)']]
        sous_total = Decimal('0')
        for p in lst:
            sous_total += p.montant or Decimal('0')
            data.append([
                p.numero_recu, p.date_paiement.strftime('%d/%m/%Y'),
                p.eleve.matricule or '', f"{p.eleve.prenom} {p.eleve.nom}",
                p.type_paiement.nom if p.type_paiement else '',
                p.mode_paiement.nom if p.mode_paiement else '',
                _fmt_gnf(p.montant),
            ])
        data.append(['', '', '', '', '', 'SOUS-TOTAL', _fmt_gnf(sous_total)])
        elements.append(Table(
            data, repeatRows=1,
            colWidths=[3.2 * cm, 2.4 * cm, 3.0 * cm, 7.4 * cm, 4.4 * cm, 3.6 * cm, 3.4 * cm],
            style=style_tableau(len(data))))
    if not d['paiements_par_classe']:
        elements.append(Paragraph("Aucun paiement validé sur la période.", styles['Normal']))

    # ── Section 2: Retards par classe ─────────────────────────────────
    elements.append(PageBreak())
    elements.append(Paragraph("2. RETARDS DE PAIEMENT PAR CLASSE", section))
    for classe in d['classes']:
        lst = d['retards_par_classe'].get(classe.id)
        if not lst:
            continue
        total_cl = sum((e.retard or Decimal('0')) for e in lst)
        elements.append(Paragraph(
            f"{classe.nom} — {len(lst)} élève(s) en retard — {_fmt_gnf(total_cl)} GNF", classe_style))
        data = [['Matricule', 'Élève', 'Exigible à ce jour (GNF)', 'Payé + remises (GNF)', 'Retard (GNF)']]
        for e in lst:
            paye_eff = (e.exigible or Decimal('0')) - (e.retard or Decimal('0'))
            data.append([
                e.eleve.matricule or '', f"{e.eleve.prenom} {e.eleve.nom}",
                _fmt_gnf(e.exigible), _fmt_gnf(paye_eff), _fmt_gnf(e.retard),
            ])
        data.append(['', 'TOTAL CLASSE', '', '', _fmt_gnf(total_cl)])
        elements.append(Table(
            data, repeatRows=1,
            colWidths=[3.6 * cm, 9.4 * cm, 5.0 * cm, 5.0 * cm, 4.4 * cm],
            style=style_tableau(len(data))))
    if not d['retards_par_classe']:
        elements.append(Paragraph("Aucun retard de paiement.", styles['Normal']))

    # ── Section 3: Relances par classe ────────────────────────────────
    elements.append(PageBreak())
    elements.append(Paragraph("3. RELANCES PAR CLASSE", section))
    canaux = dict(Relance.CANAL_CHOICES)
    statuts_rel = dict(Relance.STATUT_CHOICES)
    for classe in d['classes']:
        lst = d['relances_par_classe'].get(classe.id)
        if not lst:
            continue
        elements.append(Paragraph(f"{classe.nom} — {len(lst)} relance(s)", classe_style))
        data = [['Date', 'Matricule', 'Élève', 'Canal', 'Statut', 'Solde estimé (GNF)']]
        for r in lst:
            data.append([
                r.date_creation.strftime('%d/%m/%Y'),
                r.eleve.matricule or '', f"{r.eleve.prenom} {r.eleve.nom}",
                canaux.get(r.canal, r.canal), statuts_rel.get(r.statut, r.statut),
                _fmt_gnf(r.solde_estime),
            ])
        data.append(['', '', '', '', 'TOTAL', _fmt_gnf(sum((r.solde_estime or Decimal('0')) for r in lst))])
        elements.append(Table(
            data, repeatRows=1,
            colWidths=[2.6 * cm, 3.4 * cm, 9.0 * cm, 3.6 * cm, 3.6 * cm, 5.2 * cm],
            style=style_tableau(len(data))))
    if not d['relances_par_classe']:
        elements.append(Paragraph("Aucune relance enregistrée sur la période.", styles['Normal']))

    # ── Section 4: Rapprochement par mode ─────────────────────────────
    elements.append(PageBreak())
    elements.append(Paragraph("4. STATISTIQUES DE RAPPROCHEMENT PAR MODE DE PAIEMENT", section))
    if d['modes_ordonnes']:
        data = [['Mode de paiement', 'Nombre de paiements', 'Montant total (GNF)', '% du total']]
        for mode in d['modes_ordonnes']:
            s = d['stats_modes'][mode]
            pct = (s['total'] / d['total_paiements'] * 100) if d['total_paiements'] else 0
            data.append([mode, str(s['nb']), _fmt_gnf(s['total']), f"{pct:.1f}%"])
        data.append(['TOTAL GÉNÉRAL', str(d['nb_paiements']), _fmt_gnf(d['total_paiements']), '100%'])
        elements.append(Table(data, repeatRows=1,
                              colWidths=[8.0 * cm, 5.0 * cm, 6.0 * cm, 4.0 * cm],
                              style=style_tableau(len(data))))

        # Matrice classe x mode
        elements.append(Paragraph("Répartition par classe et par mode", classe_style))
        entete = ['Classe'] + d['modes_ordonnes'] + ['Total classe']
        data = [entete]
        for classe in d['classes']:
            if classe.id not in d['paiements_par_classe']:
                continue
            ligne = [classe.nom]
            total_cl = Decimal('0')
            for mode in d['modes_ordonnes']:
                v = d['matrice'].get((classe.id, mode), Decimal('0'))
                total_cl += v
                ligne.append(_fmt_gnf(v) if v else '-')
            ligne.append(_fmt_gnf(total_cl))
            data.append(ligne)
        totaux = ['TOTAL']
        for mode in d['modes_ordonnes']:
            totaux.append(_fmt_gnf(d['stats_modes'][mode]['total']))
        totaux.append(_fmt_gnf(d['total_paiements']))
        data.append(totaux)
        nb_col = len(entete)
        largeur_dispo = 27.5 - 6.0
        elements.append(Table(
            data, repeatRows=1,
            colWidths=[6.0 * cm] + [max(2.4, largeur_dispo / (nb_col - 1)) * cm] * (nb_col - 1),
            style=style_tableau(len(data))))
    else:
        elements.append(Paragraph("Aucun paiement sur la période.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    suffixe = re.sub(r'[^\w-]', '_', d['classes'][0].nom) if len(d['classes']) == 1 else 'etablissement'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="rapport_comptable_{suffixe}_{date.today().isoformat()}.pdf"')
    return response


@login_required
def export_comptabilite_excel(request):
    """Rapport comptable complet en Excel (4 feuilles)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("Module openpyxl non installé", status=500)

    d = _collecter_donnees(request)
    wb = openpyxl.Workbook()

    entete_font = Font(bold=True, color='FFFFFF')
    entete_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
    classe_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    total_font = Font(bold=True)
    centre = Alignment(horizontal='center', vertical='center')
    bordure = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    def ecrire_entetes(ws, ligne, entetes):
        for col, val in enumerate(entetes, 1):
            c = ws.cell(row=ligne, column=col, value=val)
            c.font = entete_font
            c.fill = entete_fill
            c.alignment = centre
            c.border = bordure

    def titre_feuille(ws, texte, nb_col):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_col)
        c = ws.cell(row=1, column=1, value=texte)
        c.font = Font(bold=True, size=13, color='007BFF')
        c.alignment = centre
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_col)
        c2 = ws.cell(row=2, column=1, value=_periode_str(d))
        c2.alignment = centre

    nom_ecole = d['ecole'].nom.upper() if d['ecole'] else "ÉTABLISSEMENT"

    # ── Feuille 1: Paiements ──────────────────────────────────────────
    ws = wb.active
    ws.title = 'Paiements'
    titre_feuille(ws, f"{nom_ecole} — PAIEMENTS PAR CLASSE", 7)
    ligne = 4
    ecrire_entetes(ws, ligne, ['Classe', 'N° Reçu', 'Date', 'Matricule', 'Élève', 'Type', 'Mode', 'Montant (GNF)'])
    for classe in d['classes']:
        lst = d['paiements_par_classe'].get(classe.id)
        if not lst:
            continue
        sous_total = Decimal('0')
        for p in lst:
            ligne += 1
            sous_total += p.montant or Decimal('0')
            valeurs = [classe.nom, p.numero_recu, p.date_paiement.strftime('%d/%m/%Y'),
                       p.eleve.matricule or '', f"{p.eleve.prenom} {p.eleve.nom}",
                       p.type_paiement.nom if p.type_paiement else '',
                       p.mode_paiement.nom if p.mode_paiement else '', int(p.montant or 0)]
            for col, val in enumerate(valeurs, 1):
                c = ws.cell(row=ligne, column=col, value=val)
                c.border = bordure
                if col == 8:
                    c.number_format = '#,##0'
        ligne += 1
        c = ws.cell(row=ligne, column=1, value=f"SOUS-TOTAL {classe.nom}")
        c.font = total_font
        c.fill = classe_fill
        m = ws.cell(row=ligne, column=8, value=int(sous_total))
        m.font = total_font
        m.fill = classe_fill
        m.number_format = '#,##0'
    ligne += 2
    c = ws.cell(row=ligne, column=1, value='TOTAL GÉNÉRAL')
    c.font = Font(bold=True, size=12)
    m = ws.cell(row=ligne, column=8, value=int(d['total_paiements']))
    m.font = Font(bold=True, size=12)
    m.number_format = '#,##0'
    for col, larg in zip('ABCDEFGH', [22, 14, 12, 14, 30, 18, 16, 16]):
        ws.column_dimensions[col].width = larg

    # ── Feuille 2: Retards ────────────────────────────────────────────
    ws = wb.create_sheet('Retards')
    titre_feuille(ws, f"{nom_ecole} — RETARDS DE PAIEMENT PAR CLASSE", 6)
    ligne = 4
    ecrire_entetes(ws, ligne, ['Classe', 'Matricule', 'Élève', 'Exigible à ce jour (GNF)',
                               'Payé + remises (GNF)', 'Retard (GNF)'])
    for classe in d['classes']:
        lst = d['retards_par_classe'].get(classe.id)
        if not lst:
            continue
        total_cl = Decimal('0')
        for e in lst:
            ligne += 1
            total_cl += e.retard or Decimal('0')
            paye_eff = (e.exigible or Decimal('0')) - (e.retard or Decimal('0'))
            valeurs = [classe.nom, e.eleve.matricule or '', f"{e.eleve.prenom} {e.eleve.nom}",
                       int(e.exigible or 0), int(paye_eff), int(e.retard or 0)]
            for col, val in enumerate(valeurs, 1):
                c = ws.cell(row=ligne, column=col, value=val)
                c.border = bordure
                if col >= 4:
                    c.number_format = '#,##0'
        ligne += 1
        c = ws.cell(row=ligne, column=1, value=f"TOTAL {classe.nom}")
        c.font = total_font
        c.fill = classe_fill
        m = ws.cell(row=ligne, column=6, value=int(total_cl))
        m.font = total_font
        m.fill = classe_fill
        m.number_format = '#,##0'
    ligne += 2
    c = ws.cell(row=ligne, column=1, value='TOTAL GÉNÉRAL DES RETARDS')
    c.font = Font(bold=True, size=12)
    m = ws.cell(row=ligne, column=6, value=int(d['total_retards']))
    m.font = Font(bold=True, size=12)
    m.number_format = '#,##0'
    for col, larg in zip('ABCDEF', [22, 14, 30, 20, 20, 18]):
        ws.column_dimensions[col].width = larg

    # ── Feuille 3: Relances ───────────────────────────────────────────
    ws = wb.create_sheet('Relances')
    titre_feuille(ws, f"{nom_ecole} — RELANCES PAR CLASSE", 6)
    ligne = 4
    ecrire_entetes(ws, ligne, ['Classe', 'Date', 'Matricule', 'Élève', 'Canal', 'Statut', 'Solde estimé (GNF)'])
    canaux = dict(Relance.CANAL_CHOICES)
    statuts_rel = dict(Relance.STATUT_CHOICES)
    for classe in d['classes']:
        lst = d['relances_par_classe'].get(classe.id)
        if not lst:
            continue
        for r in lst:
            ligne += 1
            valeurs = [classe.nom, r.date_creation.strftime('%d/%m/%Y'),
                       r.eleve.matricule or '', f"{r.eleve.prenom} {r.eleve.nom}",
                       canaux.get(r.canal, r.canal), statuts_rel.get(r.statut, r.statut),
                       int(r.solde_estime or 0)]
            for col, val in enumerate(valeurs, 1):
                c = ws.cell(row=ligne, column=col, value=val)
                c.border = bordure
                if col == 7:
                    c.number_format = '#,##0'
    for col, larg in zip('ABCDEFG', [22, 12, 14, 30, 14, 14, 18]):
        ws.column_dimensions[col].width = larg

    # ── Feuille 4: Rapprochement par mode ─────────────────────────────
    ws = wb.create_sheet('Rapprochement modes')
    nb_col = max(4, len(d['modes_ordonnes']) + 2)
    titre_feuille(ws, f"{nom_ecole} — RAPPROCHEMENT PAR MODE DE PAIEMENT", nb_col)
    ligne = 4
    ecrire_entetes(ws, ligne, ['Mode de paiement', 'Nombre', 'Montant total (GNF)', '% du total'])
    for mode in d['modes_ordonnes']:
        ligne += 1
        s = d['stats_modes'][mode]
        pct = float(s['total'] / d['total_paiements'] * 100) if d['total_paiements'] else 0.0
        for col, val in enumerate([mode, s['nb'], int(s['total']), round(pct, 1)], 1):
            c = ws.cell(row=ligne, column=col, value=val)
            c.border = bordure
            if col == 3:
                c.number_format = '#,##0'
    ligne += 1
    c = ws.cell(row=ligne, column=1, value='TOTAL GÉNÉRAL')
    c.font = total_font
    ws.cell(row=ligne, column=2, value=d['nb_paiements']).font = total_font
    m = ws.cell(row=ligne, column=3, value=int(d['total_paiements']))
    m.font = total_font
    m.number_format = '#,##0'

    # Matrice classe x mode
    ligne += 3
    ecrire_entetes(ws, ligne, ['Classe'] + d['modes_ordonnes'] + ['Total classe'])
    for classe in d['classes']:
        if classe.id not in d['paiements_par_classe']:
            continue
        ligne += 1
        ws.cell(row=ligne, column=1, value=classe.nom).border = bordure
        total_cl = Decimal('0')
        for i, mode in enumerate(d['modes_ordonnes'], 2):
            v = d['matrice'].get((classe.id, mode), Decimal('0'))
            total_cl += v
            c = ws.cell(row=ligne, column=i, value=int(v))
            c.border = bordure
            c.number_format = '#,##0'
        c = ws.cell(row=ligne, column=len(d['modes_ordonnes']) + 2, value=int(total_cl))
        c.border = bordure
        c.font = total_font
        c.number_format = '#,##0'
    ligne += 1
    ws.cell(row=ligne, column=1, value='TOTAL').font = total_font
    for i, mode in enumerate(d['modes_ordonnes'], 2):
        c = ws.cell(row=ligne, column=i, value=int(d['stats_modes'][mode]['total']))
        c.font = total_font
        c.number_format = '#,##0'
    c = ws.cell(row=ligne, column=len(d['modes_ordonnes']) + 2, value=int(d['total_paiements']))
    c.font = total_font
    c.number_format = '#,##0'
    ws.column_dimensions['A'].width = 24
    for i in range(2, len(d['modes_ordonnes']) + 3):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    suffixe = re.sub(r'[^\w-]', '_', d['classes'][0].nom) if len(d['classes']) == 1 else 'etablissement'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="rapport_comptable_{suffixe}_{date.today().isoformat()}.xlsx"')
    return response
