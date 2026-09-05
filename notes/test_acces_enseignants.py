from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from eleves.models import Classe, Ecole, Eleve, Responsable
from salaires.models import Enseignant, AffectationClasse
from .models import (AccesEnseignantTemporaire, ClasseNote, MatiereNote,
                     NoteMensuelle, CompositionNote, AppreciationMaternelle)


@override_settings(
    MIDDLEWARE=[m for m in settings.MIDDLEWARE if m != 'ecole_moderne.licence_middleware.LicenceMiddleware'],
    PASSWORD_HASHERS=['django.contrib.auth.hashers.MD5PasswordHasher'],
)
class AccesEnseignantsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.ecole = Ecole.objects.create(nom='École test accès', adresse='Conakry', telephone='+224620123450', directeur='Direction')
        cls.autre_ecole = Ecole.objects.create(nom='Autre école', adresse='Conakry', telephone='+224620123451', directeur='Direction')
        cls.principal = get_user_model().objects.create_user(username='principal-notes', password='test')
        profil = cls.principal.profil
        profil.ecole, profil.est_compte_principal, profil.is_validated = cls.ecole, True, True
        profil.save()
        cls.responsable = Responsable.objects.create(prenom='Parent', nom='Test', relation='PERE', telephone='+224620123452', adresse='Conakry')
        cls.classe, cls.eleve, cls.math, cls.fr = cls.creer_classe(cls.ecole, 'P', 'PRIMAIRE', 'PRIMAIRE_1')
        cls.secondaire, cls.eleve_s, cls.math_s, cls.fr_s = cls.creer_classe(cls.ecole, 'S', 'SECONDAIRE', 'COLLEGE_7')
        cls.maternelle, cls.eleve_m, cls.math_m, cls.fr_m = cls.creer_classe(cls.ecole, 'M', 'MATERNELLE', 'PETITE_SECTION')
        cls.exterieure, cls.eleve_x, cls.math_x, cls.fr_x = cls.creer_classe(cls.autre_ecole, 'P', 'PRIMAIRE', 'PRIMAIRE_1', 'X')
        cls.enseignant = cls.creer_enseignant(cls.classe)
        cls.enseignant_s = cls.creer_enseignant(cls.secondaire)
        cls.enseignant_m = cls.creer_enseignant(cls.maternelle)

    @classmethod
    def creer_classe(cls, ecole, nom, type_niveau, niveau, prefix=None):
        classe = Classe.objects.create(ecole=ecole, nom=nom, niveau=niveau, annee_scolaire='2026-2027')
        cn = ClasseNote.objects.create(ecole=ecole, nom=nom, niveau='MATERNELLE' if type_niveau == 'MATERNELLE' else niveau, niveau_enseignement=type_niveau, annee_scolaire='2026-2027')
        eleve = Eleve.objects.create(matricule=f'{prefix or nom}-001', prenom='Aminata', nom='Diallo', sexe='F', classe=classe, responsable_principal=cls.responsable, statut='ACTIF')
        math = MatiereNote.objects.create(classe=cn, nom='Mathématiques', code='MAT', coefficient=1)
        fr = MatiereNote.objects.create(classe=cn, nom='Français', code='FR', coefficient=1)
        return cn, eleve, math, fr

    @classmethod
    def creer_enseignant(cls, classe):
        enseignant = Enseignant.objects.create(nom='Enseignant ' + classe.nom, prenoms='Test', ecole=classe.ecole, type_enseignant=classe.niveau_enseignement, statut='ACTIF', salaire_fixe=100000, taux_horaire=10000, date_embauche=timezone.localdate() - timedelta(days=60), cree_par=cls.principal)
        affectation = Classe.objects.get(ecole=classe.ecole, nom=classe.nom, annee_scolaire=classe.annee_scolaire)
        AffectationClasse.objects.create(enseignant=enseignant, classe=affectation, date_debut=timezone.localdate() - timedelta(days=30), heures_par_semaine=8)
        return enseignant

    def creer_acces(self, enseignant=None, classes=None, matieres=None):
        self.client.force_login(self.principal)
        response = self.client.post(reverse('notes:gerer_acces_enseignants'), {
            'action': 'creer', 'enseignant': (enseignant or self.enseignant).pk,
            'classes': [c.pk for c in (classes or [self.classe])],
            'matieres': [m.pk for m in (matieres or [])],
            'expire_le': (timezone.now() + timedelta(days=2)).strftime('%Y-%m-%dT%H:%M'),
        })
        self.assertEqual(response.status_code, 302, response.content[:1000])
        acces = AccesEnseignantTemporaire.objects.latest('pk')
        lien = self.client.session['nouveau_lien_notes'].replace('http://testserver', '')
        return acces, lien

    def entrer(self, **kwargs):
        acces, lien = self.creer_acces(**kwargs)
        teacher = Client()
        self.assertEqual(self.poster_lien(teacher, lien).status_code, 302)
        return teacher, acces, lien

    def poster_lien(self, client, lien, **data):
        return client.post(reverse('notes:enseignant_connexion'), {'token': lien.rstrip('/').split('/')[-1], **data})

    def saisie_data(self, eleve=None, matiere=None, valeur='8.5', **extra):
        return {'mode': 'simple', 'genre': 'mensuelle', 'periode': 'OCTOBRE', 'matiere': (matiere or self.math).pk,
                f'cell_{(eleve or self.eleve).pk}_{(matiere or self.math).pk}': valeur, **extra}

    def csv_file(self, text):
        return SimpleUploadedFile('notes.csv', text.encode('utf-8'), content_type='text/csv')

    def test_creation_compte_minimal_lien_affiche_une_seule_fois(self):
        acces, lien = self.creer_acces()
        self.assertFalse(acces.utilisateur.has_usable_password())
        self.assertFalse(acces.utilisateur.is_staff)
        self.assertEqual(acces.utilisateur.profil.role, 'ENSEIGNANT')
        self.assertFalse(acces.utilisateur.profil.peut_ajouter_paiements)
        self.assertFalse(acces.utilisateur.profil.peut_gerer_notes)
        self.assertEqual(set(acces.matieres.all()), {self.math, self.fr})
        self.assertNotIn(lien.split('/')[-2], acces.empreinte_lien)
        response = self.client.get(reverse('notes:gerer_acces_enseignants'))
        self.assertContains(response, lien)
        self.assertEqual(response['Cache-Control'], 'no-store, private')
        self.assertNotContains(self.client.get(reverse('notes:gerer_acces_enseignants')), lien)

    def test_page_configuration_trois_niveaux(self):
        self.client.force_login(self.principal)
        for enseignant in [self.enseignant, self.enseignant_s, self.enseignant_m]:
            response = self.client.get(reverse('notes:gerer_acces_enseignants'), {'enseignant': enseignant.pk})
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, 'Créer le compte et son lien')

    def test_compte_non_principal_refuse_meme_role_admin(self):
        profil = self.principal.profil
        profil.est_compte_principal, profil.role = False, 'ADMIN'
        profil.save()
        self.client.force_login(self.principal)
        self.assertEqual(self.client.get(reverse('notes:gerer_acces_enseignants')).status_code, 403)

    def test_classe_exterieure_ou_expiration_lointaine_refusees(self):
        self.client.force_login(self.principal)
        for classes, days in [([self.exterieure.pk], 2), ([self.classe.pk], 100)]:
            response = self.client.post(reverse('notes:gerer_acces_enseignants'), {'action': 'creer', 'enseignant': self.enseignant.pk, 'classes': classes, 'expire_le': (timezone.now() + timedelta(days=days)).strftime('%Y-%m-%dT%H:%M')})
            self.assertEqual(response.status_code, 200)
            self.assertEqual(AccesEnseignantTemporaire.objects.count(), 0)

    def test_lien_get_ne_connecte_pas_et_post_exige_csrf(self):
        _, lien = self.creer_acces()
        teacher = Client(enforce_csrf_checks=True)
        response = teacher.get(lien)
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('_auth_user_id', teacher.session)
        self.assertEqual(response['Referrer-Policy'], 'strict-origin')
        self.assertEqual(self.poster_lien(teacher, lien).status_code, 403)
        self.assertEqual(self.poster_lien(teacher, lien, csrfmiddlewaretoken=teacher.cookies['csrftoken'].value).status_code, 302)

    def test_quatre_modes_et_modeles_excel(self):
        teacher, _, _ = self.entrer()
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil')).status_code, 200)
        for route in ['enseignant_saisie', 'enseignant_import', 'enseignant_template']:
            for mode in ['simple', 'intelligent']:
                response = teacher.get(reverse('notes:' + route, args=[self.classe.pk]), {'mode': mode})
                self.assertEqual(response.status_code, 200, (route, mode))
                if route == 'enseignant_template':
                    from openpyxl import load_workbook
                    sheet = load_workbook(BytesIO(response.content)).active
                    self.assertEqual(sheet.cell(2, 1).value, self.eleve.matricule)
                    self.assertEqual(sheet.max_row, 2)

    def test_saisie_primaire_mise_a_jour_sans_doublon_vide_et_absence(self):
        teacher, acces, _ = self.entrer()
        url = reverse('notes:enseignant_saisie', args=[self.classe.pk])
        self.assertEqual(teacher.post(url, self.saisie_data()).status_code, 302)
        self.assertEqual(teacher.post(url, self.saisie_data(valeur='9,25')).status_code, 302)
        note = NoteMensuelle.objects.get()
        self.assertEqual(note.note, Decimal('9.25'))
        self.assertEqual(note.cree_par, acces.utilisateur)
        self.assertEqual(teacher.post(url, self.saisie_data(valeur='')).status_code, 302)
        note.refresh_from_db()
        self.assertEqual(note.note, Decimal('9.25'))
        self.assertEqual(teacher.post(url, self.saisie_data(valeur='ABS')).status_code, 302)
        note.refresh_from_db()
        self.assertTrue(note.absent)
        self.assertIsNone(note.note)

    def test_validation_numerique_et_lot_atomique(self):
        teacher, _, _ = self.entrer()
        for value in ['11', '-1', 'NaN', 'Infinity', '8.333']:
            data = self.saisie_data(mode='intelligent', valeur='8')
            data[f'cell_{self.eleve.pk}_{self.fr.pk}'] = value
            self.assertEqual(teacher.post(reverse('notes:enseignant_saisie', args=[self.classe.pk]), data).status_code, 400)
            self.assertEqual(NoteMensuelle.objects.count(), 0)

    def test_secondaire_matiere_autorisee_composition(self):
        teacher, _, _ = self.entrer(enseignant=self.enseignant_s, classes=[self.secondaire], matieres=[self.math_s])
        url = reverse('notes:enseignant_saisie', args=[self.secondaire.pk])
        self.assertEqual(teacher.post(url, self.saisie_data(self.eleve_s, self.math_s, '19', genre='composition', periode='SEMESTRE_1')).status_code, 302)
        self.assertEqual(CompositionNote.objects.get().note, Decimal('19'))
        self.assertEqual(teacher.post(url, self.saisie_data(self.eleve_s, self.fr_s, '17')).status_code, 403)
        self.assertEqual(NoteMensuelle.objects.count(), 0)

    def test_appreciation_maternelle(self):
        teacher, _, _ = self.entrer(enseignant=self.enseignant_m, classes=[self.maternelle])
        url = reverse('notes:enseignant_saisie', args=[self.maternelle.pk])
        self.assertEqual(teacher.get(url).status_code, 200)
        data = self.saisie_data(self.eleve_m, self.math_m, 'A+', genre='maternelle', periode='TRIMESTRE_1')
        self.assertEqual(teacher.post(url, data).status_code, 302)
        self.assertEqual(AppreciationMaternelle.objects.get().appreciation, 'A+')

    def test_eleve_classe_et_modules_hors_perimetre_refuses(self):
        teacher, _, _ = self.entrer()
        for path in [reverse('notes:tableau_bord'), reverse('notes:gerer_acces_enseignants'), '/salaires/', '/utilisateurs/', '/notes/sauvegarder-notes/', '/notes/api/matieres-classe/']:
            self.assertEqual(teacher.get(path).status_code, 403, path)
        self.assertEqual(teacher.get(reverse('notes:enseignant_saisie', args=[self.exterieure.pk])).status_code, 404)
        self.assertEqual(teacher.post(reverse('notes:enseignant_saisie', args=[self.classe.pk]), self.saisie_data(self.eleve_x)).status_code, 403)
        self.assertEqual(NoteMensuelle.objects.count(), 0)

    def test_expiration_et_revocation_sessions_ouvertes(self):
        teacher, acces, lien = self.entrer()
        acces.expire_le = timezone.now() - timedelta(seconds=1)
        acces.save()
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil')).status_code, 403)
        self.assertNotIn('_auth_user_id', teacher.session)
        self.assertEqual(Client().get(lien).status_code, 403)
        teacher, acces, lien = self.entrer()
        self.client.post(reverse('notes:gerer_acces_enseignants'), {'action': 'revoquer', 'acces_id': acces.pk})
        self.assertEqual(teacher.post(reverse('notes:enseignant_saisie', args=[self.classe.pk]), self.saisie_data()).status_code, 403)
        self.assertEqual(self.poster_lien(Client(), lien).status_code, 403)

    def test_renouvellement_invalide_ancien_lien_et_session(self):
        teacher, acces, lien = self.entrer()
        response = self.client.post(reverse('notes:gerer_acces_enseignants'), {'action': 'renouveler', 'acces_id': acces.pk, 'expire_le': (timezone.now() + timedelta(days=3)).strftime('%Y-%m-%dT%H:%M')})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.poster_lien(Client(), lien).status_code, 403)
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil')).status_code, 403)
        nouveau = self.client.session['nouveau_lien_notes'].replace('http://testserver', '')
        self.assertEqual(self.poster_lien(Client(), nouveau).status_code, 302)

    def test_nouveau_lien_reconnecte_session_revoquee_sans_blocage(self):
        teacher, acces, _ = self.entrer()
        acces.revoque_le = timezone.now()
        acces.save(update_fields=['revoque_le'])
        _, lien = self.creer_acces()
        self.assertEqual(teacher.get(lien).status_code, 200)
        self.assertEqual(self.poster_lien(teacher, lien).status_code, 302)
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil')).status_code, 200)

    def test_compte_autre_ecole_ne_revoque_pas_acces(self):
        _, acces, _ = self.entrer()
        profil = self.principal.profil
        profil.ecole = self.autre_ecole
        profil.save()
        response = self.client.post(reverse('notes:gerer_acces_enseignants'), {'action': 'revoquer', 'acces_id': acces.pk})
        self.assertEqual(response.status_code, 404)
        acces.refresh_from_db()
        self.assertIsNone(acces.revoque_le)

    def test_enseignant_inactif_et_annee_scolaire_non_modifiable(self):
        teacher, _, _ = self.entrer()
        url = reverse('notes:enseignant_saisie', args=[self.classe.pk])
        self.assertEqual(teacher.post(url, self.saisie_data(annee_scolaire='2030-2031')).status_code, 302)
        self.assertEqual(NoteMensuelle.objects.get().annee_scolaire, self.classe.annee_scolaire)
        Enseignant.objects.filter(pk=self.enseignant.pk).update(statut='SUSPENDU')
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil')).status_code, 403)

    def test_secondaire_exige_matiere_et_matricule_malforme_ne_fait_pas_500(self):
        self.client.force_login(self.principal)
        response = self.client.post(reverse('notes:gerer_acces_enseignants'), {'action': 'creer', 'enseignant': self.enseignant_s.pk, 'classes': [self.secondaire.pk], 'expire_le': (timezone.now() + timedelta(days=1)).strftime('%Y-%m-%dT%H:%M')})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(AccesEnseignantTemporaire.objects.count(), 0)
        self.assertEqual(self.client.post(reverse('notes:gerer_acces_enseignants'), {'action': 'revoquer', 'acces_id': 'invalide'}).status_code, 403)

    def test_fin_affectation_retirer_classe(self):
        teacher, _, _ = self.entrer()
        AffectationClasse.objects.filter(enseignant=self.enseignant).update(actif=False)
        self.assertEqual(teacher.get(reverse('notes:enseignant_saisie', args=[self.classe.pk])).status_code, 404)
        self.assertNotContains(teacher.get(reverse('notes:enseignant_accueil')), 'Saisie intelligente')

    def test_force_login_ne_contourne_pas_le_lien(self):
        acces, _ = self.creer_acces()
        teacher = Client()
        teacher.force_login(acces.utilisateur)
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil')).status_code, 403)

    def test_import_simple_apercu_confirmation_et_rejeu(self):
        teacher, _, _ = self.entrer()
        url = reverse('notes:enseignant_import', args=[self.classe.pk])
        data = {'mode': 'simple', 'genre': 'mensuelle', 'periode': 'OCTOBRE', 'matiere': self.math.pk}
        response = teacher.post(url, {**data, 'fichier': self.csv_file('Matricule;Note\nP-001;9.5')})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(NoteMensuelle.objects.count(), 0)
        nonce = teacher.session['notes_import_apercu']['nonce']
        self.assertEqual(teacher.post(url, {**data, 'action': 'confirmer', 'nonce': nonce}).status_code, 302)
        self.assertEqual(NoteMensuelle.objects.get().note, Decimal('9.5'))
        self.assertEqual(teacher.post(url, {**data, 'action': 'confirmer', 'nonce': nonce}).status_code, 400)

    def test_import_intelligent_reconnaissance_et_erreurs_atomiques(self):
        teacher, _, _ = self.entrer()
        url = reverse('notes:enseignant_import', args=[self.classe.pk])
        data = {'mode': 'intelligent', 'genre': 'mensuelle', 'periode': 'OCTOBRE'}
        invalid = ['Matricule;MAT;FR\nP-001;9;12', 'Matricule;MAT\nX-001;8', 'Matricule;MAT\nP-001;8\nP-001;9', 'Matricule;HISTOIRE\nP-001;8']
        for content in invalid:
            self.assertEqual(teacher.post(url, {**data, 'fichier': self.csv_file(content)}).status_code, 400)
            self.assertNotIn('notes_import_apercu', teacher.session)
            self.assertEqual(NoteMensuelle.objects.count(), 0)
        response = teacher.post(url, {**data, 'fichier': self.csv_file('Matricule;Mathématiques;Français\nP-001;9;ABS')})
        self.assertEqual(response.status_code, 200)
        nonce = teacher.session['notes_import_apercu']['nonce']
        self.assertEqual(teacher.post(url, {**data, 'action': 'confirmer', 'nonce': nonce}).status_code, 302)
        self.assertEqual(NoteMensuelle.objects.count(), 2)
        self.assertTrue(NoteMensuelle.objects.get(matiere=self.fr).absent)

    def test_xlsx_apercu_et_formules_refusees(self):
        from openpyxl import Workbook
        teacher, _, _ = self.entrer()
        url = reverse('notes:enseignant_import', args=[self.classe.pk])
        for value, status in [(9, 200), ('=4+5', 400)]:
            book, buffer = Workbook(), BytesIO()
            book.active.append(['Matricule', 'Note'])
            book.active.append(['P-001', value])
            book.save(buffer)
            response = teacher.post(url, {'matiere': self.math.pk, 'fichier': SimpleUploadedFile('notes.xlsx', buffer.getvalue())})
            self.assertEqual(response.status_code, status)
        self.assertEqual(NoteMensuelle.objects.count(), 0)

    def test_confirmation_reverifie_matieres_autorisees(self):
        teacher, acces, _ = self.entrer()
        url = reverse('notes:enseignant_import', args=[self.classe.pk])
        data = {'mode': 'intelligent', 'genre': 'mensuelle', 'periode': 'OCTOBRE'}
        teacher.post(url, {**data, 'fichier': self.csv_file('Matricule;MAT;FR\nP-001;9;8')})
        nonce = teacher.session['notes_import_apercu']['nonce']
        acces.matieres.remove(self.fr)
        self.assertEqual(teacher.post(url, {**data, 'action': 'confirmer', 'nonce': nonce}).status_code, 403)
        self.assertEqual(NoteMensuelle.objects.count(), 0)

    @override_settings(ALLOWED_HOSTS=['www.myschoolgn.space'], CSRF_COOKIE_SECURE=True,
                       CSRF_COOKIE_HTTPONLY=True, SESSION_COOKIE_SECURE=True)
    def test_gestion_https_csrf_creation_renouvellement_revocation(self):
        from html.parser import HTMLParser

        class TokenParser(HTMLParser):
            token = None

            def handle_starttag(self, tag, attrs):
                attrs = dict(attrs)
                if tag == 'input' and attrs.get('name') == 'csrfmiddlewaretoken':
                    self.token = attrs['value']

        client = Client(enforce_csrf_checks=True)
        client.force_login(self.principal)
        host = {'secure': True, 'HTTP_HOST': 'www.myschoolgn.space'}
        url = reverse('notes:gerer_acces_enseignants') + f'?enseignant={self.enseignant.pk}'
        response = client.get(url, **host)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Referrer-Policy'], 'strict-origin')
        parser = TokenParser()
        parser.feed(response.content.decode())
        self.assertTrue(parser.token)
        data = {'action': 'creer', 'enseignant': self.enseignant.pk,
                'classes': [self.classe.pk],
                'expire_le': (timezone.now() + timedelta(days=2)).strftime('%Y-%m-%dT%H:%M'),
                'csrfmiddlewaretoken': parser.token}
        # Reproduit le navigateur HTTPS sans Origin et privé de Referer.
        self.assertEqual(client.post(url, data, **host).status_code, 403)
        self.assertEqual(AccesEnseignantTemporaire.objects.count(), 0)
        referer = 'https://www.myschoolgn.space/'
        self.assertEqual(client.post(url, data, HTTP_REFERER=referer, **host).status_code, 302)
        acces = AccesEnseignantTemporaire.objects.get()
        for action in ('renouveler', 'revoquer'):
            response = client.post(url, {**data, 'action': action, 'acces_id': acces.pk},
                                   HTTP_REFERER=referer, **host)
            self.assertEqual(response.status_code, 302)
        acces.refresh_from_db()
        self.assertIsNotNone(acces.revoque_le)
        # Le correctif n'autorise ni les jetons absents ni les sites externes.
        self.assertEqual(client.post(url, {'action': 'creer'}, HTTP_REFERER=referer, **host).status_code, 403)
        self.assertEqual(client.post(url, data, HTTP_REFERER='https://example.org/', **host).status_code, 403)

    def test_connexion_enseignant_https_referer_sans_secret(self):
        _, lien = self.creer_acces()
        teacher = Client(enforce_csrf_checks=True)
        response = teacher.get(lien, secure=True)
        self.assertEqual(response['Referrer-Policy'], 'strict-origin')
        self.assertContains(response, '<meta name="referrer" content="strict-origin">')
        data = {'token': lien.rstrip('/').split('/')[-1],
                'csrfmiddlewaretoken': teacher.cookies['csrftoken'].value}
        response = teacher.post(reverse('notes:enseignant_connexion'), data,
                                secure=True, HTTP_REFERER='https://testserver/')
        self.assertEqual(response.status_code, 302)
        self.assertEqual(teacher.get(reverse('notes:enseignant_accueil'), secure=True).status_code, 200)
