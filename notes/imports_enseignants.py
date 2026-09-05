"""Lecture bornée et sans exécution de formules des imports enseignants."""
import csv
import io
import unicodedata
from itertools import islice
from zipfile import ZipFile, BadZipFile
from xml.etree.ElementTree import ParseError

from django.core.exceptions import ValidationError

from .acces_enseignants import eleves_autorises, valider_cellules, valeur_note


def normaliser(value):
    return ''.join(c for c in unicodedata.normalize('NFKD', str(value or '').lower()) if not unicodedata.combining(c) and c.isalnum())


def lire_fichier(upload):
    if not upload or upload.size > 2 * 1024 * 1024:
        raise ValidationError('Sélectionnez un fichier CSV ou XLSX de 2 Mo maximum.')
    try:
        if upload.name.lower().endswith('.csv'):
            raw = upload.read().decode('utf-8-sig')
            dialect = csv.Sniffer().sniff(raw[:8192], delimiters=',;\t')
            rows = list(islice(csv.reader(io.StringIO(raw), dialect), 2002))
        elif upload.name.lower().endswith('.xlsx'):
            from openpyxl import load_workbook
            with ZipFile(upload) as archive:
                if len(archive.infolist()) > 1000 or sum(i.file_size for i in archive.infolist()) > 20 * 1024 * 1024:
                    raise ValidationError('Classeur trop volumineux après décompression.')
            upload.seek(0)
            book = load_workbook(upload, read_only=True, data_only=False, keep_links=False)
            try:
                sheet = book.active
                if sheet is None:
                    raise ValidationError('Le classeur ne contient aucune feuille active.')
                if (sheet.max_column or 0) > 100 or (sheet.max_row or 0) > 2001:
                    raise ValidationError('Maximum : 2 000 élèves et 100 colonnes.')
                sheet.reset_dimensions()
                rows = []
                for row in islice(sheet.iter_rows(max_col=101), 2002):
                    while row and row[-1].value is None:
                        row = row[:-1]
                    if len(row) > 100:
                        raise ValidationError('Maximum : 100 colonnes.')
                    if any(cell.data_type == 'f' for cell in row):
                        raise ValidationError('Les formules ne sont pas acceptées. Collez uniquement les valeurs.')
                    rows.append([cell.value for cell in row])
            finally:
                book.close()
        else:
            raise ValidationError('Format accepté : .csv (UTF-8) ou .xlsx.')
    except (UnicodeDecodeError, csv.Error, BadZipFile, KeyError, ValueError, OSError, ParseError) as exc:
        raise ValidationError('Fichier illisible. Utilisez le modèle fourni et enregistrez-le en CSV UTF-8 ou XLSX.') from exc
    if len(rows) < 2 or len(rows) > 2001 or any(len(row) > 100 for row in rows):
        raise ValidationError('Le fichier doit contenir un en-tête et entre 1 et 2 000 élèves, avec 100 colonnes maximum.')
    return rows


def analyser_import(upload, context):
    rows = lire_fichier(upload)
    while rows[0] and rows[0][-1] in (None, ''):
        rows[0].pop()
    headers = [normaliser(h) for h in rows[0]]
    if not all(headers) or len(headers) != len(set(headers)) or 'matricule' not in headers:
        raise ValidationError('En-têtes vides ou doublonnés, ou colonne Matricule manquante.')
    mapping, ignored = {}, {'matricule', 'nom', 'prenom', 'prenoms', 'nomcomplet'}
    for index, header in enumerate(headers):
        if header in ignored:
            continue
        if context['mode'] == 'simple':
            matches = context['colonnes'] if header in ('note', 'appreciation') else []
        else:
            matches = [m for m in context['colonnes'] if header in {
                normaliser(m.code), normaliser(m.nom), normaliser(f'{m.code} - {m.nom}'),
            }]
        if len(matches) != 1:
            raise ValidationError(f'Colonne « {rows[0][index]} » inconnue, ambiguë ou non autorisée. Utilisez le modèle de cet écran.')
        if matches[0].pk in mapping.values():
            raise ValidationError('Deux colonnes correspondent à la même matière.')
        mapping[index] = matches[0].pk
    if not mapping:
        raise ValidationError('Aucune colonne de notes trouvée.')
    eleves = {e.matricule: e for e in eleves_autorises(context['classe'])}
    matricule_column, seen, cellules, preview = headers.index('matricule'), set(), [], []
    subjects = {m.pk: m for m in context['colonnes']}
    for line, row in enumerate(rows[1:], start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        if len(row) > len(headers) and any(v is not None and str(v).strip() for v in row[len(headers):]):
            raise ValidationError(f'Ligne {line} : données sans en-tête.')
        row = list(row) + [''] * max(0, len(headers) - len(row))
        matricule = str(row[matricule_column] or '').strip()
        if matricule not in eleves:
            raise ValidationError(f'Ligne {line} : matricule inconnu dans cette classe. Aucun élève extérieur ne peut être importé.')
        if matricule in seen:
            raise ValidationError(f'Ligne {line} : matricule présent plusieurs fois.')
        seen.add(matricule)
        eleve = eleves[matricule]
        for index, mid in mapping.items():
            raw = '' if row[index] is None else str(row[index]).strip()
            cell = {'eleve': eleve.pk, 'matiere': mid, 'valeur': raw}
            try:
                valeur_note(raw, context['classe'])
            except ValidationError as exc:
                raise ValidationError(f'Ligne {line}, {subjects[mid].nom} : {exc.messages[0]}')
            if raw:
                cellules.append(cell)
                preview.append({'eleve': eleve, 'matiere': subjects[mid], 'valeur': raw})
    if not cellules or len(cellules) > 10000:
        raise ValidationError('Le fichier doit contenir entre 1 et 10 000 notes / appréciations non vides.')
    valider_cellules(context['acces'], context['classe'], cellules)
    return cellules, preview
