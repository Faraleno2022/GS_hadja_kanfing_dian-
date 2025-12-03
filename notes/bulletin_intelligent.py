"""
Système de Bulletin Intelligent avec Calculs Automatiques
Exports PDF (avec filigrane) et Excel
Conforme au système guinéen (40% cours + 60% composition)
"""

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from django.db.models import Avg, Count
from decimal import Decimal
import io
from datetime import datetime

# ReportLab pour PDF
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
from PIL import Image

# OpenPyXL pour Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from eleves.models import Eleve, Classe
from .models import ClasseNote, MatiereNote, Evaluation, NoteEleve
# Import du module centralisé pour garantir la cohérence
from .calculs_moyennes import (
    calculer_moyenne_generale_eleve,
    calculer_classement_classe,
    obtenir_mention_intelligente,
    obtenir_appreciation_intelligente,
    formater_rang_intelligent
)
from .calculs import (
    calculer_moyenne_devoirs,
    calculer_moyenne_periode,
    calculer_moyenne_cours_mensuels,
    calculer_rang
)
from .classifier import classify
from utilisateurs.utils import filter_by_user_school, user_school
from ecole_moderne.security_decorators import require_school_object


class CalculateurBulletinIntelligent:
    """Calculateur intelligent de bulletin conforme au système guinéen"""
    
    def __init__(self, eleve, classe_note, periode, systeme='TRIMESTRE'):
        self.eleve = eleve
        self.classe_note = classe_note
        self.periode = periode
        self.systeme = systeme
        self.niveau, self.serie, self.section = self._detecter_niveau_intelligent()
        
    def _detecter_niveau_intelligent(self):
        """Détecte intelligemment le niveau, la série et la section à partir du nom de classe"""
        # Essayer d'abord le champ niveau_enseignement si disponible et cohérent
        niveau_db = getattr(self.classe_note, 'niveau_enseignement', None)
        
        # Classification intelligente à partir du nom
        niveau_auto, serie, section = classify(self.classe_note.nom)
        
        # Si le niveau DB existe et est cohérent, on le garde
        if niveau_db in ['PRIMAIRE', 'MATERNELLE']:
            niveau_final = 'PRIMAIRE'
        elif niveau_db in ['COLLEGE', 'LYCEE']:
            niveau_final = 'SECONDAIRE'
        else:
            # Sinon on utilise la détection automatique
            niveau_final = niveau_auto
        
        return niveau_final, serie, section
    
    def _obtenir_mois_periode(self):
        """Retourne les mois d'une période"""
        mapping = {
            'TRIMESTRE_1': ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE'],
            'TRIMESTRE_2': ['JANVIER', 'FEVRIER', 'MARS'],
            'TRIMESTRE_3': ['AVRIL', 'MAI', 'JUIN'],
            'SEMESTRE_1': ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER'],
            'SEMESTRE_2': ['MARS', 'AVRIL', 'MAI', 'JUIN'],
        }
        return mapping.get(self.periode, [])
    
    def calculer_notes_matiere(self, matiere):
        """Calcule les notes d'une matière pour la période"""
        mois_periode = self._obtenir_mois_periode()
        
        # Récupérer les notes mensuelles
        notes_mensuelles = {}
        for mois in mois_periode:
            notes_mois_qs = NoteEleve.objects.filter(
                eleve=self.eleve,
                evaluation__matiere=matiere,
                evaluation__periode=mois,
                evaluation__type_eval='DEVOIR'
            )

            notes_mois = []
            for note_obj in notes_mois_qs:
                if note_obj.absent or note_obj.note is None:
                    notes_mois.append(Decimal('0'))
                else:
                    notes_mois.append(Decimal(str(note_obj.note)))

            if notes_mois:
                notes_mensuelles[mois.lower()] = notes_mois
        
        # Récupérer la composition
        composition = None
        compo_obj = NoteEleve.objects.filter(
            eleve=self.eleve,
            evaluation__matiere=matiere,
            evaluation__periode=self.periode,
            evaluation__type_eval='COMPOSITION'
        ).first()
        
        if compo_obj:
            if compo_obj.absent or compo_obj.note is None:
                composition = Decimal('0')
            else:
                composition = Decimal(str(compo_obj.note))
        
        # Calculer selon le niveau
        if self.niveau == 'PRIMAIRE':
            # Primaire : composition uniquement
            moyenne_periode = composition
        else:
            # Secondaire : 40% cours + 60% composition
            moyenne_cours = None
            if notes_mensuelles:
                moyenne_cours = calculer_moyenne_cours_mensuels(notes_mensuelles)
            
            moyenne_periode = calculer_moyenne_periode(
                moyenne_cours,
                composition,
                niveau=self.niveau
            )
        
        return {
            'matiere': matiere.nom,
            'coefficient': matiere.coefficient,
            'notes_mensuelles': notes_mensuelles,
            'composition': composition,
            'moyenne': moyenne_periode
        }
    
    def generer_bulletin(self):
        """Génère le bulletin complet pour la période en utilisant le module centralisé"""
        from .calculs_moyennes import detecter_niveau_scolaire
        
        # Détecter si c'est une classe de maternelle
        niveau_detecte = detecter_niveau_scolaire(self.classe_note.nom)
        est_maternelle = (niveau_detecte == 'MATERNELLE')
        
        # Récupérer toutes les matières
        matieres = MatiereNote.objects.filter(classe=self.classe_note, actif=True)
        
        # Pour la maternelle : utiliser les appréciations
        if est_maternelle:
            return self._generer_bulletin_maternelle(matieres)
        
        # UTILISER LE MODULE CENTRALISÉ (SOURCE UNIQUE)
        system_type = 'trimestre' if 'TRIMESTRE' in self.periode else 'semestre' if 'SEMESTRE' in self.periode else 'mensuel'
        result_centralized = calculer_moyenne_generale_eleve(
            self.eleve, 
            matieres, 
            self.periode, 
            system_type
        )
        
        # Extraire les données du calcul centralisé
        moyenne_generale = result_centralized['moyenne_generale']
        resultats_matieres = result_centralized['details_matieres']
        
        # Calculer le rang (optionnel - nécessite tous les élèves)
        rang = self._calculer_rang_eleve(moyenne_generale)
        
        return {
            'eleve': f"{self.eleve.prenom} {self.eleve.nom}",
            'classe': self.classe_note.nom,
            'periode': self.periode,
            'niveau': self.niveau,
            'serie': self.serie,
            'section': self.section,
            'matieres': resultats_matieres,
            'moyenne_generale': moyenne_generale,
            'mention': obtenir_mention_intelligente(moyenne_generale, self.niveau) if moyenne_generale else None,
            'appreciation': obtenir_appreciation_intelligente(moyenne_generale, self.eleve.prenom) if moyenne_generale else None,
            'rang': rang,
            'total_eleves': Eleve.objects.filter(classe=self.eleve.classe).count()
        }
    
    def _generer_bulletin_maternelle(self, matieres):
        """Génère le bulletin pour la maternelle avec appréciations"""
        from .models import AppreciationMaternelle
        from .utils_rangs import calculer_rangs_classe_periode
        
        # Récupérer les appréciations pour chaque matière
        resultats_matieres = []
        for matiere in matieres:
            appreciation_data = {
                'matiere': matiere.nom,
                'matiere_obj': matiere,
                'coefficient': 1,
                'appreciation': None,
                'appreciation_display': '-',
                'commentaire': '-',
                'absent': False
            }
            
            try:
                app_obj = AppreciationMaternelle.objects.get(
                    eleve=self.eleve,
                    matiere=matiere,
                    trimestre=self.periode,
                    annee_scolaire=self.classe_note.annee_scolaire
                )
                appreciation_data['appreciation'] = app_obj.appreciation
                appreciation_data['appreciation_display'] = app_obj.get_appreciation_display()
                appreciation_data['commentaire'] = app_obj.commentaire or self._get_observation_auto(app_obj.appreciation)
                appreciation_data['absent'] = app_obj.absent
            except AppreciationMaternelle.DoesNotExist:
                pass
            
            resultats_matieres.append(appreciation_data)
        
        # Calculer le rang et taux d'acquisition
        rangs_dict = calculer_rangs_classe_periode(self.classe_note, self.periode, use_cache=False)
        rang_info = rangs_dict.get(self.eleve.id)
        
        taux_acquisition = None
        rang = None
        mention = None
        appreciation = None
        
        if rang_info:
            taux_acquisition = float(rang_info['moyenne'])
            rang = f"{rang_info['rang']}/{rang_info['total_eleves']}"
            
            # Mention basée sur le taux
            if taux_acquisition >= 90:
                mention = 'Excellent'
            elif taux_acquisition >= 75:
                mention = 'Très Bien'
            elif taux_acquisition >= 60:
                mention = 'Bien'
            elif taux_acquisition >= 50:
                mention = 'Assez Bien'
            else:
                mention = 'À encourager'
            
            appreciation = f"Bon trimestre {self.eleve.prenom}. Continue ainsi !"
        
        return {
            'eleve': f"{self.eleve.prenom} {self.eleve.nom}",
            'classe': self.classe_note.nom,
            'periode': self.periode,
            'niveau': 'MATERNELLE',
            'serie': None,
            'section': None,
            'matieres': resultats_matieres,
            'moyenne_generale': taux_acquisition,
            'mention': mention,
            'appreciation': appreciation,
            'rang': rang,
            'total_eleves': Eleve.objects.filter(classe=self.eleve.classe).count(),
            'est_maternelle': True
        }
    
    def _get_observation_auto(self, appreciation):
        """Retourne une observation automatique basée sur l'appréciation"""
        observations = {
            'TRES_BIEN_ACQUIS': 'Excellent travail, continue ainsi !',
            'BIEN_ACQUIS': 'Bon niveau, quelques perfectionnements possibles',
            'EN_COURS': 'Des progrès à faire, persévère !',
            'NON_ACQUIS': 'Nécessite un accompagnement renforcé'
        }
        return observations.get(appreciation, '-')
    
    def _calculer_rang_eleve(self, moyenne_eleve):
        """Calcule le rang de l'élève dans la classe"""
        if not moyenne_eleve:
            return None
        
        # Récupérer tous les élèves de la classe
        eleves = Eleve.objects.filter(classe=self.eleve.classe, statut='ACTIF')
        
        # Récupérer les matières de la classe
        matieres = MatiereNote.objects.filter(classe=self.classe_note)
        
        moyennes_eleves = []
        for eleve in eleves:
            # Utiliser directement calculer_moyenne_generale_eleve pour éviter la récursion
            result = calculer_moyenne_generale_eleve(eleve, matieres, self.periode)
            moyenne = result.get('moyenne_generale') if isinstance(result, dict) else result
            
            if moyenne:
                moyennes_eleves.append({
                    'eleve_id': eleve.id,
                    'prenom': eleve.prenom,
                    'sexe': eleve.sexe,
                    'moyenne': moyenne
                })
        
        # Calculer les rangs avec accord grammatical
        eleves_classes = calculer_rang(moyennes_eleves)
        
        # Trouver le rang de notre élève
        for e in eleves_classes:
            if e['eleve_id'] == self.eleve.id:
                return e['rang']
        
        return None


def generer_pdf_avec_filigrane(bulletin_data, logo_path=None, ecole=None):
    """Génère un PDF avec le même design exact que le modèle imprimé"""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Couleurs du design (exactement comme le modèle)
    BLEU_HEADER = colors.HexColor('#4a90d9')  # Bleu de l'en-tête du tableau
    BLEU_CLAIR = colors.HexColor('#e8f4fd')   # Fond des infos élève
    BLEU_APPRECIATION = colors.HexColor('#d6eaf8')  # Fond appréciation
    VERT_MOY = colors.HexColor('#c8e6c9')     # Colonne MOY
    ROUGE_PTS = colors.HexColor('#ffcdd2')    # Colonne PTS
    ROUGE_DRAPEAU = colors.HexColor('#CE1126')
    JAUNE_DRAPEAU = colors.HexColor('#FCD116')
    VERT_DRAPEAU = colors.HexColor('#009460')
    GRIS_TOTAL = colors.HexColor('#37474f')   # Ligne total
    
    # Couleurs des mentions
    MENTION_COLORS = {
        'EXCELLENT': colors.HexColor('#1b5e20'),
        'TRÈS BIEN': colors.HexColor('#2e7d32'),
        'BIEN': colors.HexColor('#0277bd'),
        'ASSEZ BIEN': colors.HexColor('#f9a825'),
        'PASSABLE': colors.HexColor('#ef6c00'),
        'INSUFFISANT': colors.HexColor('#c62828'),
        'FAIBLE': colors.HexColor('#b71c1c'),
    }
    
    # ===== FILIGRANE (logo en transparence au centre) =====
    if logo_path:
        try:
            img_reader = ImageReader(logo_path)
            c.saveState()
            c.setFillAlpha(0.08)  # Transparence 8%
            filigrane_size = 14 * cm
            x = (width - filigrane_size) / 2
            y = (height - filigrane_size) / 2
            c.drawImage(img_reader, x, y, width=filigrane_size, height=filigrane_size, 
                       preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except Exception as e:
            print(f"Erreur filigrane: {e}")
    
    # ===== EN-TÊTE =====
    # Logo en haut à gauche
    if logo_path:
        try:
            img_reader = ImageReader(logo_path)
            c.drawImage(img_reader, 1.2*cm, height - 3.2*cm, width=2.5*cm, height=2.5*cm, 
                       preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Drapeau guinéen en haut à droite
    drapeau_x = width - 2.5*cm
    drapeau_y = height - 2.2*cm
    drapeau_w = 1.2*cm
    drapeau_h = 0.8*cm
    c.setFillColor(ROUGE_DRAPEAU)
    c.rect(drapeau_x, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setFillColor(JAUNE_DRAPEAU)
    c.rect(drapeau_x + drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setFillColor(VERT_DRAPEAU)
    c.rect(drapeau_x + 2*drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.rect(drapeau_x, drapeau_y, drapeau_w, drapeau_h, fill=0, stroke=1)
    
    # Textes de l'en-tête
    y_header = height - 1.2*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width/2, y_header, "RÉPUBLIQUE DE GUINÉE")
    
    # Devise avec couleurs
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(ROUGE_DRAPEAU)
    c.drawCentredString(width/2 - 1.8*cm, y_header, "Travail")
    c.setFillColor(JAUNE_DRAPEAU)
    c.drawCentredString(width/2, y_header, "Justice")
    c.setFillColor(VERT_DRAPEAU)
    c.drawCentredString(width/2 + 1.8*cm, y_header, "Solidarité")
    # Tirets
    c.setFillColor(colors.black)
    c.drawCentredString(width/2 - 0.9*cm, y_header, "-")
    c.drawCentredString(width/2 + 0.9*cm, y_header, "-")
    
    # MPU-A
    y_header -= 0.45*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(width/2, y_header, "MPU-A")
    
    # Nom de l'école
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 10)
    nom_ecole = ecole.nom.upper() if ecole else "GROUPE SCOLAIRE HADJA KANFING DIANÉ-SONFONIA"
    c.drawCentredString(width/2, y_header, nom_ecole)
    
    # Titre du bulletin
    y_header -= 0.6*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width/2, y_header, f"BULLETIN DE NOTES - {bulletin_data['periode'].upper()}")
    
    # Année scolaire (petite, à droite)
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor('#666666'))
    c.drawRightString(width - 1.5*cm, y_header, "Année Scolaire 2024-2025")
    
    # ===== LIGNE DE SÉPARATION NOIRE =====
    y_header -= 0.4*cm
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.5)
    c.line(1.2*cm, y_header, width - 1.2*cm, y_header)
    c.setLineWidth(1)
    
    # ===== INFORMATIONS ÉLÈVE (grille 2x3) =====
    y = y_header - 0.6*cm
    
    # Extraire nom et prénom
    eleve_nom_complet = bulletin_data.get('eleve', '')
    parties = eleve_nom_complet.split(' ', 1) if eleve_nom_complet else ['', '']
    prenom = parties[0] if len(parties) > 0 else ''
    nom = parties[1] if len(parties) > 1 else ''
    
    # Première ligne: PRÉNOM, NOM, MATRICULE (Prénom avant Nom)
    # Deuxième ligne: CLASSE, PÉRIODE, EFFECTIF
    info_data = [
        [('PRÉNOM', prenom.title()), ('NOM', nom.upper()), ('MATRICULE', bulletin_data.get('matricule', '-'))],
        [('CLASSE', bulletin_data.get('classe', '-')), ('PÉRIODE', bulletin_data.get('periode', '-')), ('EFFECTIF', f"{bulletin_data.get('total_eleves', '-')} élèves")],
    ]
    
    # Largeur totale alignée avec la ligne (de 1.2cm à width-1.2cm)
    info_total_width = width - 2.4*cm
    box_width = (info_total_width - 0.4*cm) / 3  # 3 colonnes avec espacement
    box_height = 0.9*cm
    info_margin_left = 1.2*cm
    
    for row_idx, row in enumerate(info_data):
        for col_idx, (label, value) in enumerate(row):
            x = info_margin_left + col_idx * (box_width + 0.2*cm)
            box_y = y - row_idx * (box_height + 0.15*cm)
            
            # Fond bleu clair
            c.setFillColor(BLEU_CLAIR)
            c.rect(x, box_y - box_height, box_width, box_height, fill=1, stroke=0)
            # Bordure
            c.setStrokeColor(colors.HexColor('#b3d4fc'))
            c.rect(x, box_y - box_height, box_width, box_height, fill=0, stroke=1)
            
            # Label en bleu
            c.setFillColor(colors.HexColor('#1976d2'))
            c.setFont("Helvetica-Bold", 7)
            c.drawString(x + 0.15*cm, box_y - 0.3*cm, label)
            
            # Valeur en noir
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x + 0.15*cm, box_y - 0.7*cm, str(value))
    
    # ===== TABLEAU DES NOTES =====
    y -= 2.3*cm
    
    # Largeur totale du tableau = même que la ligne de séparation (de 1.2cm à width-1.2cm)
    table_total_width = width - 2.4*cm
    margin_left = 1.2*cm
    
    # Déterminer le type de système pour le bulletin individuel
    system_type_indiv = bulletin_data.get('system_type', 'mensuel')
    periode = bulletin_data.get('periode', '')
    
    # Détecter le niveau scolaire
    from notes.calculs_moyennes import detecter_niveau_scolaire
    classe_nom = bulletin_data.get('classe', '')
    niveau_scolaire = detecter_niveau_scolaire(classe_nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    est_maternelle = (niveau_scolaire == 'MATERNELLE')
    
    # MATERNELLE/GARDERIE: Afficher les appréciations au lieu des notes
    if est_maternelle:
        _dessiner_bulletin_maternelle(c, bulletin_data, width, height, y, ecole)
        # Finaliser le PDF pour la maternelle
        c.showPage()
        c.save()
        buffer.seek(0)
        return buffer
    
    # Déterminer les mois selon la période pour trimestre/semestre
    # IMPORTANT: Le dernier mois est remplacé par "Compo" (composition)
    mois_labels = []
    if system_type_indiv == 'trimestriel':
        if 'TRIMESTRE_1' in periode or '1' in periode:
            mois_labels = ['Oct.', 'Nov.']  # Déc. = Compo
        elif 'TRIMESTRE_2' in periode or '2' in periode:
            mois_labels = ['Jan.', 'Fév.']  # Mars = Compo
        elif 'TRIMESTRE_3' in periode or '3' in periode:
            mois_labels = ['Avr.', 'Mai']   # Juin = Compo
    elif system_type_indiv == 'semestriel':
        if 'SEMESTRE_1' in periode or '1' in periode:
            mois_labels = ['Oct.', 'Nov.', 'Déc.', 'Jan.']  # Fév. = Compo
        elif 'SEMESTRE_2' in periode or '2' in periode:
            mois_labels = ['Mars', 'Avr.', 'Mai', 'Juin']   # Juil. = Compo
    
    # Adapter les colonnes selon le système ET le niveau (primaire = sans COEF)
    # Structure: MATIÈRE | [COEF] | Mois1 | Mois2 | [Mois...] | Moy.C | Compo | MOY | PTS
    if system_type_indiv == 'trimestriel' and mois_labels:
        if est_primaire:
            # Primaire Trimestre: MATIÈRE | M1 | M2 | Moy.C | Compo | MOY | PTS (sans COEF)
            header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        else:
            # Trimestre: MATIÈRE | COEF | M1 | M2 | Moy.C | Compo | MOY | PTS
            header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    elif system_type_indiv == 'semestriel' and mois_labels:
        if est_primaire:
            # Primaire Semestre: MATIÈRE | M1..M4 | Moy.C | Compo | MOY | PTS (sans COEF)
            header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        else:
            # Semestre: MATIÈRE | COEF | M1..M4 | Moy.C | Compo | MOY | PTS
            header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    else:
        if est_primaire:
            # Primaire Mensuel: MATIÈRE | NOTE | MOY | PTS (sans COEF)
            data = [['MATIÈRE', 'NOTE', 'MOY', 'PTS']]
            nb_cols = 4
        else:
            # Système mensuel: MATIÈRE | COEF | NOTE | MOY | PTS
            data = [['MATIÈRE', 'COEF', 'NOTE', 'MOY', 'PTS']]
            nb_cols = 5
    
    total_coef = 0
    total_moy = 0
    total_points = 0
    nb_matieres_avec_moy = 0
    
    for matiere in bulletin_data['matieres']:
        nom_matiere = str(matiere.get('matiere', '-'))
        if hasattr(matiere.get('matiere'), 'nom'):
            nom_matiere = matiere['matiere'].nom
        
        coef = float(matiere.get('coefficient', 1))
        moy_continue = matiere.get('moyenne_continue')
        note_compo = matiere.get('note_composition')
        moyenne = matiere.get('moyenne') or matiere.get('moyenne_calculee')
        points = matiere.get('points')
        
        if points is None and moyenne:
            points = moyenne * coef
        
        # Récupérer les moyennes mensuelles si disponibles
        moyennes_mensuelles = matiere.get('moyennes_mensuelles', [])
        
        total_coef += coef
        if moyenne:
            total_moy += moyenne
            nb_matieres_avec_moy += 1
        if points:
            total_points += points
        
        if system_type_indiv in ['trimestriel', 'semestriel'] and mois_labels:
            # Construire la ligne avec les détails mensuels
            if est_primaire:
                row = [nom_matiere]  # Sans COEF pour primaire
            else:
                row = [nom_matiere, f"{coef:.0f}"]
            
            # Ajouter les notes mensuelles
            for i, mois_label in enumerate(mois_labels):
                note_mois = '-'
                if moyennes_mensuelles and i < len(moyennes_mensuelles):
                    moy_mens = moyennes_mensuelles[i]
                    if isinstance(moy_mens, dict):
                        val = moy_mens.get('moyenne')
                        if val is not None:
                            note_mois = f"{val:.2f}"
                    elif moy_mens is not None:
                        note_mois = f"{moy_mens:.2f}"
                row.append(note_mois)
            
            # Ajouter Moy.C, Compo, MOY, PTS
            row.append(f"{moy_continue:.2f}" if moy_continue else '-')
            row.append(f"{note_compo:.2f}" if note_compo else '-')
            row.append(f"{moyenne:.2f}" if moyenne else '-')
            row.append(f"{points:.2f}" if points else '-')
            data.append(row)
        else:
            # Mensuel
            cours_str = f"{moy_continue:.2f}" if moy_continue else '-'
            moyenne_str = f"{moyenne:.2f}" if moyenne else '-'
            points_str = f"{points:.2f}" if points else '-'
            if est_primaire:
                # Primaire: sans COEF
                data.append([nom_matiere, cours_str, moyenne_str, points_str])
            else:
                data.append([nom_matiere, f"{coef:.0f}", cours_str, moyenne_str, points_str])
    
    # Ligne TOTAL
    if system_type_indiv in ['trimestriel', 'semestriel'] and mois_labels:
        if est_primaire:
            total_row = ['TOTAL']  # Sans COEF pour primaire
        else:
            total_row = ['TOTAL', f"{total_coef:.0f}"]
        total_row += ['-'] * len(mois_labels)  # Colonnes mois vides
        total_row += ['-', '-']  # Moy.C et Compo
        total_row.append(f"{total_moy:.0f}" if nb_matieres_avec_moy else '-')
        total_row.append(f"{total_points:.2f}")
        data.append(total_row)
    else:
        if est_primaire:
            # Primaire: sans COEF
            data.append(['TOTAL', '-', f"{total_moy:.0f}" if nb_matieres_avec_moy else '-', f"{total_points:.2f}"])
        else:
            data.append(['TOTAL', f"{total_coef:.0f}", '-', f"{total_moy:.0f}" if nb_matieres_avec_moy else '-', f"{total_points:.2f}"])
    
    # Calculer les largeurs de colonnes
    if system_type_indiv in ['trimestriel', 'semestriel'] and mois_labels:
        col_matiere = table_total_width * 0.22
        nb_autres_cols = nb_cols - 1
        col_autres = (table_total_width - col_matiere) / nb_autres_cols
        col_widths = [col_matiere] + [col_autres] * nb_autres_cols
    else:
        if est_primaire:
            # Primaire: 4 colonnes (sans COEF)
            col_matiere = table_total_width * 0.40
            col_autres = (table_total_width - col_matiere) / 3
            col_widths = [col_matiere, col_autres, col_autres, col_autres]
        else:
            col_matiere = table_total_width * 0.35
            col_autres = (table_total_width - col_matiere) / 4
            col_widths = [col_matiere, col_autres, col_autres, col_autres, col_autres]
    
    table = Table(data, colWidths=col_widths)
    
    # Style du tableau - base
    style = [
        # En-tête bleu
        ('BACKGROUND', (0, 0), (-1, 0), BLEU_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # Corps
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (0, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        
        # Grille
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        
        # Alternance de couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        
        # Ligne TOTAL
        ('BACKGROUND', (0, -1), (-1, -1), GRIS_TOTAL),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ]
    
    # Colonnes MOY et PTS avec couleurs (position varie selon le système et le niveau)
    if system_type_indiv in ['trimestriel', 'semestriel'] and mois_labels:
        # Colonnes dynamiques: MOY = avant-dernière colonne, PTS = dernière colonne
        col_moy = nb_cols - 2
        col_pts = nb_cols - 1
        style.extend([
            ('BACKGROUND', (col_moy, 1), (col_moy, -2), VERT_MOY),
            ('FONTNAME', (col_moy, 1), (col_moy, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (col_pts, 1), (col_pts, -2), ROUGE_PTS),
            ('FONTNAME', (col_pts, 1), (col_pts, -1), 'Helvetica-Bold'),
            # Réduire la taille de police pour les colonnes nombreuses
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
        ])
    else:
        if est_primaire:
            # Primaire: 4 colonnes: MATIÈRE | NOTE | MOY | PTS (sans COEF)
            style.extend([
                ('BACKGROUND', (2, 1), (2, -2), VERT_MOY),    # MOY = colonne 2
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (3, 1), (3, -2), ROUGE_PTS),   # PTS = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
            ])
        else:
            # 5 colonnes: MATIÈRE | COEF | NOTE | MOY | PTS
            style.extend([
                ('BACKGROUND', (3, 1), (3, -2), VERT_MOY),    # MOY = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (4, 1), (4, -2), ROUGE_PTS),   # PTS = colonne 4
                ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ])
    
    table.setStyle(TableStyle(style))
    
    # Calculer la hauteur du tableau et le dessiner aligné avec la ligne
    table_w, table_h = table.wrap(width, height)
    table.drawOn(c, margin_left, y - table_h)
    
    # ===== SECTION RÉSULTATS =====
    y = y - table_h - 0.8*cm
    
    # Trois colonnes alignées avec le tableau (de 1.2cm à width-1.2cm)
    result_total_width = width - 2.4*cm
    col_width = (result_total_width - 0.4*cm) / 3  # 3 colonnes avec espacement
    col_height = 1.4*cm
    start_x = 1.2*cm
    
    # MOYENNE GÉNÉRALE
    c.setFillColor(colors.white)
    c.rect(start_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(start_x + col_width/2, y - 0.35*cm, "MOYENNE GÉNÉRALE")
    c.setFont("Helvetica-Bold", 16)
    moy_gen = bulletin_data.get('moyenne_generale')
    moy_str = f"{moy_gen:.2f}/20" if moy_gen else '-'
    c.drawCentredString(start_x + col_width/2, y - 1*cm, moy_str)
    
    # RANG
    rang_x = start_x + col_width + 0.2*cm
    c.setFillColor(colors.white)
    c.rect(rang_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(rang_x + col_width/2, y - 0.35*cm, "RANG")
    c.setFont("Helvetica-Bold", 16)
    rang = bulletin_data.get('rang', '-')
    c.drawCentredString(rang_x + col_width/2, y - 1*cm, str(rang))
    
    # MENTION (avec badge coloré)
    mention_x = start_x + 2*(col_width + 0.2*cm)
    c.setFillColor(colors.white)
    c.rect(mention_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(mention_x + col_width/2, y - 0.35*cm, "MENTION")
    
    # Badge de mention
    mention = bulletin_data.get('mention', '-')
    mention_upper = mention.upper() if mention else '-'
    mention_color = MENTION_COLORS.get(mention_upper, colors.HexColor('#666666'))
    
    # Dessiner le badge
    badge_width = min(4*cm, col_width - 0.4*cm)
    badge_height = 0.6*cm
    badge_x = mention_x + (col_width - badge_width) / 2
    badge_y = y - 1.15*cm
    
    c.setFillColor(mention_color)
    c.roundRect(badge_x, badge_y, badge_width, badge_height, 3, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(mention_x + col_width/2, badge_y + 0.15*cm, mention_upper)
    
    # ===== APPRÉCIATION =====
    y -= col_height + 0.6*cm
    
    appreciation_height = 1.2*cm
    c.setFillColor(BLEU_APPRECIATION)
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#90caf9'))
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=0, stroke=1)
    
    c.setFillColor(colors.HexColor('#1565c0'))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(1.4*cm, y - 0.35*cm, "APPRÉCIATION DU CONSEIL DE CLASSE")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Oblique", 8)
    appreciation = bulletin_data.get('appreciation') or 'Bon travail. Continuez vos efforts.'
    if appreciation and len(appreciation) > 100:
        appreciation = appreciation[:97] + '...'
    c.drawString(1.4*cm, y - 0.85*cm, appreciation)
    
    # ===== SIGNATURES =====
    y -= appreciation_height + 0.8*cm
    
    sig_width = 6*cm
    
    # Censeur (aligné à gauche avec 1.2cm)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(1.2*cm + sig_width/2, y, "Censeur de l'établissement")
    c.setStrokeColor(colors.black)
    c.line(1.2*cm + 0.5*cm, y - 1.5*cm, 1.2*cm + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(1.2*cm + sig_width/2, y - 1.8*cm, "Signature")
    
    # Directeur Général (aligné à droite avec 1.2cm)
    dir_x = width - 1.2*cm - sig_width
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(dir_x + sig_width/2, y, "Directeur Général")
    c.line(dir_x + 0.5*cm, y - 1.5*cm, dir_x + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(dir_x + sig_width/2, y - 1.8*cm, "Signature")
    
    # ===== MÉTHODE DE CALCUL =====
    y -= 2.5*cm
    
    c.setFillColor(colors.HexColor('#f5f5f5'))
    c.rect(1.2*cm, y - 1.8*cm, width - 2.4*cm, 1.8*cm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#1976d2'))
    c.setLineWidth(2)
    c.line(1.2*cm, y, 1.2*cm, y - 1.8*cm)
    c.setLineWidth(1)
    
    c.setFillColor(colors.HexColor('#1976d2'))
    c.setFont("Helvetica-Bold", 8)
    
    # Adapter le texte selon le système (pour bulletin individuel)
    system_type_indiv = bulletin_data.get('system_type', 'mensuel')
    
    # Texte pour les points (différent pour primaire)
    points_text = "• Points : Moyenne (pas de coefficient pour le primaire)" if est_primaire else "• Points : Moyenne × Coefficient"
    
    if system_type_indiv == 'trimestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du trimestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition du trimestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• Moyenne : (Cours + Compo) / 2 (50% chacun) | {points_text}")
    elif system_type_indiv == 'semestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du semestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition (examen) du semestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• Moyenne : (Cours + Compo) / 2 (50% chacun) | {points_text}")
    else:
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Note : Moyenne des notes d'évaluations continues du mois")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moyenne : Égale à la note du cours pour le système mensuel")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    
    c.drawString(1.4*cm, y - 1.5*cm, "• Mentions : Excellent ≥18.5 | Très bien ≥16.5 | Bien ≥14.5 | Assez bien ≥12.5 | Passable ≥10 | Insuffisant <10")
    
    # ===== PIED DE PAGE avec infos dynamiques de l'école =====
    c.setFillColor(colors.HexColor('#999999'))
    c.setFont("Helvetica", 6)
    # Construire le footer avec les infos de l'école
    footer_parts = [f"© 2025 {ecole.nom if ecole else 'Myschool'}. Tous droits réservés."]
    if ecole and ecole.telephone:
        footer_parts.append(f"Tél: {ecole.telephone}")
    if ecole and ecole.email:
        footer_parts.append(ecole.email)
    c.drawCentredString(width/2, 0.8*cm, " | ".join(footer_parts))
    c.drawCentredString(width/2, 0.5*cm, 
                       f"Bulletin généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    
    c.showPage()
    c.save()
    
    buffer.seek(0)
    return buffer


def _dessiner_bulletin_maternelle(c, bulletin_data, width, height, y, ecole):
    """Dessine un bulletin spécifique pour la maternelle/garderie avec appréciations"""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table, TableStyle
    
    # Couleurs
    BLEU_HEADER = colors.HexColor('#4a90d9')
    BLEU_CLAIR = colors.HexColor('#e8f4fd')
    VERT_ACQUIS = colors.HexColor('#c8e6c9')
    JAUNE_ENCOURS = colors.HexColor('#fff9c4')
    ROUGE_NONACQUIS = colors.HexColor('#ffcdd2')
    GRIS_TOTAL = colors.HexColor('#37474f')
    
    table_total_width = width - 2.4*cm
    margin_left = 1.2*cm
    
    # Récupérer les matières/appréciations depuis bulletin_data
    matieres = bulletin_data.get('matieres', [])
    
    # En-tête du tableau
    header = ['DOMAINE D\'ACTIVITÉ', 'APPRÉCIATION', 'COMMENTAIRE']
    data = [header]
    
    # Mapping des appréciations
    APPRECIATION_DISPLAY = {
        'TRES_BIEN_ACQUIS': 'Très Bien Acquis',
        'BIEN_ACQUIS': 'Bien Acquis',
        'EN_COURS': 'En Cours d\'Acquisition',
        'NON_ACQUIS': 'Non Acquis',
    }
    
    # Observations automatiques
    OBSERVATIONS_AUTO = {
        'TRES_BIEN_ACQUIS': 'Excellent travail !',
        'BIEN_ACQUIS': 'Bon niveau',
        'EN_COURS': 'Persévère !',
        'NON_ACQUIS': 'À renforcer',
    }
    
    # Afficher les matières avec leurs appréciations
    if matieres:
        for mat in matieres:
            # Récupérer le nom de la matière
            matiere_nom = mat.get('matiere', '-')
            if isinstance(matiere_nom, str):
                nom = matiere_nom
            elif hasattr(matiere_nom, 'nom'):
                nom = matiere_nom.nom
            else:
                nom = str(matiere_nom)
            
            # Récupérer l'appréciation
            appreciation = mat.get('appreciation')
            appreciation_display_val = mat.get('appreciation_display', '-')
            commentaire = mat.get('commentaire', '')
            absent = mat.get('absent', False)
            
            if absent:
                appreciation_display = 'Absent'
                commentaire = 'Absent(e) lors de l\'évaluation'
            elif appreciation:
                appreciation_display = APPRECIATION_DISPLAY.get(appreciation, appreciation_display_val)
                if not commentaire or commentaire == '-':
                    commentaire = OBSERVATIONS_AUTO.get(appreciation, '-')
            else:
                appreciation_display = '-'
                commentaire = '-'
            
            data.append([nom, appreciation_display, commentaire])
    else:
        data.append(['Aucune activité', '-', '-'])
    
    # Largeurs de colonnes
    col_widths = [table_total_width * 0.35, table_total_width * 0.30, table_total_width * 0.35]
    
    table = Table(data, colWidths=col_widths)
    
    # Styles du tableau
    style = [
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), BLEU_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Corps - Matière et Appréciation
        ('FONTNAME', (0, 1), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (1, -1), 9),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (0, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        
        # Commentaires - Police plus petite pour éviter débordement
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica'),
        ('FONTSIZE', (2, 1), (2, -1), 7),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('LEFTPADDING', (2, 1), (2, -1), 4),
        
        # Grille
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        
        # Alternance de couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]
    
    # Colorer les appréciations selon leur valeur
    for i, row in enumerate(data[1:], start=1):
        appreciation = row[1]
        if 'Très Bien' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), VERT_ACQUIS))
        elif 'Bien Acquis' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), colors.HexColor('#a5d6a7')))
        elif 'En Cours' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), JAUNE_ENCOURS))
        elif 'Non Acquis' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), ROUGE_NONACQUIS))
    
    table.setStyle(TableStyle(style))
    
    # Dessiner le tableau
    table_w, table_h = table.wrap(width, height)
    table.drawOn(c, margin_left, y - table_h)
    
    # ===== SECTION RÉSULTATS (Taux, Rang, Appréciation) =====
    y = y - table_h - 0.8*cm
    
    # Récupérer les données
    taux_acquisition = bulletin_data.get('moyenne_generale')
    rang = bulletin_data.get('rang')
    mention = bulletin_data.get('mention') or 'Suivi continu'
    
    # Couleurs des cartes
    VERT_TAUX = colors.HexColor('#27ae60')
    BLEU_RANG = colors.HexColor('#3498db')
    VIOLET_APPRECIATION = colors.HexColor('#9b59b6')
    
    # Largeur de chaque carte (3 cartes)
    card_width = (table_total_width - 0.4*cm) / 3
    card_height = 1.8*cm
    
    # === Carte 1: TAUX D'ACQUISITION ===
    card_x = margin_left
    c.setFillColor(VERT_TAUX)
    c.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(card_x + card_width/2, y - 0.4*cm, "TAUX D'ACQUISITION")
    
    c.setFont("Helvetica-Bold", 16)
    taux_str = f"{taux_acquisition:.1f}%" if taux_acquisition else "-"
    c.drawCentredString(card_x + card_width/2, y - 1.2*cm, taux_str)
    
    # === Carte 2: RANG ===
    card_x = margin_left + card_width + 0.2*cm
    c.setFillColor(BLEU_RANG)
    c.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(card_x + card_width/2, y - 0.4*cm, "RANG")
    
    c.setFont("Helvetica-Bold", 16)
    rang_str = str(rang) if rang else "-"
    c.drawCentredString(card_x + card_width/2, y - 1.2*cm, rang_str)
    
    # === Carte 3: APPRÉCIATION ===
    card_x = margin_left + 2*(card_width + 0.2*cm)
    c.setFillColor(VIOLET_APPRECIATION)
    c.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(card_x + card_width/2, y - 0.4*cm, "APPRÉCIATION")
    
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(card_x + card_width/2, y - 1.2*cm, str(mention))
    
    # Ajuster y pour la suite
    y = y - card_height - 0.3*cm
    
    # ===== SIGNATURES =====
    y -= 1.5*cm
    
    sig_width = 6*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(1.2*cm + sig_width/2, y, "L'Éducateur(trice)")
    c.setStrokeColor(colors.black)
    c.line(1.2*cm + 0.5*cm, y - 1.5*cm, 1.2*cm + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(1.2*cm + sig_width/2, y - 1.8*cm, "Signature")
    
    dir_x = width - 1.2*cm - sig_width
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(dir_x + sig_width/2, y, "Le Directeur / La Directrice")
    c.line(dir_x + 0.5*cm, y - 1.5*cm, dir_x + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(dir_x + sig_width/2, y - 1.8*cm, "Signature")
    
    # ===== MÉTHODE D'ÉVALUATION =====
    y -= 2.5*cm
    
    c.setFillColor(colors.HexColor('#f5f5f5'))
    c.rect(1.2*cm, y - 1.5*cm, table_total_width, 1.5*cm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#1976d2'))
    c.setLineWidth(2)
    c.line(1.2*cm, y, 1.2*cm, y - 1.5*cm)
    c.setLineWidth(1)
    
    c.setFillColor(colors.HexColor('#1976d2'))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(1.4*cm, y - 0.3*cm, "MÉTHODE D'ÉVALUATION - MATERNELLE / GARDERIE")
    
    c.setFillColor(colors.HexColor('#333333'))
    c.setFont("Helvetica", 7)
    c.drawString(1.4*cm, y - 0.6*cm, "• Évaluation qualitative par compétences (pas de notes numériques)")
    c.drawString(1.4*cm, y - 0.9*cm, "• Appréciations : Très Bien Acquis | Bien Acquis | En Cours d'Acquisition | Non Acquis")
    c.drawString(1.4*cm, y - 1.2*cm, "• Suivi individualisé du développement de l'enfant")
    
    # ===== PIED DE PAGE =====
    c.setFillColor(colors.HexColor('#999999'))
    c.setFont("Helvetica", 6)
    footer_parts = [f"© 2025 {ecole.nom if ecole else 'Myschool'}. Tous droits réservés."]
    if ecole and ecole.telephone:
        footer_parts.append(f"Tél: {ecole.telephone}")
    if ecole and ecole.email:
        footer_parts.append(ecole.email)
    c.drawCentredString(width/2, 0.8*cm, " | ".join(footer_parts))


def generer_excel(bulletin_data):
    """Génère un fichier Excel du bulletin"""
    if not EXCEL_AVAILABLE:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulletin"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-tête
    ws.merge_cells('A1:F1')
    ws['A1'] = "BULLETIN DE NOTES"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informations
    ws['A3'] = f"Élève: {bulletin_data['eleve']}"
    ws['A3'].font = subtitle_font
    
    ws['A4'] = f"Classe: {bulletin_data['classe']}"
    ws['D4'] = f"Période: {bulletin_data['periode']}"
    
    # En-têtes du tableau
    row = 6
    headers = ['Matière', 'Coefficient', 'Moy. Cours', 'Composition', 'Moyenne', 'Points']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Données
    row += 1
    for matiere in bulletin_data['matieres']:
        # Moyenne cours
        moy_cours = ''
        if matiere['notes_mensuelles']:
            notes_list = []
            for notes in matiere['notes_mensuelles'].values():
                notes_list.extend(notes)
            if notes_list:
                moy_cours = float(sum(notes_list) / len(notes_list))
        
        composition = float(matiere['composition']) if matiere['composition'] else ''
        moyenne = float(matiere['moyenne']) if matiere['moyenne'] else ''
        points = float(matiere['moyenne'] * matiere['coefficient']) if matiere['moyenne'] else ''
        
        ws.cell(row=row, column=1, value=matiere['matiere']).border = border
        ws.cell(row=row, column=2, value=matiere['coefficient']).border = border
        ws.cell(row=row, column=3, value=moy_cours).border = border
        ws.cell(row=row, column=4, value=composition).border = border
        ws.cell(row=row, column=5, value=moyenne).border = border
        ws.cell(row=row, column=6, value=points).border = border
        
        row += 1
    
    # Résultats
    row += 2
    if bulletin_data['moyenne_generale']:
        ws.cell(row=row, column=1, value="Moyenne Générale:").font = subtitle_font
        ws.cell(row=row, column=2, value=float(bulletin_data['moyenne_generale']))
        
        row += 1
        ws.cell(row=row, column=1, value="Rang:").font = subtitle_font
        rang_affiche = bulletin_data['rang'] if bulletin_data['rang'] else "-"
        ws.cell(row=row, column=2, value=rang_affiche)
        
        row += 1
        ws.cell(row=row, column=1, value="Mention:").font = subtitle_font
        ws.cell(row=row, column=2, value=bulletin_data['mention'])
        
        row += 1
        ws.cell(row=row, column=1, value="Appréciation:").font = subtitle_font
        ws.merge_cells(f'B{row}:F{row}')
        ws.cell(row=row, column=2, value=bulletin_data['appreciation'])
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_intelligent_view(request, eleve_id, classe_note_id, periode):
    """Vue pour afficher le bulletin intelligent"""
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Déterminer le système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    
    # Calculer le bulletin
    calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
    bulletin_data = calculateur.generer_bulletin()
    
    return render(request, 'notes/bulletin_intelligent.html', {
        'bulletin': bulletin_data,
        'eleve': eleve,
        'classe_note': classe_note,
        'periode': periode
    })


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_intelligent_pdf(request, eleve_id, classe_note_id, periode):
    """Génère le bulletin en PDF avec filigrane"""
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Déterminer le système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    
    # Déterminer le type de système pour l'affichage
    system_type = 'mensuel'
    if 'TRIMESTRE' in periode:
        system_type = 'trimestriel'
    elif 'SEMESTRE' in periode:
        system_type = 'semestriel'
    
    # Calculer le bulletin
    calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
    bulletin_data = calculateur.generer_bulletin()
    
    # Ajouter le system_type et la période
    bulletin_data['system_type'] = system_type
    bulletin_data['periode'] = periode
    
    # Pour trimestre/semestre, enrichir avec les moyennes mensuelles détaillées
    if system_type in ['trimestriel', 'semestriel']:
        from notes.utils_moyennes_mensuelles import calculer_bulletin_avec_details_mensuels
        
        periode_type_detail = 'trimestre' if system_type == 'trimestriel' else 'semestre'
        matieres_enrichies = []
        
        for mat_data in bulletin_data.get('matieres', []):
            mat_obj = mat_data.get('matiere')
            if mat_obj:
                try:
                    details_mensuels = calculer_bulletin_avec_details_mensuels(
                        eleve, mat_obj, periode_type_detail, periode
                    )
                    mat_data['moyennes_mensuelles'] = details_mensuels.get('moyennes_mensuelles', [])
                    mat_data['note_composition'] = details_mensuels.get('note_composition')
                    mat_data['moyenne_continue'] = details_mensuels.get('moyenne_continue')
                except:
                    pass
            matieres_enrichies.append(mat_data)
        
        bulletin_data['matieres'] = matieres_enrichies
    
    # Chemin du logo
    ecole = eleve.classe.ecole
    logo_path = None
    if hasattr(ecole, 'logo') and ecole.logo:
        logo_path = ecole.logo.path
    
    # Ajouter le matricule aux données du bulletin
    bulletin_data['matricule'] = eleve.matricule
    
    # Générer le PDF avec l'école
    pdf_buffer = generer_pdf_avec_filigrane(bulletin_data, logo_path, ecole)
    
    # Réponse HTTP
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    filename = f"bulletin_{eleve.nom}_{eleve.prenom}_{periode}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_intelligent_excel(request, eleve_id, classe_note_id, periode):
    """Génère le bulletin en Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export n'est pas disponible", status=500)
    
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Déterminer le système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    
    # Calculer le bulletin
    calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
    bulletin_data = calculateur.generer_bulletin()
    
    # Générer l'Excel
    excel_buffer = generer_excel(bulletin_data)
    
    if not excel_buffer:
        return HttpResponse("Erreur lors de la génération Excel", status=500)
    
    # Réponse HTTP
    response = HttpResponse(
        excel_buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"bulletin_{eleve.nom}_{eleve.prenom}_{periode}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def bulletins_classe_pdf(request, classe_note_id, periode):
    """Génère tous les bulletins d'une classe en un seul PDF - VERSION OPTIMISÉE"""
    import re
    
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Récupérer tous les élèves de la classe
    classe_eleve = Classe.objects.filter(
        nom=classe_note.nom,
        annee_scolaire=classe_note.annee_scolaire,
        ecole=classe_note.ecole
    ).first()
    
    if not classe_eleve:
        return HttpResponse("Classe non trouvée", status=404)
    
    eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom'))
    
    if not eleves:
        return HttpResponse("Aucun élève dans cette classe", status=404)
    
    # Déterminer le système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    
    # Chemin du logo
    ecole = classe_note.ecole
    logo_path = None
    if hasattr(ecole, 'logo') and ecole.logo:
        logo_path = ecole.logo.path
    
    # ===== OPTIMISATION: Pré-calculer le classement pour tous les élèves =====
    from notes.calculs_moyennes import calculer_classement_classe
    
    # Récupérer les matières de la classe (MatiereNote.classe est une FK vers ClasseNote)
    matieres = MatiereNote.objects.filter(classe=classe_note)
    
    # Déterminer le type de système
    system_type = 'mensuel'
    if 'TRIMESTRE' in periode:
        system_type = 'trimestriel'
    elif 'SEMESTRE' in periode:
        system_type = 'semestriel'
    
    # Calculer le classement une seule fois pour toute la classe
    classement_result = calculer_classement_classe(eleves, matieres, periode, system_type)
    
    # Extraire les données pré-calculées
    rang_map = classement_result.get('rang_map', {})
    moyennes_map = classement_result.get('moyennes_par_eleve', {})
    details_map = classement_result.get('details_par_eleve', {})
    
    total_eleves = len(eleves)
    
    # Pré-charger le logo une seule fois
    logo_reader = None
    if logo_path:
        try:
            logo_reader = ImageReader(logo_path)
        except:
            pass
    
    # ===== Générer un seul PDF multi-pages =====
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    
    # Import pour les détails mensuels (trimestre/semestre)
    from notes.utils_moyennes_mensuelles import calculer_bulletin_avec_details_mensuels
    from notes.calculs_moyennes import detecter_niveau_scolaire
    
    # Déterminer le type de période pour les détails mensuels
    periode_type_detail = None
    if system_type == 'trimestriel':
        periode_type_detail = 'trimestre'
    elif system_type == 'semestriel':
        periode_type_detail = 'semestre'
    
    # Détecter le niveau scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe_note.nom)
    est_maternelle = (niveau_scolaire == 'MATERNELLE')
    
    for idx, eleve in enumerate(eleves):
        try:
            # Pour la maternelle, toujours utiliser le calculateur spécifique
            if est_maternelle:
                calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
                bulletin_data = calculateur.generer_bulletin()
                bulletin_data['matricule'] = eleve.matricule
                bulletin_data['total_eleves'] = total_eleves
                bulletin_data['system_type'] = system_type
            # Utiliser les données pré-calculées si disponibles
            elif eleve.id in details_map:
                details = details_map[eleve.id]
                matieres_data = details.get('details_matieres', [])
                
                # Pour trimestre/semestre, enrichir avec les moyennes mensuelles détaillées
                if periode_type_detail and matieres:
                    matieres_enrichies = []
                    for mat_data in matieres_data:
                        mat_obj = mat_data.get('matiere')
                        if mat_obj:
                            # Récupérer les détails mensuels
                            try:
                                details_mensuels = calculer_bulletin_avec_details_mensuels(
                                    eleve, mat_obj, periode_type_detail, periode
                                )
                                mat_data['moyennes_mensuelles'] = details_mensuels.get('moyennes_mensuelles', [])
                                mat_data['note_composition'] = details_mensuels.get('note_composition')
                                mat_data['moyenne_continue'] = details_mensuels.get('moyenne_continue')
                            except:
                                pass
                        matieres_enrichies.append(mat_data)
                    matieres_data = matieres_enrichies
                
                bulletin_data = {
                    'eleve': f"{eleve.prenom} {eleve.nom}",
                    'classe': classe_note.nom,
                    'periode': periode,
                    'system_type': system_type,
                    'matieres': matieres_data,
                    'moyenne_generale': details.get('moyenne_generale'),
                    'total_points': details.get('total_points'),
                    'total_coefficients': details.get('total_coefficients'),
                    'rang': rang_map.get(eleve.id, '-'),
                    'mention': obtenir_mention_intelligente(details.get('moyenne_generale'), niveau_scolaire),
                    'matricule': eleve.matricule,
                    'total_eleves': total_eleves,
                }
            else:
                # Fallback: calculer si pas dans le cache
                calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
                bulletin_data = calculateur.generer_bulletin()
                bulletin_data['rang'] = rang_map.get(eleve.id, '-')
                bulletin_data['matricule'] = eleve.matricule
                bulletin_data['total_eleves'] = total_eleves
                bulletin_data['system_type'] = system_type
            
            # Dessiner le bulletin sur la page courante (passer le logo pré-chargé)
            _dessiner_bulletin_page(c, bulletin_data, logo_path, ecole, logo_reader)
            
            # Nouvelle page sauf pour le dernier élève
            if idx < len(eleves) - 1:
                c.showPage()
                
        except Exception as e:
            print(f"Erreur pour {eleve.nom} {eleve.prenom}: {str(e)}")
            # Ajouter une page vide avec message d'erreur
            c.setFont("Helvetica", 12)
            c.drawString(2*cm, A4[1]/2, f"Erreur pour {eleve.nom} {eleve.prenom}")
            if idx < len(eleves) - 1:
                c.showPage()
            continue
    
    c.save()
    buffer.seek(0)
    
    # Nettoyer le nom de fichier
    nom_classe_clean = re.sub(r'[^\w\s-]', '', classe_note.nom).replace(' ', '_')
    filename = f"bulletins_{nom_classe_clean}_{periode}.pdf"
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def _dessiner_bulletin_page(c, bulletin_data, logo_path, ecole, logo_reader=None):
    """Dessine un bulletin sur la page courante du canvas - fonction interne optimisée"""
    width, height = A4
    
    # Couleurs du design (constantes)
    BLEU_HEADER = colors.HexColor('#4a90d9')
    BLEU_CLAIR = colors.HexColor('#e8f4fd')
    BLEU_APPRECIATION = colors.HexColor('#d6eaf8')
    VERT_MOY = colors.HexColor('#c8e6c9')
    ROUGE_PTS = colors.HexColor('#ffcdd2')
    ROUGE_DRAPEAU = colors.HexColor('#CE1126')
    JAUNE_DRAPEAU = colors.HexColor('#FCD116')
    VERT_DRAPEAU = colors.HexColor('#009460')
    GRIS_TOTAL = colors.HexColor('#37474f')
    
    MENTION_COLORS = {
        'EXCELLENT': colors.HexColor('#1b5e20'),
        'TRÈS BIEN': colors.HexColor('#2e7d32'),
        'BIEN': colors.HexColor('#0277bd'),
        'ASSEZ BIEN': colors.HexColor('#f9a825'),
        'PASSABLE': colors.HexColor('#ef6c00'),
        'INSUFFISANT': colors.HexColor('#c62828'),
        'FAIBLE': colors.HexColor('#b71c1c'),
    }
    
    # Utiliser le logo pré-chargé ou le charger si nécessaire
    img_reader = logo_reader
    if img_reader is None and logo_path:
        try:
            img_reader = ImageReader(logo_path)
        except:
            pass
    
    # ===== FILIGRANE =====
    if img_reader:
        try:
            c.saveState()
            c.setFillAlpha(0.08)
            filigrane_size = 14 * cm
            x = (width - filigrane_size) / 2
            y = (height - filigrane_size) / 2
            c.drawImage(img_reader, x, y, width=filigrane_size, height=filigrane_size, 
                       preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except:
            pass
    
    # ===== EN-TÊTE =====
    if img_reader:
        try:
            c.drawImage(img_reader, 1.2*cm, height - 3.2*cm, width=2.5*cm, height=2.5*cm, 
                       preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Drapeau guinéen
    drapeau_x = width - 2.5*cm
    drapeau_y = height - 2.2*cm
    drapeau_w = 1.2*cm
    drapeau_h = 0.8*cm
    c.setFillColor(ROUGE_DRAPEAU)
    c.rect(drapeau_x, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setFillColor(JAUNE_DRAPEAU)
    c.rect(drapeau_x + drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setFillColor(VERT_DRAPEAU)
    c.rect(drapeau_x + 2*drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.rect(drapeau_x, drapeau_y, drapeau_w, drapeau_h, fill=0, stroke=1)
    
    # Textes en-tête
    y_header = height - 1.2*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width/2, y_header, "RÉPUBLIQUE DE GUINÉE")
    
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(ROUGE_DRAPEAU)
    c.drawCentredString(width/2 - 1.8*cm, y_header, "Travail")
    c.setFillColor(JAUNE_DRAPEAU)
    c.drawCentredString(width/2, y_header, "Justice")
    c.setFillColor(VERT_DRAPEAU)
    c.drawCentredString(width/2 + 1.8*cm, y_header, "Solidarité")
    c.setFillColor(colors.black)
    c.drawCentredString(width/2 - 0.9*cm, y_header, "-")
    c.drawCentredString(width/2 + 0.9*cm, y_header, "-")
    
    y_header -= 0.45*cm
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(width/2, y_header, "MPU-A")
    
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 10)
    nom_ecole = ecole.nom.upper() if ecole else "GROUPE SCOLAIRE"
    c.drawCentredString(width/2, y_header, nom_ecole)
    
    y_header -= 0.6*cm
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width/2, y_header, f"BULLETIN DE NOTES - {bulletin_data['periode'].upper()}")
    
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor('#666666'))
    c.drawRightString(width - 1.2*cm, y_header, "Année Scolaire 2024-2025")
    
    # Ligne de séparation
    y_header -= 0.4*cm
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.5)
    c.line(1.2*cm, y_header, width - 1.2*cm, y_header)
    c.setLineWidth(1)
    
    # ===== INFORMATIONS ÉLÈVE =====
    y = y_header - 0.6*cm
    
    eleve_nom_complet = bulletin_data.get('eleve', '')
    parties = eleve_nom_complet.split(' ', 1) if eleve_nom_complet else ['', '']
    prenom = parties[0] if len(parties) > 0 else ''
    nom = parties[1] if len(parties) > 1 else ''
    
    info_data = [
        [('PRÉNOM', prenom.title()), ('NOM', nom.upper()), ('MATRICULE', bulletin_data.get('matricule', '-'))],
        [('CLASSE', bulletin_data.get('classe', '-')), ('PÉRIODE', bulletin_data.get('periode', '-')), ('EFFECTIF', f"{bulletin_data.get('total_eleves', '-')} élèves")],
    ]
    
    info_total_width = width - 2.4*cm
    box_width = (info_total_width - 0.4*cm) / 3
    box_height = 0.9*cm
    
    for row_idx, row in enumerate(info_data):
        for col_idx, (label, value) in enumerate(row):
            x = 1.2*cm + col_idx * (box_width + 0.2*cm)
            box_y = y - row_idx * (box_height + 0.15*cm)
            
            c.setFillColor(BLEU_CLAIR)
            c.rect(x, box_y - box_height, box_width, box_height, fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor('#b3d4fc'))
            c.rect(x, box_y - box_height, box_width, box_height, fill=0, stroke=1)
            
            c.setFillColor(colors.HexColor('#1976d2'))
            c.setFont("Helvetica-Bold", 7)
            c.drawString(x + 0.15*cm, box_y - 0.3*cm, label)
            
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x + 0.15*cm, box_y - 0.7*cm, str(value))
    
    # ===== TABLEAU DES NOTES =====
    y -= 2.3*cm
    
    table_total_width = width - 2.4*cm
    margin_left = 1.2*cm
    
    # Déterminer le type de système
    system_type = bulletin_data.get('system_type', 'mensuel')
    periode = bulletin_data.get('periode', '')
    
    # Détecter le niveau scolaire
    from notes.calculs_moyennes import detecter_niveau_scolaire
    classe_nom = bulletin_data.get('classe', '')
    niveau_scolaire = detecter_niveau_scolaire(classe_nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    est_maternelle = (niveau_scolaire == 'MATERNELLE')
    
    # MATERNELLE/GARDERIE: Afficher les appréciations au lieu des notes
    if est_maternelle:
        _dessiner_bulletin_maternelle(c, bulletin_data, width, height, y, ecole)
        # Ne pas faire c.save() ici - c'est géré par la fonction appelante
        return
    
    # Déterminer les mois selon la période pour trimestre/semestre
    # IMPORTANT: Le dernier mois est remplacé par "Compo" (composition)
    mois_labels = []
    if system_type == 'trimestriel':
        if 'TRIMESTRE_1' in periode or '1' in periode:
            mois_labels = ['Oct.', 'Nov.']  # Déc. = Compo
        elif 'TRIMESTRE_2' in periode or '2' in periode:
            mois_labels = ['Jan.', 'Fév.']  # Mars = Compo
        elif 'TRIMESTRE_3' in periode or '3' in periode:
            mois_labels = ['Avr.', 'Mai']   # Juin = Compo
    elif system_type == 'semestriel':
        if 'SEMESTRE_1' in periode or '1' in periode:
            mois_labels = ['Oct.', 'Nov.', 'Déc.', 'Jan.']  # Fév. = Compo
        elif 'SEMESTRE_2' in periode or '2' in periode:
            mois_labels = ['Mars', 'Avr.', 'Mai', 'Juin']   # Juil. = Compo
    
    # Adapter les colonnes selon le système ET le niveau (primaire = sans COEF ni PTS)
    # Structure: MATIÈRE | [COEF] | Mois1 | Mois2 | [Mois...] | Moy.C | Compo | MOY | [PTS]
    if system_type == 'trimestriel' and mois_labels:
        if est_primaire:
            header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY']
        else:
            header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    elif system_type == 'semestriel' and mois_labels:
        if est_primaire:
            header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY']
        else:
            header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    else:
        if est_primaire:
            data = [['MATIÈRE', 'NOTE', 'MOY']]
            nb_cols = 3
        else:
            data = [['MATIÈRE', 'COEF', 'NOTE', 'MOY', 'PTS']]
            nb_cols = 5
    
    total_coef = 0
    total_moy = 0
    total_points = 0
    nb_matieres_avec_moy = 0
    
    for matiere in bulletin_data['matieres']:
        nom_matiere = str(matiere.get('matiere', '-'))
        if hasattr(matiere.get('matiere'), 'nom'):
            nom_matiere = matiere['matiere'].nom
        
        coef = float(matiere.get('coefficient', 1))
        moy_continue = matiere.get('moyenne_continue')
        note_compo = matiere.get('note_composition')
        moyenne = matiere.get('moyenne') or matiere.get('moyenne_calculee')
        points = matiere.get('points')
        
        if points is None and moyenne:
            points = moyenne * coef
        
        # Récupérer les moyennes mensuelles si disponibles
        moyennes_mensuelles = matiere.get('moyennes_mensuelles', [])
        
        total_coef += coef
        if moyenne:
            total_moy += moyenne
            nb_matieres_avec_moy += 1
        if points:
            total_points += points
        
        if system_type in ['trimestriel', 'semestriel'] and mois_labels:
            # Construire la ligne avec les détails mensuels
            if est_primaire or est_maternelle:
                row = [nom_matiere]  # Sans COEF pour primaire/maternelle
            else:
                row = [nom_matiere, f"{coef:.0f}"]
            
            # Ajouter les notes mensuelles
            for i, mois_label in enumerate(mois_labels):
                note_mois = '-'
                if moyennes_mensuelles and i < len(moyennes_mensuelles):
                    moy_mens = moyennes_mensuelles[i]
                    if isinstance(moy_mens, dict):
                        val = moy_mens.get('moyenne')
                        if val is not None:
                            note_mois = f"{val:.2f}"
                    elif moy_mens is not None:
                        note_mois = f"{moy_mens:.2f}"
                row.append(note_mois)
            
            # Ajouter Moy.C, Compo, MOY (et PTS seulement pour collège/lycée)
            row.append(f"{moy_continue:.2f}" if moy_continue else '-')
            row.append(f"{note_compo:.2f}" if note_compo else '-')
            row.append(f"{moyenne:.2f}" if moyenne else '-')
            if not (est_primaire or est_maternelle):
                row.append(f"{points:.2f}" if points else '-')
            data.append(row)
        else:
            # Mensuel
            cours_str = f"{moy_continue:.2f}" if moy_continue else '-'
            moyenne_str = f"{moyenne:.2f}" if moyenne else '-'
            points_str = f"{points:.2f}" if points else '-'
            if est_primaire or est_maternelle:
                data.append([nom_matiere, cours_str, moyenne_str])
            else:
                data.append([nom_matiere, f"{coef:.0f}", cours_str, moyenne_str, points_str])
    
    # Ligne TOTAL
    if system_type in ['trimestriel', 'semestriel'] and mois_labels:
        if est_primaire or est_maternelle:
            total_row = ['TOTAL']
        else:
            total_row = ['TOTAL', f"{total_coef:.0f}"]
        total_row += ['-'] * len(mois_labels)  # Colonnes mois vides
        total_row += ['-', '-']  # Moy.C et Compo
        total_row.append(f"{total_moy:.0f}" if nb_matieres_avec_moy else '-')
        if not (est_primaire or est_maternelle):
            total_row.append(f"{total_points:.2f}")
        data.append(total_row)
    else:
        if est_primaire or est_maternelle:
            data.append(['TOTAL', '-', f"{total_moy:.0f}" if nb_matieres_avec_moy else '-'])
        else:
            data.append(['TOTAL', f"{total_coef:.0f}", '-', f"{total_moy:.0f}" if nb_matieres_avec_moy else '-', f"{total_points:.2f}"])
    
    # Calculer les largeurs de colonnes
    if system_type in ['trimestriel', 'semestriel'] and mois_labels:
        # Répartition: MATIÈRE 25%, reste divisé équitablement
        col_matiere = table_total_width * 0.22
        nb_autres_cols = nb_cols - 1
        col_autres = (table_total_width - col_matiere) / nb_autres_cols
        col_widths = [col_matiere] + [col_autres] * nb_autres_cols
    else:
        if est_primaire or est_maternelle:
            # Primaire/Maternelle: 3 colonnes (MATIÈRE, NOTE, MOY)
            col_matiere = table_total_width * 0.50
            col_autres = (table_total_width - col_matiere) / 2
            col_widths = [col_matiere, col_autres, col_autres]
        else:
            col_matiere = table_total_width * 0.35
            col_autres = (table_total_width - col_matiere) / 4
            col_widths = [col_matiere, col_autres, col_autres, col_autres, col_autres]
    
    table = Table(data, colWidths=col_widths)
    
    # Styles de base
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), BLEU_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (0, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('BACKGROUND', (0, -1), (-1, -1), GRIS_TOTAL),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ]
    
    # Colonnes MOY et PTS avec couleurs (position varie selon le système et le niveau)
    if system_type in ['trimestriel', 'semestriel'] and mois_labels:
        # Colonnes dynamiques: MOY = avant-dernière colonne, PTS = dernière colonne
        col_moy = nb_cols - 2
        col_pts = nb_cols - 1
        style.extend([
            ('BACKGROUND', (col_moy, 1), (col_moy, -2), VERT_MOY),
            ('FONTNAME', (col_moy, 1), (col_moy, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (col_pts, 1), (col_pts, -2), ROUGE_PTS),
            ('FONTNAME', (col_pts, 1), (col_pts, -1), 'Helvetica-Bold'),
            # Réduire la taille de police pour les colonnes nombreuses
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
        ])
    else:
        if est_primaire:
            # Primaire: 4 colonnes: MATIÈRE | NOTE | MOY | PTS (sans COEF)
            style.extend([
                ('BACKGROUND', (2, 1), (2, -2), VERT_MOY),    # MOY = colonne 2
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (3, 1), (3, -2), ROUGE_PTS),   # PTS = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
            ])
        else:
            # 5 colonnes: MATIÈRE | COEF | NOTE | MOY | PTS
            style.extend([
                ('BACKGROUND', (3, 1), (3, -2), VERT_MOY),    # MOY = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (4, 1), (4, -2), ROUGE_PTS),   # PTS = colonne 4
                ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ])
    
    table.setStyle(TableStyle(style))
    table_w, table_h = table.wrap(width, height)
    table.drawOn(c, margin_left, y - table_h)
    
    # ===== SECTION RÉSULTATS =====
    y = y - table_h - 0.8*cm
    
    result_total_width = width - 2.4*cm
    col_width = (result_total_width - 0.4*cm) / 3
    col_height = 1.4*cm
    start_x = 1.2*cm
    
    # MOYENNE GÉNÉRALE
    c.setFillColor(colors.white)
    c.rect(start_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(start_x + col_width/2, y - 0.35*cm, "MOYENNE GÉNÉRALE")
    c.setFont("Helvetica-Bold", 16)
    moy_gen = bulletin_data.get('moyenne_generale')
    moy_str = f"{moy_gen:.2f}/20" if moy_gen else '-'
    c.drawCentredString(start_x + col_width/2, y - 1*cm, moy_str)
    
    # RANG
    rang_x = start_x + col_width + 0.2*cm
    c.setFillColor(colors.white)
    c.rect(rang_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(rang_x + col_width/2, y - 0.35*cm, "RANG")
    c.setFont("Helvetica-Bold", 16)
    rang = bulletin_data.get('rang', '-')
    c.drawCentredString(rang_x + col_width/2, y - 1*cm, str(rang))
    
    # MENTION
    mention_x = start_x + 2*(col_width + 0.2*cm)
    c.setFillColor(colors.white)
    c.rect(mention_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(mention_x + col_width/2, y - 0.35*cm, "MENTION")
    
    mention = bulletin_data.get('mention', '-')
    mention_upper = mention.upper() if mention else '-'
    mention_color = MENTION_COLORS.get(mention_upper, colors.HexColor('#666666'))
    
    badge_width = min(4*cm, col_width - 0.4*cm)
    badge_height = 0.6*cm
    badge_x = mention_x + (col_width - badge_width) / 2
    badge_y = y - 1.15*cm
    
    c.setFillColor(mention_color)
    c.roundRect(badge_x, badge_y, badge_width, badge_height, 3, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(mention_x + col_width/2, badge_y + 0.15*cm, mention_upper)
    
    # ===== APPRÉCIATION =====
    y -= col_height + 0.6*cm
    
    appreciation_height = 1.2*cm
    c.setFillColor(BLEU_APPRECIATION)
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#90caf9'))
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=0, stroke=1)
    
    c.setFillColor(colors.HexColor('#1565c0'))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(1.4*cm, y - 0.35*cm, "APPRÉCIATION DU CONSEIL DE CLASSE")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Oblique", 8)
    appreciation = bulletin_data.get('appreciation') or 'Bon travail. Continuez vos efforts.'
    if appreciation and len(appreciation) > 100:
        appreciation = appreciation[:97] + '...'
    c.drawString(1.4*cm, y - 0.85*cm, appreciation)
    
    # ===== SIGNATURES =====
    y -= appreciation_height + 0.8*cm
    
    sig_width = 6*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(1.2*cm + sig_width/2, y, "Censeur de l'établissement")
    c.setStrokeColor(colors.black)
    c.line(1.2*cm + 0.5*cm, y - 1.5*cm, 1.2*cm + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(1.2*cm + sig_width/2, y - 1.8*cm, "Signature")
    
    dir_x = width - 1.2*cm - sig_width
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(dir_x + sig_width/2, y, "Directeur Général")
    c.line(dir_x + 0.5*cm, y - 1.5*cm, dir_x + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(dir_x + sig_width/2, y - 1.8*cm, "Signature")
    
    # ===== MÉTHODE DE CALCUL =====
    y -= 2.5*cm
    
    c.setFillColor(colors.HexColor('#f5f5f5'))
    c.rect(1.2*cm, y - 1.8*cm, width - 2.4*cm, 1.8*cm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#1976d2'))
    c.setLineWidth(2)
    c.line(1.2*cm, y, 1.2*cm, y - 1.8*cm)
    c.setLineWidth(1)
    
    c.setFillColor(colors.HexColor('#1976d2'))
    c.setFont("Helvetica-Bold", 8)
    
    # Texte pour les points (différent pour primaire)
    points_text = "• Points : Moyenne (pas de coefficient pour le primaire)" if est_primaire else "• Points : Moyenne × Coefficient"
    
    # Adapter le texte selon le système
    if system_type == 'trimestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du trimestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition du trimestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• Moyenne : (Cours + Compo) / 2 (50% chacun) | {points_text}")
    elif system_type == 'semestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du semestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition (examen) du semestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• Moyenne : (Cours + Compo) / 2 (50% chacun) | {points_text}")
    else:
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Note : Moyenne des notes d'évaluations continues du mois")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moyenne : Égale à la note du cours pour le système mensuel")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    
    c.drawString(1.4*cm, y - 1.5*cm, "• Mentions : Excellent ≥18.5 | Très bien ≥16.5 | Bien ≥14.5 | Assez bien ≥12.5 | Passable ≥10 | Insuffisant <10")
    
    # ===== PIED DE PAGE avec infos dynamiques de l'école =====
    c.setFillColor(colors.HexColor('#999999'))
    c.setFont("Helvetica", 6)
    # Construire le footer avec les infos de l'école
    footer_parts = [f"© 2025 {ecole.nom if ecole else 'Myschool'}. Tous droits réservés."]
    if ecole and ecole.telephone:
        footer_parts.append(f"Tél: {ecole.telephone}")
    if ecole and ecole.email:
        footer_parts.append(ecole.email)
    c.drawCentredString(width/2, 0.8*cm, " | ".join(footer_parts))
    c.drawCentredString(width/2, 0.5*cm, f"Bulletin généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
