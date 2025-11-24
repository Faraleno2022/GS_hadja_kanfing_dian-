from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Avg, Count, Sum
from django.core.paginator import Paginator
from django.utils import timezone
from django.template.loader import render_to_string
from django.db import IntegrityError
from decimal import Decimal
import json
from eleves.models import Classe as ClasseEleve
from utilisateurs.utils import filter_by_user_school, user_school
from ecole_moderne.security_decorators import admin_required, require_school_object
from utilisateurs.permissions import any_permission_required, can_manage_notes
from .forms import ClasseNoteForm, MatiereNoteForm, EvaluationForm, NoteEleveForm
from .models import ClasseNote, MatiereNote, Evaluation, NoteEleve, NoteMensuelle
from eleves.models import Eleve
from decimal import Decimal
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
import json
import os
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm
from .calculs_intelligent import (
    obtenir_mention_intelligente,
    obtenir_appreciation_intelligente,
    formater_rang_intelligent,
)
# from .utils import (
#     semester_course_avg,
#     semester_compo_avg,
#     semester_avg,
#     course_month_avg,
#     compo_month_avg,
#     monthly_avg,
# )

# Groupes de niveaux pour l'affichage
PRIMAIRE = {
    'PRIMAIRE_1', 'PRIMAIRE_2', 'PRIMAIRE_3', 'PRIMAIRE_4', 'PRIMAIRE_5', 'PRIMAIRE_6'
}
COLLEGE = {
    'COLLEGE_7', 'COLLEGE_8', 'COLLEGE_9', 'COLLEGE_10'
}
LYCEE = {
    'LYCEE_11', 'LYCEE_12', 'TERMINALE'
}

def _draw_school_header(c, ecole, *, y_start, margin, page_width):
    """Dessine un en-tête officiel (centré) avec logo, nom en MAJUSCULES, coordonnées et encadré.
    Retourne la nouvelle coordonnée y après dessin."""
    from reportlab.lib import colors
    y = y_start
    # En-tête national
    center_x = page_width / 2
    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(center_x, y, "République de Guinée")
    y -= 12
    c.setFont('Helvetica-Oblique', 10)
    # Dessiner la devise avec couleurs par mot: Travail (rouge), Justice (jaune), Solidarité (vert)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.lib import colors
    parts = [
        ("Travail", colors.red),
        (" - ", colors.black),
        ("Justice", colors.yellow),
        (" - ", colors.black),
        ("Solidarité", colors.green),
    ]
    total_w = sum(pdfmetrics.stringWidth(t, 'Helvetica-Oblique', 10) for t, _ in parts)
    start_x = center_x - (total_w / 2)
    x = start_x
    for text, col in parts:
        c.setFillColor(col)
        c.drawString(x, y, text)
        x += pdfmetrics.stringWidth(text, 'Helvetica-Oblique', 10)
    c.setFillColor(colors.black)
    y -= 12
    c.setFont('Helvetica', 10)
    c.drawCentredString(center_x, y, "Ministère de l’Enseignement Pré-Universitaire et de l’Alphabétisation")
    y -= 12
    # Abréviations sur 3 lignes (centrées)
    c.setFont('Helvetica-Bold', 10)
    y -= 6
    ire = getattr(ecole, 'ire', None) or ''
    dpe = getattr(ecole, 'dpe', None) or ''
    desee = getattr(ecole, 'desee', None) or ''
    c.drawCentredString(center_x, y, f"IRE: {ire}")
    y -= 12
    c.drawCentredString(center_x, y, f"DPE: {dpe}")
    y -= 12
    c.drawCentredString(center_x, y, f"DESEE: {desee}")
    y -= 16
    # Espace supplémentaire pour descendre le premier cadre du bulletin
    y -= 30

    # Mémoriser la position du haut du cadre pour le dessiner après le contenu
    frame_top = y
    box_height = 60

    # Logo (gauche) si disponible
    logo_path = None
    try:
        if hasattr(ecole, 'logo') and getattr(ecole.logo, 'path', None) and os.path.exists(ecole.logo.path):
            logo_path = ecole.logo.path
    except Exception:
        logo_path = None
    if logo_path:
        try:
            c.drawImage(logo_path, margin + 8, y - 62, width=54, height=54, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    # Texte centré
    top_line_y = y + 12
    school_name = (getattr(ecole, 'nom', '') or 'ÉCOLE').upper()
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(center_x, top_line_y, school_name)

    c.setFont('Helvetica', 10)
    adresse = getattr(ecole, 'adresse', None) or ''
    telephone = getattr(ecole, 'telephone', None) or ''
    email = getattr(ecole, 'email', None) or ''
    directeur = getattr(ecole, 'directeur', None) or ''

    # Helper: wrap centered text within available width
    from reportlab.pdfbase import pdfmetrics
    def draw_wrapped_centered(text, y_pos, max_width, line_height=12):
        words = text.split()
        lines = []
        cur = ''
        for w in words:
            test = (cur + ' ' + w).strip()
            if pdfmetrics.stringWidth(test, 'Helvetica', 10) <= max_width:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        for ln in lines:
            c.drawCentredString(center_x, y_pos, ln)
            y_pos -= line_height
        return y_pos

    # Centrer les informations (adresse/contacts/directeur) au milieu du cadre
    line_y = top_line_y - 30
    # Contacts en gris léger pour hiérarchie visuelle
    try:
        c.setFillGray(0.3)
    except Exception:
        pass
    if adresse:
        # keep inside box: reduce available width a bit
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Adresse: {adresse}", line_y, avail_w)
    # Afficher téléphone et email sur des lignes séparées pour éviter le débordement
    if telephone:
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Tél: {telephone}", line_y, avail_w)
    if email:
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Email: {email}", line_y, avail_w)
    if directeur:
        c.drawCentredString(center_x, line_y, f"Directeur: {directeur}")
    # Rétablir la couleur par défaut
    try:
        c.setFillGray(0.0)
    except Exception:
        pass

    # Dessiner le cadre maintenant que le contenu est placé
    # Le cadre commence au-dessus du nom de l'école pour avoir plus d'espace en haut
    frame_start_y = top_line_y - 8  # En dessous du nom de l'école
    adjusted_box_height = box_height  # Utiliser la hauteur fixe définie
    c.setLineWidth(1)
    c.setStrokeColor(colors.black)
    c.roundRect(margin, frame_start_y - adjusted_box_height, page_width - 2*margin, adjusted_box_height, 6, stroke=1, fill=0)

    # Retourner y en dessous du cadre
    y = y - box_height - 8
    # Ligne séparatrice légère
    c.setFillColor(colors.grey)
    c.rect(margin, y, page_width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 8
    return y

@login_required
def tableau_bord(request):
    """Tableau de bord des notes: liste les classes par groupe de niveaux.
    Filtré par l'école de l'utilisateur (sauf admin).
    """
    classes_qs = filter_by_user_school(ClasseEleve.objects.all().order_by('niveau', 'nom'), request.user, 'ecole')

    def group_classes(qs):
        primaire, college, lycee = [], [], []
        for c in qs:
            if c.niveau in PRIMAIRE:
                primaire.append(c)
            elif c.niveau in COLLEGE:
                college.append(c)
            elif c.niveau in LYCEE:
                lycee.append(c)
        return primaire, college, lycee

    primaire, college, lycee = group_classes(classes_qs)

    context = {
        'classes_primaire': primaire,
        'classes_college': college,
        'classes_lycee': lycee,
    }
    return render(request, 'notes/tableau_bord.html', context)

@admin_required
def creer_classe(request, niveau):
    """Créer une classe pour un niveau donné dans l'école de l'utilisateur."""
    if request.method == 'POST':
        form = ClasseNotesForm(request.POST, niveau_initial=niveau)
        if form.is_valid():
            classe = form.save(commit=False)
            classe.ecole = user_school(request.user)
            if classe.ecole is None:
                messages.error(request, "Aucune école associée à votre compte.")
                return redirect('notes:tableau_bord')
            classe.save()
            messages.success(request, f"Classe '{classe.nom}' créée avec succès.")
            return redirect('notes:tableau_bord')
    else:
        form = ClasseNotesForm(niveau_initial=niveau)
    return render(request, 'notes/classe_form.html', {'form': form, 'niveau': niveau})

@admin_required
def supprimer_classe(request, classe_id):
    """Supprimer une classe si elle appartient à l'école de l'utilisateur et qu'elle est vide."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    if request.method == 'POST':
        if hasattr(classe, 'eleves') and classe.eleves.exists():
            messages.error(request, "Impossible de supprimer une classe qui contient des élèves.")
            return redirect('notes:tableau_bord')
        classe.delete()
        messages.success(request, "Classe supprimée avec succès.")
        return redirect('notes:tableau_bord')
    return render(request, 'notes/confirm_delete.html', {
        'objet': classe,
        'message': "Confirmez-vous la suppression de cette classe ? (Cette action est irréversible)",
        'action_url': reverse('notes:supprimer_classe', args=[classe.id])
    })

@can_manage_notes
def matieres_classe(request, classe_id):
    """Liste optimisée et gestion des matières d'une classe avec cache."""
    from django.core.cache import cache
    
    # Cache de la classe
    classe_cache_key = f'classe_{classe_id}_{request.user.id}'
    classe = cache.get(classe_cache_key)
    
    if classe is None:
        classe = get_object_or_404(
            filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), 
            pk=classe_id
        )
        cache.set(classe_cache_key, classe, 300)  # 5 minutes
    
    # Cache des matières
    matieres_cache_key = f'matieres_classe_{classe_id}'
    matieres = cache.get(matieres_cache_key)
    
    if matieres is None:
        matieres = list(
            MatiereClasse.objects
            .filter(classe=classe, ecole=classe.ecole)
            .select_related('classe', 'ecole')
            .prefetch_related('evaluations')
            .order_by('nom')
        )
        cache.set(matieres_cache_key, matieres, 180)  # 3 minutes
    
    return render(request, 'notes/matieres_classe.html', {
        'classe': classe,
        'matieres': matieres,
    })

@admin_required
def creer_matiere(request, classe_id):
    """Créer une matière pour une classe donnée."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    if request.method == 'POST':
        form = MatiereClasseForm(request.POST)
        if form.is_valid():
            mat = form.save(commit=False)
            mat.classe = classe
            mat.ecole = classe.ecole
            try:
                mat.save()
                messages.success(request, f"Matière '{mat.nom}' ajoutée.")
                return redirect('notes:matieres_classe', classe.id)
            except Exception as e:
                messages.error(request, f"Erreur lors de la création: {e}")
    else:
        form = MatiereClasseForm()
    return render(request, 'notes/matiere_form.html', {'form': form, 'classe': classe})

@admin_required
def supprimer_matiere(request, pk):
    """Supprimer une matière de classe."""
    matiere = get_object_or_404(filter_by_user_school(MatiereClasse.objects.select_related('classe', 'ecole'), request.user, 'ecole'), pk=pk)
    if request.method == 'POST':
        classe_id = matiere.classe_id
        matiere.delete()
        messages.success(request, "Matière supprimée.")
        return redirect('notes:matieres_classe', classe_id)
    return render(request, 'notes/confirm_delete.html', {
        'objet': matiere,
        'message': "Confirmez-vous la suppression de cette matière ?",
        'action_url': reverse('notes:supprimer_matiere', args=[matiere.id])
    })

@admin_required
def creer_evaluation(request, classe_id, matiere_id):
    """Créer une évaluation pour une classe/matière donnée."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)
    if request.method == 'POST':
        form = EvaluationForm(request.POST)
        if form.is_valid():
            ev = form.save(commit=False)
            ev.ecole = classe.ecole
            ev.classe = classe
            ev.matiere = matiere
            ev.annee_scolaire = getattr(classe, 'annee_scolaire', None)
            ev.cree_par = request.user
            ev.save()
            messages.success(request, f"Évaluation '{ev.titre}' créée pour {classe.nom} — {matiere.nom}.")
            return redirect('notes:saisie_notes', evaluation_id=ev.id)
    else:
        form = EvaluationForm()
    return render(request, 'notes/evaluation_form.html', {
        'form': form,
        'classe': classe,
        'matiere': matiere,
    })

@can_manage_notes
def saisie_notes(request, evaluation_id):
    """Saisie en masse des notes par matricule pour une évaluation.
    Format par ligne: MATRICULE;NOTE
    """
    # Récupération évaluation dans le périmètre école
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    if request.method == 'POST':
        form = NotesBulkForm(request.POST)
        if form.is_valid():
            donnees = form.cleaned_data['donnees']
            lignes = [l.strip() for l in donnees.splitlines() if l.strip()]
            ok, erreurs, maj, crees = 0, [], 0, 0
            # Restreindre aux élèves de la classe + école de l'évaluation
            eleves_qs = Eleve.objects.select_related('classe', 'classe__ecole').filter(classe=evaluation.classe)
            eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
            # Index par matricule (upper)
            index_mat = { (e.matricule or '').strip().upper(): e for e in eleves_qs }
            for i, ligne in enumerate(lignes, start=1):
                parts = [p.strip() for p in ligne.split(';')]
                if len(parts) < 2:
                    erreurs.append(f"Ligne {i}: format invalide (attendu MATRICULE;NOTE)")
                    continue
                matricule, note_txt = parts[0].upper(), parts[1].replace(',', '.').strip()
                obs = ''  # Pas d'observation requise
                if matricule not in index_mat:
                    erreurs.append(f"Ligne {i}: matricule inconnu pour la classe ({matricule})")
                    continue
                # Parse note 0..20
                try:
                    val = Decimal(note_txt)
                except Exception:
                    erreurs.append(f"Ligne {i}: note invalide '{note_txt}'")
                    continue
                if val < 0 or val > 20:
                    erreurs.append(f"Ligne {i}: la note doit être entre 0 et 20 (reçu {val})")
                    continue
                eleve = index_mat[matricule]
                # Créer/mettre à jour la note
                obj, created = Note.objects.update_or_create(
                    evaluation=evaluation,
                    eleve=eleve,
                    defaults={
                        'ecole': evaluation.ecole,
                        'classe': evaluation.classe,
                        'matiere': evaluation.matiere,
                        'matricule': eleve.matricule or matricule,
                        'note': val,
                        'observation': obs or None,
                        'saisie_par': request.user,
                    }
                )
                if created:
                    crees += 1
                else:
                    maj += 1
                ok += 1
            if ok:
                messages.success(request, f"{ok} note(s) traitée(s) — {crees} créée(s), {maj} mise(s) à jour.")
            if erreurs:
                messages.warning(request, "\n".join(erreurs[:10]) + ("\n…" if len(erreurs) > 10 else ''))
            return redirect('notes:saisie_notes', evaluation_id=evaluation.id)
    else:
        form = NotesBulkForm()
    # Préparer un export des élèves de la classe avec matricules pour aide à la saisie
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    # Notes existantes pour cette évaluation (map JSON par élève)
    notes_qs = evaluation.notes.select_related('eleve')
    try:
        import json as _json
        notes_existantes_map = {
            n.eleve_id: {
                'note': float(n.note) if getattr(n, 'note', None) is not None else None,
                'absent': bool(getattr(n, 'absent', False)),
            }
            for n in notes_qs
        }
        notes_existantes_json = _json.dumps(notes_existantes_map)
    except Exception:
        notes_existantes_json = '{}'

    return render(request, 'notes/saisir_notes.html', {
        'evaluation': evaluation,
        'form': form,
        'eleves': eleves,
        'notes_existantes_json': notes_existantes_json,
    })

@can_manage_notes
def saisie_notes_individuelle(request, evaluation_id: int):
    """Affiche la saisie individuelle des notes pour une évaluation donnée."""
    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )

    # Élèves de la classe (dans le périmètre école)
    eleves_qs = Eleve.objects.select_related('classe').filter(classe=evaluation.classe)
    eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')

    # Notes existantes pour cette évaluation
    notes_map = {
        n.eleve_id: n for n in Note.objects.filter(evaluation=evaluation).select_related('eleve')
    }

    eleves_context = []
    notes_saisies = 0
    for e in eleves_qs.order_by('prenom', 'nom'):
        n = notes_map.get(e.id)
        note_val = n.note if n else None
        if n:
            notes_saisies += 1
        eleves_context.append({
            'id': e.id,
            'nom': e.nom,
            'prenom': e.prenom,
            'matricule': e.matricule,
            'date_naissance': e.date_naissance,
            'photo': getattr(e, 'photo', None),
            'note_actuelle': note_val,
            'appreciation_actuelle': n.appreciation_finale if n else None,
        })

    return render(request, 'notes/saisie_notes_individuelle.html', {
        'evaluation': evaluation,
        'eleves': eleves_context,
        'notes_saisies': notes_saisies,
    })

@can_manage_notes
def saisie_notes_simple(request, evaluation_id):
    """Interface simplifiée et intuitive pour la saisie des notes avec tableau interactif."""
    # Récupération évaluation dans le périmètre école
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    
    if request.method == 'POST':
        # Récupérer toutes les notes soumises
        ok, erreurs, maj, crees = 0, [], 0, 0
        
        # Restreindre aux élèves de la classe + école de l'évaluation
        eleves_qs = Eleve.objects.select_related('classe', 'classe__ecole').filter(classe=evaluation.classe)
        eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
        
        for eleve in eleves_qs:
            note_key = f'note_{eleve.id}'
            note_value = request.POST.get(note_key, '').strip()
            
            # Si pas de note saisie, passer au suivant
            if not note_value:
                continue
            
            try:
                val = Decimal(note_value.replace(',', '.'))
            except Exception:
                erreurs.append(f"{eleve.nom} {eleve.prenom}: note invalide '{note_value}'")
                continue
            
            if val < 0 or val > 20:
                erreurs.append(f"{eleve.nom} {eleve.prenom}: la note doit être entre 0 et 20 (reçu {val})")
                continue
            
            # Créer/mettre à jour la note
            obj, created = Note.objects.update_or_create(
                evaluation=evaluation,
                eleve=eleve,
                defaults={
                    'ecole': evaluation.ecole,
                    'classe': evaluation.classe,
                    'matiere': evaluation.matiere,
                    'matricule': eleve.matricule or '',
                    'note': val,
                    'observation': None,
                    'saisie_par': request.user,
                }
            )
            if created:
                crees += 1
            else:
                maj += 1
            ok += 1
        
        if ok:
            messages.success(request, f"✅ {ok} note(s) enregistrée(s) avec succès — {crees} créée(s), {maj} mise(s) à jour.")
        if erreurs:
            messages.warning(request, "⚠️ Erreurs:\n" + "\n".join(erreurs[:10]) + ("\n…" if len(erreurs) > 10 else ''))
        
        return redirect('notes:saisie_notes_simple', evaluation_id=evaluation.id)
    
    # GET: Afficher le formulaire
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    
    # Notes existantes pour cette évaluation
    notes_existantes = evaluation.notes.select_related('eleve').order_by('eleve__nom', 'eleve__prenom')
    
    return render(request, 'notes/saisie_notes_simple.html', {
        'evaluation': evaluation,
        'eleves': eleves,
        'notes_existantes': notes_existantes,
    })

@require_POST
@can_manage_notes
def ajax_sauvegarder_note(request):
    """Enregistre ou met à jour une note individuelle (JSON)."""
    try:
        data = json.loads(request.body.decode('utf-8'))
        evaluation_id = int(data.get('evaluation_id'))
        eleve_id = int(data.get('eleve_id'))
        note_val = Decimal(str(data.get('note')))
    except Exception:
        return JsonResponse({'success': False, 'error': "Payload invalide"}, status=400)

    if note_val < 0 or note_val > 20:
        return JsonResponse({'success': False, 'error': "La note doit être entre 0 et 20"}, status=400)

    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )
    eleve = get_object_or_404(
        filter_by_user_school(Eleve.objects.select_related('classe'), request.user, 'classe__ecole'),
        pk=eleve_id,
        classe=evaluation.classe,
    )

    obj, created = Note.objects.update_or_create(
        evaluation=evaluation,
        eleve=eleve,
        defaults={
            'ecole': evaluation.ecole,
            'classe': evaluation.classe,
            'matiere': evaluation.matiere,
            'matricule': eleve.matricule or '',
            'note': note_val,
            'saisie_par': request.user,
        }
    )

    total_notes = Note.objects.filter(evaluation=evaluation).count()
    return JsonResponse({'success': True, 'created': created, 'total_notes': total_notes})

@require_POST
@can_manage_notes
def ajax_sauvegarder_notes_masse(request):
    """Enregistre plusieurs notes d'un coup (JSON)."""
    try:
        data = json.loads(request.body.decode('utf-8'))
        evaluation_id = int(data.get('evaluation_id'))
        notes = data.get('notes') or []
    except Exception:
        return JsonResponse({'success': False, 'error': "Payload invalide"}, status=400)

    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )

    # Préparer les élèves de la classe
    eleves_qs = Eleve.objects.filter(classe=evaluation.classe)
    eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
    index_eleves = {e.id: e for e in eleves_qs}

    saved = 0
    for item in notes:
        try:
            eleve_id = int(item.get('eleve_id'))
            note_val = Decimal(str(item.get('note')))
        except Exception:
            continue
        if note_val < 0 or note_val > 20:
            continue
        eleve = index_eleves.get(eleve_id)
        if not eleve:
            continue
        Note.objects.update_or_create(
            evaluation=evaluation,
            eleve=eleve,
            defaults={
                'ecole': evaluation.ecole,
                'classe': evaluation.classe,
                'matiere': evaluation.matiere,
                'matricule': eleve.matricule or '',
                'note': note_val,
                'saisie_par': request.user,
            }
        )
        saved += 1

    total_notes = Note.objects.filter(evaluation=evaluation).count()
    return JsonResponse({'success': True, 'saved_count': saved, 'total_notes': total_notes})

@require_POST
@can_manage_notes
def ajax_supprimer_note(request):
    """Supprime la note d'un élève pour une évaluation (JSON)."""
    try:
        data = json.loads(request.body.decode('utf-8'))
        evaluation_id = int(data.get('evaluation_id'))
        eleve_id = int(data.get('eleve_id'))
    except Exception:
        return JsonResponse({'success': False, 'error': "Payload invalide"}, status=400)

    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )
    eleve = get_object_or_404(
        filter_by_user_school(Eleve.objects.select_related('classe'), request.user, 'classe__ecole'),
        pk=eleve_id,
        classe=evaluation.classe,
    )

    Note.objects.filter(evaluation=evaluation, eleve=eleve).delete()
    total_notes = Note.objects.filter(evaluation=evaluation).count()
    return JsonResponse({'success': True, 'total_notes': total_notes})

@can_manage_notes
def evaluations_matiere(request, classe_id, matiere_id):
    """Liste des évaluations d'une matière pour une classe, avec accès rapide à la saisie et à l'affichage des notes."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)
    evaluations = (
        Evaluation.objects.filter(classe=classe, matiere=matiere)
        .order_by('-date', '-id')
    )
    return render(request, 'notes/evaluations_matiere.html', {
        'classe': classe,
        'matiere': matiere,
        'evaluations': evaluations,
    })

@can_manage_notes
def evaluation_detail(request, evaluation_id):
    """Affiche un tableau des élèves de la classe avec leurs notes (ou vide si non saisie)."""
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    # Élèves de la classe
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    # Index des notes
    notes_map = {n.eleve_id: n for n in evaluation.notes.select_related('eleve')}
    rows = []
    for e in eleves:
        n = notes_map.get(e.id)
        rows.append({
            'eleve': e,
            'matricule': e.matricule,
            'note': getattr(n, 'note', None),
            'observation': getattr(n, 'observation', ''),
        })
    return render(request, 'notes/evaluation_detail.html', {
        'evaluation': evaluation,
        'rows': rows,
    })

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_pdf(request, classe_id: int, eleve_id: int, trimestre: str = "T1"):
    """Génère un bulletin de notes PDF pour un élève donné et un trimestre (T1/T2/T3).
    Calcule la moyenne par matière (pondérée par coefficient d'évaluation) et la moyenne générale (pondérée par coefficient de matière).
    """
    # Sécuriser l'accès à la classe / élève
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Matières définies pour la classe
    # Note: MatiereNote est lié à ClasseNote, pas ClasseEleve
    # Nous devons trouver la ClasseNote correspondante
    from notes.models import ClasseNote
    
    # Mapping spécial pour certaines classes (ClasseEleve ID → ClasseNote ID)
    mapping_inverse = {
        8: 59,   # ClasseEleve '11ème série littéraire' → ClasseNote '11ème Série littéraire'
        56: 61,  # ClasseEleve '12ÈME ANNÉE' → ClasseNote '12ème Année'
    }
    
    try:
        # Essayer le mapping spécial d'abord
        if classe.id in mapping_inverse:
            classe_note = ClasseNote.objects.filter(id=mapping_inverse[classe.id]).first()
        else:
            # Sinon chercher par nom
            classe_note = ClasseNote.objects.filter(
                nom=classe.nom,
                annee_scolaire=classe.annee_scolaire,
                ecole=classe.ecole
            ).first()
        
        if classe_note:
            matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
        else:
            matieres = []
    except Exception:
        matieres = []

    # Récupérer les évaluations de la classe/matière pour le trimestre
    # Note: Evaluation est lié à matiere (pas classe), utilise periode (pas trimestre)
    # Conversion trimestre format court vers format long
    trimestre_mapping = {'T1': 'TRIMESTRE_1', 'T2': 'TRIMESTRE_2', 'T3': 'TRIMESTRE_3', 'S1': 'SEMESTRE_1', 'S2': 'SEMESTRE_2'}
    periode_longue = trimestre_mapping.get(trimestre, trimestre)
    
    evals_by_matiere = {}
    for mat in matieres:
        evals = Evaluation.objects.filter(matiere=mat, periode=periode_longue).order_by('date_evaluation', 'id')
        evals_by_matiere[mat.id] = list(evals)

    # Récupérer les notes de l'élève pour ces évaluations
    # Filtrer par matières de la période (evaluation__classe n'existe pas)
    notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode=periode_longue).select_related('evaluation', 'evaluation__matiere')}

    # Calculs des moyennes par matière
    lignes = []
    somme_moyennes_coef = Decimal('0')
    somme_coef_matieres = Decimal('0')

    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        if not evals:
            moy_mat = None
        else:
            num = Decimal('0')
            den = Decimal('0')
            for ev in evals:
                n = notes_by_eval.get(ev.id)
                c = Decimal(ev.coefficient or 1)
                if n is None or n.note is None:
                    # Absence ou note manquante = 0
                    num += Decimal('0') * c
                else:
                    num += Decimal(n.note) * c
                den += c
            moy_mat = (num / den).quantize(Decimal('0.01')) if den > 0 else None

        # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
        if moy_mat is None:
            moy_mat = Decimal('0')
        somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
        somme_coef_matieres += Decimal(mat.coefficient or 1)

        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moyenne': moy_mat,
        })

    moyenne_generale = None
    if somme_coef_matieres > 0:
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01'))

    # Moyennes de classe par matière (pondérées par coeffs d'évaluations)
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        total_num = Decimal('0'); total_den = Decimal('0')
        for ev in evals:
            # toutes les notes de l'évaluation pour la classe
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                total_num += Decimal(n.note) * cc
                total_den += cc
        moyennes_classe_par_matiere[mat.id] = (total_num / total_den).quantize(Decimal('0.01')) if total_den > 0 else None

    # Classement (rang): calculer la moyenne générale de tous les élèves
    eleves_classe = Eleve.objects.filter(classe=classe).only('id')
    moyennes_generales = []  # list of (eleve_id, moyenne_generale)
    for e in eleves_classe:
        # Filtrer par matières de la période (evaluation__classe n'existe pas)
        notes_by_eval_e = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=e, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode=periode_longue)}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            # moyenne matière élève
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval_e.get(ev.id)
                cc = Decimal(ev.coefficient or 1)
                if not nn or nn.note is None:
                    # Absence ou note manquante = 0
                    num += Decimal('0') * cc
                else:
                    num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num / den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            mg = (s_num / s_den)
        else:
            mg = None
        if mg is not None:
            moyennes_generales.append((e.id, mg))
    # Calculer les rangs avec calculer_rang_intelligent
    from .calculs_intelligent import calculer_rang_intelligent
    
    rang = None
    total_eleves_ayant_moyenne = len(moyennes_generales)
    
    if moyennes_generales:
        # Préparer les données pour calculer_rang_intelligent
        moyennes_pour_rang = []
        for eid, mg in moyennes_generales:
            e_obj = eleves.get(id=eid)
            moyennes_pour_rang.append({
                'eleve_id': eid,
                'prenom': e_obj.prenom,
                'nom': e_obj.nom,
                'sexe': getattr(e_obj, 'sexe', 'M'),
                'moyenne': mg
            })
        
        # Calculer les rangs
        resultats_rangs = calculer_rang_intelligent(moyennes_pour_rang)
        
        # Trouver le rang de notre élève
        for r in resultats_rangs:
            if r['eleve_id'] == eleve.id:
                rang = r.get('rang_num')
                break

    # Mention selon barème simple (modifiable)
    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('16'):
            return "Très Bien"
        if avg >= Decimal('14'):
            return "Bien"
        if avg >= Decimal('12'):
            return "Assez Bien"
        if avg >= Decimal('10'):
            return "Passable"
        return "Insuffisant"
    mention = mention_for(moyenne_generale)

    # Génération PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_{eleve.matricule}_{trimestre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane standard si disponible (spécifique à l'école)
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass

    margin = 2 * cm
    y = height - margin

    # En-tête avec logo et coordonnées
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin de notes — {trimestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )")
    y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}")
    y -= 12
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16

    # Tableau entêtes
    c.setFont('Helvetica-Bold', 12)
    headers = ["Matière", "Coef.", "Moyenne /20", "Moy. classe"]
    colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt)
        x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 10

    c.setFont('Helvetica', 11)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            except Exception:
                pass
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        moy_txt = '-' if row['moyenne'] is None else f"{row['moyenne']}"
        c.drawString(x, y, moy_txt); x += colw[2]
        # moyenne de classe
        mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
        mc_txt = '-' if mc is None else f"{mc}"
        c.drawString(x, y, mc_txt)
        y -= 14

    # Séparateur
    y -= 6
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16

    # Moyenne générale + Rang + Mention
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
    y -= 16
    c.setFont('Helvetica', 12)
    if rang is not None:
        # Format intelligent du rang avec accord grammatical
        from .calculs_intelligent import formater_rang_intelligent
        sexe = getattr(eleve, 'sexe', 'M') or 'M'
        rang_str = formater_rang_intelligent(rang, sexe, total_eleves_ayant_moyenne)
        c.drawString(margin, y, f"Rang: {rang_str}")
        y -= 14
    if mention:
        c.drawString(margin, y, f"Mention: {mention}")
        y -= 18

    # Pied de page
    # Signatures
    c.setFont('Helvetica', 11)
    sig_y = margin + 50
    c.drawString(margin, sig_y, "Prof. principal:")
    c.line(margin + 120, sig_y-2, margin + 250, sig_y-2)
    c.drawString(margin + 280, sig_y, "Chef d'établ.:")
    c.line(margin + 380, sig_y-2, margin + 510, sig_y-2)
    c.drawString(margin, sig_y - 28, "Parent/Tuteur:")
    c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)

    c.setFont('Helvetica-Oblique', 10)
    c.setFillColor(colors.darkgrey)
    from datetime import datetime
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response

@admin_required
def bulletins_mensuels_classe_pdf(request, classe_id: int, mois: int):
    """Génère en un seul PDF les bulletins mensuels de tous les élèves d'une classe (Collège/Lycée)."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = Eleve.objects.select_related('classe').filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_classe_mois_{mois:02d}_{classe.nom}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    mois_label = [None,'Janvier','Février','Mars','Avril','Mai','Juin','Juillet','Août','Septembre','Octobre','Novembre','Décembre'][mois] if 1 <= mois <= 12 else f"Mois {mois}"

    # Pré-calcul du classement mensuel (moy. générale mensuelle pondérée par coef matière)
    classement_list = []  # list of (eleve, moy_mensuelle)
    for e in eleves:
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            mm = monthly_avg(e, mat, annee_scolaire, mois, mode='weighted')
            if mm is not None:
                s_num += mm * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        avg = (s_num / s_den).quantize(Decimal('0.01')) if s_den > 0 else None
        if avg is not None:
            classement_list.append((e, avg))
    classement_list.sort(key=lambda t: t[1], reverse=True)

    # Page 1: Couverture
    margin = 2 * cm
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 20); c.drawCentredString(width/2, y-10, f"Bulletins mensuels — {mois_label}"); y -= 50
    c.setFont('Helvetica', 13); c.drawCentredString(width/2, y, f"Classe: {classe.nom}"); y -= 20
    c.drawCentredString(width/2, y, f"Année scolaire: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Oblique', 11); c.setFillColorRGB(0.3,0.3,0.3)
    c.drawCentredString(width/2, y, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    try:
        c.setFillColorRGB(0,0,0)
    except Exception:
        pass
    c.showPage()

    # Page 2: Table des matières
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 16); c.drawCentredString(width/2, y-10, f"Table des matières — {mois_label}"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, "Table des matières"); y -= 18
    c.setFont('Helvetica', 11)
    start_page = 3  # 1 page de couverture + 1 page de TOC, puis 1 page par élève
    page_no = start_page
    for e in eleves:
        c.drawString(margin, y, f"{e.prenom} {e.nom}")
        c.drawRightString(width - margin, y, f"p. {page_no}")
        y -= 14
        page_no += 1
        if y < margin + 40:
            c.showPage()
            try:
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            except Exception:
                pass
            y = height - margin
    c.showPage()

    for eleve in eleves:
        # Filigrane
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
        except Exception:
            pass
        y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin mensuel — {mois_label}"); y -= 40
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        # Calculs
        lignes = []
        somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
        for mat in matieres:
            moy_cours = course_month_avg(eleve, mat, annee_scolaire, mois)
            moy_compo = compo_month_avg(eleve, mat, annee_scolaire, mois)
            moy_mois = monthly_avg(eleve, mat, annee_scolaire, mois, mode='weighted')
            if moy_mois is not None:
                somme_moyennes_coef += moy_mois * Decimal(mat.coefficient or 1)
                somme_coef_matieres += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom,'coef_matiere': mat.coefficient,'moy_cours': moy_cours,'moy_compo': moy_compo,'moy_mois': moy_mois})
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

        # Entêtes
        c.setFont('Helvetica-Bold', 12)
        headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", "Moy. mois"]
        colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
        x = margin
        for i, htxt in enumerate(headers):
            c.drawString(x, y, htxt); x += colw[i]
        y -= 14
        c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 11)
        for row in lignes:
            if y < margin + 60:
                c.showPage(); y = height - margin
                try:
                    from ecole_moderne.pdf_utils import draw_logo_watermark
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
                except Exception:
                    pass
            x = margin
            c.drawString(x, y, str(row['matiere'])); x += colw[0]
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[2]
            c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[3]
            c.drawString(x, y, '-' if row['moy_mois'] is None else f"{row['moy_mois']}")
            y -= 14
        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        c.setFont('Helvetica-Bold', 13)
        c.drawString(margin, y, f"Moyenne générale mensuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20");
        c.showPage()
    c.save();
    return response

@admin_required
def bulletins_semestre_classe_pdf(request, classe_id: int, semestre: int = 1):
    """Génère en un seul PDF les bulletins semestriels de tous les élèves d'une classe (Collège/Lycée)."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = Eleve.objects.select_related('classe').filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_classe_semestre{semestre}_{classe.nom}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul des moyennes générales pour classement
    moyenne_map = {}
    for e in eleves:
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            moy_sem = semester_avg(e, mat, annee_scolaire, semestre, mode='weighted')
            if moy_sem is not None:
                s_num += moy_sem * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        moyenne_map[e.id] = (s_num / s_den).quantize(Decimal('0.01')) if s_den > 0 else None
    classement_list = [(e, moyenne_map.get(e.id)) for e in eleves]
    classement_list = [(e, m) for (e, m) in classement_list if m is not None]
    classement_list.sort(key=lambda t: t[1], reverse=True)
    rang_map = {e.id: idx for idx, (e, _) in enumerate(classement_list, start=1)}

    # Page 1: Couverture
    margin = 2 * cm
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 20); c.drawCentredString(width/2, y-10, f"Bulletins semestriels — S{semestre}"); y -= 50
    c.setFont('Helvetica', 13); c.drawCentredString(width/2, y, f"Classe: {classe.nom}"); y -= 20
    c.drawCentredString(width/2, y, f"Année scolaire: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Oblique', 11); c.setFillColorRGB(0.3,0.3,0.3)
    c.drawCentredString(width/2, y, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    try:
        c.setFillColorRGB(0,0,0)
    except Exception:
        pass
    c.showPage()

    # Page 2: Table des matières
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 16); c.drawCentredString(width/2, y-10, f"Table des matières — S{semestre}"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, "Table des matières"); y -= 18
    c.setFont('Helvetica', 11)
    start_page = 3
    page_no = start_page
    for e in eleves:
        c.drawString(margin, y, f"{e.prenom} {e.nom}")
        c.drawRightString(width - margin, y, f"p. {page_no}")
        y -= 14
        page_no += 1
        if y < margin + 40:
            c.showPage()
            try:
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            except Exception:
                pass
            y = height - margin
    c.showPage()

    for eleve in eleves:
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
        except Exception:
            pass
        y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin semestriel — S{semestre}"); y -= 40
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        # Calculs
        lignes = []
        somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
        for mat in matieres:
            moy_cours = semester_course_avg(eleve, mat, annee_scolaire, semestre)
            moy_compo = semester_compo_avg(eleve, mat, annee_scolaire, semestre)
            moy_sem = semester_avg(eleve, mat, annee_scolaire, semestre, mode='weighted')
            if moy_sem is not None:
                somme_moyennes_coef += moy_sem * Decimal(mat.coefficient or 1)
                somme_coef_matieres += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom,'coef_matiere': mat.coefficient,'moy_cours': moy_cours,'moy_compo': moy_compo,'moy_sem': moy_sem})
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

        # Entêtes
        c.setFont('Helvetica-Bold', 12)
        headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", f"Moy. S{semestre}"]
        colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
        x = margin
        for i, htxt in enumerate(headers):
            c.drawString(x, y, htxt); x += colw[i]
        y -= 14
        c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 11)
        for row in lignes:
            if y < margin + 60:
                c.showPage(); y = height - margin
                try:
                    from ecole_moderne.pdf_utils import draw_logo_watermark
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
                except Exception:
                    pass
            x = margin
            c.drawString(x, y, str(row['matiere'])); x += colw[0]
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[2]
            c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[3]
            c.drawString(x, y, '-' if row['moy_sem'] is None else f"{row['moy_sem']}")
            y -= 14
        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        c.setFont('Helvetica-Bold', 13)
        c.drawString(margin, y, f"Moyenne générale semestrielle: {moyenne_generale if moyenne_generale is not None else '-'} / 20");
        c.showPage()
    c.save();
    return response

def course_month_avg(eleve, matiere, annee_scolaire, mois):
    """Calcule la moyenne des devoirs pour un mois donné"""
    from django.db.models import Q
    
    periode_str = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }.get(mois, 'OCTOBRE')
    
    # Récupérer les évaluations de type DEVOIR pour ce mois
    evals = Evaluation.objects.filter(
        matiere=matiere,
        periode=periode_str,
        type_evaluation__in=['DEVOIR', 'CONTROLE', 'INTERROGATION']
    )
    
    total = Decimal('0')
    count = 0
    
    for ev in evals:
        try:
            n = NoteEleve.objects.get(eleve=eleve, evaluation=ev)
            if n.absent or n.note is None:
                total += Decimal('0')  # Absence = 0
                count += 1
            else:
                total += Decimal(str(n.note))
                count += 1
        except NoteEleve.DoesNotExist:
            total += Decimal('0')  # Pas de note = 0
            count += 1
    
    return (total / count).quantize(Decimal('0.01')) if count > 0 else None

def compo_month_avg(eleve, matiere, annee_scolaire, mois):
    """Calcule la moyenne de composition pour un mois donné"""
    from django.db.models import Q
    
    periode_str = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }.get(mois, 'OCTOBRE')
    
    # Récupérer les évaluations de type COMPOSITION pour ce mois
    evals = Evaluation.objects.filter(
        matiere=matiere,
        periode=periode_str,
        type_evaluation__in=['COMPOSITION', 'EXAMEN']
    )
    
    total = Decimal('0')
    count = 0
    
    for ev in evals:
        try:
            n = NoteEleve.objects.get(eleve=eleve, evaluation=ev)
            if n.absent or n.note is None:
                total += Decimal('0')  # Absence = 0
                count += 1
            else:
                total += Decimal(str(n.note))
                count += 1
        except NoteEleve.DoesNotExist:
            total += Decimal('0')  # Pas de note = 0
            count += 1
    
    return (total / count).quantize(Decimal('0.01')) if count > 0 else None

def monthly_avg(eleve, matiere, annee_scolaire, mois, mode='weighted'):
    """Calcule la moyenne mensuelle (pondérée ou simple)"""
    moy_cours = course_month_avg(eleve, matiere, annee_scolaire, mois)
    moy_compo = compo_month_avg(eleve, matiere, annee_scolaire, mois)
    
    if mode == 'weighted':
        # Formule pondérée : (moy_cours + 2 * moy_compo) / 3
        if moy_cours is not None and moy_compo is not None:
            return ((moy_cours + moy_compo * 2) / 3).quantize(Decimal('0.01'))
        elif moy_compo is not None:
            return moy_compo
        elif moy_cours is not None:
            return moy_cours
    else:
        # Moyenne simple
        if moy_cours is not None and moy_compo is not None:
            return ((moy_cours + moy_compo) / 2).quantize(Decimal('0.01'))
        elif moy_compo is not None:
            return moy_compo
        elif moy_cours is not None:
            return moy_cours
    
    return None

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_mensuel_pdf(request, classe_id: int, eleve_id: int, mois: int):
    """Bulletin mensuel Collège/Lycée pour un élève.
    Colonnes: Matière, Coef., Moy. cours (mois), Moy. compo (mois), Moy. mensuelle (pondérée 2:1 si compo présente).
    """
    # Sécuriser classe/élève
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Matières actives
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    for mat in matieres:
        moy_cours = course_month_avg(eleve, mat, annee_scolaire, mois)
        moy_compo = compo_month_avg(eleve, mat, annee_scolaire, mois)
        moy_mois = monthly_avg(eleve, mat, annee_scolaire, mois, mode='weighted')
        if moy_mois is not None:
            somme_moyennes_coef += moy_mois * Decimal(mat.coefficient or 1)
            somme_coef_matieres += Decimal(mat.coefficient or 1)
        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moy_cours': moy_cours,
            'moy_compo': moy_compo,
            'moy_mois': moy_mois,
        })

    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # Calcul du rang de l'élève dans la classe
    rang_str = "-"
    total_eleves = 0
    if moyenne_generale is not None:
        # Récupérer tous les élèves de la classe
        eleves_classe = Eleve.objects.filter(classe=classe).select_related('classe')
        moyennes_classe = []
        
        for e in eleves_classe:
            somme_e = Decimal('0')
            somme_coef_e = Decimal('0')
            
            for mat in matieres:
                # Utiliser monthly_avg pour calculer la moyenne mensuelle
                moy_mois_e = monthly_avg(e, mat, annee_scolaire, mois, mode='weighted')
                if moy_mois_e is not None:
                    somme_e += moy_mois_e * Decimal(mat.coefficient or 1)
                    somme_coef_e += Decimal(mat.coefficient or 1)
            
            if somme_coef_e > 0:
                moy_gen_e = (somme_e / somme_coef_e).quantize(Decimal('0.01'))
                moyennes_classe.append((e.id, moy_gen_e))
        
        # Calculer les rangs avec calculer_rang_intelligent
        from .calculs_intelligent import calculer_rang_intelligent
        
        total_eleves = len(moyennes_classe)
        rang_str = "-"
        
        if moyennes_classe:
            # Préparer les données pour calculer_rang_intelligent
            moyennes_pour_rang = []
            for eid, moy in moyennes_classe:
                e_obj = eleves_classe.get(id=eid)
                moyennes_pour_rang.append({
                    'eleve_id': eid,
                    'prenom': e_obj.prenom,
                    'nom': e_obj.nom,
                    'sexe': getattr(e_obj, 'sexe', 'M'),
                    'moyenne': moy
                })
            
            # Calculer les rangs
            resultats_rangs = calculer_rang_intelligent(moyennes_pour_rang)
            
            # Trouver le rang de notre élève
            for r in resultats_rangs:
                if r['eleve_id'] == eleve.id:
                    rang_str = r.get('rang', '-')
                    break

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_mensuel_{eleve.matricule}_{mois:02d}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass

    margin = 2 * cm
    y = height - margin

    # En-tête
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    mois_label = [
        None, 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
    ][mois] if 1 <= mois <= 12 else f"Mois {mois}"
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin mensuel — {mois_label}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Entêtes colonnes
    c.setFont('Helvetica-Bold', 12)
    headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", "Moy. mois"]
    colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt); x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10

    c.setFont('Helvetica', 11)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            except Exception:
                pass
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[2]
        c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[3]
        c.drawString(x, y, '-' if row['moy_mois'] is None else f"{row['moy_mois']}")
        y -= 14

    # Séparateur
    y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Moyenne générale et rang
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale mensuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20"); y -= 18
    c.drawString(margin, y, f"Rang: {rang_str}"); y -= 18
    if total_eleves > 0:
        c.setFont('Helvetica', 11)
        c.drawString(margin, y, f"Effectif: {total_eleves} élèves"); y -= 18

    # Pied
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey)
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_semestre_pdf(request, classe_id: int, eleve_id: int, semestre: int = 1):
    """Génère un bulletin semestriel (S1 ou S2) pour Collège/Lycée.
    Règle de pondération: ((Moy. composition * 2) + Moy. cours) / 3.
    """
    # Sécuriser classe/élève
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Matières actives
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    for mat in matieres:
        moy_cours = semester_course_avg(eleve, mat, annee_scolaire, semestre)
        moy_compo = semester_compo_avg(eleve, mat, annee_scolaire, semestre)
        moy_sem = semester_avg(eleve, mat, annee_scolaire, semestre, mode='weighted')
        if moy_sem is not None:
            somme_moyennes_coef += moy_sem * Decimal(mat.coefficient or 1)
            somme_coef_matieres += Decimal(mat.coefficient or 1)
        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moy_cours': moy_cours,
            'moy_compo': moy_compo,
            'moy_sem': moy_sem,
        })

    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_semestre{semestre}_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass

    margin = 2 * cm
    y = height - margin

    # En-tête
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin semestriel — S{semestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Entêtes colonnes
    c.setFont('Helvetica-Bold', 12)
    headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", "Moy. S{}".format(semestre)]
    colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt); x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10

    c.setFont('Helvetica', 11)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            except Exception:
                pass
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[2]
        c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[3]
        c.drawString(x, y, '-' if row['moy_sem'] is None else f"{row['moy_sem']}")
        y -= 14

    # Séparateur
    y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Moyenne générale
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale semestrielle: {moyenne_generale if moyenne_generale is not None else '-'} / 20"); y -= 18

    # Pied
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey)
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response

@admin_required
def bulletins_classe_pdf(request, classe_id: int, trimestre: str = "T1"):
    """Génère en un seul PDF les bulletins de tous les élèves d'une classe pour un trimestre."""
    # Sécuriser la classe
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    # Élèves de la classe (dans le périmètre utilisateur)
    eleves = Eleve.objects.select_related('classe').filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')

    # Précharger matières et évaluations du trimestre
    # Note: MatiereNote est lié à ClasseNote, pas ClasseEleve
    # Nous devons trouver la ClasseNote correspondante
    from notes.models import ClasseNote
    try:
        # Essayer de trouver la ClasseNote correspondante
        classe_note = ClasseNote.objects.filter(
            nom=classe.nom,
            annee_scolaire=classe.annee_scolaire,
            ecole=classe.ecole
        ).first()
        
        if classe_note:
            matieres = list(MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom'))
        else:
            matieres = []
    except Exception:
        matieres = []
    # Conversion trimestre format court (T1, T2, T3) vers format long (TRIMESTRE_1, etc.)
    trimestre_mapping = {
        'T1': 'TRIMESTRE_1',
        'T2': 'TRIMESTRE_2',
        'T3': 'TRIMESTRE_3',
        'S1': 'SEMESTRE_1',
        'S2': 'SEMESTRE_2'
    }
    periode_longue = trimestre_mapping.get(trimestre, trimestre)
    
    evals_by_matiere = {}
    for mat in matieres:
        # Evaluation est lié à matiere, pas directement à classe
        # et utilise periode, pas trimestre
        evals_by_matiere[mat.id] = list(Evaluation.objects.filter(matiere=mat, periode=periode_longue).order_by('date_evaluation', 'id'))

    # Init PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_classe_{classe.nom}_{trimestre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul des moyennes de classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        total_num = Decimal('0'); total_den = Decimal('0')
        for ev in evals:
            for n in NoteEleve.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                total_num += Decimal(n.note) * cc
                total_den += cc
        moyennes_classe_par_matiere[mat.id] = (total_num / total_den).quantize(Decimal('0.01')) if total_den > 0 else None

    # Pré-calcul des moyennes générales par élève (pour classement/rang)
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        # Filtrer les notes par matières de la période (pas par evaluation__classe qui n'existe pas)
        notes_by_eval_e = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=e, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode=periode_longue)}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num / den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))

    # Classement
    classement = sorted(moyenne_generale_map.items(), key=lambda t: t[1], reverse=True)
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}
    total_eleves_ayant_moyenne = len(classement)

    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('16'):
            return "Très Bien"
        if avg >= Decimal('14'):
            return "Bien"
        if avg >= Decimal('12'):
            return "Assez Bien"
        if avg >= Decimal('10'):
            return "Passable"
        return "Insuffisant"

    def draw_bulletin_for_student(eleve):
        # Calcul des moyennes pour l'élève
        notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode=periode_longue).select_related('evaluation', 'evaluation__matiere')}
        lignes = []
        somme_moyennes_coef = Decimal('0')
        somme_coef_matieres = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            if not evals:
                moy_mat = None
            else:
                num = Decimal('0'); den = Decimal('0')
                for ev in evals:
                    n = notes_by_eval.get(ev.id)
                    if n is None or n.note is None:
                        continue
                    cc = Decimal(ev.coefficient or 1)
                    num += Decimal(n.note) * cc
                    den += cc
                moy_mat = (num / den).quantize(Decimal('0.01')) if den > 0 else None
            # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
            if moy_mat is None:
                moy_mat = Decimal('0')
            somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
            somme_coef_matieres += Decimal(mat.coefficient or 1)
            lignes.append({
                'matiere': mat.nom,
                'coef_matiere': mat.coefficient,
                'moyenne': moy_mat,
            })
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

        # Dessiner la page
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
        except Exception:
            pass

        margin = 2 * cm
        y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin de notes — {trimestre}"); y -= 40
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )")
        y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}")
        y -= 12
        c.setFillColor(colors.grey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 16

        c.setFont('Helvetica-Bold', 12)
        headers = ["Matière", "Coef.", "Moyenne /20", "Moy. classe"]
        colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
        x = margin
        for i, htxt in enumerate(headers):
            c.drawString(x, y, htxt); x += colw[i]
        y -= 14
        c.setFillColor(colors.lightgrey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 10
        c.setFont('Helvetica', 11)
        for row in lignes:
            if y < margin + 60:
                c.showPage()
                try:
                    from ecole_moderne.pdf_utils import draw_logo_watermark
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
                except Exception:
                    pass
                y = height - margin
            x = margin
            c.drawString(x, y, str(row['matiere'])); x += colw[0]
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            moy_txt = '-' if row['moyenne'] is None else f"{row['moyenne']}"
            c.drawString(x, y, moy_txt); x += colw[2]
            mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
            mc_txt = '-' if mc is None else f"{mc}"
            c.drawString(x, y, mc_txt)
            y -= 14

        y -= 6
        c.setFillColor(colors.grey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 16
        c.setFont('Helvetica-Bold', 13)
        c.drawString(margin, y, f"Moyenne générale: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
        y -= 16
        # Rang + Mention
        rg = rang_map.get(eleve.id)
        if rg is not None:
            from .calculs_intelligent import formater_rang_intelligent
            sexe = getattr(eleve, 'sexe', 'M') or 'M'
            rang_str = formater_rang_intelligent(rg, sexe, total_eleves_ayant_moyenne)
            c.setFont('Helvetica', 12)
            c.drawString(margin, y, f"Rang: {rang_str}")
            y -= 14
        men = mention_for(moyenne_generale)
        if men:
            c.setFont('Helvetica', 12)
            c.drawString(margin, y, f"Mention: {men}")
            y -= 16
        # Signatures
        c.setFont('Helvetica', 11)
        sig_y = margin + 50
        c.drawString(margin, sig_y, "Professeur principal:")
        c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
        c.drawString(margin + 350, sig_y, "Chef d’établissement:")
        c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
        c.drawString(margin, sig_y - 28, "Parent/Tuteur:")
        c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
        c.showPage()

    # Dessiner pour chaque élève
    for e in eleves:
        draw_bulletin_for_student(e)

    c.save()
    return response

@admin_required
def export_notes_excel(request, classe_id: int, matiere_id: int, trimestre: str = "T1"):
    """Export Excel des notes d'une classe pour une matière et un trimestre.
    Colonnes: Matricule, Élève, [colonnes de chaque évaluation], Moyenne matière.
    """
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)

    # Évaluations du trimestre pour cette matière
    evaluations = list(Evaluation.objects.filter(classe=classe, matiere=matiere, trimestre=trimestre).order_by('date', 'id'))
    # Élèves
    eleves = Eleve.objects.filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')

    # Notes indexées par (eleve_id, evaluation_id)
    notes = Note.objects.filter(evaluation__in=evaluations, eleve__in=eleves)
    notes_map = {(n.eleve_id, n.evaluation_id): n for n in notes}

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("openpyxl requis (pip install openpyxl)", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{matiere.nom} {trimestre}"

    # En-tête
    headers = ["Matricule", "Élève"] + [ev.titre or f"Eval {i+1}" for i, ev in enumerate(evaluations)] + ["Moyenne /20"]
    ws.append(headers)

    # Lignes
    from decimal import Decimal as D
    for e in eleves:
        row = [e.matricule, f"{e.prenom} {e.nom}"]
        num = D('0'); den = D('0')
        for ev in evaluations:
            n = notes_map.get((e.id, ev.id))
            if n and n.note is not None:
                row.append(float(n.note))
                c = D(ev.coefficient or 1)
                num += D(n.note) * c
                den += c
            else:
                row.append(None)
        moy = float((num/den)) if den > 0 else None
        row.append(moy)
        ws.append(row)

    # Styles simples: largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 2 else 12

    # Réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"notes_{classe.nom}_{matiere.nom}_{trimestre}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def _moyenne_generale_semestrielle(eleve, matieres, annee_scolaire, semestre: int) -> Decimal | None:
    s_num = Decimal('0'); s_den = Decimal('0')
    for mat in matieres:
        m = semester_avg(eleve, mat, annee_scolaire, semestre, mode='weighted')
        if m is not None:
            s_num += m * Decimal(mat.coefficient or 1)
            s_den += Decimal(mat.coefficient or 1)
    return (s_num / s_den).quantize(Decimal('0.01')) if s_den > 0 else None

@admin_required
def export_admis_semestre_excel(request, classe_id: int, semestre: int = 1):
    """Export Excel de la liste des admis (moyenne générale semestrielle >= 10) pour une classe."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('prenom', 'nom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True))
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    results = []  # (eleve, moyenne)
    for e in eleves:
        mg = _moyenne_generale_semestrielle(e, matieres, annee_scolaire, semestre)
        if mg is not None and mg >= Decimal('10'):
            results.append((e, mg))
    # Trier par moyenne desc, puis nom
    results.sort(key=lambda t: (-(t[1]), t[0].nom, t[0].prenom))

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("openpyxl requis (pip install openpyxl)", status=500)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Admis_S{semestre}"
    ws.append(["Rang", "Matricule", "Élève", f"Moyenne S{semestre}"])
    for idx, (e, avg) in enumerate(results, start=1):
        ws.append([idx, e.matricule, f"{e.prenom} {e.nom}", float(avg)])
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col != 4 else 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"admis_S{semestre}_{classe.nom}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@admin_required
def export_admis_semestre_pdf(request, classe_id: int, semestre: int = 1):
    """Export PDF de la liste des admis (moyenne générale semestrielle >= 10) pour une classe."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('prenom', 'nom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True))
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    rows = []  # (eleve, moyenne)
    for e in eleves:
        mg = _moyenne_generale_semestrielle(e, matieres, annee_scolaire, semestre)
        if mg is not None and mg >= Decimal('10'):
            rows.append((e, mg))
    rows.sort(key=lambda t: (-(t[1]), t[0].nom, t[0].prenom))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"admis_S{semestre}_{classe.nom}.pdf".replace(' ', '_')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    margin = 2*cm
    # En-tête standard + titre
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(width/2, y-10, f"Liste des admis — S{semestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 16
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 12

    # En-têtes de colonnes
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin, y, "Rang");
    c.drawString(margin + 60, y, "Matricule");
    c.drawString(margin + 170, y, "Élève");
    c.drawRightString(width - margin, y, f"Moy. S{semestre}"); y -= 14
    c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 8

    c.setFont('Helvetica', 11)
    for idx, (e, avg) in enumerate(rows, start=1):
        if y < margin + 40:
            c.showPage(); y = height - margin
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            except Exception:
                pass
            if getattr(classe, 'ecole', None):
                y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
            c.setFont('Helvetica-Bold', 12)
            c.drawString(margin, y, "Rang");
            c.drawString(margin + 60, y, "Matricule");
            c.drawString(margin + 170, y, "Élève");
            c.drawRightString(width - margin, y, f"Moy. S{semestre}"); y -= 14
            c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 8
            c.setFont('Helvetica', 11)
        c.drawString(margin, y, str(idx))
        c.drawString(margin + 60, y, e.matricule or '-')
        c.drawString(margin + 170, y, f"{e.prenom} {e.nom}")
        c.drawRightString(width - margin, y, f"{avg}")
        y -= 14

    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey)
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.showPage(); c.save()
    return response
def _collect_evals_all_trimestres(classe, matieres):
    """Retourne un dict {matiere_id: [evaluations sur T1+T2+T3]} triées par date."""
    evals_by_matiere = {}
    for mat in matieres:
        evals_by_matiere[mat.id] = list(Evaluation.objects.filter(classe=classe, matiere=mat, trimestre__in=["T1", "T2", "T3"]).order_by('date', 'id'))
    return evals_by_matiere

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_annuel_pdf(request, classe_id: int, eleve_id: int):
    """Bulletin annuel PDF (T1+T2+T3 cumulés) avec moyennes par matière, moyenne générale, rang, mention, signatures."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Trouver la ClasseNote correspondante pour MatiereNote
    from notes.models import ClasseNote
    
    # Mapping spécial pour certaines classes (ClasseEleve ID → ClasseNote ID)
    mapping_inverse = {
        8: 59,   # ClasseEleve '11ème série littéraire' → ClasseNote '11ème Série littéraire'
        56: 61,  # ClasseEleve '12ÈME ANNÉE' → ClasseNote '12ème Année'
    }
    
    try:
        # Essayer le mapping spécial d'abord
        if classe.id in mapping_inverse:
            classe_note = ClasseNote.objects.filter(id=mapping_inverse[classe.id]).first()
        else:
            # Sinon chercher par nom
            classe_note = ClasseNote.objects.filter(
                nom=classe.nom,
                annee_scolaire=classe.annee_scolaire,
                ecole=classe.ecole
            ).first()
        
        if classe_note:
            matieres = list(MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom'))
        else:
            matieres = []
    except Exception:
        matieres = []
    
    evals_by_matiere = _collect_evals_all_trimestres(classe, matieres)

    # Calculs élève
    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    # Filtrer les notes par matières et par toutes les périodes trimestrielles
    notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        num = Decimal('0'); den = Decimal('0')
        for ev in evals:
            n = notes_by_eval.get(ev.id)
            if not n or n.note is None:
                continue
            cc = Decimal(ev.coefficient or 1)
            num += Decimal(n.note) * cc
            den += cc
        moy_mat = (num/den).quantize(Decimal('0.01')) if den > 0 else None
        # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
        if moy_mat is None:
            moy_mat = Decimal('0')
        somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
        somme_coef_matieres += Decimal(mat.coefficient or 1)
        lignes.append({'matiere': mat.nom, 'coef_matiere': mat.coefficient, 'moyenne': moy_mat})
    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # Moyennes de classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        tnum = Decimal('0'); tden = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                tnum += Decimal(n.note) * cc
                tden += cc
        moyennes_classe_par_matiere[mat.id] = (tnum/tden).quantize(Decimal('0.01')) if tden > 0 else None

    # Classement annuel
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe), request.user, 'classe__ecole')
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        # Filtrer les notes par matières et par toutes les périodes trimestrielles
        notes_e = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=e, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num/den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))

    classement = sorted(moyenne_generale_map.items(), key=lambda t: t[1], reverse=True)
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}
    rang = rang_map.get(eleve.id)

    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('16'): return "Très Bien"
        if avg >= Decimal('14'): return "Bien"
        if avg >= Decimal('12'): return "Assez Bien"
        if avg >= Decimal('10'): return "Passable"
        return "Insuffisant"
    mention = mention_for(moyenne_generale)

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_annuel_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    margin = 2*cm; y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, "Bulletin de notes — Annuel"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom} (Matricule: {eleve.matricule or '-'})"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    c.setFont('Helvetica-Bold', 12)
    headers = ["Matière", "Coef.", "Moy. annuelle", "Moy. classe"]
    colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
    x = margin
    for i, h in enumerate(headers): c.drawString(x, y, h); x += colw[i]
    y -= 14; c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
    c.setFont('Helvetica', 11)
    for row in lignes:
        if y < margin + 60:
            c.showPage();
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
            y = height - margin
        x = margin
        c.drawString(x, y, row['matiere']); x += colw[0]
        c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moyenne'] is None else f"{row['moyenne']}"); x += colw[2]
        mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
        c.drawString(x, y, '-' if mc is None else f"{mc}")
        y -= 14

    y -= 6
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale annuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
    y -= 16
    if rang is not None:
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Rang annuel: {rang} / {len(classement)}")
        y -= 14
    men = mention
    if men:
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Mention: {men}")
        y -= 16
    # Signatures
    c.setFont('Helvetica', 11); sig_y = margin + 50
    c.drawString(margin, sig_y, "Professeur principal:"); c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
    c.drawString(margin + 350, sig_y, "Chef d’établissement:"); c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
    c.drawString(margin, sig_y - 28, "Parent/Tuteur:"); c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
    from datetime import datetime
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey); c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.showPage(); c.save(); return response

@admin_required
def bulletins_annuels_classe_pdf(request, classe_id: int):
    """Bulletins annuels (T1+T2+T3) pour tous les élèves d'une classe en un seul PDF."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('nom','prenom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom'))
    evals_by_matiere = _collect_evals_all_trimestres(classe, matieres)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_annuels_{classe.nom}.pdf".replace(' ','_')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul moyennes classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        tnum = Decimal('0'); tden = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                tnum += Decimal(n.note) * cc
                tden += cc
        moyennes_classe_par_matiere[mat.id] = (tnum/tden).quantize(Decimal('0.01')) if tden > 0 else None

    # Classement annuel
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        # Filtrer les notes par matières et par toutes les périodes trimestrielles
        notes_e = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=e, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num/den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))
    classement = sorted(moyenne_generale_map.items(), key=lambda t: t[1], reverse=True)
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}

    def mention_for(avg: Decimal | None) -> str:
        if avg is None: return ""
        if avg >= Decimal('16'): return "Très Bien"
        if avg >= Decimal('14'): return "Bien"
        if avg >= Decimal('12'): return "Assez Bien"
        if avg >= Decimal('10'): return "Passable"
        return "Insuffisant"

    def draw_for_student(eleve):
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
        except Exception:
            pass
        margin = 2*cm; y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, "Bulletin de notes — Annuel"); y -= 40
        c.setFont('Helvetica', 12); c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom} (Matricule: {eleve.matricule or '-'})"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        c.setFont('Helvetica-Bold', 12)
        headers = ["Matière", "Coef.", "Moy. annuelle", "Moy. classe"]
        colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
        x = margin
        for i, h in enumerate(headers): c.drawString(x, y, h); x += colw[i]
        y -= 14; c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 11)

        # Lignes
        lignes = []
        somme_moy_coef = Decimal('0'); somme_coef = Decimal('0')
        # Filtrer les notes par matières et par toutes les périodes trimestrielles
        notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            moy_mat = (num/den).quantize(Decimal('0.01')) if den > 0 else None
            if moy_mat is not None:
                somme_moy_coef += moy_mat * Decimal(mat.coefficient or 1)
                somme_coef += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom, 'coef_matiere': mat.coefficient, 'moyenne': moy_mat})

        for row in lignes:
            if y < margin + 60:
                c.showPage();
                try:
                    from ecole_moderne.pdf_utils import draw_logo_watermark
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
                except Exception:
                    pass
                y = height - margin
            x = margin
            c.drawString(x, y, row['matiere']); x += colw[0]
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moyenne'] is None else f"{row['moyenne']}"); x += colw[2]
            mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
            c.drawString(x, y, '-' if mc is None else f"{mc}")
            y -= 14

        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        mg = (somme_moy_coef / somme_coef).quantize(Decimal('0.01')) if somme_coef > 0 else None
        c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, f"Moyenne générale annuelle: {mg if mg is not None else '-'} / 20"); y -= 16
        rg = rang_map.get(eleve.id)
        if rg is not None: c.setFont('Helvetica', 12); c.drawString(margin, y, f"Rang annuel: {rg} / {len(classement)}"); y -= 14
        men = mention_for(mg)
        if men: c.setFont('Helvetica', 12); c.drawString(margin, y, f"Mention: {men}"); y -= 16
        # Signatures
        c.setFont('Helvetica', 11); sig_y = margin + 50
        c.drawString(margin, sig_y, "Professeur principal:"); c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
        c.drawString(margin + 350, sig_y, "Chef d’établissement:"); c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
        c.drawString(margin, sig_y - 28, "Parent/Tuteur:"); c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
        c.showPage()

    for e in eleves:
        draw_for_student(e)

    c.save(); return response

@login_required
@require_school_object(model=ClasseEleve, pk_kwarg='classe_id', field_path='ecole')
def classement_classe(request, classe_id: int, trimestre: str = "T1"):
    """Affiche le classement des élèves d'une classe pour un trimestre donné."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Calculer le classement
    eleves = classe.eleves.filter(statut='actif').order_by('prenom', 'nom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere', 'evaluation')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        # Regrouper les notes par matière avec pondération par coefficient d'évaluation
        matieres_notes = {}
        for note_obj in notes:
            matiere = note_obj.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef_matiere': matiere.coefficient,
                    'notes_ponderees': []  # (note, coef_eval)
                }
            # Stocker la note avec son coefficient d'évaluation (absence = 0)
            coef_eval = Decimal(note_obj.evaluation.coefficient or 1)
            if note_obj.absent or note_obj.note is None:
                # Absence = 0
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal('0'), coef_eval)
                )
            else:
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal(note_obj.note), coef_eval)
                )
        
        # Calculer la moyenne pondérée par matière, puis la moyenne générale
        for matiere_data in matieres_notes.values():
            if matiere_data['notes_ponderees']:
                # Moyenne pondérée par coefficient d'évaluation
                num = Decimal('0')
                den = Decimal('0')
                for note_val, coef_eval in matiere_data['notes_ponderees']:
                    num += note_val * coef_eval
                    den += coef_eval
                if den > 0:
                    moyenne_matiere = (num / den).quantize(Decimal('0.01'))
                    # Pondération par coefficient de matière
                    somme_moy_coef += moyenne_matiere * Decimal(matiere_data['coef_matiere'])
                    somme_coef += Decimal(matiere_data['coef_matiere'])
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Ajouter les rangs
    for i, item in enumerate(classement):
        item['rang'] = i + 1
    
    # Calculer la moyenne de classe
    moyenne_classe = None
    if classement:
        total_moyennes = sum(item['moyenne'] for item in classement)
        moyenne_classe = round(total_moyennes / len(classement), 2)
    
    context = {
        'classe': classe,
        'trimestre': trimestre,
        'classement': classement,
        'total_eleves': len(classement),
        'moyenne_classe': moyenne_classe
    }
    
    return render(request, 'notes/classement_classe.html', context)

@admin_required
def classement_classe_pdf(request, classe_id: int, trimestre: str = "T1"):
    """Export PDF du classement d'une classe."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Récupérer le classement (même logique que la vue HTML)
    eleves = classe.eleves.filter(statut='actif').order_by('prenom', 'nom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere', 'evaluation')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        # Regrouper les notes par matière avec pondération par coefficient d'évaluation
        matieres_notes = {}
        for note_obj in notes:
            matiere = note_obj.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef_matiere': matiere.coefficient,
                    'notes_ponderees': []  # (note, coef_eval)
                }
            # Stocker la note avec son coefficient d'évaluation (absence = 0)
            coef_eval = Decimal(note_obj.evaluation.coefficient or 1)
            if note_obj.absent or note_obj.note is None:
                # Absence = 0
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal('0'), coef_eval)
                )
            else:
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal(note_obj.note), coef_eval)
                )
        
        # Calculer la moyenne pondérée par matière, puis la moyenne générale
        for matiere_data in matieres_notes.values():
            if matiere_data['notes_ponderees']:
                # Moyenne pondérée par coefficient d'évaluation
                num = Decimal('0')
                den = Decimal('0')
                for note_val, coef_eval in matiere_data['notes_ponderees']:
                    num += note_val * coef_eval
                    den += coef_eval
                if den > 0:
                    moyenne_matiere = (num / den).quantize(Decimal('0.01'))
                    # Pondération par coefficient de matière
                    somme_moy_coef += moyenne_matiere * Decimal(matiere_data['coef_matiere'])
                    somme_coef += Decimal(matiere_data['coef_matiere'])
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="classement_{classe.nom}_{trimestre}.pdf"'
    
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    
    margin = 2 * cm
    y = height - margin
    
    # En-tête avec logo et coordonnées
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    
    y -= 20
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(width/2, y, f"Classement de la classe {classe.nom} - {trimestre}")
    y -= 40
    
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Année scolaire: {getattr(classe, 'annee_scolaire', '')}")
    y -= 16
    c.drawString(margin, y, f"Total d'élèves classés: {len(classement)}")
    y -= 20
    
    # En-têtes du tableau
    c.setFont('Helvetica-Bold', 12)
    headers = ["Rang", "Nom et Prénom", "Matricule", "Moyenne", "Mention"]
    colw = [2*cm, 6*cm, 3*cm, 2.5*cm, 3*cm]
    x = margin
    for i, h in enumerate(headers):
        c.drawString(x, y, h)
        x += colw[i]
    
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 10
    
    # Données du classement
    c.setFont('Helvetica', 11)
    for i, item in enumerate(classement):
        if y < margin + 60:
            c.showPage()
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
            y = height - margin
        
        x = margin
        c.drawString(x, y, str(i + 1))  # Rang
        x += colw[0]
        c.drawString(x, y, f"{item['eleve'].prenom} {item['eleve'].nom}")  # Nom
        x += colw[1]
        c.drawString(x, y, item['eleve'].matricule or '-')  # Matricule
        x += colw[2]
        c.drawString(x, y, f"{item['moyenne']}")  # Moyenne
        x += colw[3]
        c.drawString(x, y, item['mention'] or '-')  # Mention
        
        y -= 14
    
    # Pied de page
    c.setFont('Helvetica-Oblique', 10)
    c.setFillColor(colors.darkgrey)
    from datetime import datetime
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    c.save()
    return response

@admin_required
def classement_classe_excel(request, classe_id: int, trimestre: str = "T1"):
    """Export Excel du classement d'une classe."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Récupérer le classement (même logique que les autres vues)
    eleves = classe.eleves.filter(statut='actif').order_by('prenom', 'nom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere', 'evaluation')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        # Regrouper les notes par matière avec pondération par coefficient d'évaluation
        matieres_notes = {}
        for note_obj in notes:
            matiere = note_obj.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef_matiere': matiere.coefficient,
                    'notes_ponderees': []  # (note, coef_eval)
                }
            # Stocker la note avec son coefficient d'évaluation (absence = 0)
            coef_eval = Decimal(note_obj.evaluation.coefficient or 1)
            if note_obj.absent or note_obj.note is None:
                # Absence = 0
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal('0'), coef_eval)
                )
            else:
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal(note_obj.note), coef_eval)
                )
        
        # Calculer la moyenne pondérée par matière, puis la moyenne générale
        for matiere_data in matieres_notes.values():
            if matiere_data['notes_ponderees']:
                # Moyenne pondérée par coefficient d'évaluation
                num = Decimal('0')
                den = Decimal('0')
                for note_val, coef_eval in matiere_data['notes_ponderees']:
                    num += note_val * coef_eval
                    den += coef_eval
                if den > 0:
                    moyenne_matiere = (num / den).quantize(Decimal('0.01'))
                    # Pondération par coefficient de matière
                    somme_moy_coef += moyenne_matiere * Decimal(matiere_data['coef_matiere'])
                    somme_coef += Decimal(matiere_data['coef_matiere'])
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Classement {classe.nom} {trimestre}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = ["Rang", "Nom", "Prénom", "Matricule", "Moyenne", "Mention"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Données
    for row, item in enumerate(classement, 2):
        ws.cell(row=row, column=1, value=row-1)  # Rang
        ws.cell(row=row, column=2, value=item['eleve'].nom)
        ws.cell(row=row, column=3, value=item['eleve'].prenom)
        ws.cell(row=row, column=4, value=item['eleve'].matricule or '-')
        ws.cell(row=row, column=5, value=float(item['moyenne']))
        ws.cell(row=row, column=6, value=item['mention'] or '-')
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="classement_{classe.nom}_{trimestre}.xlsx"'
    
    wb.save(response)
    return response

@login_required
@require_school_object(model=ClasseEleve, pk_kwarg='classe_id', field_path='ecole')
def cartes_scolaires_classe(request, classe_id):
    """Interface pour générer les cartes scolaires d'une classe"""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Filtrage par école pour non-admin
    if not request.user.is_superuser:
        ecole_user = user_school(request.user)
        if ecole_user and classe.ecole != ecole_user:
            messages.error(request, "Accès non autorisé à cette classe.")
            return redirect('notes:tableau_bord')
    
    # Récupérer tous les élèves de la classe
    eleves = classe.eleves.filter(statut='ACTIF').order_by('prenom', 'nom')
    
    context = {
        'classe': classe,
        'eleves': eleves,
        'nb_eleves': eleves.count(),
    }
    
    return render(request, 'notes/cartes_scolaires.html', context)

@login_required
@require_school_object(model=ClasseEleve, pk_kwarg='classe_id', field_path='ecole')
def cartes_scolaires_pdf(request, classe_id):
    """Génère les cartes scolaires PDF pour une classe"""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Filtrage par école pour non-admin
    if not request.user.is_superuser:
        ecole_user = user_school(request.user)
        if ecole_user and classe.ecole != ecole_user:
            messages.error(request, "Accès non autorisé à cette classe.")
            return redirect('notes:tableau_bord')
    
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io
    except ImportError:
        messages.error(request, "ReportLab requis pour générer le PDF")
        return redirect('notes:cartes_scolaires_classe', classe_id=classe_id)
    
    # Récupérer les élèves
    eleves = classe.eleves.filter(statut='ACTIF').order_by('prenom', 'nom')
    
    if not eleves.exists():
        messages.warning(request, "Aucun élève actif dans cette classe.")
        return redirect('notes:cartes_scolaires_classe', classe_id=classe_id)
    
    # Créer le PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane standardisé (logo centré, rotation légère, opacité 4%)
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass
    
    # Configuration des cartes (10 cartes par page - 2x5)
    card_width = 8.5 * cm
    card_height = 5.4 * cm  # Format carte de crédit
    margin = 0.3 * cm
    spacing_x = 0.2 * cm  # Espacement horizontal réduit
    spacing_y = 0.15 * cm  # Espacement vertical réduit
    
    # Position des cartes sur la page (2 colonnes x 5 lignes)
    positions = [
        # Colonne gauche
        (margin, height - margin - card_height),  # Ligne 1
        (margin, height - margin - 2 * card_height - spacing_y),  # Ligne 2
        (margin, height - margin - 3 * card_height - 2 * spacing_y),  # Ligne 3
        (margin, height - margin - 4 * card_height - 3 * spacing_y),  # Ligne 4
        (margin, height - margin - 5 * card_height - 4 * spacing_y),  # Ligne 5
        # Colonne droite
        (margin + card_width + spacing_x, height - margin - card_height),  # Ligne 1
        (margin + card_width + spacing_x, height - margin - 2 * card_height - spacing_y),  # Ligne 2
        (margin + card_width + spacing_x, height - margin - 3 * card_height - 2 * spacing_y),  # Ligne 3
        (margin + card_width + spacing_x, height - margin - 4 * card_height - 3 * spacing_y),  # Ligne 4
        (margin + card_width + spacing_x, height - margin - 5 * card_height - 4 * spacing_y),  # Ligne 5
    ]
    
    card_count = 0
    
    for eleve in eleves:
        # Nouvelle page si nécessaire
        if card_count > 0 and card_count % 10 == 0:
            c.showPage()
            # Filigrane standardisé sur nouvelle page
            try:
                from ecole_moderne.pdf_utils import draw_logo_watermark
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
            except Exception:
                pass
        
        # Position de la carte actuelle
        pos_x, pos_y = positions[card_count % 10]
        
        # Dessiner l'arrière-plan blanc de la carte d'abord
        c.setFillColor(colors.white)
        c.rect(pos_x, pos_y, card_width, card_height, fill=1, stroke=0)
        
        # Ajouter un filigrane sur chaque carte individuelle
        c.saveState()
        try:
            from django.contrib.staticfiles import finders
            import os
            from django.conf import settings
            # Logo d'école prioritaire
            school_logo = getattr(getattr(classe, 'ecole', None), 'logo', None)
            logo_path = None
            try:
                if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                    logo_path = school_logo.path
            except Exception:
                logo_path = None
            if not logo_path:
                # Fallback: logo statique
                logo_path = finders.find('logos/logo.png') or os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
            
            if logo_path and os.path.exists(logo_path):
                # Calculer la position centrale de la carte
                center_x = pos_x + card_width / 2
                center_y = pos_y + card_height / 2
                
                # Appliquer la transformation (rotation et opacité)
                c.translate(center_x, center_y)
                c.rotate(30)  # Rotation de 30 degrés
                
                # Taille du filigrane (plus petit pour s'adapter à la carte)
                watermark_size = min(card_width, card_height) * 0.6
                
                # Dessiner le logo en filigrane avec opacité réduite
                c.setFillAlpha(0.08)  # Opacité très faible
                c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                          watermark_size, watermark_size, preserveAspectRatio=True)
        except Exception:
            pass
        c.restoreState()
        
        # Dessiner le cadre de la carte
        c.setStrokeColor(colors.black)
        c.setLineWidth(2)
        c.rect(pos_x, pos_y, card_width, card_height, fill=0, stroke=1)
        
        # Logo de l'école (en haut à droite)
        logo_size = 1.2 * cm
        logo_x = pos_x + card_width - logo_size - 0.2 * cm
        logo_y = pos_y + card_height - logo_size - 0.1 * cm
        
        try:
            import os
            from django.conf import settings
            from django.contrib.staticfiles import finders
            # Priorité au logo de l'école
            school_logo = getattr(getattr(classe, 'ecole', None), 'logo', None)
            logo_path = None
            try:
                if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                    logo_path = school_logo.path
            except Exception:
                logo_path = None
            if not logo_path:
                logo_path = finders.find('logos/logo.png') or os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
            if logo_path and os.path.exists(logo_path):
                c.drawImage(logo_path, logo_x, logo_y, logo_size, logo_size, preserveAspectRatio=True)
            else:
                # Aucun logo trouvé
                c.setFont('Helvetica', 6)
                c.setFillColor(colors.grey)
                c.drawCentredString(logo_x + logo_size/2, logo_y + logo_size/2, "LOGO")
        except Exception as e:
            # En cas d'erreur, afficher simplement le texte d'erreur
            c.setFont('Helvetica', 6)
            c.setFillColor(colors.red)
            c.drawCentredString(logo_x + logo_size/2, logo_y + logo_size/2, "ERREUR")
        
        # En-tête école (centré sur toute la largeur de la carte)
        c.setFillColor(colors.darkblue)
        ecole_nom = classe.ecole.nom.upper() if classe.ecole else "ÉCOLE"
        # Ajuster la taille de police selon la longueur du nom
        if len(ecole_nom) > 40:
            c.setFont('Helvetica-Bold', 6)
        elif len(ecole_nom) > 30:
            c.setFont('Helvetica-Bold', 7)
        else:
            c.setFont('Helvetica-Bold', 8)
        # Centrer le nom au milieu de la carte (meilleur équilibre visuel)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 0.8*cm, ecole_nom)
        
        # Titre "CARTE SCOLAIRE"
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(colors.red)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 1.3*cm, "CARTE SCOLAIRE")
        
        # Année scolaire
        annee_actuelle = datetime.now().year
        annee_scolaire = f"{annee_actuelle}-{annee_actuelle + 1}"
        c.setFont('Helvetica', 7)
        c.setFillColor(colors.black)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 1.7*cm, f"Année: {annee_scolaire}")
        
        # Photo de l'élève (côté gauche) — encore remontée pour un meilleur équilibre visuel
        photo_size = 2.2 * cm
        photo_x = pos_x + 0.3 * cm
        photo_y = pos_y + 1.2 * cm
        
        # Dessiner le cadre de la photo
        c.setStrokeColor(colors.grey)
        c.setLineWidth(1)
        c.rect(photo_x, photo_y, photo_size, photo_size)
        
        # Afficher la photo de l'élève si elle existe
        if eleve.photo and hasattr(eleve.photo, 'path'):
            try:
                import os
                from reportlab.lib.utils import ImageReader
                from PIL import Image
                
                # Vérifier que le fichier photo existe
                if os.path.exists(eleve.photo.path):
                    # Ouvrir et redimensionner l'image
                    with Image.open(eleve.photo.path) as img:
                        # Convertir en RGB si nécessaire
                        if img.mode != 'RGB':
                            img = img.convert('RGB')
                        
                        # Calculer les dimensions pour maintenir le ratio
                        img_width, img_height = img.size
                        ratio = min(photo_size / (img_width * 72/96), photo_size / (img_height * 72/96))
                        
                        new_width = img_width * ratio * 72/96
                        new_height = img_height * ratio * 72/96
                        
                        # Centrer l'image dans le cadre
                        img_x = photo_x + (photo_size - new_width) / 2
                        img_y = photo_y + (photo_size - new_height) / 2
                        
                        # Dessiner l'image
                        c.drawImage(ImageReader(img), img_x, img_y, new_width, new_height)
                else:
                    # Fichier photo introuvable
                    c.setFont('Helvetica', 7)
                    c.setFillColor(colors.red)
                    c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.1*cm, "PHOTO")
                    c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.3*cm, "MANQUANTE")
            except Exception as e:
                # Erreur lors du traitement de l'image
                c.setFont('Helvetica', 7)
                c.setFillColor(colors.red)
                c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.1*cm, "ERREUR")
                c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.3*cm, "PHOTO")
        else:
            # Pas de photo définie
            c.setFont('Helvetica', 8)
            c.setFillColor(colors.grey)
            c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.2*cm, "PHOTO")
        
        # Informations élève (côté droit)
        info_x = pos_x + 3.2 * cm
        info_y = pos_y + card_height - 2.2 * cm
        
        c.setFillColor(colors.black)
        
        # Nom et prénom
        c.setFont('Helvetica-Bold', 9)
        nom_complet = f"{eleve.nom} {eleve.prenom}".upper()
        if len(nom_complet) > 25:
            nom_complet = nom_complet[:25] + "..."
        c.drawString(info_x, info_y, nom_complet)
        info_y -= 0.4 * cm
        
        # Matricule
        c.setFont('Helvetica', 7)
        c.drawString(info_x, info_y, f"Matricule: {eleve.matricule}")
        info_y -= 0.3 * cm
        
        # Classe
        c.drawString(info_x, info_y, f"Classe: {classe.nom}")
        info_y -= 0.3 * cm
        
        # Date de naissance
        if eleve.date_naissance:
            date_naiss = eleve.date_naissance.strftime('%d/%m/%Y')
            c.drawString(info_x, info_y, f"Né(e) le: {date_naiss}")
            info_y -= 0.3 * cm
        
        # Contact responsable
        if hasattr(eleve, 'responsable_principal') and eleve.responsable_principal:
            resp = eleve.responsable_principal
            if resp.telephone:
                c.drawString(info_x, info_y, f"Contact: {resp.telephone}")
        
        # Pied de carte
        c.setFont('Helvetica', 6)
        c.setFillColor(colors.grey)
        c.drawString(pos_x + 0.2*cm, pos_y + 0.2*cm, "Cette carte est strictement personnelle")
        
        card_count += 1
    
    c.save()
    
    # Préparer la réponse
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"cartes_scolaires_{classe.nom.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response

@login_required
def carte_eleve_pdf(request, matricule):
    """Génère la carte scolaire d'un élève spécifique par son matricule"""
    from django.contrib import messages
    from eleves.models import Eleve
    
    # Récupérer l'élève par matricule
    try:
        eleve = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').get(matricule=matricule)
        
        # Vérifier les permissions (sauf pour superuser)
        if not request.user.is_superuser:
            if hasattr(request.user, 'profil') and request.user.profil.ecole:
                if eleve.classe.ecole != request.user.profil.ecole:
                    messages.error(request, "Vous n'avez pas accès à cet élève.")
                    return redirect('eleves:liste_eleves')
            else:
                messages.error(request, "Accès non autorisé.")
                return redirect('eleves:liste_eleves')
                
    except Eleve.DoesNotExist:
        messages.error(request, f"Aucun élève trouvé avec le matricule: {matricule}")
        return redirect('eleves:liste_eleves')
    
    # Créer le PDF avec une seule carte
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane standardisé (logo centré, rotation légère, opacité 4%)
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)
    except Exception:
        pass
    
    # Configuration de la carte (centrée sur la page) - Format carte bancaire standard
    card_width = 8.6 * cm  # 86mm
    card_height = 5.4 * cm  # 54mm
    
    # Position centrée sur la page
    pos_x = (width - card_width) / 2
    pos_y = (height - card_height) / 2
    
    # Dessiner l'arrière-plan blanc de la carte d'abord
    c.setFillColor(colors.white)
    c.rect(pos_x, pos_y, card_width, card_height, fill=1, stroke=0)
    
    # Ajouter un filigrane sur la carte individuelle
    c.saveState()
    try:
        from django.contrib.staticfiles import finders
        import os
        from django.conf import settings
        # Priorité au logo de l'école de l'élève
        school_logo = getattr(getattr(eleve.classe, 'ecole', None), 'logo', None)
        logo_path = None
        try:
            if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                logo_path = school_logo.path
        except Exception:
            logo_path = None
        if not logo_path:
            logo_path = finders.find('logos/logo.png') or os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
        
        if logo_path and os.path.exists(logo_path):
            # Calculer la position centrale de la carte
            center_x = pos_x + card_width / 2
            center_y = pos_y + card_height / 2
            
            # Appliquer la transformation (rotation et opacité)
            c.translate(center_x, center_y)
            c.rotate(25)  # Rotation de 25 degrés
            
            # Taille du filigrane (agrandi pour être plus visible)
            watermark_size = min(card_width, card_height) * 0.7
            
            # Dessiner le logo en filigrane avec opacité visible
            c.setFillAlpha(0.12)  # Opacité visible (comme les tickets)
            c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                      watermark_size, watermark_size, preserveAspectRatio=True)
    except Exception:
        pass
    c.restoreState()
    
    # Extraire les couleurs du logo de l'école
    primary_color = '#10b981'
    light_color = '#d1fae5'
    
    try:
        if logo_path and os.path.exists(logo_path):
            from eleves.views import _extraire_couleurs_logo
            primary_color, light_color = _extraire_couleurs_logo(logo_path)
    except:
        pass
    
    # Design moderne avec couleurs personnalisées
    # Formes géométriques décoratives en arrière-plan
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.15)
    c.circle(pos_x + card_width - 1 * cm, pos_y + card_height - 1 * cm, 2.5 * cm, stroke=0, fill=1)
    c.circle(pos_x + 0.8 * cm, pos_y + 0.8 * cm, 1.8 * cm, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Bordure moderne avec coins arrondis
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(pos_x + 0.15 * cm, pos_y + 0.15 * cm, card_width - 0.3 * cm, card_height - 0.3 * cm, 8, stroke=1, fill=0)
    
    # Bande décorative en haut avec forme ondulée
    c.setFillColor(colors.HexColor(primary_color))
    path = c.beginPath()
    path.moveTo(pos_x + 0.2 * cm, pos_y + card_height - 0.2 * cm)
    path.lineTo(pos_x + card_width - 0.2 * cm, pos_y + card_height - 0.2 * cm)
    path.lineTo(pos_x + card_width - 0.2 * cm, pos_y + card_height - 1.2 * cm)
    # Courbe ondulée
    path.curveTo(pos_x + card_width * 0.75, pos_y + card_height - 1.15 * cm, 
                 pos_x + card_width * 0.5, pos_y + card_height - 1.25 * cm, 
                 pos_x + card_width * 0.25, pos_y + card_height - 1.15 * cm)
    path.curveTo(pos_x + card_width * 0.15, pos_y + card_height - 1.12 * cm, 
                 pos_x + 0.2 * cm, pos_y + card_height - 1.15 * cm, 
                 pos_x + 0.2 * cm, pos_y + card_height - 1.2 * cm)
    path.close()
    c.drawPath(path, fill=1, stroke=0)
    
    # Nom de l'école dans l'en-tête
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(colors.white)
    school_name = (eleve.classe.ecole.nom if eleve.classe.ecole else "École").upper()
    if len(school_name) > 35:
        school_name = school_name[:32] + "..."
    c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 0.6 * cm, school_name)
    
    # Sous-titre "CARTE SCOLAIRE"
    c.setFont("Helvetica-Bold", 7)
    c.setFillAlpha(0.85)
    c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 0.95 * cm, "CARTE SCOLAIRE")
    c.setFillAlpha(1)
    
    # Année scolaire (petit badge)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.15)
    c.roundRect(pos_x + 0.25 * cm, pos_y + card_height - 1.25 * cm, 2 * cm, 0.35 * cm, 3, stroke=0, fill=1)
    c.setFillAlpha(1)
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor(primary_color))
    current_year = timezone.now().year
    next_year = current_year + 1
    annee_scolaire = f"{current_year}-{next_year}"
    c.drawCentredString(pos_x + 1.25 * cm, pos_y + card_height - 1.15 * cm, annee_scolaire)
    
    # Photo de l'élève avec bordures arrondies (à gauche) - agrandi
    photo_width = 2.2 * cm
    photo_height = 2.2 * cm
    photo_x = pos_x + 0.3 * cm
    photo_y = pos_y + card_height/2 - photo_height/2
    
    # Afficher la photo ou placeholder
    try:
        print(f"DEBUG Carte - Élève: {eleve.prenom} {eleve.nom}")
        print(f"DEBUG Carte - Photo field: {eleve.photo}")
        print(f"DEBUG Carte - Has photo: {bool(eleve.photo)}")
        
        if eleve.photo:
            print(f"DEBUG Carte - Photo path: {eleve.photo.path if hasattr(eleve.photo, 'path') else 'NO PATH'}")
            if hasattr(eleve.photo, 'path'):
                print(f"DEBUG Carte - File exists: {os.path.exists(eleve.photo.path)}")
        
        if eleve.photo and hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
            from PIL import Image, ImageDraw
            
            print(f"DEBUG Carte - Chargement photo: {eleve.photo.path}")
            # Ouvrir l'image
            img = Image.open(eleve.photo.path)
            print(f"DEBUG Carte - Image mode: {img.mode}, size: {img.size}")
            
            # Convertir en RGB si nécessaire
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Redimensionner l'image
            size = (int(photo_width * 28.35), int(photo_height * 28.35))  # Conversion cm vers pixels
            img = img.resize(size, Image.Resampling.LANCZOS)
            
            # Créer un masque avec coins arrondis
            mask = Image.new('L', size, 0)
            draw = ImageDraw.Draw(mask)
            # Rectangle avec coins arrondis (radius = 10% de la largeur)
            radius = int(size[0] * 0.15)
            draw.rounded_rectangle([(0, 0), size], radius=radius, fill=255)
            
            # Appliquer le masque
            output = Image.new('RGBA', size, (255, 255, 255, 0))
            output.paste(img, (0, 0))
            output.putalpha(mask)
            
            # Sauvegarder temporairement
            temp_buffer = io.BytesIO()
            output.save(temp_buffer, format='PNG')
            temp_buffer.seek(0)
            
            # Ombre portée
            c.setFillColor(colors.HexColor('#000000'))
            c.setFillAlpha(0.15)
            c.roundRect(photo_x + 0.05 * cm, photo_y - 0.05 * cm, photo_width, photo_height, 0.3 * cm, stroke=0, fill=1)
            c.setFillAlpha(1)
            
            # Dessiner la photo sur le PDF
            c.drawImage(temp_buffer, photo_x, photo_y, width=photo_width, height=photo_height, mask='auto')
            
            # Bordure colorée autour de la photo
            c.setStrokeColor(colors.HexColor(primary_color))
            c.setLineWidth(3)
            c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=1, fill=0)
                
        else:
            # Placeholder si pas de photo
            c.setFillColor(colors.HexColor(light_color))
            c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=0, fill=1)
            c.setStrokeColor(colors.HexColor(primary_color))
            c.setLineWidth(3)
            c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=1, fill=0)
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.HexColor(primary_color))
            c.drawCentredString(photo_x + photo_width/2, photo_y + photo_height/2 - 0.1 * cm, "PHOTO")
            
    except Exception as e:
        # Placeholder en cas d'erreur
        print(f"Erreur photo carte scolaire: {e}")
        c.setFillColor(colors.HexColor(light_color))
        c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor(primary_color))
        c.setLineWidth(3)
        c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=1, fill=0)
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawCentredString(photo_x + photo_width/2, photo_y + photo_height/2 - 0.1 * cm, "PHOTO")
    
    # Zone d'information avec fond subtil (poussée vers la droite)
    info_box_x = photo_x + photo_width + 0.4 * cm
    info_box_y = pos_y + 0.4 * cm
    info_box_width = card_width - (photo_x + photo_width + 0.6 * cm - pos_x)
    info_box_height = card_height - 1.7 * cm
    
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.08)
    c.roundRect(info_box_x, info_box_y, info_box_width, info_box_height, 6, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Informations de l'élève (poussées vers la droite)
    info_x = info_box_x + 0.2 * cm
    info_y_start = pos_y + card_height/2 + 0.5 * cm
    
    # Nom complet (en majuscules et gras)
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.HexColor('#1f2937'))
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    if len(nom_complet) > 16:
        nom_complet = nom_complet[:13] + "..."
    c.drawString(info_x, info_y_start, nom_complet)
    
    # Ligne décorative sous le nom
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(1.5)
    c.line(info_x, info_y_start - 0.1 * cm, info_x + 3 * cm, info_y_start - 0.1 * cm)
    
    # Informations détaillées avec espacement
    y_info = info_y_start - 0.55 * cm
    c.setFillColor(colors.HexColor('#374151'))
    
    # Matricule avec espacement
    c.setFont("Helvetica-Bold", 9)
    c.drawString(info_x, y_info, "N° : ")
    c.setFont("Helvetica", 9)
    c.drawString(info_x + 0.9 * cm, y_info, eleve.matricule)
    
    y_info -= 0.45 * cm
    # Classe avec espacement
    c.setFont("Helvetica-Bold", 9)
    c.drawString(info_x, y_info, "Classe : ")
    c.setFont("Helvetica", 9)
    c.drawString(info_x + 1.3 * cm, y_info, eleve.classe.nom)
    
    y_info -= 0.45 * cm
    # Date de naissance avec espacement
    if eleve.date_naissance:
        date_naiss = eleve.date_naissance.strftime("%d/%m/%Y")
        c.setFont("Helvetica-Bold", 9)
        c.drawString(info_x, y_info, "Né(e) le : ")
        c.setFont("Helvetica", 8)
        c.drawString(info_x + 1.5 * cm, y_info, date_naiss)
    
    y_info -= 0.45 * cm
    # Téléphone responsable avec espacement
    if eleve.responsable_principal and eleve.responsable_principal.telephone:
        tel = eleve.responsable_principal.telephone[:14]
        c.setFont("Helvetica-Bold", 9)
        c.drawString(info_x, y_info, "Tél : ")
        c.setFont("Helvetica", 8)
        c.drawString(info_x + 0.9 * cm, y_info, tel)
    
    # Pied de page moderne
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.05)
    c.roundRect(pos_x + 0.3 * cm, pos_y + 0.15 * cm, card_width - 0.6 * cm, 0.35 * cm, 4, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(pos_x + card_width/2, pos_y + 0.23 * cm, 
                     "Cette carte est strictement personnelle")
    
    # Finaliser le PDF
    c.showPage()
    c.save()
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"carte_scolaire_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def statistiques(request):
    """Statistiques globales de l'école"""
    from eleves.models import Ecole
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else Ecole.objects.first()
    
    # Récupérer les classes disponibles
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('nom')
    
    # Classe sélectionnée
    classe_id = request.GET.get('classe_id')
    classe_selectionnee = None
    if classe_id:
        try:
            classe_selectionnee = classes.get(id=classe_id)
        except ClasseNote.DoesNotExist:
            pass
    
    # Statistiques globales de l'école
    total_eleves = Eleve.objects.filter(statut='ACTIF').count()
    total_classes = ClasseEleve.objects.all().count()
    
    # Période sélectionnée
    periode = request.GET.get('periode', 'TRIMESTRE_1')
    
    # Initialiser les statistiques
    nb_evalues = 0
    nb_non_evalues = 0
    nb_non_admis = 0
    nb_a_suivre = 0
    nb_excellents = 0
    nb_precaution = 0
    eleves_non_admis = []
    eleves_a_suivre = []
    eleves_excellents = []
    eleves_precaution = []
    recommandations = []
    
    # Si une classe est sélectionnée, calculer les statistiques
    if classe_selectionnee and periode:
        from decimal import Decimal
        
        # Récupérer les élèves de la classe
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_selectionnee.nom,
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole
        ).first()
        
        if classe_eleve:
            eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
        else:
            eleves = Eleve.objects.none()
        
        # Récupérer les matières de la classe
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True)
        
        if eleves.exists() and matieres.exists():
            # Pour chaque élève, calculer sa moyenne
            for eleve in eleves:
                total_points = Decimal('0')
                total_coefficients = Decimal('0')
                has_notes = False
                
                for matiere in matieres:
                    # Récupérer les évaluations de la période
                    evaluations = Evaluation.objects.filter(
                        matiere=matiere,
                        periode=periode
                    )
                    
                    if evaluations.exists():
                        # Calculer la moyenne de la matière
                        total_devoirs = Decimal('0')
                        count_devoirs = 0
                        total_compo = Decimal('0')
                        count_compo = 0
                        
                        for evaluation in evaluations:
                            try:
                                note_obj = NoteEleve.objects.get(eleve=eleve, evaluation=evaluation)
                                if note_obj.note is not None and not note_obj.absent:
                                    has_notes = True
                                    if evaluation.type_evaluation in ['COMPOSITION', 'EXAMEN']:
                                        total_compo += Decimal(str(note_obj.note))
                                        count_compo += 1
                                    else:
                                        total_devoirs += Decimal(str(note_obj.note))
                                        count_devoirs += 1
                            except NoteEleve.DoesNotExist:
                                pass
                        
                        # Calculer la moyenne de la matière
                        moyenne_continue = total_devoirs / count_devoirs if count_devoirs > 0 else None
                        note_composition = total_compo / count_compo if count_compo > 0 else None
                        
                        moyenne_matiere = None
                        if moyenne_continue is not None and note_composition is not None:
                            # Formule corrigée : (Moyenne Continue + Composition) / 2 (poids égal)
                            moyenne_matiere = (moyenne_continue + note_composition) / 2
                        elif note_composition is not None:
                            moyenne_matiere = note_composition
                        elif moyenne_continue is not None:
                            moyenne_matiere = moyenne_continue
                        
                        if moyenne_matiere is not None:
                            total_points += moyenne_matiere * matiere.coefficient
                            total_coefficients += matiere.coefficient
                
                # Calculer la moyenne générale
                if has_notes and total_coefficients > 0:
                    moyenne_generale = float(total_points / total_coefficients)
                    nb_evalues += 1
                    
                    # Classifier l'élève
                    eleve_data = {
                        'eleve': eleve,
                        'moyenne': round(moyenne_generale, 2)
                    }
                    
                    if moyenne_generale < 10:
                        nb_non_admis += 1
                        eleves_non_admis.append(eleve_data)
                    elif moyenne_generale < 12:
                        nb_a_suivre += 1
                        eleves_a_suivre.append(eleve_data)
                    elif moyenne_generale < 14:
                        nb_precaution += 1
                        eleves_precaution.append(eleve_data)
                    else:
                        nb_excellents += 1
                        eleves_excellents.append(eleve_data)
                else:
                    nb_non_evalues += 1
            
            # Calculer les taux
            total_eleves_classe = eleves.count()
            taux_reussite = round((nb_evalues - nb_non_admis) / nb_evalues * 100, 1) if nb_evalues > 0 else 0
            taux_echec = round(nb_non_admis / nb_evalues * 100, 1) if nb_evalues > 0 else 0
            
            # Générer des recommandations
            if nb_non_admis > 0:
                recommandations.append({
                    'type': 'DANGER',
                    'message': f'{nb_non_admis} élève(s) en difficulté (moyenne < 10/20). Mise en place de soutien scolaire recommandée.',
                    'couleur': 'danger'
                })
            
            if nb_a_suivre > 0:
                recommandations.append({
                    'type': 'WARNING',
                    'message': f'{nb_a_suivre} élève(s) à suivre (moyenne entre 10 et 12/20). Accompagnement personnalisé conseillé.',
                    'couleur': 'warning'
                })
            
            if nb_excellents > 0:
                recommandations.append({
                    'type': 'SUCCESS',
                    'message': f'{nb_excellents} élève(s) excellent(s) (moyenne ≥ 14/20). Félicitations !',
                    'couleur': 'success'
                })
            
            if nb_non_evalues > 0:
                recommandations.append({
                    'type': 'INFO',
                    'message': f'{nb_non_evalues} élève(s) non évalué(s) pour cette période.',
                    'couleur': 'info'
                })
            
            if not recommandations:
                recommandations.append({
                    'type': 'INFO',
                    'message': 'Statistiques calculées avec succès.',
                    'couleur': 'info'
                })
        else:
            recommandations.append({
                'type': 'WARNING',
                'message': 'Aucune donnée disponible pour cette classe et cette période.',
                'couleur': 'warning'
            })
    else:
        recommandations.append({
            'type': 'INFO',
            'message': 'Sélectionnez une classe et une période pour voir les statistiques.',
            'couleur': 'info'
        })
    
    context = {
        'titre_page': 'Statistiques de l\'École',
        'ecole': ecole,
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'periode': periode,
        'total_eleves': total_eleves,
        'total_classes': total_classes,
        'nb_evalues': nb_evalues,
        'nb_non_evalues': nb_non_evalues,
        'nb_non_admis': nb_non_admis,
        'nb_a_suivre': nb_a_suivre,
        'nb_excellents': nb_excellents,
        'nb_precaution': nb_precaution,
        'total_echecs': nb_non_admis,
        'taux_reussite': round((nb_evalues - nb_non_admis) / nb_evalues * 100, 1) if nb_evalues > 0 else 0,
        'taux_echec': round(nb_non_admis / nb_evalues * 100, 1) if nb_evalues > 0 else 0,
        'eleves_non_admis': eleves_non_admis,
        'eleves_a_suivre': eleves_a_suivre,
        'eleves_excellents': eleves_excellents,
        'eleves_precaution': eleves_precaution,
        'strategies': [],
        'recommandations': recommandations,
        'periodes': [
            # Mois
            ('OCTOBRE', 'Octobre'),
            ('NOVEMBRE', 'Novembre'),
            ('DECEMBRE', 'Décembre'),
            ('JANVIER', 'Janvier'),
            ('FEVRIER', 'Février'),
            ('MARS', 'Mars'),
            ('AVRIL', 'Avril'),
            ('MAI', 'Mai'),
            ('JUIN', 'Juin'),
            # Trimestres
            ('TRIMESTRE_1', '1er Trimestre'),
            ('TRIMESTRE_2', '2ème Trimestre'),
            ('TRIMESTRE_3', '3ème Trimestre'),
            # Semestres
            ('SEMESTRE_1', '1er Semestre'),
            ('SEMESTRE_2', '2ème Semestre'),
        ]
    }
    
    return render(request, 'notes/statistiques.html', context)

@login_required
def gerer_classes(request):
    """Gérer les classes - Liste et ajout"""
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole).order_by('-date_creation')
    else:
        classes = ClasseNote.objects.all().order_by('-date_creation')
    
    # Statistiques
    total_classes = classes.count()
    classes_actives = classes.filter(actif=True).count()
    
    # Traitement du formulaire
    form = ClasseNoteForm(ecole=ecole)
    if request.method == 'POST':
        form = ClasseNoteForm(request.POST, ecole=ecole)
        if form.is_valid():
            classe = form.save(commit=False)
            if ecole:
                classe.ecole = ecole
            classe.cree_par = request.user
            
            # Vérifier si la classe existe déjà
            classe_existante = ClasseNote.objects.filter(
                ecole=classe.ecole,
                nom=classe.nom,
                annee_scolaire=classe.annee_scolaire
            ).first()
            
            if classe_existante:
                messages.error(request, f'❌ La classe "{classe.nom}" existe déjà pour l\'année scolaire {classe.annee_scolaire}.')
            else:
                try:
                    classe.save()
                    messages.success(request, f'✅ Classe "{classe.nom}" créée avec succès!')
                    return redirect('notes:gerer_classes')
                except IntegrityError:
                    messages.error(request, f'❌ Erreur: La classe "{classe.nom}" existe déjà pour cette année scolaire.')
        else:
            messages.error(request, '❌ Veuillez corriger les erreurs dans le formulaire.')
    
    context = {
        'titre_page': 'Gestion des Classes',
        'classes': classes,
        'total_classes': total_classes,
        'classes_actives': classes_actives,
        'form': form,
    }
    
    return render(request, 'notes/gerer_classes.html', context)

@login_required
def modifier_classe(request, classe_id):
    """Modifier une classe"""
    classe = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Vérifier que l'utilisateur a accès à cette classe
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Sécurité : Vérifier que la classe appartient à l'école de l'utilisateur
    if ecole and classe.ecole != ecole:
        messages.error(request, "❌ Vous n'avez pas accès à cette classe")
        return redirect('notes:gerer_classes')
    
    if request.method == 'POST':
        form = ClasseNoteForm(request.POST, instance=classe, ecole=ecole)
        if form.is_valid():
            # Sauvegarder sans commit pour garder l'école
            instance = form.save(commit=False)
            # S'assurer que l'école reste la même
            instance.ecole = classe.ecole
            instance.save()
            messages.success(request, f'✅ Classe "{classe.nom}" modifiée avec succès!')
            return redirect('notes:gerer_classes')
        else:
            # Afficher les erreurs détaillées
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'❌ {field}: {error}')
    else:
        form = ClasseNoteForm(instance=classe, ecole=ecole)
    
    context = {
        'titre_page': 'Modifier une Classe',
        'form': form,
        'classe': classe,
    }
    
    return render(request, 'notes/modifier_classe.html', context)

@login_required
def supprimer_classe(request, classe_id):
    """Supprimer une classe avec vérification des données liées"""
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        
        # Sécurité : Vérifier que la classe appartient à l'école de l'utilisateur
        user_profil = getattr(request.user, 'profil', None)
        ecole = user_profil.ecole if user_profil else None
        if ecole and classe.ecole != ecole:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
        
        # Vérifier s'il y a des matières ou notes liées
        has_matieres = MatiereNote.objects.filter(classe=classe).exists()
        has_evaluations = Evaluation.objects.filter(matiere__classe=classe).exists()
        
        if has_matieres or has_evaluations:
            # Désactiver au lieu de supprimer
            classe.actif = False
            classe.save()
            return JsonResponse({
                'success': True,
                'message': f'Classe "{classe.nom}" désactivée (contient des données)'
            })
        else:
            # Supprimer si pas de données
            nom_classe = classe.nom
            classe.delete()
            return JsonResponse({
                'success': True,
                'message': f'Classe "{nom_classe}" supprimée avec succès'
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def gerer_matieres(request):
    """Gérer les matières par classe"""
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Filtres
    classe_id = request.GET.get('classe_id')
    classe_selectionnee = None
    matieres = []
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee).order_by('nom')
    
    # Traitement du formulaire d'ajout
    form = MatiereNoteForm()
    if request.method == 'POST' and classe_selectionnee:
        form = MatiereNoteForm(request.POST)
        if form.is_valid():
            matiere = form.save(commit=False)
            matiere.classe = classe_selectionnee
            matiere.cree_par = request.user
            
            # Gestion des coefficients selon le niveau
            from .calculs_moyennes import detecter_niveau_scolaire
            niveau = detecter_niveau_scolaire(classe_selectionnee.nom)
            
            if niveau == 'MATERNELLE':
                # MATERNELLE: Pas de coefficient (pas de notes numériques)
                matiere.coefficient = None
            elif niveau == 'PRIMAIRE':
                # PRIMAIRE: Coefficient = 1.0 (pas de pondération)
                matiere.coefficient = 1.0
            # COLLEGE/LYCEE: Garder le coefficient saisi par l'utilisateur
            
            matiere.save()
            messages.success(request, f'✅ Matière "{matiere.nom}" ajoutée avec succès!')
            return redirect(f'/notes/matieres/?classe_id={classe_id}')
        else:
            messages.error(request, '❌ Veuillez corriger les erreurs dans le formulaire.')
    
    # Détecter le niveau de la classe sélectionnée pour le template
    niveau_classe = None
    if classe_selectionnee:
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_classe = detecter_niveau_scolaire(classe_selectionnee.nom)
    
    context = {
        'titre_page': 'Gestion des Matières',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'niveau_classe': niveau_classe,
        'matieres': matieres,
        'form': form,
    }
    
    return render(request, 'notes/gerer_matieres.html', context)

@login_required
def modifier_matiere(request, matiere_id):
    """Modifier une matière"""
    matiere = get_object_or_404(MatiereNote, pk=matiere_id)
    
    # Sécurité : Vérifier que la matière appartient à l'école de l'utilisateur
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    if ecole and matiere.classe.ecole != ecole:
        messages.error(request, "❌ Vous n'avez pas accès à cette matière")
        return redirect('notes:gerer_matieres')
    classe_id = matiere.classe.id
    
    if request.method == 'POST':
        form = MatiereNoteForm(request.POST, instance=matiere)
        if form.is_valid():
            instance = form.save(commit=False)
            
            # Gestion des coefficients selon le niveau
            from .calculs_moyennes import detecter_niveau_scolaire
            niveau = detecter_niveau_scolaire(matiere.classe.nom)
            
            if niveau == 'MATERNELLE':
                # MATERNELLE: Pas de coefficient (pas de notes numériques)
                instance.coefficient = None
            elif niveau == 'PRIMAIRE':
                # PRIMAIRE: Coefficient = 1.0 (pas de pondération)
                instance.coefficient = 1.0
            # COLLEGE/LYCEE: Garder le coefficient saisi par l'utilisateur
            
            instance.save()
            messages.success(request, f'✅ Matière "{matiere.nom}" modifiée avec succès!')
            return redirect(f'/notes/matieres/?classe_id={classe_id}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'❌ {field}: {error}')
    else:
        form = MatiereNoteForm(instance=matiere)
    
    context = {
        'titre_page': 'Modifier une Matière',
        'form': form,
        'matiere': matiere,
        'classe_id': classe_id,
    }
    
    return render(request, 'notes/modifier_matiere.html', context)

@login_required
def supprimer_matiere(request, matiere_id):
    """Supprimer une matière avec vérification des données liées"""
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)
        
        # Sécurité : Vérifier que la matière appartient à l'école de l'utilisateur
        user_profil = getattr(request.user, 'profil', None)
        ecole = user_profil.ecole if user_profil else None
        if ecole and matiere.classe.ecole != ecole:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
        
        classe_id = matiere.classe.id
        
        # Vérifier s'il y a des évaluations ou notes liées
        has_evaluations = Evaluation.objects.filter(matiere=matiere).exists()
        has_notes = NoteEleve.objects.filter(evaluation__matiere=matiere).exists()
        
        if has_evaluations or has_notes:
            # Désactiver au lieu de supprimer
            matiere.actif = False
            matiere.save()
            return JsonResponse({
                'success': True,
                'message': f'Matière "{matiere.nom}" désactivée (contient des données)',
                'classe_id': classe_id
            })
        else:
            # Supprimer si pas de données
            nom_matiere = matiere.nom
            matiere.delete()
            return JsonResponse({
                'success': True,
                'message': f'Matière "{nom_matiere}" supprimée avec succès',
                'classe_id': classe_id
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def charger_matieres_defaut(request, classe_id):
    """Charger les matières par défaut pour une classe"""
    from notes.matieres_defaut import charger_matieres_pour_classe
    
    classe = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Sécurité : Vérifier les permissions
    user_profil = getattr(request.user, 'profil', None)
    if user_profil and user_profil.ecole and classe.ecole != user_profil.ecole:
        messages.error(request, "❌ Vous n'avez pas accès à cette classe")
        return redirect('notes:gerer_matieres')
    
    # Charger les matières par défaut
    nombre_creees, nombre_existantes, erreurs = charger_matieres_pour_classe(classe, request.user)
    
    # Messages de retour
    if nombre_creees > 0:
        messages.success(request, f'✅ {nombre_creees} matière(s) créée(s) avec succès pour {classe.nom}')
    
    if nombre_existantes > 0:
        messages.info(request, f'ℹ️ {nombre_existantes} matière(s) existaient déjà')
    
    if erreurs:
        for erreur in erreurs:
            messages.warning(request, f'⚠️ {erreur}')
    
    if nombre_creees == 0 and nombre_existantes == 0:
        messages.warning(request, 'Aucune matière par défaut disponible pour ce niveau')
    
    return redirect('notes:gerer_matieres')

@login_required
def gerer_evaluations(request):
    """Gérer les évaluations"""
    return render(request, 'notes/gerer_evaluations.html', {'titre_page': 'Gestion des Évaluations'})

@login_required
def creer_evaluation(request):
    """Créer une évaluation"""
    messages.info(request, 'Fonction en cours de développement')
    return redirect('notes:gerer_evaluations')

@login_required
def gerer_eleves(request):
    """Gérer les élèves - Consultation par classe"""
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes de notes
    if ecole:
        classes_notes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes_notes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Filtres
    classe_id = request.GET.get('classe_id')
    classe_selectionnee = None
    eleves = []
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        
        # Trouver la classe d'élèves correspondante
        classe_eleve = None
        
        # Utiliser l'école de la classe sélectionnée (pas celle de l'utilisateur)
        ecole_classe = classe_selectionnee.ecole
        
        # Méthode 1 : Correspondance exacte par nom, année scolaire ET école
        try:
            # Utiliser filter().first() au lieu de get() pour éviter MultipleObjectsReturned
            classe_eleve = ClasseEleve.objects.filter(
                nom=classe_selectionnee.nom,
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole_classe  # Utiliser l'école de la classe, pas celle de l'utilisateur
            ).first()
        except Exception:
            pass
        
        # Méthode 2 : Recherche approximative améliorée
        if not classe_eleve:
            # Extraire les chiffres et mots-clés du nom
            nom_recherche = classe_selectionnee.nom.lower()
            # Nettoyer le nom
            nom_nettoye = nom_recherche.replace('série', '').replace('année', '').replace('ème', '').replace('eme', '').strip()
            
            # Chercher dans les classes de la même année scolaire ET de la même école
            classes_similaires = ClasseEleve.objects.filter(
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole_classe  # Filtrer par l'école de la classe
            )
            
            # Essayer de trouver une correspondance
            for classe_candidate in classes_similaires:
                nom_candidate = classe_candidate.nom.lower()
                # Vérifier si le nom nettoyé est contenu dans le nom de la classe candidate
                if nom_nettoye in nom_candidate or any(mot in nom_candidate for mot in nom_nettoye.split() if len(mot) > 2):
                    classe_eleve = classe_candidate
                    break
        
        # Méthode 3 : Recherche par niveau si disponible
        if not classe_eleve and hasattr(classe_selectionnee, 'niveau'):
            niveau_map = {
                'MATERNELLE': ['maternelle', 'petite', 'moyenne', 'grande'],
                'PRIMAIRE': ['cp', 'ce', 'cm', 'primaire', '1ère', '2ème', '3ème', '4ème', '5ème', '6ème'],
                'COLLEGE': ['7ème', '8ème', '9ème', '10ème', 'collège', 'college'],
                'LYCEE': ['11ème', '12ème', 'lycée', 'lycee', 'terminale', 'première', 'seconde']
            }
            
            niveau = classe_selectionnee.niveau
            mots_cles = niveau_map.get(niveau, [])
            nom_classe_lower = classe_selectionnee.nom.lower()
            
            # Filtrer par année scolaire ET école
            for classe_candidate in ClasseEleve.objects.filter(
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole_classe  # Filtrer par l'école de la classe
            ):
                nom_candidate_lower = classe_candidate.nom.lower()
                # Vérifier si des mots-clés du niveau sont présents dans les deux noms
                if any(mot in nom_classe_lower and mot in nom_candidate_lower for mot in mots_cles):
                    classe_eleve = classe_candidate
                    break
        
        # Récupérer les élèves si une classe a été trouvée
        if classe_eleve:
            eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
    
    # Calculer les statistiques
    total_eleves = len(eleves)
    eleves_avec_notes = 0  # TODO: Calculer le nombre d'élèves avec des notes
    
    # Informations de débogage pour l'utilisateur
    classe_eleve_trouvee = None
    classes_disponibles = []
    if classe_selectionnee and not eleves:
        # Lister toutes les classes d'élèves disponibles pour aider au diagnostic
        # Filtrer par l'école de la classe sélectionnée
        classes_disponibles = list(ClasseEleve.objects.filter(
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
        ).values_list('nom', flat=True))
    
    context = {
        'titre_page': 'Gestion des Élèves',
        'classes': classes_notes,
        'classe_selectionnee': classe_selectionnee,
        'eleves': eleves,
        'total_eleves': total_eleves,
        'eleves_avec_notes': eleves_avec_notes,
        'classes_disponibles': classes_disponibles,
    }
    
    return render(request, 'notes/gerer_eleves.html', context)

@login_required
def saisir_notes(request):
    """Saisir les notes"""
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Paramètres de sélection
    classe_id = request.GET.get('classe_id')
    periode_classement = request.GET.get('periode', '')
    matiere_id = request.GET.get('matiere_id')
    type_note = request.GET.get('type_note', '')
    periode = request.GET.get('periode', '')
    system_type = request.GET.get('system_type', 'semestre')
    
    classe_selectionnee = None
    matiere_selectionnee = None
    matieres = []
    eleves = []
    evaluations = []
    niveau_enseignement = 'SECONDAIRE'
    
    # Types de notes disponibles (selon le niveau)
    types_notes_disponibles = [
        ('mensuelle', 'Note Mensuelle'),
        ('trimestrielle', 'Note Trimestrielle'),
        ('semestrielle', 'Note Semestrielle'),
        ('composition', 'Composition'),
        ('appreciation', 'Appréciation'),
    ]
    
    # Périodes disponibles par défaut
    periodes_disponibles = []
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        niveau_enseignement = classe_selectionnee.niveau_enseignement
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
        
        # Déterminer les périodes disponibles selon le type de note
        if type_note == 'appreciation':
            # Pour les appréciations : trimestres
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        elif type_note == 'mensuelle':
            # Pour les notes mensuelles : 9 mois de l'année scolaire
            periodes_disponibles = [
                ('OCTOBRE', 'Octobre'),
                ('NOVEMBRE', 'Novembre'),
                ('DECEMBRE', 'Décembre'),
                ('JANVIER', 'Janvier'),
                ('FEVRIER', 'Février'),
                ('MARS', 'Mars'),
                ('AVRIL', 'Avril'),
                ('MAI', 'Mai'),
                ('JUIN', 'Juin'),
            ]
        elif type_note == 'trimestrielle':
            # Pour les notes trimestrielles
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        elif type_note == 'semestrielle':
            # Pour les notes semestrielles
            periodes_disponibles = [
                ('SEMESTRE_1', '1er Semestre'),
                ('SEMESTRE_2', '2ème Semestre'),
            ]
        elif type_note == 'composition':
            # Pour les compositions : selon le système choisi
            if system_type == 'semestre':
                periodes_disponibles = [
                    ('SEMESTRE_1', '1er Semestre'),
                    ('SEMESTRE_2', '2ème Semestre'),
                ]
            else:  # trimestre
                periodes_disponibles = [
                    ('TRIMESTRE_1', '1er Trimestre'),
                    ('TRIMESTRE_2', '2ème Trimestre'),
                    ('TRIMESTRE_3', '3ème Trimestre'),
                ]
        else:
            # Par défaut : trimestres
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        
        if matiere_id:
            matiere_selectionnee = get_object_or_404(MatiereNote, pk=matiere_id)
            if periode:
                evaluations = Evaluation.objects.filter(matiere=matiere_selectionnee, periode=periode).order_by('date_evaluation')
            else:
                evaluations = Evaluation.objects.none()
            
            # Vérifier si des notes existent déjà pour cette période
            notes_existantes_count = 0
            if evaluations.exists():
                notes_existantes_count = NoteEleve.objects.filter(
                    evaluation__in=evaluations
                ).count()
            
            # Récupérer les élèves
            try:
                # Mapping spécial pour les classes avec noms différents (même que consulter_notes)
                mapping_classes = {
                    61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
                    59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
                }
                
                if classe_selectionnee.id in mapping_classes:
                    classe_eleve = ClasseEleve.objects.filter(
                        id=mapping_classes[classe_selectionnee.id]
                    ).first()
                else:
                    # Utiliser filter().first() au lieu de get() pour éviter MultipleObjectsReturned
                    classe_eleve = ClasseEleve.objects.filter(
                        nom=classe_selectionnee.nom,
                        annee_scolaire=classe_selectionnee.annee_scolaire,
                        ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
                    ).first()
                
                if classe_eleve:
                    eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
                else:
                    # Recherche approximative
                    nom_recherche = classe_selectionnee.nom.lower().replace('série', '').replace('année', '').strip()
                    classes_similaires = ClasseEleve.objects.filter(
                        nom__icontains=nom_recherche,
                        ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
                    )
                    if classes_similaires.count() >= 1:
                        classe_eleve = classes_similaires.first()
                        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
            except Exception:
                pass
    
    # Préparer les informations sur les notes existantes
    notes_deja_saisies = False
    nombre_notes_existantes = 0
    notes_existantes_json = '{}'
    if matiere_selectionnee and periode and evaluations.exists():
        qs_notes = NoteEleve.objects.filter(
            evaluation__in=evaluations
        ).select_related('eleve')
        nombre_notes_existantes = qs_notes.count()
        notes_deja_saisies = nombre_notes_existantes > 0
        try:
            import json as _json
            notes_map = {}
            for n in qs_notes:
                try:
                    # Dernière valeur écrase la précédente si plusieurs evals
                    notes_map[n.eleve_id] = {
                        'note': float(n.note) if getattr(n, 'note', None) is not None else None,
                        'absent': bool(getattr(n, 'absent', False)),
                        'appreciation': getattr(n, 'appreciation_finale', None),
                        'commentaire': getattr(n, 'commentaire', None),
                    }
                except Exception:
                    continue
            notes_existantes_json = _json.dumps(notes_map)
        except Exception:
            notes_existantes_json = '{}'
    
    # Déterminer la note maximale selon le niveau
    note_sur = 20  # Par défaut
    if classe_selectionnee:
        if niveau_enseignement == 'PRIMAIRE' or 'PRIMAIRE' in classe_selectionnee.niveau:
            note_sur = 10
        else:
            note_sur = 20
    
    context = {
        'titre_page': 'Saisie des Notes',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'matieres': matieres,
        'matiere_selectionnee': matiere_selectionnee,
        'eleves': eleves,
        'evaluations': evaluations,
        'type_note': type_note,
        'periode': periode,
        'periodes_disponibles': periodes_disponibles,
        'types_notes_disponibles': types_notes_disponibles,
        'system_type': system_type,
        'niveau_enseignement': niveau_enseignement,
        'notes_existantes_json': notes_existantes_json,
        'notes_deja_saisies': notes_deja_saisies,
        'nombre_notes_existantes': nombre_notes_existantes,
        'note_sur': note_sur,
    }
    
    return render(request, 'notes/saisir_notes.html', context)

@login_required
def liste_saisie_pdf(request):
    """Générer un PDF de la liste de saisie des notes"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    periode = request.GET.get('periode')
    type_note = request.GET.get('type_note', '')
    
    if not all([classe_id, matiere_id, periode]):
        return HttpResponse("Paramètres manquants", status=400)
    
    classe = get_object_or_404(ClasseNote, pk=classe_id)
    matiere = get_object_or_404(MatiereNote, pk=matiere_id)
    
    # Déterminer le type de notation selon le niveau
    niveau_enseignement = classe.niveau_enseignement
    is_maternelle = niveau_enseignement == 'MATERNELLE'
    is_primaire = niveau_enseignement == 'PRIMAIRE' or 'PRIMAIRE' in classe.niveau
    is_appreciation = type_note == 'appreciation'
    
    # Déterminer la note maximale
    if is_primaire:
        note_sur = 10
    else:
        note_sur = 20
    
    # Récupérer les élèves avec mapping spécial (même logique que saisir_notes et consulter_notes)
    mapping_classes = {
        61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
        59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
    }
    
    if classe.id in mapping_classes:
        classe_eleve = ClasseEleve.objects.filter(
            id=mapping_classes[classe.id]
        ).first()
    else:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom,
            annee_scolaire=classe.annee_scolaire,
            ecole=classe.ecole
        ).first()
    
    if classe_eleve:
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
    else:
        eleves = []
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#007bff'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Titre
    elements.append(Paragraph(f"Liste de Saisie - {classe.nom}", title_style))
    elements.append(Paragraph(f"Matière: {matiere.nom} | Période: {periode}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # En-tête du tableau selon le type
    if is_appreciation:
        # Pour les appréciations (Maternelle)
        data = [['N°', 'Matricule', 'Prénom', 'Nom', 'Appréciation', 'Commentaire', 'Absent']]
        col_widths = [1*cm, 3*cm, 4*cm, 4*cm, 4*cm, 5*cm, 2*cm]
    else:
        # Pour les notes (Primaire /10 ou Secondaire /20)
        data = [['N°', 'Matricule', 'Prénom', 'Nom', f'Note /{note_sur}', 'Absent', 'Observations']]
        col_widths = [1*cm, 3*cm, 4*cm, 4*cm, 2*cm, 2*cm, 6*cm]
    
    for idx, eleve in enumerate(eleves, 1):
        if is_appreciation:
            data.append([
                str(idx),
                eleve.matricule,
                eleve.prenom,
                eleve.nom,
                '',  # Appréciation à remplir
                '',  # Commentaire
                ''   # Absent à cocher
            ])
        else:
            data.append([
                str(idx),
                eleve.matricule,
                eleve.prenom,
                eleve.nom,
                '',  # Note à remplir
                '',  # Absent à cocher
                ''   # Observations
            ])
    
    # Style du tableau
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    
    # Retourner la réponse
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="liste_saisie_{classe.nom}_{matiere.code}.pdf"'
    
    return response

@login_required
def sauvegarder_notes(request):
    """Sauvegarder les notes saisies avec support des transactions"""
    from django.http import JsonResponse
    from eleves.models import Eleve
    from django.db import transaction
    import json
    from decimal import Decimal, InvalidOperation
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        notes_data = data.get('notes', [])
        evaluation_id = data.get('evaluation_id')
        matiere_id = data.get('matiere_id')
        periode = data.get('periode')
        
        # Validation des paramètres
        if not all([matiere_id, periode]):
            return JsonResponse({'success': False, 'error': 'Paramètres manquants (matière ou période)'}, status=400)
        
        # Récupérer ou créer l'évaluation
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)
        
        # Sécurité : Vérifier que la matière appartient à l'école de l'utilisateur
        user_profil = getattr(request.user, 'profil', None)
        ecole = user_profil.ecole if user_profil else None
        if ecole and matiere.classe.ecole != ecole:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
        
        # Créer l'évaluation si elle n'existe pas
        if not evaluation_id:
            evaluation, created = Evaluation.objects.get_or_create(
                matiere=matiere,
                periode=periode,
                defaults={
                    'date_evaluation': timezone.now().date(),
                    'note_sur': 20 if matiere.classe.niveau_enseignement == 'SECONDAIRE' else 10,
                    'coefficient': matiere.coefficient,
                }
            )
            logger.info(f"Évaluation {'créée' if created else 'récupérée'}: {evaluation.id}")
        else:
            evaluation = get_object_or_404(Evaluation, pk=evaluation_id)
        
        notes_sauvegardees = 0
        notes_modifiees = 0
        erreurs = []
        notes_details = []
        
        # Utiliser une transaction pour garantir l'intégrité des données
        with transaction.atomic():
            for note_data in notes_data:
                try:
                    eleve_id = note_data.get('eleve_id')
                    absent = note_data.get('absent', False)
                    
                    if not eleve_id:
                        continue
                    
                    eleve = Eleve.objects.get(pk=eleve_id)
                    
                    # Traiter selon le type de note
                    if 'appreciation' in note_data:
                        # Appréciation (pour maternelle) - utiliser AppreciationMaternelle
                        from notes.models import AppreciationMaternelle
                        
                        appreciation = note_data.get('appreciation', '').strip()
                        commentaire = note_data.get('commentaire', '').strip()
                        
                        if not appreciation and not absent:
                            continue
                        
                        # Déterminer le trimestre depuis la période
                        trimestre = periode if periode.startswith('TRIMESTRE') else 'TRIMESTRE_1'
                        
                        note_obj, created = AppreciationMaternelle.objects.update_or_create(
                            eleve=eleve,
                            matiere=matiere,
                            trimestre=trimestre,
                            annee_scolaire=matiere.classe.annee_scolaire,
                            defaults={
                                'appreciation': appreciation if appreciation else None,
                                'commentaire': commentaire if commentaire else None,
                                'absent': absent,
                                'cree_par': request.user,
                            }
                        )
                    else:
                        # Note numérique
                        note_value = str(note_data.get('note', '')).strip()
                        
                        # Valider la note
                        note_decimal = None
                        if note_value and not absent:
                            try:
                                note_decimal = Decimal(note_value.replace(',', '.'))
                                if note_decimal < 0 or note_decimal > evaluation.note_sur:
                                    erreurs.append(f"{eleve.nom} {eleve.prenom}: Note invalide (doit être entre 0 et {evaluation.note_sur})")
                                    continue
                            except (InvalidOperation, ValueError, TypeError):
                                erreurs.append(f"{eleve.nom} {eleve.prenom}: Format de note invalide")
                                continue
                        elif not absent:
                            # Pas de note et pas absent, on ignore
                            continue
                        
                        # Créer ou mettre à jour la note
                        note_obj, created = NoteEleve.objects.update_or_create(
                            eleve=eleve,
                            evaluation=evaluation,
                            defaults={
                                'note': note_decimal if not absent else None,
                                'absent': absent,
                                'cree_par': request.user,
                            }
                        )
                    
                    if created:
                        notes_sauvegardees += 1
                    else:
                        notes_modifiees += 1
                    
                    # Ajouter les détails de la note sauvegardée
                    note_detail = {
                        'eleve_id': eleve_id,
                        'eleve_nom': f"{eleve.nom} {eleve.prenom}",
                        'absent': note_obj.absent,
                        'created': created
                    }
                    
                    # Ajouter les champs spécifiques selon le type d'objet
                    if hasattr(note_obj, 'appreciation'):
                        # AppreciationMaternelle
                        note_detail['appreciation'] = note_obj.appreciation
                        note_detail['note'] = None
                    else:
                        # NoteEleve
                        note_detail['note'] = float(note_obj.note) if note_obj.note else None
                        note_detail['appreciation'] = None
                    
                    notes_details.append(note_detail)
                    
                except Eleve.DoesNotExist:
                    erreurs.append(f"Élève ID {eleve_id} introuvable")
                except Exception as e:
                    logger.error(f"Erreur lors de la sauvegarde de la note pour l'élève {eleve_id}: {str(e)}")
                    erreurs.append(f"Erreur: {str(e)}")
        
        # Préparer la réponse
        total_notes = notes_sauvegardees + notes_modifiees
        message_parts = []
        
        if notes_sauvegardees > 0:
            message_parts.append(f"{notes_sauvegardees} nouvelle(s) note(s) ajoutée(s)")
        if notes_modifiees > 0:
            message_parts.append(f"{notes_modifiees} note(s) modifiée(s)")
        
        message = " et ".join(message_parts) if message_parts else "Aucune note à sauvegarder"
        
        response_data = {
            'success': True,
            'notes_sauvegardees': notes_sauvegardees,
            'notes_modifiees': notes_modifiees,
            'total': total_notes,
            'message': f'✅ {message}',
            'notes_details': notes_details,
            'evaluation_id': evaluation.id
        }
        
        if erreurs:
            response_data['erreurs'] = erreurs
            response_data['message'] += f' ⚠️ {len(erreurs)} erreur(s) détectée(s)'
        
        logger.info(f"Sauvegarde terminée: {total_notes} notes traitées, {len(erreurs)} erreurs")
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de la sauvegarde des notes: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur serveur: {str(e)}'}, status=500)

@login_required
def supprimer_notes(request):
    """Supprimer les notes d'une évaluation ou d'une période spécifique"""
    from django.http import JsonResponse
    from eleves.models import Eleve
    from django.db import transaction
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        matiere_id = data.get('matiere_id')
        periode = data.get('periode')
        eleve_ids = data.get('eleve_ids', [])  # Liste optionnelle d'élèves spécifiques
        
        # Validation des paramètres
        if not all([matiere_id, periode]):
            return JsonResponse({'success': False, 'error': 'Paramètres manquants (matière ou période)'}, status=400)
        
        # Récupérer la matière
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)
        
        # Récupérer les évaluations correspondantes
        evaluations = Evaluation.objects.filter(matiere=matiere, periode=periode)
        
        if not evaluations.exists():
            return JsonResponse({'success': False, 'error': 'Aucune évaluation trouvée pour cette période'}, status=404)
        
        notes_supprimees = 0
        
        # Utiliser une transaction pour garantir l'intégrité des données
        with transaction.atomic():
            # Construire la requête de suppression
            notes_query = NoteEleve.objects.filter(evaluation__in=evaluations)
            
            # Si des élèves spécifiques sont fournis, filtrer par élève
            if eleve_ids:
                notes_query = notes_query.filter(eleve_id__in=eleve_ids)
            
            # Compter et supprimer
            notes_supprimees = notes_query.count()
            notes_query.delete()
            
            logger.info(f"Suppression de {notes_supprimees} notes pour la matière {matiere.nom}, période {periode}")
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {notes_supprimees} note(s) supprimée(s) avec succès',
            'notes_supprimees': notes_supprimees
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression des notes: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur serveur: {str(e)}'}, status=500)

@login_required
def consulter_notes(request):
    """Consulter les notes - Vue complète par classe"""
    from decimal import Decimal
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Paramètres de sélection
    classe_id = request.GET.get('classe_id')
    periode_classement = request.GET.get('periode', '')
    
    classe_selectionnee = None
    matieres = []
    eleves_toutes_notes = []
    niveau_enseignement = 'SECONDAIRE'
    periodes_disponibles = []
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        niveau_enseignement = classe_selectionnee.niveau_enseignement
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
        
        # Déterminer les périodes disponibles (toutes les périodes possibles)
        periodes_disponibles = [
            # Mois
            ('OCTOBRE', 'Octobre'),
            ('NOVEMBRE', 'Novembre'),
            ('DECEMBRE', 'Décembre'),
            ('JANVIER', 'Janvier'),
            ('FEVRIER', 'Février'),
            ('MARS', 'Mars'),
            ('AVRIL', 'Avril'),
            ('MAI', 'Mai'),
            ('JUIN', 'Juin'),
            # Trimestres
            ('TRIMESTRE_1', '1er Trimestre'),
            ('TRIMESTRE_2', '2ème Trimestre'),
            ('TRIMESTRE_3', '3ème Trimestre'),
            # Semestres
            ('SEMESTRE_1', '1er Semestre'),
            ('SEMESTRE_2', '2ème Semestre'),
        ]
        
        # Récupérer les élèves
        try:
            # Mapping spécial pour les classes avec noms différents
            mapping_classes = {
                61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
                59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
            }
            
            if classe_selectionnee.id in mapping_classes:
                classe_eleve = ClasseEleve.objects.filter(
                    id=mapping_classes[classe_selectionnee.id]
                ).first()
            else:
                # Utiliser filter().first() au lieu de get() pour éviter MultipleObjectsReturned
                classe_eleve = ClasseEleve.objects.filter(
                    nom=classe_selectionnee.nom,
                    annee_scolaire=classe_selectionnee.annee_scolaire,
                    ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
                ).first()
            
            if classe_eleve:
                eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
            else:
                eleves = []
            
            # Pour chaque élève, récupérer toutes ses notes pour toutes les matières
            for eleve in eleves:
                notes_par_matiere = {}
                somme_moy_coef = Decimal('0')
                somme_coef = Decimal('0')
                
                for matiere in matieres:
                    # Récupérer toutes les évaluations de cette matière
                    evaluations = Evaluation.objects.filter(matiere=matiere)
                    if periode_classement:
                        evaluations = evaluations.filter(periode=periode_classement)
                    
                    notes_matiere = {
                        'evaluations': [],
                        'notes': [],
                        'moyenne': None
                    }
                    
                    # Calculer la moyenne pondérée par coefficient d'évaluation
                    total_pondere = Decimal('0')
                    total_coef_eval = Decimal('0')
                    
                    for evaluation in evaluations:
                        try:
                            note_obj = NoteEleve.objects.get(eleve=eleve, evaluation=evaluation)
                            notes_matiere['evaluations'].append(evaluation)
                            notes_matiere['notes'].append({
                                'evaluation': evaluation,
                                'note': note_obj.note,
                                'absent': note_obj.absent,
                            })
                            # Compter les absences comme 0
                            coef_eval = Decimal(str(evaluation.coefficient or 1))
                            if note_obj.absent or note_obj.note is None:
                                # Absence = 0
                                total_pondere += Decimal('0') * coef_eval
                            else:
                                total_pondere += Decimal(str(note_obj.note)) * coef_eval
                            total_coef_eval += coef_eval
                        except NoteEleve.DoesNotExist:
                            notes_matiere['evaluations'].append(evaluation)
                            notes_matiere['notes'].append({
                                'evaluation': evaluation,
                                'note': None,
                                'absent': False,
                            })
                            # Pas de note = 0
                            coef_eval = Decimal(str(evaluation.coefficient or 1))
                            total_pondere += Decimal('0') * coef_eval
                            total_coef_eval += coef_eval
                    
                    # Calculer la moyenne pondérée de la matière
                    if total_coef_eval > 0:
                        moyenne_matiere = total_pondere / total_coef_eval
                        notes_matiere['moyenne'] = round(float(moyenne_matiere), 2)
                        
                        # Ajouter à la moyenne générale pondérée par coefficient de matière
                        coef_matiere = Decimal(str(matiere.coefficient or 1))
                        somme_moy_coef += moyenne_matiere * coef_matiere
                        somme_coef += coef_matiere
                    
                    notes_par_matiere[matiere.id] = notes_matiere
                
                # Calculer la moyenne générale
                moyenne_generale = None
                if somme_coef > 0:
                    moyenne_generale = round(float(somme_moy_coef / somme_coef), 2)
                
                eleves_toutes_notes.append({
                    'eleve': eleve,
                    'notes_par_matiere': notes_par_matiere,
                    'moyenne_generale': moyenne_generale,
                })
                
        except ClasseEleve.DoesNotExist:
            pass
    
    # Utiliser la fonction centralisée pour garantir cohérence totale avec les bulletins
    from .utils_rangs import calculer_rangs_classe_periode
    
    # Calculer les rangs avec la fonction centralisée
    # Utiliser la période sélectionnée ou OCTOBRE par défaut
    if periodes_disponibles and classe_selectionnee:
        if periode_classement:
            periode_pour_rang = periode_classement  # Utiliser la période sélectionnée
        else:
            # Utiliser OCTOBRE par défaut (période la plus courante)
            periode_pour_rang = 'OCTOBRE'
            # Si OCTOBRE n'est pas disponible, prendre la première période
            periodes_codes = [p[0] for p in periodes_disponibles]
            if 'OCTOBRE' not in periodes_codes:
                periode_pour_rang = periodes_disponibles[0][0]
        
        rangs_dict = calculer_rangs_classe_periode(classe_selectionnee, periode_pour_rang)
        
        # Attribuer les rangs aux élèves
        for eleve_data in eleves_toutes_notes:
            eleve_id = eleve_data['eleve'].id
            rang_info = rangs_dict.get(eleve_id)
            if rang_info:
                eleve_data['rang'] = rang_info['rang']
            else:
                eleve_data['rang'] = '-'
    else:
        # Pas de période disponible, pas de rang
        for eleve_data in eleves_toutes_notes:
            eleve_data['rang'] = '-'
    
    # Trier par rang pour l'affichage (élèves avec rang d'abord)
    eleves_avec_rang = [e for e in eleves_toutes_notes if e['rang'] != '-']
    eleves_sans_rang = [e for e in eleves_toutes_notes if e['rang'] == '-']
    eleves_avec_rang.sort(key=lambda x: x['moyenne_generale'], reverse=True)
    eleves_toutes_notes = eleves_avec_rang + eleves_sans_rang
    
    # Récupérer toutes les évaluations pour les en-têtes
    evaluations_par_matiere = {}
    for matiere in matieres:
        evaluations_par_matiere[matiere.id] = Evaluation.objects.filter(
            matiere=matiere
        ).order_by('periode', 'date_evaluation')
    periode_classement = periodes_disponibles[0] if periodes_disponibles else None
    
    context = {
        'titre_page': 'Consultation des Notes',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'matieres': matieres,
        'periodes_disponibles': periodes_disponibles,
        'periode_classement': periode_classement,
        'eleves_toutes_notes': eleves_toutes_notes,
        'evaluations_par_matiere': evaluations_par_matiere,
        'niveau_enseignement': niveau_enseignement,
    }
    
    return render(request, 'notes/consulter_notes.html', context)

@login_required
def bulletin_guineen(request):
    """Bulletin guinéen"""
    return render(request, 'notes/bulletin_guineen.html', {'titre_page': 'Bulletin Guinéen'})

@login_required
def bulletin_dynamique(request):
    """Bulletin dynamique - Génération de bulletins personnalisés"""
    from decimal import Decimal
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Paramètres de sélection
    classe_id = request.GET.get('classe_id')
    eleve_id = request.GET.get('eleve_id')
    periode = request.GET.get('periode', '')
    system_type = request.GET.get('system_type', 'trimestre')  # mensuel, trimestre, semestre, annuel
    
    classe_selectionnee = None
    eleves = []
    eleve_selectionne = None
    matieres = []
    niveau_enseignement = 'SECONDAIRE'
    periodes_disponibles = []
    bulletin_data = None
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        niveau_enseignement = classe_selectionnee.niveau_enseignement
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
        
        # Déterminer les périodes disponibles selon le système
        if system_type == 'mensuel':
            periodes_disponibles = [
                ('OCTOBRE', 'Octobre'),
                ('NOVEMBRE', 'Novembre'),
                ('DECEMBRE', 'Décembre'),
                ('JANVIER', 'Janvier'),
                ('FEVRIER', 'Février'),
                ('MARS', 'Mars'),
                ('AVRIL', 'Avril'),
                ('MAI', 'Mai'),
                ('JUIN', 'Juin'),
            ]
        elif system_type == 'trimestre':
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        elif system_type == 'semestre':
            periodes_disponibles = [
                ('SEMESTRE_1', '1er Semestre'),
                ('SEMESTRE_2', '2ème Semestre'),
            ]
        elif system_type == 'annuel':
            periodes_disponibles = [
                ('ANNUEL', 'Bulletin Annuel'),
            ]
        
        # Récupérer les élèves de la classe
        # Essayer d'abord une correspondance exacte
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_selectionnee.nom,
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole
        ).first()
        
        if not classe_eleve:
            # Si pas de correspondance exacte, essayer avec insensibilité à la casse
            classe_eleve = ClasseEleve.objects.filter(
                nom__iexact=classe_selectionnee.nom,
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=classe_selectionnee.ecole
            ).first()
        
        if classe_eleve:
            eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
        else:
            try:
                # Recherche approximative
                pass
                eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
            except (ClasseEleve.DoesNotExist, ClasseEleve.MultipleObjectsReturned):
                eleves = []
        
        # Préparer un bulletin vide avec toutes les matières (dès la sélection de la classe)
        if matieres:
            # Déterminer le titre de la période
            titre_periode = ''
            for code, libelle in periodes_disponibles:
                if code == periode:
                    titre_periode = libelle
                    break
            
            # Déterminer le type de bulletin
            type_bulletin = system_type
            
            bulletin_data = {
                'eleve': None,
                'classe': classe_selectionnee,
                'periode': periode,
                'system_type': system_type,
                'type_bulletin': type_bulletin,
                'titre_periode': titre_periode,
                'titre_moyenne': 'Moyenne Continue',
                'titre_composition': 'Composition',
                'matieres_notes': [],
                'moyenne_generale': None,
                'rang': None,
                'mention': None,
                'appreciation': '',
                'appreciation_generale': '',
                'effectif': len(eleves),
                'mois_libelle': titre_periode,
            }
            
            # Initialiser eleve_selectionne
            eleve_selectionne = None
            
            # Si un élève est sélectionné
            if eleve_id:
                eleve_selectionne = get_object_or_404(Eleve, pk=eleve_id)
                bulletin_data['eleve'] = eleve_selectionne
            
            # Pour chaque matière, préparer la structure
            total_points = Decimal('0')
            total_coefficients = Decimal('0')
            
            for matiere in matieres:
                # Initialiser les variables
                moyenne_continue = None
                note_composition = None
                moyennes_mensuelles = []
                
                # Si une période est sélectionnée, récupérer les données
                if periode and eleve_selectionne:
                    if system_type in ['trimestre', 'semestre']:
                        # NOUVEAU: Utiliser les moyennes mensuelles détaillées
                        from .utils_moyennes_mensuelles import calculer_bulletin_avec_details_mensuels
                        
                        data_matiere = calculer_bulletin_avec_details_mensuels(
                            eleve_selectionne, matiere, system_type, periode
                        )
                        
                        moyenne_continue = data_matiere['moyenne_continue']
                        note_composition = data_matiere['note_composition']
                        moyennes_mensuelles = data_matiere['moyennes_mensuelles']
                        
                    elif system_type == 'mensuel':
                        # NOUVEAU: Pour les bulletins mensuels, utiliser NoteMensuelle
                        try:
                            note_mensuelle = NoteMensuelle.objects.get(
                                eleve=eleve_selectionne,
                                matiere=matiere,
                                mois=periode,
                                annee_scolaire=classe_selectionnee.annee_scolaire
                            )
                            if not note_mensuelle.absent and note_mensuelle.note is not None:
                                moyenne_continue = float(note_mensuelle.note)
                        except NoteMensuelle.DoesNotExist:
                            pass
                    else:
                        # ANCIEN: Système annuel ou autres - garder l'ancien calcul
                        # Chercher les évaluations de cette matière
                        evaluations = Evaluation.objects.filter(
                            matiere=matiere,
                            periode=periode
                        ).order_by('date_evaluation')
                        
                        # Si pas d'évaluation trouvée, chercher par nom de matière
                        if not evaluations.exists():
                            from django.db.models import Q
                            evaluations = Evaluation.objects.filter(
                                Q(matiere__nom=matiere.nom) &
                                Q(matiere__classe=classe_selectionnee) &
                                Q(periode=periode)
                            ).order_by('date_evaluation')
                        
                        # Séparer devoirs/contrôles (moyenne continue) et compositions
                        total_devoirs = Decimal('0')
                        count_devoirs = 0
                        total_compo = Decimal('0')
                        count_compo = 0
                        
                        for evaluation in evaluations:
                            try:
                                note_obj = NoteEleve.objects.get(eleve=eleve_selectionne, evaluation=evaluation)
                                if note_obj.note is not None and not note_obj.absent:
                                    # Déterminer le type d'évaluation
                                    if evaluation.type_evaluation in ['COMPOSITION', 'EXAMEN']:
                                        total_compo += Decimal(str(note_obj.note))
                                        count_compo += 1
                                    else:
                                        # DEVOIR, CONTROLE, INTERROGATION
                                        total_devoirs += Decimal(str(note_obj.note))
                                        count_devoirs += 1
                            except NoteEleve.DoesNotExist:
                                pass
                        
                        # Calculer les moyennes
                        if count_devoirs > 0:
                            moyenne_continue = round(float(total_devoirs / count_devoirs), 2)
                        
                        if count_compo > 0:
                            note_composition = round(float(total_compo / count_compo), 2)
                
                # Calculer la moyenne de la matière selon le système guinéen
                # Si trimestre/semestre: moyenne = (moyenne_continue + composition) / 2 (poids égal)
                # Si mensuel: moyenne = moyenne_continue uniquement (pas de composition)
                moyenne_matiere = None
                
                if system_type == 'mensuel':
                    moyenne_matiere = moyenne_continue
                elif moyenne_continue is not None and note_composition is not None:
                    # Formule corrigée : (Moyenne Continue + Composition) / 2 (poids égal)
                    moyenne_matiere = round((moyenne_continue + note_composition) / 2, 2)
                elif note_composition is not None:
                    # Seulement la composition
                    moyenne_matiere = note_composition
                elif moyenne_continue is not None:
                    # Seulement la moyenne continue
                    moyenne_matiere = moyenne_continue
                
                # Calculer les points
                points = None
                if moyenne_matiere is not None:
                    points = round(moyenne_matiere * float(matiere.coefficient), 2)
                    total_points += Decimal(str(moyenne_matiere)) * matiere.coefficient
                    total_coefficients += matiere.coefficient
                
                # Préparer les notes pour l'affichage
                notes_matiere = []
                if system_type in ['trimestre', 'semestre']:
                    # NOUVEAU: Inclure les moyennes mensuelles détaillées
                    if moyennes_mensuelles:
                        # Ajouter les moyennes mensuelles
                        for moy_mens in moyennes_mensuelles:
                            notes_matiere.append({
                                'note': moy_mens['moyenne'],
                                'absent': moy_mens['absent'],
                                'libelle': moy_mens['libelle'],
                                'type': 'mensuelle'
                            })
                    
                    # Ajouter la moyenne continue calculée
                    notes_matiere.append({
                        'note': moyenne_continue,
                        'absent': False,
                        'libelle': 'Moy. Continue',
                        'type': 'continue'
                    })
                    
                    # Ajouter la composition
                    notes_matiere.append({
                        'note': note_composition,
                        'absent': False,
                        'libelle': 'Composition',
                        'type': 'composition'
                    })
                    
                elif system_type == 'mensuel':
                    notes_matiere = [
                        {'note': moyenne_continue, 'absent': False, 'libelle': 'Moyenne', 'type': 'mensuelle'}
                    ]
                
                bulletin_data['matieres_notes'].append({
                    'matiere': matiere,
                    'notes': notes_matiere,
                    'moyennes_mensuelles': moyennes_mensuelles,  # NOUVEAU: Détails mensuels
                    'moyenne_continue': moyenne_continue,
                    'note_composition': note_composition,
                    'moyenne': moyenne_matiere,
                    'coefficient': matiere.coefficient,
                    'points': points,
                    'total': points,  # Alias pour compatibilité
                })
            
            # Ajouter les totaux au bulletin
            bulletin_data['total_points'] = round(float(total_points), 2) if total_points > 0 else None
            bulletin_data['total_coefficients'] = float(total_coefficients) if total_coefficients > 0 else None
            
            # Calculer la moyenne générale (seulement si élève sélectionné)
            if eleve_selectionne and total_coefficients > 0:
                moyenne_generale = round(float(total_points / total_coefficients), 2)
                bulletin_data['moyenne_generale'] = moyenne_generale
                
                # Essayer d'abord de récupérer depuis le Classement (plus précis)
                from .models import Classement
                classement = Classement.objects.filter(
                    eleve=eleve_selectionne,
                    classe=classe_selectionnee,
                    periode=periode,
                    annee_scolaire=classe_selectionnee.annee_scolaire
                ).first()
                
                if classement:
                    # Utiliser les données du classement (plus précis)
                    bulletin_data['moyenne_generale'] = float(classement.moyenne_generale)
                    bulletin_data['mention'] = classement.mention
                    bulletin_data['appreciation'] = classement.appreciation
                    bulletin_data['rang'] = classement.rang_formate
                    bulletin_data['total_points'] = float(classement.total_points)
                    bulletin_data['total_coefficients'] = float(classement.total_coefficients)
                else:
                    # Sinon, calculer à la volée (fallback)
                    from decimal import Decimal as _Dec
                    moyenne_dec = _Dec(str(moyenne_generale))
                    bulletin_data['mention'] = obtenir_mention_intelligente(moyenne_dec)
                    bulletin_data['appreciation'] = obtenir_appreciation_intelligente(moyenne_dec, eleve_selectionne.prenom)
                    
                    # Calculer le rang
                    # Récupérer toutes les moyennes de la classe pour cette période
                    if periode:
                        moyennes_classe_pairs = []  # (eleve_id, moyenne_float)
                        for eleve_classe in eleves:
                            total_eleve = Decimal('0')
                            total_coef_eleve = Decimal('0')
                            
                            for matiere in matieres:
                                # Récupérer les évaluations de cette matière pour la période
                                evals = Evaluation.objects.filter(
                                    matiere=matiere,
                                    periode=periode
                                )
                                
                                # Si pas d'évaluation, chercher par nom (matière recréée)
                                if not evals.exists():
                                    from django.db.models import Q
                                    evals = Evaluation.objects.filter(
                                        Q(matiere__nom=matiere.nom) &
                                        Q(matiere__classe=classe_selectionnee) &
                                        Q(periode=periode)
                                    )
                                
                                # Séparer devoirs et compositions
                                total_dev = Decimal('0')
                                count_dev = 0
                                total_comp = Decimal('0')
                                count_comp = 0
                                
                                for ev in evals:
                                    try:
                                        n = NoteEleve.objects.get(eleve=eleve_classe, evaluation=ev)
                                        # Traiter les absences comme 0 (harmonisation avec le classement)
                                        note_value = Decimal(str(n.note)) if n.note is not None and not n.absent else Decimal('0')
                                        
                                        if ev.type_evaluation in ['COMPOSITION', 'EXAMEN']:
                                            total_comp += note_value
                                            count_comp += 1
                                        else:
                                            total_dev += note_value
                                            count_dev += 1
                                    except NoteEleve.DoesNotExist:
                                        # Pas de note = 0
                                        if ev.type_evaluation in ['COMPOSITION', 'EXAMEN']:
                                            count_comp += 1
                                        else:
                                            count_dev += 1
                                
                                # Calculer la moyenne de la matière
                                moy_cont = total_dev / count_dev if count_dev > 0 else None
                                moy_comp = total_comp / count_comp if count_comp > 0 else None
                                moy_mat = None
                                
                                if system_type == 'mensuel':
                                    moy_mat = moy_cont
                                elif moy_cont is not None and moy_comp is not None:
                                    moy_mat = (moy_cont + moy_comp * 2) / 3
                                elif moy_comp is not None:
                                    moy_mat = moy_comp
                                elif moy_cont is not None:
                                    moy_mat = moy_cont
                                
                                # Si l'élève a une moyenne dans cette matière
                                if moy_mat is not None:
                                    total_eleve += moy_mat * matiere.coefficient
                                    total_coef_eleve += matiere.coefficient
                            
                            if total_coef_eleve > 0:
                                moy_eleve = float(total_eleve / total_coef_eleve)
                                moyennes_classe_pairs.append((eleve_classe.id, moy_eleve))
                        
                        # Trier et attribuer les rangs avec gestion des ex-aequo
                        moyennes_classe_pairs.sort(key=lambda p: p[1], reverse=True)
                        rang_map = {}
                        prev_moy = None
                        prev_rank = None
                        for idx, (eid, mg) in enumerate(moyennes_classe_pairs, start=1):
                            if prev_moy is not None and abs(mg - prev_moy) < 0.01:
                                # ex-aequo: même rang que le précédent
                                rang_map[eid] = prev_rank
                            else:
                                rang_map[eid] = idx
                                prev_rank = idx
                                prev_moy = mg

                        rang_num = rang_map.get(eleve_selectionne.id)
                        
                        if rang_num is not None:
                            sexe = getattr(eleve_selectionne, 'sexe', 'M') or 'M'
                            bulletin_data['rang'] = formater_rang_intelligent(rang_num, sexe, len(moyennes_classe_pairs))
                        else:
                            bulletin_data['rang'] = "-"
                    else:
                        bulletin_data['rang'] = "-"
    
    context = {
        'titre_page': 'Bulletin Dynamique',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'eleves': eleves,
        'eleve_selectionne': eleve_selectionne,
        'matieres': matieres,
        'periodes_disponibles': periodes_disponibles,
        'periode': periode,
        'periode_selectionnee': periode,  # Alias pour le template
        'system_type': system_type,
        'niveau_enseignement': niveau_enseignement,
        'bulletin_data': bulletin_data,
        'ecole': ecole,
        'annee_scolaire': classe_selectionnee.annee_scolaire if classe_selectionnee else '',
    }
    
    return render(request, 'notes/bulletin_dynamique.html', context)

@login_required
def bulletin_dynamique_pdf(request):
    """Générer le bulletin dynamique en PDF"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.contrib import messages
    from django.shortcuts import redirect
    import tempfile
    import os
    
    # Détection du système et choix du générateur PDF
    use_weasyprint = True
    try:
        from weasyprint import HTML, CSS
        from weasyprint.text.fonts import FontConfiguration
    except (ImportError, OSError) as e:
        # Sur Windows, WeasyPrint peut nécessiter GTK+
        # Utiliser ReportLab comme alternative
        use_weasyprint = False
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch, mm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            messages.error(request, "Aucun générateur PDF disponible. Installez weasyprint ou reportlab.")
            return redirect('notes:bulletin_dynamique')
    
    # Récupérer les mêmes paramètres que bulletin_dynamique
    classe_id = request.GET.get('classe_id')
    eleve_id = request.GET.get('eleve_id')
    periode = request.GET.get('periode', '')
    system_type = request.GET.get('system_type', 'trimestre')
    
    # Validation des paramètres
    if not classe_id or not eleve_id or not periode:
        messages.error(request, "❌ Veuillez sélectionner une classe, un élève et une période avant de générer le bulletin PDF.")
        return redirect('notes:bulletin_dynamique')
    
    # Si WeasyPrint n'est pas disponible, rediriger vers l'alternative
    if not use_weasyprint:
        messages.warning(request, "⚠️ WeasyPrint non disponible sur ce système. Utilisez l'export de bulletins de classe à la place.")
        return redirect(f'/notes/bulletins/classe/pdf/?classe_id={classe_id}&periode={periode}&system_type={system_type}')
    
    # Réutiliser la logique de bulletin_dynamique pour obtenir les données
    from decimal import Decimal
    from .calculs_moyennes import calculer_moyenne_generale_eleve, calculer_classement_classe
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
    eleve_selectionne = get_object_or_404(Eleve, pk=eleve_id)
    matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
    
    # Récupérer les élèves pour calculer le rang avec mapping spécial
    mapping_classes = {
        61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
        59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
    }
    
    if classe_selectionnee.id in mapping_classes:
        classe_eleve = ClasseEleve.objects.filter(
            id=mapping_classes[classe_selectionnee.id]
        ).first()
    else:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_selectionnee.nom,
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole
        ).first()
    
    if classe_eleve:
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
    else:
        eleves = []
    
    # Déterminer le titre de la période
    periodes_map = {
        'OCTOBRE': 'Octobre', 'NOVEMBRE': 'Novembre', 'DECEMBRE': 'Décembre',
        'JANVIER': 'Janvier', 'FEVRIER': 'Février', 'MARS': 'Mars',
        'AVRIL': 'Avril', 'MAI': 'Mai', 'JUIN': 'Juin',
        'TRIMESTRE_1': '1er Trimestre', 'TRIMESTRE_2': '2ème Trimestre', 'TRIMESTRE_3': '3ème Trimestre',
        'SEMESTRE_1': '1er Semestre', 'SEMESTRE_2': '2ème Semestre',
        'ANNUEL': 'Bulletin Annuel'
    }
    titre_periode = periodes_map.get(periode, periode)
    
    # Préparer les données du bulletin (logique simplifiée)
    bulletin_data = {
        'eleve': eleve_selectionne,
        'classe': classe_selectionnee,
        'periode': periode,
        'system_type': system_type,
        'titre_periode': titre_periode,
        'matieres_notes': [],
        'moyenne_generale': None,
        'rang': None,
        'mention': None,
        'effectif': len(eleves),
        'mois_libelle': titre_periode,
    }
    
    # UTILISER LE MODULE CENTRALISÉ (SOURCE UNIQUE)
    result_centralized = calculer_moyenne_generale_eleve(eleve_selectionne, matieres, periode, system_type)
    
    # Extraire les données du calcul centralisé
    bulletin_data['moyenne_generale'] = result_centralized['moyenne_generale']
    bulletin_data['total_points'] = result_centralized['total_points']
    bulletin_data['total_coefficients'] = result_centralized['total_coefficients']
    
    # Préparer les données des matières avec moyennes mensuelles détaillées
    for detail in result_centralized['details_matieres']:
        matiere = detail['matiere']
        
        # NOUVEAU: Récupérer les moyennes mensuelles détaillées pour trimestre/semestre
        moyennes_mensuelles = []
        if system_type in ['trimestre', 'semestre']:
            from .utils_moyennes_mensuelles import calculer_bulletin_avec_details_mensuels
            
            data_matiere = calculer_bulletin_avec_details_mensuels(
                eleve_selectionne, matiere, system_type, periode
            )
            moyennes_mensuelles = data_matiere['moyennes_mensuelles']
        
        # Préparer les notes pour l'affichage
        notes_matiere = []
        if system_type in ['trimestre', 'semestre']:
            # NOUVEAU: Inclure les moyennes mensuelles détaillées
            if moyennes_mensuelles:
                # Ajouter les moyennes mensuelles
                for moy_mens in moyennes_mensuelles:
                    notes_matiere.append({
                        'note': moy_mens['moyenne'],
                        'absent': moy_mens['absent'],
                        'libelle': moy_mens['libelle'],
                        'type': 'mensuelle'
                    })
            
            # Ajouter la moyenne continue calculée
            notes_matiere.append({
                'note': detail['moyenne_continue'],
                'absent': False,
                'libelle': 'Moy. Continue',
                'type': 'continue'
            })
            
            # Ajouter la composition
            notes_matiere.append({
                'note': detail['note_composition'],
                'absent': False,
                'libelle': 'Composition',
                'type': 'composition'
            })
            
        elif system_type == 'mensuel':
            notes_matiere = [
                {'note': detail['moyenne_continue'], 'absent': False, 'libelle': 'Moyenne', 'type': 'mensuelle'}
            ]
        
        bulletin_data['matieres_notes'].append({
            'matiere': matiere,
            'notes': notes_matiere,
            'moyennes_mensuelles': moyennes_mensuelles,  # NOUVEAU: Détails mensuels
            'moyenne_continue': detail['moyenne_continue'],
            'note_composition': detail['note_composition'],
            'moyenne': detail['moyenne'],
            'coefficient': detail['coefficient'],
            'points': detail['points'],
            'total': detail['points'],
        })
    
    # Les valeurs sont déjà extraites du module centralisé
    total_points = Decimal(str(result_centralized['total_points'])) if result_centralized['total_points'] else Decimal('0')
    total_coefficients = Decimal(str(result_centralized['total_coefficients'])) if result_centralized['total_coefficients'] else Decimal('0')
    
    # Calculer le rang en utilisant le module centralisé pour tous les élèves
    if eleves and bulletin_data['moyenne_generale'] is not None:
        # Calculer le classement avec le module centralisé (SOURCE UNIQUE)
        classement_complet = calculer_classement_classe(eleves, matieres, periode, system_type)
        
        # Récupérer le rang de l'élève depuis le classement centralisé
        rang_map = classement_complet.get('rang_map', {})
        rang_num = rang_map.get(eleve_selectionne.id)
        
        if rang_num:
            # Formater le rang avec accord grammatical
            from .calculs_moyennes import formater_rang_intelligent
            sexe = getattr(eleve_selectionne, 'sexe', 'M') or 'M'
            bulletin_data['rang'] = formater_rang_intelligent(rang_num, sexe, classement_complet['total_eleves'])
        else:
            bulletin_data['rang'] = '-'
    
    # Déterminer la mention avec le module centralisé
    if bulletin_data['moyenne_generale'] is not None:
        from .calculs_moyennes import obtenir_mention_intelligente, obtenir_appreciation_intelligente
        bulletin_data['mention'] = obtenir_mention_intelligente(bulletin_data['moyenne_generale'])
        bulletin_data['appreciation'] = obtenir_appreciation_intelligente(bulletin_data['moyenne_generale'], eleve_selectionne.prenom)
    
    # Préparer les images en base64 (même logique que pour l'export de classe)
    import base64
    logo_base64 = None
    photo_base64 = None
    
    if ecole and getattr(ecole, 'logo', None):
        try:
            logo_path = ecole.logo.path
            with open(logo_path, 'rb') as img_file:
                logo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
        except Exception:
            pass
    
    if getattr(eleve_selectionne, 'photo', None):
        try:
            photo_path = eleve_selectionne.photo.path
            with open(photo_path, 'rb') as img_file:
                photo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
        except Exception:
            pass

    # CSS identique à celui utilisé pour bulletins_dynamiques_classe_pdf
    css_string = """
    @page {
        size: A4;
        margin: 10mm 10mm 10mm 10mm;
    }
    body {
        font-family: 'Arial', 'Helvetica', sans-serif;
        margin: 0;
        padding: 0;
        font-size: 12px;
    }
    .bulletin-container {
        background: white;
        width: 190mm;
        height: 277mm;
        padding: 5mm;
        position: relative;
        page-break-after: always;
        page-break-inside: avoid;
        overflow: hidden;
        box-sizing: border-box;
    }
    .bulletin-container:last-child {
        page-break-after: auto;
    }
    /* Filigrane amélioré */
    .watermark {
        position: fixed;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%) rotate(-30deg);
        opacity: 0.08;
        width: 600px;
        height: 600px;
        z-index: -1;
        pointer-events: none;
        filter: grayscale(100%);
        mix-blend-mode: multiply;
    }
    .header-section, .info-grid, .notes-table, .resultats-section,
    .appreciation-section, .signatures-section, .calcul-explanation-section, .footer-section {
        position: relative;
        z-index: 1;
    }
    """

    # Contexte pour le template "single" (même que pour l'export de classe)
    context = {
        'bulletin_data': bulletin_data,
        'classe_selectionnee': classe_selectionnee,
        'ecole': ecole,
        'annee_scolaire': classe_selectionnee.annee_scolaire,
        'system_type': system_type,
        'logo_base64': logo_base64,
        'photo_base64': photo_base64,
    }

    # Générer le HTML complet avec le même fragment que pour l'export de classe
    bulletin_html = render_to_string('notes/bulletin_dynamique_single.html', context)
    full_html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            {css_string}
        </style>
    </head>
    <body>
        {bulletin_html}
    </body>
    </html>
    '''

    # Générer le PDF avec WeasyPrint (même moteur que pour l'export de classe)
    try:
        from weasyprint import HTML, CSS
        from weasyprint.text.fonts import FontConfiguration

        font_config = FontConfiguration()

        response = HttpResponse(content_type='application/pdf')
        filename = f"bulletin_{eleve_selectionne.nom}_{eleve_selectionne.prenom}_{periode}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        HTML(string=full_html).write_pdf(
            response,
            stylesheets=[CSS(string=css_string, font_config=font_config)],
            font_config=font_config
        )

        return response

    except ImportError:
        # Si WeasyPrint n'est pas installé, retourner le HTML avec un message
        return HttpResponse(
            "<h1>WeasyPrint non installé</h1>"
            "<p>Pour générer des PDF, installez WeasyPrint:</p>"
            "<pre>pip install weasyprint</pre>"
            "<p>En attendant, utilisez le bouton 'Imprimer' et sélectionnez 'Enregistrer au format PDF' dans votre navigateur.</p>"
            f"<p><a href='/notes/bulletins/?classe_id={classe_id}&eleve_id={eleve_id}&periode={periode}&system_type={system_type}'>Retour au bulletin</a></p>",
            status=501
        )
    except Exception as e:
        # En cas d'erreur, retourner un message d'erreur
        return HttpResponse(
            f"Erreur lors de la génération du PDF: {str(e)}<br><br>"
            f"<a href='/notes/bulletins/?classe_id={classe_id}&eleve_id={eleve_id}&periode={periode}&system_type={system_type}'>Retour au bulletin</a>",
            status=500
        )

@login_required
def saisie_notes_simple(request):
    """Saisie notes simple"""
    return render(request, 'notes/saisie_notes_simple.html', {'titre_page': 'Saisie Notes Simple'})

@login_required
def sauvegarder_notes_guineen(request):
    """Sauvegarder notes guinéen"""
    from django.http import JsonResponse
    return JsonResponse({'success': True, 'message': 'Fonction en cours de développement'})

@login_required
def sauvegarder_appreciations_maternelle(request):
    """Sauvegarder appréciations maternelle"""
    from django.http import JsonResponse
    return JsonResponse({'success': True, 'message': 'Fonction en cours de développement'})

@login_required
def bulletins_dynamiques_classe_pdf(request):
    """Générer tous les bulletins d'une classe en un seul PDF avec système dynamique et filigrane"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from decimal import Decimal
    import tempfile
    import os
    import sys
    from django.contrib import messages
    from django.shortcuts import redirect
    
    # Détection du système et choix du générateur PDF
    use_weasyprint = True
    try:
        from weasyprint import HTML, CSS
        from weasyprint.text.fonts import FontConfiguration
    except (ImportError, OSError) as e:
        # Sur Windows, WeasyPrint peut nécessiter GTK+
        # Utiliser ReportLab comme alternative
        use_weasyprint = False
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch, mm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            messages.error(request, "Aucun générateur PDF disponible. Installez weasyprint ou reportlab.")
            return redirect('notes:bulletin_dynamique')
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode', '')
    system_type = request.GET.get('system_type', 'trimestre')
    
    # Validation des paramètres
    if not classe_id or not periode:
        messages.error(request, "❌ Veuillez sélectionner une classe et une période avant de générer les bulletins PDF.")
        return redirect('notes:bulletin_dynamique')
    
    # Convertir classe_id en entier
    try:
        classe_id = int(classe_id)
    except (ValueError, TypeError):
        messages.error(request, f"❌ ID de classe invalide: {classe_id}")
        return redirect('notes:bulletin_dynamique')
    
    # Récupérer les informations de l'école et de la classe
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Filtrage par école pour la sécurité (sauf pour les admins)
    if not request.user.is_superuser:
        from utilisateurs.utils import filter_by_user_school
        classe_selectionnee = get_object_or_404(
            filter_by_user_school(ClasseNote.objects.all(), request.user, 'ecole'), 
            pk=classe_id
        )
    else:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
    
    matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
    
    # Récupérer la classe élève correspondante avec mapping spécial
    mapping_classes = {
        61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
        59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
    }
    
    if classe_selectionnee.id in mapping_classes:
        classe_eleve = ClasseEleve.objects.filter(
            id=mapping_classes[classe_selectionnee.id]
        ).first()
    else:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_selectionnee.nom,
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole
        ).first()
        
        if not classe_eleve:
            classe_eleve = ClasseEleve.objects.filter(
                nom__iexact=classe_selectionnee.nom,
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=classe_selectionnee.ecole
            ).first()
    
    if not classe_eleve:
        messages.error(request, f"❌ Aucune classe élève correspondante trouvée pour ClasseNote ID {classe_selectionnee.id} ({classe_selectionnee.nom}).")
        return redirect('notes:bulletin_dynamique')
    
    # Récupérer tous les élèves de la classe
    eleves = Eleve.objects.filter(
        classe=classe_eleve, 
        statut='ACTIF'
    ).order_by('prenom', 'nom')
    
    if not eleves:
        messages.warning(request, "⚠️ Aucun élève actif dans cette classe.")
        return redirect('notes:bulletin_dynamique')
    
    # Déterminer le titre de la période
    periodes_map = {
        'OCTOBRE': 'Octobre', 'NOVEMBRE': 'Novembre', 'DECEMBRE': 'Décembre',
        'JANVIER': 'Janvier', 'FEVRIER': 'Février', 'MARS': 'Mars',
        'AVRIL': 'Avril', 'MAI': 'Mai', 'JUIN': 'Juin',
        'TRIMESTRE_1': '1er Trimestre', 'TRIMESTRE_2': '2ème Trimestre', 
        'TRIMESTRE_3': '3ème Trimestre',
        'SEMESTRE_1': '1er Semestre', 'SEMESTRE_2': '2ème Semestre',
        'ANNUEL': 'Bulletin Annuel'
    }
    titre_periode = periodes_map.get(periode, periode)
    
    # Préparer le CSS pour WeasyPrint avec filigrane amélioré et polices augmentées
    css_string = """
    @page {
        size: A4;
        margin: 10mm 10mm 10mm 10mm;
    }
    body {
        font-family: 'Arial', 'Helvetica', sans-serif;
        margin: 0;
        padding: 0;
        font-size: 12px;
    }
    .bulletin-container {
        background: white;
        width: 190mm;
        height: 277mm;
        padding: 5mm;
        position: relative;
        page-break-after: always;
        page-break-inside: avoid;
        overflow: hidden;
        box-sizing: border-box;
    }
    .bulletin-container:last-child {
        page-break-after: auto;
    }
    /* Filigrane amélioré */
    .watermark {
        position: fixed;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%) rotate(-30deg);
        opacity: 0.08;
        width: 600px;
        height: 600px;
        z-index: -1;
        pointer-events: none;
        filter: grayscale(100%);
        mix-blend-mode: multiply;
    }
    .header-section, .info-grid, .notes-table, .resultats-section, 
    .appreciation-section, .signatures-section, .calcul-explanation-section, .footer-section {
        position: relative;
        z-index: 1;
    }
    /* Styles du bulletin */
    .header-section {
        text-align: center;
        border-bottom: 3px solid #000;
        padding-bottom: 8px;
        margin-bottom: 10px;
    }
    .logo-left {
        position: absolute;
        top: 0;
        left: 0;
        width: 80px;
        height: 80px;
        object-fit: contain;
    }
    .photo-right {
        position: absolute;
        top: 0;
        right: 0;
        width: 85px;
        height: 100px;
        border: 2px solid #333;
        object-fit: cover;
    }
    /* Drapeau guinéen */
    .drapeau-guinee {
        position: absolute;
        top: 5px;
        right: 95px;
        width: 35px;
        height: 22px;
        display: flex;
        border: 1px solid #999;
    }
    .drapeau-rouge {
        width: 33.33%;
        background-color: #CE1126;
    }
    .drapeau-jaune {
        width: 33.33%;
        background-color: #FCD116;
    }
    .drapeau-vert {
        width: 33.33%;
        background-color: #009460;
    }
    .header-section h1 {
        font-size: 18px;
        margin: 0;
        font-weight: bold;
    }
    /* Devise nationale */
    .devise-nationale {
        font-size: 14px;
        font-weight: bold;
        margin: 3px 0;
        letter-spacing: 0.5px;
    }
    .devise-rouge {
        color: #CE1126;
    }
    .devise-jaune {
        color: #FCD116;
    }
    .devise-vert {
        color: #009460;
    }
    .ministere-education {
        font-size: 13px;
        font-weight: bold;
        color: #333;
        margin: 3px 0;
    }
    .nom-ecole {
        font-size: 15px;
        font-weight: bold;
        color: #000;
        margin: 4px 0;
        text-transform: uppercase;
    }
    .header-section h2 {
        font-size: 16px;
        margin: 5px 0;
    }
    .header-section p {
        font-size: 12px;
        margin: 3px 0;
        color: #666;
    }
    .info-grid {
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 8px;
        margin: 10px 0;
        font-size: 12px;
    }
    .info-item {
        background: #f9f9f9;
        padding: 6px 8px;
        border-radius: 3px;
        border-left: 3px solid #007bff;
    }
    .info-item strong {
        color: #007bff;
        display: block;
        font-size: 11px;
        margin-bottom: 3px;
    }
    .notes-table {
        width: 100%;
        border-collapse: collapse;
        margin: 12px 0;
    }
    .notes-table th {
        background: #007bff;
        color: white;
        padding: 8px 5px;
        font-size: 11px;
        text-align: center;
        border: 1px solid #0056b3;
    }
    .notes-table td {
        padding: 6px;
        font-size: 11px;
        text-align: center;
        border: 1px solid #ddd;
    }
    .notes-table tbody tr:nth-child(even) {
        background: #f9f9f9;
    }
    .notes-table .matiere-col {
        text-align: left;
        padding-left: 6px;
    }
    .notes-table tfoot {
        background: #333;
        color: white;
        font-weight: bold;
    }
    .resultats-section {
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 10px;
        margin: 15px 0;
        text-align: center;
    }
    .resultat-card {
        background: #f8f9fa;
        border-radius: 5px;
        padding: 10px;
        border: 1px solid #dee2e6;
    }
    .resultat-card h3 {
        font-size: 14px;
        color: #495057;
        margin: 0 0 5px 0;
    }
    .resultat-card .value {
        font-size: 20px;
        font-weight: bold;
        color: #007bff;
    }
    .mention-badge {
        display: inline-block;
        padding: 8px 20px;
        border-radius: 20px;
        font-weight: bold;
        font-size: 16px;
        margin-top: 5px;
        text-transform: uppercase;
    }
    .mention-excellent { background: #146c43; color: white; }
    .mention-tres-bien { background: #28a745; color: white; }
    .mention-bien { background: #17a2b8; color: white; }
    .mention-assez-bien { background: #ffc107; color: #000; }
    .mention-passable { background: #fd7e14; color: white; }
    .mention-faible { background: #e55353; color: white; }
    .mention-insuffisant { background: #dc3545; color: white; }
    .appreciation-section {
        background: #f0f7ff;
        border-left: 4px solid #007bff;
        padding: 12px;
        margin: 15px 0;
    }
    .appreciation-section h3 {
        font-size: 14px;
        color: #007bff;
        margin: 0 0 5px 0;
    }
    .appreciation-section p {
        font-size: 13px;
        color: #333;
        margin: 0;
        font-style: italic;
    }
    .signatures-section {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 60px;
        margin: 20px 0;
        padding: 0 40px;
    }
    .signature-box {
        text-align: center;
    }
    .signature-box .title {
        font-size: 13px;
        font-weight: bold;
        margin-bottom: 3px;
    }
    .signature-box .space {
        height: 40px;
        border-bottom: 1px solid #333;
        margin: 3px 0;
    }
    .signature-box div {
        font-size: 12px;
        color: #666;
    }
    .calcul-explanation-section {
        background: #fafafa;
        padding: 10px;
        margin: 10px 0;
        border: 1px solid #ddd;
        border-radius: 3px;
        font-size: 11px;
    }
    .calcul-explanation-section h4 {
        font-size: 10px;
        margin: 5px 0 2px 0;
        padding: 2px;
        background: #f0f0f0;
        border-left: 3px solid #0066cc;
        color: #333;
    }
    .calcul-explanation-section div {
        font-size: 10px;
        padding: 2px 4px;
        background: #fafafa;
        border: 1px solid #ddd;
        border-radius: 2px;
    }
    .calcul-explanation-section p {
        margin: 2px 0;
        color: #555;
    }
    .footer-section {
        margin-top: 15px;
        padding-top: 10px;
        border-top: 1px solid #ddd;
        font-size: 12px;
        color: #666;
    }
    .footer-section p {
        margin: 2px 0;
    }
    """
    
    # Variables pour le calcul des moyennes de classe
    all_moyennes_classe = []
    
    # Générer le bulletin pour chaque élève
    for index, eleve in enumerate(eleves, start=1):
        # Préparer les données du bulletin pour cet élève
        bulletin_data = {
            'eleve': eleve,
            'classe': classe_selectionnee,
            'periode': periode,
            'system_type': system_type,
            'titre_periode': titre_periode,
            'matieres_notes': [],
            'moyenne_generale': None,
            'rang': None,
            'mention': None,
            'appreciation': '',
            'effectif': len(eleves),
        }
        
        # Calculer les notes pour chaque matière
        total_points = Decimal('0')
        total_coefficients = Decimal('0')
        
        for matiere in matieres:
            moyenne_continue = None
            note_composition = None
            
            # Récupérer les évaluations
            evaluations = Evaluation.objects.filter(
                matiere=matiere,
                periode=periode
            ).order_by('date_evaluation')
            
            if not evaluations.exists():
                from django.db.models import Q
                evaluations = Evaluation.objects.filter(
                    Q(matiere__nom=matiere.nom) &
                    Q(matiere__classe=classe_selectionnee) &
                    Q(periode=periode)
                ).order_by('date_evaluation')
            
            # Calculer moyennes continue et composition
            total_devoirs = Decimal('0')
            count_devoirs = 0
            total_compo = Decimal('0')
            count_compo = 0
            
            for evaluation in evaluations:
                try:
                    note_obj = NoteEleve.objects.get(eleve=eleve, evaluation=evaluation)
                    if note_obj.note is not None and not note_obj.absent:
                        if evaluation.type_evaluation in ['COMPOSITION', 'EXAMEN']:
                            total_compo += Decimal(str(note_obj.note))
                            count_compo += 1
                        else:
                            total_devoirs += Decimal(str(note_obj.note))
                            count_devoirs += 1
                except NoteEleve.DoesNotExist:
                    pass
            
            if count_devoirs > 0:
                moyenne_continue = round(float(total_devoirs / count_devoirs), 2)
            
            if count_compo > 0:
                note_composition = round(float(total_compo / count_compo), 2)
            
            # Calculer la moyenne de la matière
            moyenne_matiere = None
            if system_type == 'mensuel':
                moyenne_matiere = moyenne_continue
            elif moyenne_continue is not None and note_composition is not None:
                # Formule corrigée : (Moyenne Continue + Composition) / 2 (poids égal)
                moyenne_matiere = round((moyenne_continue + note_composition) / 2, 2)
            elif note_composition is not None:
                moyenne_matiere = note_composition
            elif moyenne_continue is not None:
                moyenne_matiere = moyenne_continue
            
            # Calculer les points
            points = None
            if moyenne_matiere is not None:
                points = round(moyenne_matiere * float(matiere.coefficient), 2)
                total_points += Decimal(str(moyenne_matiere)) * matiere.coefficient
                total_coefficients += matiere.coefficient
            
            # Préparer les notes pour l'affichage
            notes_matiere = []
            if system_type in ['trimestre', 'semestre']:
                notes_matiere = [
                    {'note': moyenne_continue, 'absent': False},
                    {'note': note_composition, 'absent': False}
                ]
            elif system_type == 'mensuel':
                notes_matiere = [
                    {'note': moyenne_continue, 'absent': False}
                ]
            
            bulletin_data['matieres_notes'].append({
                'matiere': matiere,
                'notes': notes_matiere,
                'moyenne': moyenne_matiere,
                'coefficient': matiere.coefficient,
                'points': points,
            })
        
        bulletin_data['total_points'] = round(float(total_points), 2) if total_points > 0 else None
        bulletin_data['total_coefficients'] = float(total_coefficients) if total_coefficients > 0 else None
        
        # Calculer la moyenne générale
        if total_coefficients > 0:
            moyenne_generale = round(float(total_points / total_coefficients), 2)
            bulletin_data['moyenne_generale'] = moyenne_generale
            all_moyennes_classe.append((eleve.id, moyenne_generale))
            
            # Déterminer la mention et l'appréciation
            from decimal import Decimal as _Dec
            moyenne_dec = _Dec(str(moyenne_generale))
            bulletin_data['mention'] = obtenir_mention_intelligente(moyenne_dec)
            bulletin_data['appreciation'] = obtenir_appreciation_intelligente(moyenne_dec, eleve.prenom)
        else:
            bulletin_data['moyenne_generale'] = 0
            bulletin_data['mention'] = 'Non évalué'
            bulletin_data['appreciation'] = 'Aucune note disponible pour cette période.'
    
    # Calculer les rangs avec calculer_rang_intelligent
    from .calculs_intelligent import calculer_rang_intelligent
    from decimal import Decimal as _Dec
    
    rang_map = {}
    
    if all_moyennes_classe:
        # Préparer les données pour calculer_rang_intelligent
        moyennes_pour_rang = []
        for eid, mg in all_moyennes_classe:
            e_obj = eleves.get(id=eid)
            moyennes_pour_rang.append({
                'eleve_id': eid,
                'prenom': e_obj.prenom,
                'nom': e_obj.nom,
                'sexe': getattr(e_obj, 'sexe', 'M'),
                'moyenne': _Dec(str(mg))
            })
        
        # Utiliser la fonction centralisée pour garantir cohérence avec le classement
        from .utils_rangs import calculer_rangs_classe_periode
        
        rangs_dict = calculer_rangs_classe_periode(classe_selectionnee, periode)
        
        # Créer le dictionnaire de rangs (déjà formaté sans le total)
        for eleve_id, info in rangs_dict.items():
            rang_map[eleve_id] = info['rang']
    
    # Générer le HTML pour tous les bulletins
    bulletins_html = []
    
    for eleve in eleves:
        # Recréer les données du bulletin avec le rang
        bulletin_data = {
            'eleve': eleve,
            'classe': classe_selectionnee,
            'periode': periode,
            'system_type': system_type,
            'titre_periode': titre_periode,
            'matieres_notes': [],
            'moyenne_generale': None,
            'rang': None,
            'mention': None,
            'appreciation': '',
            'effectif': len(eleves),
        }
        
        # Recalculer les notes (code identique à ci-dessus)
        total_points = Decimal('0')
        total_coefficients = Decimal('0')
        
        for matiere in matieres:
            moyenne_continue = None
            note_composition = None
            
            evaluations = Evaluation.objects.filter(
                matiere=matiere,
                periode=periode
            ).order_by('date_evaluation')
            
            if not evaluations.exists():
                from django.db.models import Q
                evaluations = Evaluation.objects.filter(
                    Q(matiere__nom=matiere.nom) &
                    Q(matiere__classe=classe_selectionnee) &
                    Q(periode=periode)
                ).order_by('date_evaluation')
            
            total_devoirs = Decimal('0')
            count_devoirs = 0
            total_compo = Decimal('0')
            count_compo = 0
            
            for evaluation in evaluations:
                try:
                    note_obj = NoteEleve.objects.get(eleve=eleve, evaluation=evaluation)
                    if note_obj.note is not None and not note_obj.absent:
                        if evaluation.type_evaluation in ['COMPOSITION', 'EXAMEN']:
                            total_compo += Decimal(str(note_obj.note))
                            count_compo += 1
                        else:
                            total_devoirs += Decimal(str(note_obj.note))
                            count_devoirs += 1
                except NoteEleve.DoesNotExist:
                    pass
            
            if count_devoirs > 0:
                moyenne_continue = round(float(total_devoirs / count_devoirs), 2)
            
            if count_compo > 0:
                note_composition = round(float(total_compo / count_compo), 2)
            
            moyenne_matiere = None
            if system_type == 'mensuel':
                moyenne_matiere = moyenne_continue
            elif moyenne_continue is not None and note_composition is not None:
                # Formule corrigée : (Moyenne Continue + Composition) / 2 (poids égal)
                moyenne_matiere = round((moyenne_continue + note_composition) / 2, 2)
            elif note_composition is not None:
                moyenne_matiere = note_composition
            elif moyenne_continue is not None:
                moyenne_matiere = moyenne_continue
            
            points = None
            if moyenne_matiere is not None:
                points = round(moyenne_matiere * float(matiere.coefficient), 2)
                total_points += Decimal(str(moyenne_matiere)) * matiere.coefficient
                total_coefficients += matiere.coefficient
            
            notes_matiere = []
            if system_type in ['trimestre', 'semestre']:
                notes_matiere = [
                    {'note': moyenne_continue, 'absent': False},
                    {'note': note_composition, 'absent': False}
                ]
            elif system_type == 'mensuel':
                notes_matiere = [
                    {'note': moyenne_continue, 'absent': False}
                ]
            
            bulletin_data['matieres_notes'].append({
                'matiere': matiere,
                'notes': notes_matiere,
                'moyenne': moyenne_matiere,
                'coefficient': matiere.coefficient,
                'points': points,
            })
        
        bulletin_data['total_points'] = round(float(total_points), 2) if total_points > 0 else None
        bulletin_data['total_coefficients'] = float(total_coefficients) if total_coefficients > 0 else None
        
        if total_coefficients > 0:
            moyenne_generale = round(float(total_points / total_coefficients), 2)
            bulletin_data['moyenne_generale'] = moyenne_generale
            
            from decimal import Decimal as _Dec
            moyenne_dec = _Dec(str(moyenne_generale))
            bulletin_data['mention'] = obtenir_mention_intelligente(moyenne_dec)
            bulletin_data['appreciation'] = obtenir_appreciation_intelligente(moyenne_dec, eleve.prenom)
            
            # Ajouter le rang (déjà formaté dans rang_map)
            rang_str = rang_map.get(eleve.id)
            if rang_str:
                bulletin_data['rang'] = rang_str
            else:
                bulletin_data['rang'] = "-"
        else:
            bulletin_data['moyenne_generale'] = 0
            bulletin_data['mention'] = 'Non évalué'
            bulletin_data['appreciation'] = 'Aucune note disponible pour cette période.'
            bulletin_data['rang'] = "-"
        
        # Préparer les images en base64 pour le PDF
        import base64
        from django.conf import settings
        
        logo_base64 = None
        photo_base64 = None
        
        # Encoder le logo de l'école en base64
        if ecole and ecole.logo:
            try:
                logo_path = ecole.logo.path
                with open(logo_path, 'rb') as img_file:
                    logo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
            except:
                pass
        
        # Encoder la photo de l'élève en base64
        if eleve.photo:
            try:
                photo_path = eleve.photo.path
                with open(photo_path, 'rb') as img_file:
                    photo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
            except:
                pass
        
        # Contexte pour le template
        context = {
            'bulletin_data': bulletin_data,
            'classe_selectionnee': classe_selectionnee,
            'ecole': ecole,
            'annee_scolaire': classe_selectionnee.annee_scolaire,
            'system_type': system_type,
            'logo_base64': logo_base64,
            'photo_base64': photo_base64,
        }
        
        # Générer le HTML pour ce bulletin
        bulletin_html = render_to_string('notes/bulletin_dynamique_single.html', context)
        bulletins_html.append(bulletin_html)
    
    # Créer un template HTML complet avec tous les bulletins
    full_html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            {css_string}
        </style>
    </head>
    <body>
        {''.join(bulletins_html)}
    </body>
    </html>
    '''
    
    # Créer le PDF final
    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_{classe_selectionnee.nom}_{titre_periode.replace(' ', '_')}_{system_type}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    if use_weasyprint:
        # Utiliser WeasyPrint (préféré)
        # Configuration des fonts
        font_config = FontConfiguration()
        
        # Générer le PDF avec WeasyPrint
        HTML(string=full_html).write_pdf(
            response,
            stylesheets=[CSS(string=css_string, font_config=font_config)],
            font_config=font_config
        )
    else:
        # Utiliser ReportLab comme alternative (sur Windows sans GTK+)
        # Note: ReportLab ne supporte pas le HTML/CSS avancé
        # On va utiliser la fonction existante bulletins_classe_pdf qui utilise déjà ReportLab
        messages.warning(request, "WeasyPrint non disponible. Utilisation de ReportLab comme alternative. Le rendu peut différer légèrement.")
        
        # Pour les périodes mensuelles, utiliser une approche différente
        # car bulletins_classe_pdf ne supporte que les trimestres/semestres
        if periode in ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']:
            # Pour les périodes mensuelles, rediriger vers la consultation
            messages.info(request, f"⚠️ Les bulletins PDF pour les périodes mensuelles ne sont pas encore disponibles. Utilisez la consultation des notes.")
            return redirect(f'/notes/consulter/?classe_id={classe_id}&periode={periode}')
        else:
            # Convertir la période en format trimestre pour la fonction existante
            periode_mapping = {
                'TRIMESTRE_1': 'T1',
                'TRIMESTRE_2': 'T2', 
                'TRIMESTRE_3': 'T3',
                'SEMESTRE_1': 'S1',
                'SEMESTRE_2': 'S2'
            }
            trimestre = periode_mapping.get(periode, 'T1')  # Par défaut T1
            
            from . import views
            # Utiliser l'ID de ClasseEleve, pas ClasseNote
            classe_eleve_id = classe_eleve.id if classe_eleve else classe_id
            return views.bulletins_classe_pdf(request, int(classe_eleve_id), trimestre)
    
    return response
