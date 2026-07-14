"""
Export complet des notes par matière en PDF et Excel
Inclut toutes les notes de chaque élève pour chaque matière
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from decimal import Decimal
import io
import logging
import re

logger = logging.getLogger(__name__)

from .models import ClasseNote, MatiereNote, NoteMensuelle, CompositionNote, AppreciationMaternelle
from eleves.models import Eleve, Classe as ClasseEleve


def get_notes_eleves_par_matiere(classe, periode, eleves, matieres, est_maternelle=False):
    """
    Récupère toutes les notes des élèves par matière pour une période donnée.
    Retourne un dictionnaire structuré avec les données.
    Pour la maternelle, récupère les appréciations (A+, A, B+, etc.) au lieu des notes numériques.
    """
    from .utils_rangs import calculer_rangs_classe_periode
    
    # Périodes
    periodes_mensuelles = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
    periodes_trimestrielles = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
    periodes_semestrielles = ['SEMESTRE_1', 'SEMESTRE_2']
    periodes_annuelles = {'ANNUEL_TRIM': 'annuel_trimestriel', 'ANNUEL_SEM': 'annuel_semestriel'}

    # Récupérer les rangs centralisés
    rangs_dict = calculer_rangs_classe_periode(classe, periode, use_cache=False)

    # Résultat annuel : moyennes annuelles par matière calculées en batch (rapide)
    annuel_map = {}
    if periode in periodes_annuelles and not est_maternelle:
        from .calculs_moyennes import calculer_moyennes_classe_annuelle_optimise
        res_annuel = calculer_moyennes_classe_annuelle_optimise(
            eleves, matieres, periodes_annuelles[periode]
        )
        for eid, data in res_annuel.items():
            annuel_map[eid] = {
                det['matiere'].id: det.get('moyenne_annuelle')
                for det in data.get('details_matieres', [])
            }
    
    resultats = []
    
    for eleve in eleves:
        notes_eleve = {
            'eleve': eleve,
            'notes_par_matiere': {},
            'moyenne_generale': None,
            'rang': '-'
        }
        
        # Récupérer rang et moyenne depuis la source centralisée
        rang_info = rangs_dict.get(eleve.id)
        if rang_info:
            notes_eleve['moyenne_generale'] = float(rang_info['moyenne'])
            notes_eleve['rang'] = rang_info['rang']
            notes_eleve['rang_num'] = rang_info['rang_num']
        else:
            notes_eleve['rang_num'] = 9999
        
        # Récupérer les notes par matière
        for matiere in matieres:
            note_value = None
            absent = False
            appreciation = None  # Pour maternelle
            
            # Pour la maternelle : récupérer les appréciations
            if est_maternelle:
                try:
                    appr_obj = AppreciationMaternelle.objects.get(
                        eleve=eleve,
                        matiere=matiere,
                        trimestre=periode,
                        annee_scolaire=classe.annee_scolaire
                    )
                    appreciation = appr_obj.appreciation
                    absent = appr_obj.absent
                except AppreciationMaternelle.DoesNotExist:
                    pass
            elif periode in periodes_mensuelles:
                # Notes mensuelles
                try:
                    note_obj = NoteMensuelle.objects.get(
                        eleve=eleve,
                        matiere=matiere,
                        mois=periode,
                        annee_scolaire=classe.annee_scolaire
                    )
                    note_value = float(note_obj.note) if note_obj.note is not None else None
                    absent = note_obj.absent
                except NoteMensuelle.DoesNotExist:
                    pass
            elif periode in periodes_trimestrielles or periode in periodes_semestrielles:
                # Notes de composition
                try:
                    note_obj = CompositionNote.objects.get(
                        eleve=eleve,
                        matiere=matiere,
                        periode=periode,
                        annee_scolaire=classe.annee_scolaire
                    )
                    note_value = float(note_obj.note) if note_obj.note is not None else None
                    absent = note_obj.absent
                except CompositionNote.DoesNotExist:
                    pass
            elif periode in periodes_annuelles:
                # Moyenne annuelle de la matière (calculée en batch plus haut)
                moy_ann = annuel_map.get(eleve.id, {}).get(matiere.id)
                note_value = float(moy_ann) if moy_ann is not None else None

            notes_eleve['notes_par_matiere'][matiere.id] = {
                'note': note_value,
                'absent': absent,
                'appreciation': appreciation,  # Pour maternelle
                'coefficient': float(matiere.coefficient) if matiere.coefficient else 1.0
            }
        
        resultats.append(notes_eleve)
    
    # Trier par rang
    resultats.sort(key=lambda x: x.get('rang_num', 9999))
    
    return resultats


@login_required
def exporter_notes_complet_excel(request):
    """Exporter toutes les notes par matière en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Module openpyxl non installé", status=500)
    
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode', 'OCTOBRE')
    
    if not classe_id:
        return HttpResponse("Paramètre classe_id manquant", status=400)
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = list(MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom'))
        
        # Récupérer les élèves
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom, annee_scolaire=classe.annee_scolaire, ecole=classe.ecole
        ).first()
        
        if not classe_eleve:
            return HttpResponse("Classe élèves non trouvée", status=404)
        
        eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom'))
        
        # Détecter le niveau scolaire
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_scolaire = detecter_niveau_scolaire(classe.nom)
        est_primaire = (niveau_scolaire == 'PRIMAIRE')
        est_maternelle = (niveau_scolaire == 'MATERNELLE')
        
        # Récupérer les notes (avec appréciations pour maternelle)
        resultats = get_notes_eleves_par_matiere(classe, periode, eleves, matieres, est_maternelle=est_maternelle)
        
        # Créer le workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notes Complètes"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        matiere_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        red_font = Font(color="FF0000", bold=True)
        
        # Récupérer les informations de l'école
        ecole = classe.ecole
        
        # En-tête école
        ws.merge_cells('A1:' + get_column_letter(5 + len(matieres)) + '1')
        nom_ecole = ecole.nom.upper() if ecole else "ÉCOLE"
        ws['A1'] = nom_ecole
        ws['A1'].font = Font(bold=True, size=14, color="007BFF")
        ws['A1'].alignment = center_align
        
        # Titre
        periode_label = {'ANNUEL_TRIM': 'RÉSULTAT ANNUEL (Trimestres)', 'ANNUEL_SEM': 'RÉSULTAT ANNUEL (Semestres)'}.get(periode, periode)
        ws.merge_cells('A2:' + get_column_letter(5 + len(matieres)) + '2')
        ws['A2'] = f"TABLEAU DES NOTES - {classe.nom} - {periode_label}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = center_align
        
        # Sous-titre
        ws.merge_cells('A3:' + get_column_letter(5 + len(matieres)) + '3')
        ws['A3'] = f"Année scolaire: {classe.annee_scolaire} | Effectif: {len(eleves)} élèves"
        ws['A3'].font = Font(size=10)
        ws['A3'].alignment = center_align
        
        # En-têtes du tableau (ligne 5)
        row = 5
        headers = ['N°', 'Matricule', 'Prénom', 'Nom']
        
        # Ajouter les matières comme en-têtes
        for matiere in matieres:
            coef = matiere.coefficient if matiere.coefficient else 1
            headers.append(f"{matiere.nom}\n(Coef: {coef})")
        
        headers.extend(['Moyenne', 'Rang'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            if col <= 4 or col > 4 + len(matieres):
                cell.fill = header_fill
            else:
                cell.fill = matiere_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Données des élèves
        for idx, r in enumerate(resultats, 1):
            row += 1
            eleve = r['eleve']
            
            # Colonnes fixes
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=1).alignment = center_align
            
            ws.cell(row=row, column=2, value=eleve.matricule or '').border = thin_border
            ws.cell(row=row, column=2).alignment = center_align
            
            ws.cell(row=row, column=3, value=eleve.prenom or '').border = thin_border
            ws.cell(row=row, column=3).alignment = left_align
            
            ws.cell(row=row, column=4, value=eleve.nom or '').border = thin_border
            ws.cell(row=row, column=4).alignment = left_align
            
            # Notes par matière
            col = 5
            for matiere in matieres:
                note_info = r['notes_par_matiere'].get(matiere.id, {})
                note = note_info.get('note')
                absent = note_info.get('absent', False)
                appreciation = note_info.get('appreciation')  # Pour maternelle
                
                if absent:
                    cell_value = 'ABS'
                elif est_maternelle and appreciation:
                    # Afficher l'appréciation pour la maternelle
                    cell_value = appreciation
                elif note is not None:
                    cell_value = f"{note:.2f}"
                else:
                    cell_value = '-'
                
                cell = ws.cell(row=row, column=col, value=cell_value)
                cell.border = thin_border
                cell.alignment = center_align
                
                # Colorer en rouge si note < 10 (ou < 5 pour primaire) - pas pour maternelle
                if not est_maternelle:
                    seuil = 5 if est_primaire else 10
                    if note is not None and note < seuil:
                        cell.font = red_font
                
                col += 1
            
            # Moyenne générale
            moy = r.get('moyenne_generale')
            if est_maternelle:
                moy_cell = ws.cell(row=row, column=col, value=f"{moy:.1f}%" if moy else '-')
            else:
                moy_cell = ws.cell(row=row, column=col, value=f"{moy:.2f}" if moy else '-')
            moy_cell.border = thin_border
            moy_cell.alignment = center_align
            seuil_moy = 5 if est_primaire else 10
            if moy is not None and moy < seuil_moy:
                moy_cell.font = red_font
            
            # Rang
            rang_cell = ws.cell(row=row, column=col + 1, value=r.get('rang', '-'))
            rang_cell.border = thin_border
            rang_cell.alignment = center_align
        
        # Statistiques par matière
        row += 2
        ws.cell(row=row, column=1, value="STATISTIQUES").font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=1, value="Moyenne").font = Font(bold=True)
        col = 5
        for matiere in matieres:
            notes_matiere = [r['notes_par_matiere'].get(matiere.id, {}).get('note') 
                           for r in resultats 
                           if r['notes_par_matiere'].get(matiere.id, {}).get('note') is not None]
            if notes_matiere:
                moy_mat = sum(notes_matiere) / len(notes_matiere)
                ws.cell(row=row, column=col, value=f"{moy_mat:.2f}")
            else:
                ws.cell(row=row, column=col, value='-')
            col += 1
        
        # Moyenne générale de la classe
        moyennes_gen = [r['moyenne_generale'] for r in resultats if r['moyenne_generale']]
        if moyennes_gen:
            ws.cell(row=row, column=col, value=f"{sum(moyennes_gen)/len(moyennes_gen):.2f}")
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        for i, matiere in enumerate(matieres, 5):
            ws.column_dimensions[get_column_letter(i)].width = 12
        ws.column_dimensions[get_column_letter(5 + len(matieres))].width = 10
        ws.column_dimensions[get_column_letter(6 + len(matieres))].width = 10
        
        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        nom_clean = re.sub(r'[^\w\s-]', '', classe.nom).replace(' ', '_')
        filename = f"notes_completes_{nom_clean}_{periode}.xlsx"
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.exception("Erreur export Excel")
        return HttpResponse("Une erreur est survenue lors de l'export Excel.", status=500)


@login_required
def exporter_notes_complet_pdf(request):
    """Exporter toutes les notes par matière en PDF"""
    from reportlab.lib.pagesizes import A3, A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode', 'OCTOBRE')
    
    if not classe_id:
        return HttpResponse("Paramètre classe_id manquant", status=400)
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = list(MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom'))
        
        # Récupérer les élèves
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom, annee_scolaire=classe.annee_scolaire, ecole=classe.ecole
        ).first()
        
        if not classe_eleve:
            return HttpResponse("Classe élèves non trouvée", status=404)
        
        eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom'))
        
        # Détecter le niveau scolaire
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_scolaire = detecter_niveau_scolaire(classe.nom)
        est_primaire = (niveau_scolaire == 'PRIMAIRE')
        est_maternelle = (niveau_scolaire == 'MATERNELLE')
        
        # Récupérer les notes (avec appréciations pour maternelle)
        resultats = get_notes_eleves_par_matiere(classe, periode, eleves, matieres, est_maternelle=est_maternelle)
        
        nb_matieres = len(matieres)
        page_size = landscape(A3 if nb_matieres >= 8 else A4)

        # Créer le PDF en paysage
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=page_size,
            topMargin=0.5*cm, 
            bottomMargin=0.5*cm,
            leftMargin=0.5*cm,
            rightMargin=0.5*cm
        )
        elements = []
        styles = getSampleStyleSheet()
        
        # Récupérer les informations de l'école
        ecole = classe.ecole
        
        # Styles personnalisés
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#007bff'),
            alignment=TA_CENTER,
            spaceAfter=4
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=8
        )
        
        # En-tête
        nom_ecole = ecole.nom.upper() if ecole else "ÉCOLE"
        periode_label = {'ANNUEL_TRIM': 'RÉSULTAT ANNUEL (Trimestres)', 'ANNUEL_SEM': 'RÉSULTAT ANNUEL (Semestres)'}.get(periode, periode)
        elements.append(Paragraph(f"<b>{nom_ecole}</b>", title_style))
        elements.append(Paragraph(
            f"TABLEAU DES NOTES - {classe.nom} - {periode_label} | Année: {classe.annee_scolaire}",
            subtitle_style
        ))
        
        # Calculer les largeurs de colonnes dynamiquement
        page_width = page_size[0] - 1*cm  # Largeur disponible
        
        # Colonnes fixes: N°, Matricule, Élève, Moyenne, Rang
        fixed_cols_width = 1.0*cm + 2.0*cm + 7.0*cm + 1.5*cm + 1.5*cm
        remaining_width = page_width - fixed_cols_width
        matiere_col_width = remaining_width / nb_matieres if nb_matieres > 0 else 2*cm
        
        # Limiter la largeur des colonnes matières
        matiere_col_width = min(matiere_col_width, 2.8*cm)
        matiere_col_width = max(matiere_col_width, 1.4*cm)
        
        header_cell_style = ParagraphStyle(
            'TableHeaderCell',
            parent=styles['Normal'],
            fontSize=7.2,
            leading=7.8,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
            wordWrap='CJK'
        )

        # Construire les en-têtes
        headers = ['N°', 'Matricule', 'Élève']
        for matiere in matieres:
            headers.append(Paragraph(matiere.nom, header_cell_style))
        headers.extend(['Moy.', 'Rang'])
        
        # Construire les données
        name_style = ParagraphStyle(
            'StudentName',
            parent=styles['Normal'],
            fontSize=7.2,
            leading=8.0,
            alignment=TA_LEFT,
            splitLongWords=0
        )

        data = [headers]
        
        for idx, r in enumerate(resultats, 1):
            eleve = r['eleve']
            nom_complet = f"{eleve.prenom or ''} {eleve.nom or ''}".strip()
            row = [
                str(idx),
                eleve.matricule or '',
                nom_complet if len(nom_complet) <= 18 else Paragraph(nom_complet, name_style)
            ]
            
            # Notes par matière
            for matiere in matieres:
                note_info = r['notes_par_matiere'].get(matiere.id, {})
                note = note_info.get('note')
                absent = note_info.get('absent', False)
                appreciation = note_info.get('appreciation')  # Pour maternelle
                
                if absent:
                    row.append('ABS')
                elif est_maternelle and appreciation:
                    # Afficher l'appréciation pour la maternelle
                    row.append(appreciation)
                elif note is not None:
                    row.append(f"{note:.1f}")
                else:
                    row.append('-')
            
            # Moyenne et rang
            moy = r.get('moyenne_generale')
            if est_maternelle:
                # Pour maternelle, afficher en pourcentage
                row.append(f"{moy:.1f}%" if moy else '-')
            else:
                row.append(f"{moy:.2f}" if moy else '-')
            row.append(r.get('rang', '-'))
            
            data.append(row)
        
        # Construire les largeurs de colonnes
        col_widths = [1.0*cm, 2.0*cm, 7.0*cm]
        col_widths.extend([matiere_col_width] * nb_matieres)
        col_widths.extend([1.5*cm, 1.5*cm])
        
        # Créer le tableau
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Styles du tableau
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.4),
            ('FONTSIZE', (0, 1), (-1, -1), 7.2),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Nom complet aligné à gauche
        ]
        
        # Colorer les notes insuffisantes en rouge
        seuil = 5 if est_primaire else 10
        for row_idx, r in enumerate(resultats, 1):
            # Notes par matière
            for mat_idx, matiere in enumerate(matieres):
                note_info = r['notes_par_matiere'].get(matiere.id, {})
                note = note_info.get('note')
                if note is not None and note < seuil:
                    col_idx = 3 + mat_idx
                    style_commands.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.red))
            
            # Moyenne générale
            moy = r.get('moyenne_generale')
            if moy is not None and moy < seuil:
                moy_col = 3 + len(matieres)
                style_commands.append(('TEXTCOLOR', (moy_col, row_idx), (moy_col, row_idx), colors.red))
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Statistiques
        elements.append(Spacer(1, 0.5*cm))
        moyennes_gen = [r['moyenne_generale'] for r in resultats if r['moyenne_generale']]
        if moyennes_gen:
            moy_classe = sum(moyennes_gen) / len(moyennes_gen)
            stats_text = f"Effectif: {len(resultats)} | Moyenne classe: {moy_classe:.2f} | "
            stats_text += f"Max: {max(moyennes_gen):.2f} | Min: {min(moyennes_gen):.2f}"
            elements.append(Paragraph(stats_text, styles['Normal']))
        
        # Construire le PDF
        doc.build(elements)
        buffer.seek(0)
        
        nom_clean = re.sub(r'[^\w\s-]', '', classe.nom).replace(' ', '_')
        filename = f"notes_completes_{nom_clean}_{periode}.pdf"
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.exception("Erreur export PDF")
        return HttpResponse("Une erreur est survenue lors de l'export PDF.", status=500)
