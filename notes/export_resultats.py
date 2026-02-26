"""
Export des résultats de classe en PDF et Excel
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import io
import re

from .models import ClasseNote, MatiereNote
from eleves.models import Eleve, Classe as ClasseEleve


@login_required
def exporter_resultats_pdf(request):
    """Exporter les résultats d'une classe en PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode', '')
    
    if not classe_id:
        return HttpResponse("Paramètre classe_id manquant", status=400)
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom')
        
        # Récupérer les élèves
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom, annee_scolaire=classe.annee_scolaire, ecole=classe.ecole
        ).first()
        
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom') if classe_eleve else []
        
        # IMPORTANT: Récupérer les moyennes et rangs depuis la source centralisée
        # Ne PAS recalculer pour garantir la cohérence avec consulter_notes et bulletins
        from .utils_rangs import calculer_rangs_classe_periode
        
        # Si pas de période spécifiée, utiliser OCTOBRE par défaut
        periode_calcul = periode if periode else 'OCTOBRE'
        
        # Récupérer les rangs et moyennes depuis la source centralisée
        rangs_dict = calculer_rangs_classe_periode(classe, periode_calcul, use_cache=False)
        
        # Construire la liste des résultats à partir des données centralisées
        resultats = []
        for eleve in eleves:
            rang_info = rangs_dict.get(eleve.id)
            if rang_info:
                resultats.append({
                    'eleve': eleve,
                    'moyenne': float(rang_info['moyenne']),
                    'rang': rang_info['rang'],
                    'rang_num': rang_info['rang_num']
                })
            else:
                resultats.append({
                    'eleve': eleve,
                    'moyenne': 0.0,
                    'rang': '-',
                    'rang_num': 9999
                })
        
        # Trier par rang numérique (déjà calculé par utils_rangs)
        resultats.sort(key=lambda x: x['rang_num'])
        
        # Mettre à jour la période pour l'affichage
        periode = periode_calcul
        
        # Créer le PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Récupérer les informations de l'école
        ecole = classe.ecole
        
        # Styles personnalisés
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=2
        )
        
        school_name_style = ParagraphStyle(
            'SchoolName',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#007bff'),
            alignment=TA_CENTER,
            spaceAfter=4
        )
        
        title_style = ParagraphStyle(
            'Title', 
            parent=styles['Heading1'], 
            fontSize=16, 
            textColor=colors.HexColor('#007bff'), 
            spaceAfter=8, 
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        # Détecter le niveau scolaire
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_scolaire = detecter_niveau_scolaire(classe.nom)
        est_maternelle = (niveau_scolaire == 'MATERNELLE')
        est_primaire = (niveau_scolaire == 'PRIMAIRE')
        
        # En-tête simplifié - Nom de l'école uniquement
        nom_ecole = ecole.nom if ecole else "École"
        elements.append(Paragraph(f"<b>{nom_ecole.upper()}</b>", school_name_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Titre du document
        elements.append(Paragraph(f"RÉSULTATS DE LA CLASSE - {classe.nom}", title_style))
        elements.append(Paragraph(f"Période: {periode or 'Année complète'} | Année scolaire: {classe.annee_scolaire}", subtitle_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Tableau des résultats (Prénom avant Nom)
        # En-tête adapté selon le niveau
        if est_maternelle:
            data = [['Rang', 'Matricule', 'Prénom', 'Nom', 'Acquisition (%)', 'Appréciation']]
        elif est_primaire:
            data = [['Rang', 'Matricule', 'Prénom', 'Nom', 'Moyenne /10', 'Mention']]
        else:
            data = [['Rang', 'Matricule', 'Prénom', 'Nom', 'Moyenne /20', 'Mention']]
        lignes_non_admis = []  # Pour stocker les indices des lignes avec moyenne < 10
        
        for idx, r in enumerate(resultats):
            moy = r['moyenne']
            
            # Mentions adaptées selon le type de classe
            if est_maternelle:
                # Pour la maternelle : mentions basées sur le taux d'acquisition
                if moy >= 90:
                    mention = 'Excellent'
                elif moy >= 75:
                    mention = 'Très Bien'
                elif moy >= 60:
                    mention = 'Bien'
                elif moy >= 50:
                    mention = 'Assez Bien'
                else:
                    mention = 'À encourager'
                    lignes_non_admis.append(idx + 1)
            elif est_primaire:
                # Pour le primaire : mentions sur 10
                if moy >= 9:
                    mention = 'Excellent'
                elif moy >= 8:
                    mention = 'Très Bien'
                elif moy >= 7:
                    mention = 'Bien'
                elif moy >= 6:
                    mention = 'Assez Bien'
                elif moy >= 5:
                    mention = 'Passable'
                elif moy >= 4:
                    mention = 'Insuffisant'
                    lignes_non_admis.append(idx + 1)
                elif moy >= 3:
                    mention = 'Faible'
                    lignes_non_admis.append(idx + 1)
                else:
                    mention = 'Très faible'
                    lignes_non_admis.append(idx + 1)
            else:
                # Pour le secondaire : mentions sur 20
                if moy >= 18:
                    mention = 'Excellent'
                elif moy >= 16:
                    mention = 'Très Bien'
                elif moy >= 14:
                    mention = 'Bien'
                elif moy >= 12:
                    mention = 'Assez Bien'
                elif moy >= 10:
                    mention = 'Passable'
                elif moy >= 8:
                    mention = 'Insuffisant'
                    lignes_non_admis.append(idx + 1)
                elif moy >= 6:
                    mention = 'Faible'
                    lignes_non_admis.append(idx + 1)
                else:
                    mention = 'Très faible'
                    lignes_non_admis.append(idx + 1)  # +1 car la ligne 0 est l'en-tête
            
            # Format de la moyenne/acquisition
            if est_maternelle:
                moy_display = f"{moy:.1f}%" if moy else '-'
            else:
                moy_display = f"{moy:.2f}" if moy else '-'
            
            data.append([
                r['rang'], 
                r['eleve'].matricule or '', 
                r['eleve'].prenom or '',  # Prénom avant Nom
                r['eleve'].nom or '', 
                moy_display, 
                mention if moy else '-'
            ])
        
        # Style du tableau
        table = Table(data, colWidths=[2*cm, 3*cm, 5*cm, 5*cm, 3*cm, 3*cm])
        
        # Styles de base
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('ALIGN', (2, 1), (3, -1), 'LEFT'),
        ]
        
        # Ajouter le style rouge pour les élèves non admis (moyenne < 10)
        for ligne in lignes_non_admis:
            # Colonne 4 = Moyenne, Colonne 5 = Mention
            style_commands.append(('TEXTCOLOR', (4, ligne), (5, ligne), colors.red))
            style_commands.append(('FONTNAME', (4, ligne), (5, ligne), 'Helvetica-Bold'))
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Statistiques détaillées
        elements.append(Spacer(1, 0.5*cm))
        
        moyennes_valides = [r['moyenne'] for r in resultats if r['moyenne']]
        
        if moyennes_valides:
            note_sur = 10 if est_primaire else (100 if est_maternelle else 20)
            seuil_admis = 5 if est_primaire else (50 if est_maternelle else 10)
            suffixe = '%' if est_maternelle else f'/{note_sur}'
            fmt = '.1f' if est_maternelle else '.2f'
            
            moy_classe = sum(moyennes_valides) / len(moyennes_valides)
            
            # Statistiques par sexe
            filles = [r for r in resultats if r['eleve'].sexe == 'F' and r['moyenne']]
            garcons = [r for r in resultats if r['eleve'].sexe == 'M' and r['moyenne']]
            nb_filles = len(filles)
            nb_garcons = len(garcons)
            moy_filles = sum(r['moyenne'] for r in filles) / nb_filles if nb_filles else 0
            moy_garcons = sum(r['moyenne'] for r in garcons) / nb_garcons if nb_garcons else 0
            
            # Admis / Non admis
            nb_admis = len([r for r in resultats if r['moyenne'] and r['moyenne'] >= seuil_admis])
            nb_non_admis = len([r for r in resultats if r['moyenne'] and r['moyenne'] < seuil_admis])
            taux_reussite = (nb_admis / len(moyennes_valides) * 100) if moyennes_valides else 0
            
            # Titre section statistiques
            stats_title_style = ParagraphStyle(
                'StatsTitle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#007bff'),
                spaceAfter=6,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("STATISTIQUES DE LA CLASSE", stats_title_style))
            elements.append(Spacer(1, 0.2*cm))
            
            # Tableau statistiques
            stats_data = [
                ['Statistique', 'Valeur'],
                ['Ont composé', str(len(moyennes_valides))],
                ['Filles', str(nb_filles)],
                ['Garçons', str(nb_garcons)],
                ['Moyenne de classe', f"{moy_classe:{fmt}}{suffixe}"],
                ['Moyenne des filles', f"{moy_filles:{fmt}}{suffixe}" if nb_filles else '-'],
                ['Moyenne des garçons', f"{moy_garcons:{fmt}}{suffixe}" if nb_garcons else '-'],
                ['Meilleure moyenne', f"{max(moyennes_valides):{fmt}}{suffixe}"],
                ['Plus faible moyenne', f"{min(moyennes_valides):{fmt}}{suffixe}"],
                ['Admis', f"{nb_admis} ({taux_reussite:.1f}%)"],
                ['Non admis', f"{nb_non_admis} ({100 - taux_reussite:.1f}%)"],
            ]
            
            stats_table = Table(stats_data, colWidths=[8*cm, 6*cm])
            stats_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                # Non admis en rouge
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.red),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ])
            stats_table.setStyle(stats_style)
            elements.append(stats_table)
        
        # Construire le PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Nom de fichier nettoyé
        nom_clean = re.sub(r'[^\w\s-]', '', classe.nom).replace(' ', '_')
        filename = f"resultats_{nom_clean}_{periode or 'annee'}.pdf"
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur export PDF: {str(e)}", status=500)


@login_required  
def exporter_resultats_excel(request):
    """Exporter les résultats d'une classe en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("Module openpyxl non installé", status=500)
    
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode', '')
    
    if not classe_id:
        return HttpResponse("Paramètre classe_id manquant", status=400)
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom')
        
        # Récupérer les élèves
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom, annee_scolaire=classe.annee_scolaire, ecole=classe.ecole
        ).first()
        
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom') if classe_eleve else []
        
        # IMPORTANT: Récupérer les moyennes et rangs depuis la source centralisée
        # Ne PAS recalculer pour garantir la cohérence avec consulter_notes et bulletins
        from .utils_rangs import calculer_rangs_classe_periode
        
        # Si pas de période spécifiée, utiliser OCTOBRE par défaut
        periode_calcul = periode if periode else 'OCTOBRE'
        
        # Récupérer les rangs et moyennes depuis la source centralisée
        rangs_dict = calculer_rangs_classe_periode(classe, periode_calcul, use_cache=False)
        
        # Construire la liste des résultats à partir des données centralisées
        resultats = []
        for eleve in eleves:
            rang_info = rangs_dict.get(eleve.id)
            if rang_info:
                resultats.append({
                    'eleve': eleve,
                    'moyenne': float(rang_info['moyenne']),
                    'rang': rang_info['rang'],
                    'rang_num': rang_info['rang_num']
                })
            else:
                resultats.append({
                    'eleve': eleve,
                    'moyenne': 0.0,
                    'rang': '-',
                    'rang_num': 9999
                })
        
        # Trier par rang numérique (déjà calculé par utils_rangs)
        resultats.sort(key=lambda x: x['rang_num'])
        
        # Mettre à jour la période pour l'affichage
        periode = periode_calcul
        
        # Créer le workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats"
        
        # Récupérer les informations de l'école
        ecole = classe.ecole
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # En-tête avec informations de l'école
        ws.merge_cells('A1:F1')
        ws['A1'] = "RÉPUBLIQUE DE GUINÉE"
        ws['A1'].font = Font(bold=True, size=10)
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:F2')
        ws['A2'] = "Travail - Justice - Solidarité"
        ws['A2'].font = Font(italic=True, size=9)
        ws['A2'].alignment = center_align
        
        ws.merge_cells('A3:F3')
        nom_ecole = ecole.nom.upper() if ecole else "ÉCOLE"
        ws['A3'] = nom_ecole
        ws['A3'].font = Font(bold=True, size=14, color="007BFF")
        ws['A3'].alignment = center_align
        
        # Adresse et contact
        ws.merge_cells('A4:F4')
        if ecole:
            adresse_parts = []
            if ecole.adresse:
                adresse_parts.append(ecole.adresse)
            if ecole.telephone:
                adresse_parts.append(f"Tél: {ecole.tous_telephones}")
            if ecole.email:
                adresse_parts.append(f"Email: {ecole.email}")
            ws['A4'] = " | ".join(adresse_parts) if adresse_parts else ""
        ws['A4'].font = Font(size=9)
        ws['A4'].alignment = center_align
        
        # Ligne vide
        ws.merge_cells('A5:F5')
        ws['A5'] = "─" * 50
        ws['A5'].alignment = center_align
        
        # Titre du document
        ws.merge_cells('A6:F6')
        ws['A6'] = f"RÉSULTATS DE LA CLASSE - {classe.nom}"
        ws['A6'].font = Font(bold=True, size=14)
        ws['A6'].alignment = center_align
        
        ws.merge_cells('A7:F7')
        ws['A7'] = f"Période: {periode or 'Année complète'} | Année scolaire: {classe.annee_scolaire}"
        ws['A7'].alignment = center_align
        
        # Détecter le niveau scolaire
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_scolaire = detecter_niveau_scolaire(classe.nom)
        est_maternelle = (niveau_scolaire == 'MATERNELLE')
        est_primaire = (niveau_scolaire == 'PRIMAIRE')
        
        # En-têtes du tableau (décalés à la ligne 9) - Prénom avant Nom
        if est_maternelle:
            headers = ['Rang', 'Matricule', 'Prénom', 'Nom', 'Acquisition (%)', 'Appréciation']
        elif est_primaire:
            headers = ['Rang', 'Matricule', 'Prénom', 'Nom', 'Moyenne /10', 'Mention']
        else:
            headers = ['Rang', 'Matricule', 'Prénom', 'Nom', 'Moyenne /20', 'Mention']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Données (décalées à partir de la ligne 10)
        for row_idx, r in enumerate(resultats, 10):
            moy = r['moyenne']
            
            # Mentions adaptées selon le type de classe
            if est_maternelle:
                if moy >= 90:
                    mention = 'Excellent'
                elif moy >= 75:
                    mention = 'Très Bien'
                elif moy >= 60:
                    mention = 'Bien'
                elif moy >= 50:
                    mention = 'Assez Bien'
                else:
                    mention = 'À encourager'
                moy_display = f"{moy:.1f}%" if moy else '-'
            elif est_primaire:
                # Pour le primaire : mentions sur 10
                if moy >= 9:
                    mention = 'Excellent'
                elif moy >= 8:
                    mention = 'Très Bien'
                elif moy >= 7:
                    mention = 'Bien'
                elif moy >= 6:
                    mention = 'Assez Bien'
                elif moy >= 5:
                    mention = 'Passable'
                elif moy >= 4:
                    mention = 'Insuffisant'
                elif moy >= 3:
                    mention = 'Faible'
                else:
                    mention = 'Très faible'
                moy_display = f"{moy:.2f}" if moy else '-'
            else:
                # Pour le secondaire : mentions sur 20
                if moy >= 18:
                    mention = 'Excellent'
                elif moy >= 16:
                    mention = 'Très Bien'
                elif moy >= 14:
                    mention = 'Bien'
                elif moy >= 12:
                    mention = 'Assez Bien'
                elif moy >= 10:
                    mention = 'Passable'
                elif moy >= 8:
                    mention = 'Insuffisant'
                elif moy >= 6:
                    mention = 'Faible'
                else:
                    mention = 'Très faible'
                moy_display = f"{moy:.2f}" if moy else '-'
            
            data_row = [
                r['rang'], 
                r['eleve'].matricule or '', 
                r['eleve'].prenom or '',  # Prénom avant Nom
                r['eleve'].nom or '', 
                moy_display, 
                mention if moy else '-'
            ]
            
            # Déterminer si non admis
            seuil_admis = 5 if est_primaire else (50 if est_maternelle else 10)
            est_non_admis = moy < seuil_admis if moy else False
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [1, 5, 6]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                # Mention et moyenne en rouge pour les non admis
                if est_non_admis and col in [5, 6]:
                    cell.font = Font(bold=True, color="FF0000")
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Statistiques détaillées
        last_row = len(resultats) + 11
        moyennes_valides = [r['moyenne'] for r in resultats if r['moyenne']]
        if moyennes_valides:
            note_sur = 10 if est_primaire else (100 if est_maternelle else 20)
            seuil_admis = 5 if est_primaire else (50 if est_maternelle else 10)
            suffixe = '%' if est_maternelle else f'/{note_sur}'
            fmt = '.1f' if est_maternelle else '.2f'
            
            moy_classe = sum(moyennes_valides) / len(moyennes_valides)
            
            # Statistiques par sexe
            filles = [r for r in resultats if r['eleve'].sexe == 'F' and r['moyenne']]
            garcons = [r for r in resultats if r['eleve'].sexe == 'M' and r['moyenne']]
            nb_filles = len(filles)
            nb_garcons = len(garcons)
            moy_filles = sum(r['moyenne'] for r in filles) / nb_filles if nb_filles else 0
            moy_garcons = sum(r['moyenne'] for r in garcons) / nb_garcons if nb_garcons else 0
            
            nb_admis = len([r for r in resultats if r['moyenne'] and r['moyenne'] >= seuil_admis])
            nb_non_admis = len([r for r in resultats if r['moyenne'] and r['moyenne'] < seuil_admis])
            taux_reussite = (nb_admis / len(moyennes_valides) * 100) if moyennes_valides else 0
            
            # Titre statistiques
            stats_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
            ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)
            cell_title = ws.cell(row=last_row, column=1, value="STATISTIQUES DE LA CLASSE")
            cell_title.font = Font(bold=True, color="FFFFFF", size=11)
            cell_title.fill = stats_fill
            cell_title.alignment = center_align
            ws.cell(row=last_row, column=2).fill = stats_fill
            
            stats_rows = [
                ('Ont composé', str(len(moyennes_valides))),
                ('Filles', str(nb_filles)),
                ('Garçons', str(nb_garcons)),
                ('Moyenne de classe', f"{moy_classe:{fmt}}{suffixe}"),
                ('Moyenne des filles', f"{moy_filles:{fmt}}{suffixe}" if nb_filles else '-'),
                ('Moyenne des garçons', f"{moy_garcons:{fmt}}{suffixe}" if nb_garcons else '-'),
                ('Meilleure moyenne', f"{max(moyennes_valides):{fmt}}{suffixe}"),
                ('Plus faible moyenne', f"{min(moyennes_valides):{fmt}}{suffixe}"),
                ('Admis', f"{nb_admis} ({taux_reussite:.1f}%)"),
                ('Non admis', f"{nb_non_admis} ({100 - taux_reussite:.1f}%)"),
            ]
            
            for i, (label, value) in enumerate(stats_rows, 1):
                row_num = last_row + i
                c1 = ws.cell(row=row_num, column=1, value=label)
                c1.font = Font(bold=True)
                c1.border = thin_border
                c1.alignment = left_align
                c2 = ws.cell(row=row_num, column=2, value=value)
                c2.border = thin_border
                c2.alignment = center_align
                # Non admis en rouge
                if label == 'Non admis':
                    c1.font = Font(bold=True, color="FF0000")
                    c2.font = Font(bold=True, color="FF0000")
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nom de fichier nettoyé
        nom_clean = re.sub(r'[^\w\s-]', '', classe.nom).replace(' ', '_')
        filename = f"resultats_{nom_clean}_{periode or 'annee'}.xlsx"
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur export Excel: {str(e)}", status=500)
