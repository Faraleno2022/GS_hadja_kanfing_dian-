"""
Module pour exporter les classements par classe (Excel et PDF)
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from decimal import Decimal
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import io
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics

from .models import ClasseNote, MatiereNote, NoteMensuelle, CompositionNote
from eleves.models import Eleve, Classe as ClasseEleve
from .calculs_moyennes import calculer_moyenne_generale_eleve, calculer_classement_classe, detecter_niveau_scolaire


def formater_rang(rang, sexe=None):
    """
    Formate le rang avec l'accord grammatical selon le sexe
    
    Règles:
    - 1er pour les garçons (sexe='M')
    - 1ère pour les filles (sexe='F')
    - 2ème, 3ème, 4ème, etc. pour tous les autres rangs (pas de distinction de sexe)
    
    Args:
        rang: Le numéro de rang (int ou str) ou rang déjà formaté (ex: "1er", "2ème ex")
        sexe: 'M' pour masculin, 'F' pour féminin (utilisé uniquement pour le rang 1)
    
    Returns:
        str: Le rang formaté (ex: "1er", "1ère", "2ème", "3ème", etc.)
    """
    if rang == '-' or rang is None:
        return '-'
    
    # Si le rang est déjà formaté (vérifier plus précisément)
    if isinstance(rang, str):
        # Si le rang contient "/" (ancien format), extraire le numéro
        if '/' in rang:
            import re
            # Extraire la partie numérique avant "/"
            match = re.match(r'(\d+)', rang)
            if match:
                rang_num = int(match.group(1))
            else:
                rang_num = int(rang.split('/')[0])
        # Patterns de rangs déjà formatés (sans "/")
        elif rang.endswith('er') or rang.endswith('ère') or 'ème' in rang or ' ex' in rang:
            return rang
        else:
            # Convertir en entier si c'est un nombre
            try:
                rang_num = int(rang)
            except ValueError:
                return rang
    else:
        # Si c'est déjà un nombre
        rang_num = int(rang)
    
    # Cas spécial pour le rang 1 - accord selon le sexe
    if rang_num == 1:
        if sexe == 'F':
            return "1ère"
        else:
            return "1er"
    
    # Pour tous les autres rangs (2, 3, 4, ...), on utilise "ème" sans distinction de sexe
    return f"{rang_num}ème"


@login_required
def exporter_classement_classe(request):
    """
    Exporter le classement d'une classe avec rang, nom complet, matricule et moyenne
    Format: Excel (.xlsx)
    """
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    type_note = request.GET.get('type_note', 'mensuelle')
    periode = request.GET.get('periode', '')
    
    if not classe_id:
        return HttpResponse("Classe non spécifiée", status=400)
    
    # Récupérer la classe
    classe_note = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Récupérer la classe élève correspondante avec mapping spécial (même logique que les autres vues)
    try:
        # Mapping spécial pour les classes avec noms différents
        mapping_classes = {
            61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
            59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
        }
        
        if classe_note.id in mapping_classes:
            classe_eleve = ClasseEleve.objects.filter(
                id=mapping_classes[classe_note.id]
            ).first()
        else:
            # Essai 1: Correspondance exacte
            classe_eleve = ClasseEleve.objects.filter(
                nom=classe_note.nom,
                annee_scolaire=classe_note.annee_scolaire,
                ecole=classe_note.ecole
            ).first()
            
            # Essai 2: Correspondance insensible à la casse
            if not classe_eleve:
                classe_eleve = ClasseEleve.objects.filter(
                    nom__iexact=classe_note.nom,
                    annee_scolaire=classe_note.annee_scolaire,
                    ecole=classe_note.ecole
                ).first()
            
            # Essai 3: Recherche par mots-clés (ex: "12ème Série scientifique" → "12" + "SCIENCES")
            if not classe_eleve:
                # Extraire le niveau (ex: "12")
                import re
                match = re.search(r'(\d+)', classe_note.nom)
                if match:
                    niveau_num = match.group(1)
                    
                    # Chercher d'abord avec l'école
                    classes_possibles = ClasseEleve.objects.filter(
                        nom__icontains=niveau_num,
                        annee_scolaire=classe_note.annee_scolaire,
                        ecole=classe_note.ecole
                    )
                    
                    # Si aucune classe trouvée avec l'école, chercher sans filtrer par école
                    if not classes_possibles.exists():
                        classes_possibles = ClasseEleve.objects.filter(
                            nom__icontains=niveau_num,
                            annee_scolaire=classe_note.annee_scolaire
                        )
                    
                    classe_eleve = classes_possibles.first()
                    
                    # Si plusieurs classes trouvées, essayer d'affiner avec les mots-clés
                    if classe_eleve and classes_possibles.count() > 1:
                        # Chercher des mots-clés spécifiques dans le nom de la classe
                        if 'scientifique' in classe_note.nom.lower() or 'science' in classe_note.nom.lower():
                            for c in classes_possibles:
                                if 'SCIENCE' in c.nom.upper():
                                    classe_eleve = c
                                    break
                        elif 'littéraire' in classe_note.nom.lower() or 'lettre' in classe_note.nom.lower():
                            for c in classes_possibles:
                                if 'LETTRE' in c.nom.upper():
                                    classe_eleve = c
                                    break
        
        if not classe_eleve:
            return HttpResponse(
                f"Classe élève non trouvée pour '{classe_note.nom}' ({classe_note.annee_scolaire}). "
                f"Vérifiez que les noms de classes correspondent entre le système de notes et le système d'élèves.",
                status=404
            )
        
        # Récupérer les élèves actifs
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom')
        
        if not eleves.exists():
            return HttpResponse(
                f"Aucun élève actif trouvé dans la classe '{classe_eleve.nom}'",
                status=404
            )
    except Exception as e:
        return HttpResponse(f"Erreur lors de la récupération des élèves: {str(e)}", status=500)
    
    # Préparer les données de classement
    if matiere_id:
        classement_data, titre_export = _generer_classement_matiere(
            eleves, classe_note, matiere_id, type_note, periode
        )
    else:
        classement_data, titre_export = _generer_classement_general(
            eleves, classe_note, type_note, periode
        )
    
    # Détecter le niveau scolaire pour adapter l'affichage
    niveau_scolaire = detecter_niveau_scolaire(classe_note.nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classement"
    
    # Styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = titre_export
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Date d'export
    ws.merge_cells('A2:E2')
    date_cell = ws['A2']
    date_cell.value = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    date_cell.alignment = Alignment(horizontal='center')
    date_cell.font = Font(italic=True, size=10)
    ws.row_dimensions[2].height = 20
    
    # En-têtes (adapter selon le niveau)
    moyenne_header = 'Moyenne /10' if est_primaire else 'Moyenne /20'
    headers = ['Rang', 'Matricule', 'Nom Complet', 'Sexe', moyenne_header]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[4].height = 25
    
    # Données
    for row_num, eleve_data in enumerate(classement_data, 5):
        # Rang
        rang_cell = ws.cell(row=row_num, column=1)
        rang_value = eleve_data['rang']
        sexe = eleve_data.get('sexe', 'M')  # Par défaut masculin si non spécifié
        
        # Formater le rang avec l'accord grammatical
        rang_formate = formater_rang(rang_value, sexe)
        
        if rang_value == 1:
            rang_cell.value = f"🥇 {rang_formate}"
            rang_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        elif rang_value == 2:
            rang_cell.value = f"🥈 {rang_formate}"
            rang_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        elif rang_value == 3:
            rang_cell.value = f"🥉 {rang_formate}"
            rang_cell.fill = PatternFill(start_color="FFE7D9", end_color="FFE7D9", fill_type="solid")
        else:
            rang_cell.value = rang_formate
        
        rang_cell.alignment = Alignment(horizontal='center', vertical='center')
        rang_cell.border = border
        rang_cell.font = Font(bold=True, size=11)
        
        # Matricule
        matricule_cell = ws.cell(row=row_num, column=2)
        matricule_cell.value = eleve_data['matricule']
        matricule_cell.alignment = Alignment(horizontal='center', vertical='center')
        matricule_cell.border = border
        
        # Nom complet
        nom_cell = ws.cell(row=row_num, column=3)
        nom_cell.value = eleve_data['nom_complet']
        nom_cell.alignment = Alignment(horizontal='left', vertical='center')
        nom_cell.border = border
        nom_cell.font = Font(size=11)
        
        # Sexe
        sexe_cell = ws.cell(row=row_num, column=4)
        sexe_val = eleve_data.get('sexe', '')
        sexe_cell.value = 'Fille' if sexe_val == 'F' else 'Garçon'
        sexe_cell.alignment = Alignment(horizontal='center', vertical='center')
        sexe_cell.border = border
        if sexe_val == 'F':
            sexe_cell.font = Font(color='C0008A', size=10)
        else:
            sexe_cell.font = Font(color='0055CC', size=10)

        # Moyenne
        moyenne_cell = ws.cell(row=row_num, column=5)
        if eleve_data['moyenne'] is not None:
            moyenne_cell.value = eleve_data['moyenne']
            moyenne_cell.number_format = '0.00'
            
            # Coloration selon la moyenne
            moyenne_val = eleve_data['moyenne']
            if moyenne_val >= 16:
                moyenne_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif moyenne_val >= 14:
                moyenne_cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
            elif moyenne_val >= 10:
                moyenne_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            else:
                moyenne_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        else:
            # Afficher clairement pourquoi pas de moyenne
            if eleve_data.get('absent'):
                moyenne_cell.value = "Absent"
            elif eleve_data.get('pas_de_notes'):
                moyenne_cell.value = "Pas de notes"
            else:
                moyenne_cell.value = "Non saisi"
            moyenne_cell.font = Font(italic=True, color="999999")
        
        moyenne_cell.alignment = Alignment(horizontal='center', vertical='center')
        moyenne_cell.border = border
        if eleve_data['moyenne'] is not None:
            moyenne_cell.font = Font(bold=True, size=11)
        
        ws.row_dimensions[row_num].height = 20
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    
    # Statistiques enrichies
    seuil_admission = 5 if est_primaire else 10
    eleves_avec_moyenne = [e for e in classement_data if e.get('moyenne') is not None]
    eleves_sans_notes = [e for e in classement_data if e.get('moyenne') is None]
    nb_total = len(classement_data)
    nb_filles = sum(1 for e in classement_data if e.get('sexe') == 'F')
    nb_garcons = nb_total - nb_filles
    nb_admis = sum(1 for e in eleves_avec_moyenne if e['moyenne'] >= seuil_admission)
    nb_redoubles = sum(1 for e in eleves_avec_moyenne if e['moyenne'] < seuil_admission)
    taux_reussite = round((nb_admis / len(eleves_avec_moyenne)) * 100, 1) if eleves_avec_moyenne else 0

    if eleves_avec_moyenne or eleves_sans_notes:
        stats_row = row_num + 2

        ws.merge_cells(f'A{stats_row}:E{stats_row}')
        stats_title = ws[f'A{stats_row}']
        stats_title.value = "STATISTIQUES DE LA CLASSE"
        stats_title.font = Font(bold=True, size=12, color='FFFFFF')
        stats_title.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        stats_title.alignment = Alignment(horizontal='center')

        stats_data = [
            ('Effectif total :', nb_total),
            ('Nombre de filles :', nb_filles),
            ('Nombre de garçons :', nb_garcons),
            ('Élèves évalués :', len(eleves_avec_moyenne)),
            ('Élèves non évalués :', len(eleves_sans_notes)),
            ('Admis (moy ≥ {}) :'.format(seuil_admission), nb_admis),
            ('Redoublants (moy < {}) :'.format(seuil_admission), nb_redoubles),
            ('Taux de réussite :', f"{taux_reussite} %"),
        ]

        if eleves_avec_moyenne:
            moyennes = [e['moyenne'] for e in eleves_avec_moyenne]
            moyenne_classe = sum(moyennes) / len(moyennes)
            note_max = max(moyennes)
            note_min = min(moyennes)
            base = '/10' if est_primaire else '/20'
            stats_data.extend([
                ('Moyenne de classe :', f"{moyenne_classe:.2f}{base}"),
                ('Note maximale :', f"{note_max:.2f}{base}"),
                ('Note minimale :', f"{note_min:.2f}{base}"),
            ])

        for i, (label, value) in enumerate(stats_data, stats_row + 1):
            label_cell = ws.cell(row=i, column=2)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='right')

            value_cell = ws.cell(row=i, column=3)
            value_cell.value = value
            value_cell.alignment = Alignment(horizontal='left')

    # Avertissement si des élèves n'ont pas de notes
    if eleves_sans_notes:
        warning_row = ws.max_row + 2
        ws.merge_cells(f'A{warning_row}:E{warning_row}')
        warning_cell = ws[f'A{warning_row}']
        warning_cell.value = f" ATTENTION: {len(eleves_sans_notes)} élève(s) n'ont pas de notes pour cette période"
        warning_cell.font = Font(color="FF0000", bold=True)
        warning_cell.alignment = Alignment(horizontal='center')
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"Classement_{classe_note.nom}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def _generer_classement_matiere(eleves, classe_note, matiere_id, type_note, periode):
    """Générer le classement pour une matière spécifique"""
    matiere = get_object_or_404(MatiereNote, pk=matiere_id)
    classement_data = []
    
    for eleve in eleves:
        note_value = None
        absent = False
        pas_de_notes = True
        
        # Essayer d'abord avec NoteEleve et Evaluation (système moderne)
        from .models import Evaluation, NoteEleve
        
        # Pour les périodes mensuelles, chercher les évaluations de cette période
        if periode and periode in ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']:
            # Système mensuel : chercher une évaluation pour cette période
            evaluations = Evaluation.objects.filter(
                matiere=matiere,
                periode=periode
            )
        else:
            # Autres périodes : chercher toutes les évaluations
            evaluations = Evaluation.objects.filter(matiere=matiere)
            if periode:
                evaluations = evaluations.filter(periode=periode)
        
        if evaluations.exists():
            # Calculer la moyenne pondérée des évaluations pour cette matière
            total_pondere = Decimal('0')
            total_coef_eval = Decimal('0')
            notes_trouvees = False
            toutes_absentes = True
            
            for evaluation in evaluations:
                try:
                    note_obj = NoteEleve.objects.get(eleve=eleve, evaluation=evaluation)
                    if note_obj.absent:
                        toutes_absentes = True
                    if note_obj.note is not None and not note_obj.absent:
                        coef_eval = Decimal(str(evaluation.coefficient or 1))
                        total_pondere += Decimal(str(note_obj.note)) * coef_eval
                        total_coef_eval += coef_eval
                        notes_trouvees = True
                        toutes_absentes = False
                except NoteEleve.DoesNotExist:
                    pass
            
            if total_coef_eval > 0:
                note_value = float(total_pondere / total_coef_eval)
                absent = toutes_absentes and not notes_trouvees
                pas_de_notes = False
        
        # Si pas de notes trouvées, garder note_value = None et pas_de_notes = True
        # Ne plus essayer NoteMensuelle/CompositionNote comme fallback
        
        classement_data.append({
            'matricule': eleve.matricule or 'N/A',
            'nom_complet': f"{eleve.prenom} {eleve.nom}",
            'moyenne': note_value,
            'absent': absent,
            'pas_de_notes': pas_de_notes,
            'sexe': eleve.sexe  # Ajouter le sexe pour l'accord grammatical
        })
    
    # Trier et attribuer les rangs
    classement_data = _calculer_rangs(classement_data)
    
    titre_export = f"Classement - {classe_note.nom} - {matiere.nom}"
    if periode:
        titre_export += f" - {periode}"
    
    return classement_data, titre_export


def _generer_classement_general(eleves, classe_note, type_note, periode):
    """Générer le classement général en utilisant le module centralisé"""
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True)
    
    # UTILISER LE MODULE CENTRALISÉ pour garantir la cohérence
    # Déterminer le system_type selon la période
    if periode in ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']:
        system_type = 'mensuel'
    elif periode and 'TRIMESTRE' in periode:
        system_type = 'trimestre'
    elif periode and 'SEMESTRE' in periode:
        system_type = 'semestre'
    else:
        system_type = 'annuel'
    
    # Calculer avec le module centralisé (SOURCE UNIQUE)
    classement_complet = calculer_classement_classe(eleves, matieres, periode, system_type)
    
    # Extraire les données du classement centralisé
    classement_data = []
    
    for eleve in eleves:
        # Récupérer les détails depuis le calcul centralisé
        details_eleve = classement_complet.get('details_par_eleve', {}).get(eleve.id)
        
        if details_eleve:
            # Utiliser les données du module centralisé (SOURCE UNIQUE)
            moyenne_generale = details_eleve.get('moyenne_generale')
            classement_data.append({
                'eleve_id': eleve.id,
                'prenom': eleve.prenom,
                'nom': eleve.nom,
                'matricule': eleve.matricule or 'N/A',
                'nom_complet': f"{eleve.prenom} {eleve.nom}",
                'moyenne': round(moyenne_generale, 2) if moyenne_generale else None,
                'absent': False,
                'pas_de_notes': False,
                'sexe': eleve.sexe
            })
        else:
            classement_data.append({
                'eleve_id': eleve.id,
                'prenom': eleve.prenom,
                'nom': eleve.nom,
                'matricule': eleve.matricule or 'N/A',
                'nom_complet': f"{eleve.prenom} {eleve.nom}",
                'moyenne': None,
                'absent': False,
                'pas_de_notes': True,
                'sexe': eleve.sexe
            })
    
    # Trier et attribuer les rangs
    classement_data = _calculer_rangs(classement_data)
    
    titre_export = f"Classement Général - {classe_note.nom}"
    if periode:
        titre_export += f" - {periode}"
    
    return classement_data, titre_export


def _calculer_rangs(classement_data):
    """Calculer les rangs avec gestion des ex-aequo en utilisant calculer_rang_intelligent"""
    from .calculs_intelligent import calculer_rang_intelligent
    from decimal import Decimal
    
    # Séparer élèves avec et sans notes
    eleves_avec_notes = [e for e in classement_data if e['moyenne'] is not None]
    eleves_sans_notes = [e for e in classement_data if e['moyenne'] is None]
    
    if eleves_avec_notes:
        # Préparer les données pour calculer_rang_intelligent
        moyennes_pour_rang = []
        for e in eleves_avec_notes:
            moyennes_pour_rang.append({
                'eleve_id': e.get('eleve_id'),
                'prenom': e.get('prenom', ''),
                'nom': e.get('nom', ''),
                'sexe': e.get('sexe', 'M'),
                'moyenne': Decimal(str(e['moyenne'])) if e['moyenne'] is not None else Decimal('0')
            })
        
        # Calculer les rangs avec calculer_rang_intelligent
        resultats_rangs = calculer_rang_intelligent(moyennes_pour_rang)
        
        # Créer un dictionnaire eleve_id -> rang
        rangs_dict = {}
        for r in resultats_rangs:
            rangs_dict[r['eleve_id']] = {
                'rang': r.get('rang'),  # Déjà formaté sans le total par calculer_rang_intelligent
                'rang_num': r.get('rang_num')
            }
        
        # Appliquer les rangs aux données
        for e in eleves_avec_notes:
            eleve_id = e.get('eleve_id')
            if eleve_id in rangs_dict:
                e['rang'] = rangs_dict[eleve_id]['rang']
                e['rang_num'] = rangs_dict[eleve_id]['rang_num']
        
        # Trier par rang_num
        eleves_avec_notes.sort(key=lambda x: x.get('rang_num', 999))
    
    # Marquer les élèves sans notes
    for eleve_note in eleves_sans_notes:
        eleve_note['rang'] = '-'
        eleve_note['rang_num'] = None
    
    # Reconstruire la liste triée
    return eleves_avec_notes + eleves_sans_notes


# =============================================================================
# EXPORT PDF
# =============================================================================

def _draw_school_header_classement(c, ecole, *, y_start, margin, page_width):
    """Dessine un en-tête officiel pour le classement"""
    y = y_start
    center_x = page_width / 2
    
    # En-tête national
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(center_x, y, "République de Guinée")
    y -= 10
    
    # Devise avec couleurs
    c.setFont('Helvetica-Oblique', 9)
    parts = [
        ("Travail", colors.red),
        (" - ", colors.black),
        ("Justice", colors.yellow),
        (" - ", colors.black),
        ("Solidarité", colors.green),
    ]
    total_w = sum(pdfmetrics.stringWidth(t, 'Helvetica-Oblique', 9) for t, _ in parts)
    start_x = center_x - (total_w / 2)
    x = start_x
    for text, col in parts:
        c.setFillColor(col)
        c.drawString(x, y, text)
        x += pdfmetrics.stringWidth(text, 'Helvetica-Oblique', 9)
    c.setFillColor(colors.black)
    y -= 10
    
    # Ministère
    c.setFont('Helvetica', 9)
    c.drawCentredString(center_x, y, "Ministère de l'Enseignement Pré-Universitaire et de l'Alphabétisation")
    y -= 10
    
    # IRE, DPE, DESEE
    c.setFont('Helvetica-Bold', 9)
    ire = getattr(ecole, 'ire', None) or ''
    dpe = getattr(ecole, 'dpe', None) or ''
    desee = getattr(ecole, 'desee', None) or ''
    if ire:
        c.drawCentredString(center_x, y, f"IRE: {ire}")
        y -= 10
    if dpe:
        c.drawCentredString(center_x, y, f"DPE: {dpe}")
        y -= 10
    if desee:
        c.drawCentredString(center_x, y, f"DESEE: {desee}")
        y -= 12
    
    # Cadre avec informations de l'école
    frame_top = y
    box_height = 50
    
    # Logo (gauche) si disponible
    logo_path = None
    try:
        if hasattr(ecole, 'logo') and getattr(ecole.logo, 'path', None) and os.path.exists(ecole.logo.path):
            logo_path = ecole.logo.path
    except Exception:
        logo_path = None
    if logo_path:
        try:
            c.drawImage(logo_path, margin + 5, y - 48, width=40, height=40, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    
    # Nom de l'école
    school_name = (getattr(ecole, 'nom', '') or 'ÉCOLE').upper()
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(center_x, y - 8, school_name)
    
    # Contacts (sans l'adresse)
    c.setFont('Helvetica', 8)
    c.setFillGray(0.3)
    telephone = getattr(ecole, 'telephone', None) or ''
    email = getattr(ecole, 'email', None) or ''
    
    line_y = y - 22
    # Adresse supprimée sur demande de l'utilisateur
    # Afficher téléphone et email sur des lignes séparées pour éviter le débordement
    if telephone:
        c.drawCentredString(center_x, line_y, f"Tél: {telephone}")
        line_y -= 12
    if email:
        c.drawCentredString(center_x, line_y, f"Email: {email}")
        line_y -= 12
    
    c.setFillGray(0.0)
    
    # Cadre supprimé sur demande de l'utilisateur
    
    y = y - box_height - 8
    return y


def _draw_watermark(c, ecole, page_width, page_height):
    """Dessine le logo en filigrane au centre de la page"""
    logo_path = None
    try:
        if hasattr(ecole, 'logo') and getattr(ecole.logo, 'path', None) and os.path.exists(ecole.logo.path):
            logo_path = ecole.logo.path
    except Exception:
        logo_path = None
    
    if logo_path:
        try:
            # Logo en filigrane au centre
            watermark_size = 300
            x = (page_width - watermark_size) / 2
            y = (page_height - watermark_size) / 2
            
            c.saveState()
            c.setFillAlpha(0.08)  # Transparence
            c.drawImage(logo_path, x, y, width=watermark_size, height=watermark_size, 
                       preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except Exception:
            pass


@login_required
def exporter_classement_classe_pdf(request):
    """
    Exporter le classement d'une classe en PDF avec en-tête officiel, logo et filigrane
    """
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    type_note = request.GET.get('type_note', 'mensuelle')
    periode = request.GET.get('periode', '')
    
    if not classe_id:
        return HttpResponse("Classe non spécifiée", status=400)
    
    # Récupérer la classe
    classe_note = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Récupérer la classe élève correspondante avec mapping spécial (même logique que les autres vues)
    try:
        # Mapping spécial pour les classes avec noms différents
        mapping_classes = {
            61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
            59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
        }
        
        if classe_note.id in mapping_classes:
            classe_eleve = ClasseEleve.objects.filter(
                id=mapping_classes[classe_note.id]
            ).first()
        else:
            # Essai 1: Correspondance exacte
            classe_eleve = ClasseEleve.objects.filter(
                nom=classe_note.nom,
                annee_scolaire=classe_note.annee_scolaire,
                ecole=classe_note.ecole
            ).first()
            
            # Essai 2: Correspondance insensible à la casse
            if not classe_eleve:
                classe_eleve = ClasseEleve.objects.filter(
                    nom__iexact=classe_note.nom,
                    annee_scolaire=classe_note.annee_scolaire,
                    ecole=classe_note.ecole
                ).first()
            
            # Essai 3: Recherche par mots-clés
            if not classe_eleve:
                import re
                match = re.search(r'(\d+)', classe_note.nom)
                if match:
                    niveau_num = match.group(1)
                    
                    # Chercher d'abord avec l'école
                    classes_possibles = ClasseEleve.objects.filter(
                        nom__icontains=niveau_num,
                        annee_scolaire=classe_note.annee_scolaire,
                        ecole=classe_note.ecole
                    )
                    
                    # Si aucune classe trouvée avec l'école, chercher sans filtrer par école
                    if not classes_possibles.exists():
                        classes_possibles = ClasseEleve.objects.filter(
                            nom__icontains=niveau_num,
                            annee_scolaire=classe_note.annee_scolaire
                        )
                    
                    classe_eleve = classes_possibles.first()
                    
                    # Si plusieurs classes trouvées, essayer d'affiner avec les mots-clés
                    if classe_eleve and classes_possibles.count() > 1:
                        if 'scientifique' in classe_note.nom.lower() or 'science' in classe_note.nom.lower():
                            for c in classes_possibles:
                                if 'SCIENCE' in c.nom.upper():
                                    classe_eleve = c
                                    break
                        elif 'littéraire' in classe_note.nom.lower() or 'lettre' in classe_note.nom.lower():
                            for c in classes_possibles:
                                if 'LETTRE' in c.nom.upper():
                                    classe_eleve = c
                                    break
        
        if not classe_eleve:
            return HttpResponse(
                f"Classe élève non trouvée pour '{classe_note.nom}' ({classe_note.annee_scolaire}).",
                status=404
            )
        
        # Récupérer les élèves actifs
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom')
        
        if not eleves.exists():
            return HttpResponse(
                f"Aucun élève actif trouvé dans la classe '{classe_eleve.nom}'",
                status=404
            )
    except Exception as e:
        return HttpResponse(f"Erreur lors de la récupération des élèves: {str(e)}", status=500)
    
    # Préparer les données de classement
    if matiere_id:
        classement_data, titre_export = _generer_classement_matiere(
            eleves, classe_note, matiere_id, type_note, periode
        )
    else:
        classement_data, titre_export = _generer_classement_general(
            eleves, classe_note, type_note, periode
        )
    
    # Détecter le niveau scolaire pour adapter l'affichage
    niveau_scolaire = detecter_niveau_scolaire(classe_note.nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    
    # Créer le PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    page_width, page_height = A4
    margin = 2*cm
    
    # Filigrane
    _draw_watermark(c, classe_note.ecole, page_width, page_height)
    
    # En-tête
    y = page_height - 1.5*cm
    y = _draw_school_header_classement(c, classe_note.ecole, y_start=y, margin=margin, page_width=page_width)
    
    # Titre du document
    y -= 15
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(page_width/2, y, titre_export)
    y -= 15
    
    # Type de période (composition, mensuelle, etc.)
    c.setFont('Helvetica', 11)
    c.setFillColorRGB(0.2, 0.2, 0.2)
    type_periode_text = ""
    if type_note == 'mensuelle':
        type_periode_text = f"{periode} - Notes Mensuelles" if periode else "Notes Mensuelles"
    elif type_note == 'composition':
        if periode and 'TRIMESTRE' in periode.upper():
            type_periode_text = f"{periode} - Composition Trimestrielle"
        elif periode and 'SEMESTRE' in periode.upper():
            type_periode_text = f"{periode} - Composition Semestrielle"
        else:
            type_periode_text = f"{periode} - Composition" if periode else "Composition"
    elif type_note == 'evaluation':
        type_periode_text = f"{periode} - Évaluations" if periode else "Évaluations"
    else:
        type_periode_text = periode if periode else type_note
    
    if type_periode_text:
        c.drawCentredString(page_width/2, y, type_periode_text)
        y -= 12
    
    c.setFillColorRGB(0, 0, 0)  # Retour au noir
    
    # Date d'export
    c.setFont('Helvetica-Oblique', 9)
    c.drawCentredString(page_width/2, y, f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    y -= 20
    
    # En-têtes du tableau
    col_widths = [2*cm, 3*cm, 7*cm, 2*cm, 3*cm]  # Rang, Matricule, Nom, Sexe, Moyenne
    col_x = [margin]
    for w in col_widths[:-1]:
        col_x.append(col_x[-1] + w)

    # Fond gris pour les en-têtes
    c.setFillColorRGB(0.2, 0.3, 0.4)
    c.rect(margin, y-15, sum(col_widths), 15, fill=1, stroke=0)

    # Texte des en-têtes (adapter selon le niveau)
    c.setFillColorRGB(1, 1, 1)
    c.setFont('Helvetica-Bold', 10)
    moyenne_header = 'Moy /10' if est_primaire else 'Moy /20'
    headers = ['Rang', 'Matricule', 'Nom Complet', 'Sexe', moyenne_header]
    for i, header in enumerate(headers):
        c.drawString(col_x[i] + 0.2*cm, y - 10, header)

    c.setFillColorRGB(0, 0, 0)
    y -= 20
    
    # Données
    c.setFont('Helvetica', 9)
    line_height = 14
    max_lines_per_page = 35
    line_count = 0
    
    for eleve_data in classement_data:
        # Nouvelle page si nécessaire
        if line_count >= max_lines_per_page:
            c.showPage()
            
            # Redessiner le filigrane et l'en-tête sur la nouvelle page
            _draw_watermark(c, classe_note.ecole, page_width, page_height)
            y = page_height - 1.5*cm
            y = _draw_school_header_classement(c, classe_note.ecole, y_start=y, margin=margin, page_width=page_width)
            y -= 20
            
            # Redessiner les en-têtes
            c.setFillColorRGB(0.2, 0.3, 0.4)
            c.rect(margin, y-15, sum(col_widths), 15, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont('Helvetica-Bold', 10)
            for i, header in enumerate(headers):
                c.drawString(col_x[i] + 0.2*cm, y - 10, header)
            c.setFillColorRGB(0, 0, 0)
            y -= 20
            c.setFont('Helvetica', 9)
            line_count = 0
        
        # Fond alterné pour faciliter la lecture
        if line_count % 2 == 0:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(margin, y - line_height + 2, sum(col_widths), line_height, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
        
        # Rang avec accord grammatical (déjà formaté par calculer_rang_intelligent)
        rang_value = eleve_data.get('rang', '-')
        rang_num = eleve_data.get('rang_num', None)
        
        # Utiliser le rang déjà formaté (ex: "1er/18", "10ème/18")
        rang_str = rang_value if isinstance(rang_value, str) else str(rang_value)
        
        # Appliquer les couleurs selon le rang numérique
        if rang_num == 1:
            # Médaille or
            c.setFillColorRGB(1, 0.84, 0)
            c.setFont('Helvetica-Bold', 10)
        elif rang_num == 2:
            # Médaille argent
            c.setFillColorRGB(0.75, 0.75, 0.75)
            c.setFont('Helvetica-Bold', 10)
        elif rang_num == 3:
            # Médaille bronze
            c.setFillColorRGB(0.8, 0.5, 0.2)
            c.setFont('Helvetica-Bold', 10)
        elif rang_str and rang_str != '-':
            c.setFont('Helvetica', 9)
        else:
            c.setFont('Helvetica', 9)
        
        c.drawString(col_x[0] + 0.2*cm, y - 10, rang_str)
        
        # Retour à la couleur normale
        c.setFillColorRGB(0, 0, 0)
        c.setFont('Helvetica', 9)
        
        # Matricule
        matricule = eleve_data.get('matricule', 'N/A')
        c.drawString(col_x[1] + 0.2*cm, y - 10, matricule)

        # Nom complet (tronqué si trop long)
        nom_complet = eleve_data.get('nom_complet', '')
        if len(nom_complet) > 32:
            nom_complet = nom_complet[:29] + "..."
        c.drawString(col_x[2] + 0.2*cm, y - 10, nom_complet)

        # Sexe
        sexe_val = eleve_data.get('sexe', '')
        c.setFont('Helvetica', 9)
        if sexe_val == 'F':
            c.setFillColorRGB(0.75, 0, 0.55)
            c.drawString(col_x[3] + 0.2*cm, y - 10, 'F')
        else:
            c.setFillColorRGB(0, 0.33, 0.8)
            c.drawString(col_x[3] + 0.2*cm, y - 10, 'G')
        c.setFillColorRGB(0, 0, 0)

        # Moyenne avec couleur selon performance
        moyenne = eleve_data.get('moyenne')
        seuil_admission = 5 if est_primaire else 10
        if moyenne is not None:
            c.setFont('Helvetica-Bold', 9)
            if moyenne >= (8 if est_primaire else 16):
                c.setFillColorRGB(0, 0.6, 0)  # Vert
            elif moyenne >= (6 if est_primaire else 14):
                c.setFillColorRGB(0, 0.5, 0.8)  # Bleu
            elif moyenne >= seuil_admission:
                c.setFillColorRGB(0.8, 0.6, 0)  # Orange
            else:
                c.setFillColorRGB(0.8, 0, 0)  # Rouge
            c.drawString(col_x[4] + 0.2*cm, y - 10, f"{moyenne:.2f}")
            c.setFillColorRGB(0, 0, 0)
        else:
            c.setFont('Helvetica-Oblique', 8)
            c.setFillColorRGB(0.5, 0.5, 0.5)
            if eleve_data.get('absent'):
                c.drawString(col_x[4] + 0.2*cm, y - 10, "Absent")
            elif eleve_data.get('pas_de_notes'):
                c.drawString(col_x[4] + 0.2*cm, y - 10, "Pas de notes")
            else:
                c.drawString(col_x[4] + 0.2*cm, y - 10, "Non saisi")
            c.setFillColorRGB(0, 0, 0)
        
        c.setFont('Helvetica', 9)
        
        # Ligne de séparation
        c.setStrokeColorRGB(0.8, 0.8, 0.8)
        c.setLineWidth(0.5)
        c.line(margin, y - line_height + 2, margin + sum(col_widths), y - line_height + 2)
        
        y -= line_height
        line_count += 1
    
    # Statistiques
    y -= 10
    if y < 5*cm:  # Nouvelle page si pas assez de place
        c.showPage()
        _draw_watermark(c, classe_note.ecole, page_width, page_height)
        y = page_height - 3*cm
    
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin, y, "STATISTIQUES")
    y -= 15
    
    c.setFont('Helvetica', 9)
    seuil_admission = 5 if est_primaire else 10
    base_note = '/10' if est_primaire else '/20'
    eleves_avec_moyenne = [e for e in classement_data if e.get('moyenne') is not None]
    eleves_sans_notes = [e for e in classement_data if e.get('moyenne') is None]
    nb_total = len(classement_data)
    nb_filles = sum(1 for e in classement_data if e.get('sexe') == 'F')
    nb_garcons = nb_total - nb_filles
    nb_admis = sum(1 for e in eleves_avec_moyenne if e['moyenne'] >= seuil_admission)
    nb_redoubles = sum(1 for e in eleves_avec_moyenne if e['moyenne'] < seuil_admission)
    taux_reussite = round((nb_admis / len(eleves_avec_moyenne)) * 100, 1) if eleves_avec_moyenne else 0

    # Boîtes de statistiques en 2 colonnes
    box_w = (sum(col_widths) / 2) - 0.3*cm
    box_h = 14
    col_left = margin
    col_right = margin + (sum(col_widths) / 2) + 0.3*cm

    def draw_stat_box(cx, cy, label, value, r, g, b):
        c.setFillColorRGB(r, g, b)
        c.rect(cx, cy - box_h + 3, box_w, box_h, fill=1, stroke=0)
        c.setFillColorRGB(1, 1, 1)
        c.setFont('Helvetica-Bold', 8)
        c.drawString(cx + 4, cy - 4, label)
        c.setFont('Helvetica-Bold', 10)
        c.drawRightString(cx + box_w - 4, cy - 4, str(value))
        c.setFillColorRGB(0, 0, 0)

    stats_boxes = [
        (col_left,  y,      "Effectif de la classe", nb_total,                    0.13, 0.19, 0.25),
        (col_right, y,      "Filles",                 nb_filles,                   0.75, 0.00, 0.55),
        (col_left,  y-18,   "Garçons",                nb_garcons,                  0.00, 0.33, 0.80),
        (col_right, y-18,   "Élèves évalués",         len(eleves_avec_moyenne),    0.10, 0.50, 0.40),
        (col_left,  y-36,   f"Admis (≥{seuil_admission})",        nb_admis,        0.00, 0.55, 0.30),
        (col_right, y-36,   f"Redoublants (<{seuil_admission})",  nb_redoubles,    0.72, 0.15, 0.15),
        (col_left,  y-54,   "Taux de réussite",       f"{taux_reussite}%",         0.20, 0.40, 0.70),
        (col_right, y-54,   "Non évalués",            len(eleves_sans_notes),      0.50, 0.50, 0.50),
    ]

    for bx, by, lbl, val, r, g, b in stats_boxes:
        draw_stat_box(bx, by, lbl, val, r, g, b)

    y -= 72

    if eleves_avec_moyenne:
        moyennes = [e['moyenne'] for e in eleves_avec_moyenne]
        moyenne_classe = sum(moyennes) / len(moyennes)
        note_max = max(moyennes)
        note_min = min(moyennes)

        y -= 8
        c.setFont('Helvetica-Bold', 9)
        c.setFillColorRGB(0.2, 0.3, 0.4)
        for stat in [
            f"Moyenne de classe : {moyenne_classe:.2f}{base_note}",
            f"Note maximale : {note_max:.2f}{base_note}   |   Note minimale : {note_min:.2f}{base_note}",
        ]:
            c.drawString(margin + 0.5*cm, y, stat)
            y -= 13
        c.setFillColorRGB(0, 0, 0)

    # Avertissement si élèves sans notes
    if eleves_sans_notes:
        y -= 8
        c.setFillColorRGB(0.8, 0, 0)
        c.setFont('Helvetica-Bold', 10)
        c.drawString(margin, y, f"ATTENTION: {len(eleves_sans_notes)} eleve(s) sans notes pour cette periode")
        c.setFillColorRGB(0, 0, 0)

    # ── Zone de signatures ──
    # Déterminer le titre du signataire selon le niveau
    niveau_scolaire_local = detecter_niveau_scolaire(classe_note.nom)
    if niveau_scolaire_local == 'MATERNELLE':
        signataire_titre = "La Coordinatrice de la Maternelle"
    elif niveau_scolaire_local == 'PRIMAIRE':
        signataire_titre = "Le Directeur du Primaire"
    else:
        signataire_titre = "Le Censeur de l'Établissement"

    sig_y = max(y - 30, 2.5 * cm)  # au moins 2.5cm du bas, ou sous les stats

    # Si pas assez de place, nouvelle page
    if sig_y < 2.2 * cm:
        c.showPage()
        _draw_watermark(c, classe_note.ecole, page_width, page_height)
        sig_y = 4 * cm

    # Ligne de séparation fine
    c.setStrokeColorRGB(0.7, 0.7, 0.7)
    c.setLineWidth(0.5)
    c.line(margin, sig_y + 35, page_width - margin, sig_y + 35)

    # Titre de section
    c.setFont('Helvetica-Bold', 9)
    c.setFillColorRGB(0.2, 0.3, 0.4)
    c.drawString(margin, sig_y + 22, "VISA ET SIGNATURE :")
    c.setFillColorRGB(0, 0, 0)

    # Bloc signataire centré à droite
    sig_block_x = page_width - margin - 7 * cm
    sig_block_w = 7 * cm

    c.setFont('Helvetica-Bold', 10)
    c.drawCentredString(sig_block_x + sig_block_w / 2, sig_y + 20, signataire_titre)

    # Ville + date
    c.setFont('Helvetica-Oblique', 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawCentredString(sig_block_x + sig_block_w / 2, sig_y + 7,
                        f"Conakry, le {datetime.now().strftime('%d/%m/%Y')}")
    c.setFillColorRGB(0, 0, 0)

    # Ligne de signature
    c.setStrokeColorRGB(0.2, 0.2, 0.2)
    c.setLineWidth(1)
    c.line(sig_block_x, sig_y - 10, sig_block_x + sig_block_w, sig_y - 10)

    # Mention "Signature"
    c.setFont('Helvetica-Oblique', 7)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawCentredString(sig_block_x + sig_block_w / 2, sig_y - 18, "(Signature et cachet)")
    c.setFillColorRGB(0, 0, 0)

    # Finaliser
    c.save()
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"Classement_{classe_note.nom}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
