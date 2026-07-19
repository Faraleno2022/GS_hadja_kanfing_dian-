"""
Module d'importation intelligente de notes
Template Excel dynamique avec toutes les matières de la classe en colonnes
"""
import pandas as pd
from decimal import Decimal, InvalidOperation
from django.db import transaction
from io import BytesIO

from .models import ClasseNote, MatiereNote, NoteMensuelle, CompositionNote
from eleves.models import Eleve, Classe as ClasseEleve


class ImportIntelligentError(Exception):
    """Erreur lors de l'importation intelligente"""
    pass


def generer_template_intelligent(classe_id, periode, system_type='trimestre'):
    """
    Génère un template Excel avec toutes les matières de la classe en colonnes.
    Format: N° | Matricule | Prénoms | Nom | Sexe | Matière1 | Matière2 | ... | MatièreN
    
    Args:
        classe_id: ID de la ClasseNote
        periode: Période (TRIMESTRE_1, SEMESTRE_1, OCTOBRE, etc.)
        system_type: Type de système (trimestre, semestre, mensuel)
    
    Returns:
        BytesIO: Fichier Excel en mémoire
    """
    try:
        classe = ClasseNote.objects.get(id=classe_id)
    except ClasseNote.DoesNotExist:
        raise ImportIntelligentError(f"Classe non trouvée (ID: {classe_id})")
    
    # Récupérer les matières de la classe (triées par nom)
    matieres = list(MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom'))
    
    if not matieres:
        raise ImportIntelligentError(f"Aucune matière trouvée pour la classe {classe.nom}")
    
    # Trouver la classe d'élèves correspondante
    classe_eleve = _trouver_classe_eleve(classe)
    
    # Récupérer les élèves (triés par ordre alphabétique: prénom puis nom)
    eleves = []
    if classe_eleve:
        eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom'))
    
    # Détecter le niveau scolaire pour la note max
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau = detecter_niveau_scolaire(classe.nom)
    note_max = 10 if niveau == 'PRIMAIRE' else 20
    
    # Construire les données (Prénoms avant Nom, tri par prénom)
    data = {
        'N°': list(range(1, len(eleves) + 1)) if eleves else [1],
        'Matricule': [e.matricule for e in eleves] if eleves else [''],
        'Prénoms': [e.prenom for e in eleves] if eleves else [''],
        'Nom': [e.nom for e in eleves] if eleves else [''],
        'Sexe': [e.sexe or '' for e in eleves] if eleves else [''],
    }
    
    # Ajouter une colonne pour chaque matière
    for matiere in matieres:
        # Utiliser le code ou le nom tronqué comme en-tête
        header = matiere.code if matiere.code else matiere.nom[:15]
        data[header] = ['' for _ in eleves] if eleves else ['']
    
    df = pd.DataFrame(data)
    
    # Créer le fichier Excel avec mise en forme
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Notes')
        
        workbook = writer.book
        worksheet = writer.sheets['Notes']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Styles
        header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        matiere_fill = PatternFill(start_color="f39c12", end_color="f39c12", fill_type="solid")
        matiere_font = Font(bold=True, color="FFFFFF", size=9)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Largeur des colonnes fixes
        worksheet.column_dimensions['A'].width = 5   # N°
        worksheet.column_dimensions['B'].width = 15  # Matricule
        worksheet.column_dimensions['C'].width = 15  # Prénoms
        worksheet.column_dimensions['D'].width = 12  # Nom
        worksheet.column_dimensions['E'].width = 6   # Sexe
        
        # Style de l'en-tête
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if col_idx <= 5:
                # Colonnes fixes (N°, Matricule, Prénoms, Nom, Sexe)
                cell.fill = header_fill
                cell.font = header_font
            else:
                # Colonnes matières
                cell.fill = matiere_fill
                cell.font = matiere_font
                # Largeur des colonnes matières
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 10
        
        # Style des données
        for row in worksheet.iter_rows(min_row=2, max_row=len(eleves) + 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajouter une ligne d'information en haut
        worksheet.insert_rows(1)
        info_cell = worksheet.cell(row=1, column=1)
        info_cell.value = f"FICHE DE REPORT DES NOTES - {classe.nom} - {_get_periode_label(periode)} - Notes sur {note_max}"
        info_cell.font = Font(bold=True, size=12, color="1a5276")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(matieres))
        
        # Ajouter une ligne avec les noms complets des matières
        worksheet.insert_rows(2)
        legend_cell = worksheet.cell(row=2, column=1)
        legend_text = "Légende: " + " | ".join([f"{m.code if m.code else m.nom[:10]}: {m.nom}" for m in matieres])
        legend_cell.value = legend_text
        legend_cell.font = Font(italic=True, size=9, color="666666")
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5 + len(matieres))
        
        # Figer les lignes d'en-tête
        worksheet.freeze_panes = 'A4'
    
    output.seek(0)
    return output, classe, matieres, note_max


def _trouver_classe_eleve(classe_note):
    """Trouve la ClasseEleve correspondant à une ClasseNote"""
    ecole = classe_note.ecole
    
    # Méthode 1: Correspondance exacte avec année scolaire ET école
    if ecole:
        classe_eleve = ClasseEleve.objects.filter(
            nom__iexact=classe_note.nom,
            annee_scolaire=classe_note.annee_scolaire,
            ecole=ecole
        ).first()
        if classe_eleve:
            return classe_eleve
    
    # Méthode 2: Correspondance exacte avec année scolaire
    classe_eleve = ClasseEleve.objects.filter(
        nom__iexact=classe_note.nom,
        annee_scolaire=classe_note.annee_scolaire
    ).first()
    if classe_eleve:
        return classe_eleve
    
    # Méthode 3: Correspondance par nom uniquement
    classe_eleve = ClasseEleve.objects.filter(
        nom__iexact=classe_note.nom
    ).first()
    
    return classe_eleve


def _get_periode_label(periode):
    """Retourne le libellé de la période"""
    labels = {
        'OCTOBRE': 'Octobre',
        'NOVEMBRE': 'Novembre',
        'DECEMBRE': 'Décembre',
        'JANVIER': 'Janvier',
        'FEVRIER': 'Février',
        'MARS': 'Mars',
        'AVRIL': 'Avril',
        'MAI': 'Mai',
        'JUIN': 'Juin',
        'TRIMESTRE_1': '1er Trimestre',
        'TRIMESTRE_2': '2ème Trimestre',
        'TRIMESTRE_3': '3ème Trimestre',
        'SEMESTRE_1': '1er Semestre',
        'SEMESTRE_2': '2ème Semestre',
    }
    return labels.get(periode, periode)


def importer_notes_intelligent(fichier, classe_id, periode, annee_scolaire, type_notes='MENSUELLE', user=None):
    """
    Importe les notes depuis un fichier Excel avec colonnes matières dynamiques.
    
    Args:
        fichier: Fichier Excel uploadé
        classe_id: ID de la ClasseNote
        periode: Période (OCTOBRE, TRIMESTRE_1, etc.)
        annee_scolaire: Année scolaire (ex: "2025-2026")
        type_notes: Type de notes (MENSUELLE ou COMPOSITION) - AUTO-DÉTECTÉ si non spécifié
        user: Utilisateur effectuant l'import
    
    Returns:
        dict: Statistiques d'importation
    """
    # Définir les périodes mensuelles valides
    PERIODES_MENSUELLES = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
    PERIODES_TRIMESTRIELLES = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
    PERIODES_SEMESTRIELLES = ['SEMESTRE_1', 'SEMESTRE_2']
    
    # Auto-détecter le type de notes selon la période
    if periode in PERIODES_MENSUELLES:
        type_notes = 'MENSUELLE'
    elif periode in PERIODES_TRIMESTRIELLES or periode in PERIODES_SEMESTRIELLES:
        type_notes = 'COMPOSITION'
    
    try:
        classe = ClasseNote.objects.get(id=classe_id)
    except ClasseNote.DoesNotExist:
        raise ImportIntelligentError(f"Classe non trouvée (ID: {classe_id})")
    
    # Lire le fichier Excel
    try:
        df = pd.read_excel(fichier, skiprows=2)  # Sauter les 2 lignes d'info
    except Exception as e:
        raise ImportIntelligentError(f"Erreur de lecture du fichier: {e}")
    
    # Nettoyer les noms de colonnes
    df.columns = df.columns.str.strip()
    
    # Récupérer les matières de la classe
    matieres = list(MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom'))
    
    # Créer un mapping code/nom -> matière
    matieres_mapping = {}
    for m in matieres:
        if m.code:
            matieres_mapping[m.code.strip().lower()] = m
        matieres_mapping[m.nom.strip().lower()] = m
        # Ajouter aussi les versions tronquées
        matieres_mapping[m.nom[:15].strip().lower()] = m
        matieres_mapping[m.nom[:10].strip().lower()] = m
    
    # Charger les élèves de l'école de la classe (pas tous les élèves !)
    eleves_dict = {e.matricule: e for e in Eleve.objects.filter(classe__ecole=classe.ecole)}
    
    # Détecter le niveau scolaire pour la note max
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau = detecter_niveau_scolaire(classe.nom)
    note_max = 10 if niveau == 'PRIMAIRE' else 20
    
    # Statistiques
    stats = {
        'total_lignes': 0,
        'total_notes': 0,
        'importees': 0,
        'modifiees': 0,
        'erreurs': 0,
        'matieres_traitees': set(),
        'erreurs_details': []
    }
    
    # Identifier les colonnes matières (tout sauf N°, Matricule, Prénoms, Nom, Sexe)
    colonnes_fixes = ['n°', 'matricule', 'prénoms', 'prenoms', 'prénom', 'prenom', 'nom', 'sexe']
    colonnes_matieres = []
    
    for col in df.columns:
        col_lower = col.strip().lower()
        if col_lower not in colonnes_fixes:
            # Chercher la matière correspondante
            matiere = matieres_mapping.get(col_lower)
            if matiere:
                colonnes_matieres.append((col, matiere))
            else:
                # Essayer une correspondance partielle
                for key, m in matieres_mapping.items():
                    if key in col_lower or col_lower in key:
                        colonnes_matieres.append((col, m))
                        break
    
    if not colonnes_matieres:
        raise ImportIntelligentError(
            f"Aucune colonne matière reconnue. "
            f"Colonnes attendues: {', '.join([m.code or m.nom[:15] for m in matieres])}"
        )
    
    # Trouver la colonne matricule
    matricule_col = None
    for col in df.columns:
        if col.strip().lower() == 'matricule':
            matricule_col = col
            break
    
    if not matricule_col:
        raise ImportIntelligentError("Colonne 'Matricule' non trouvée dans le fichier")
    
    # Importer les notes
    with transaction.atomic():
        for index, row in df.iterrows():
            stats['total_lignes'] += 1
            
            # Récupérer le matricule
            matricule = str(row.get(matricule_col, '')).strip()
            if not matricule or pd.isna(row.get(matricule_col)):
                continue
            
            # Trouver l'élève
            eleve = eleves_dict.get(matricule)
            if not eleve:
                stats['erreurs'] += 1
                stats['erreurs_details'].append(f"Ligne {index + 4}: Matricule '{matricule}' introuvable")
                continue
            
            # Traiter chaque colonne matière
            for col_name, matiere in colonnes_matieres:
                note_value = row.get(col_name)
                
                # Ignorer les cellules vides
                if pd.isna(note_value) or note_value == '':
                    continue
                
                stats['total_notes'] += 1
                stats['matieres_traitees'].add(matiere.nom)
                
                try:
                    # Convertir la note
                    note_decimal = Decimal(str(note_value).replace(',', '.'))
                    
                    # Valider la note
                    if note_decimal < 0 or note_decimal > note_max:
                        stats['erreurs'] += 1
                        stats['erreurs_details'].append(
                            f"Ligne {index + 4}, {matiere.nom}: Note {note_value} hors limites (0-{note_max})"
                        )
                        continue
                    
                    # Créer ou mettre à jour la note
                    if type_notes == 'COMPOSITION':
                        note_obj, created = CompositionNote.objects.update_or_create(
                            eleve=eleve,
                            matiere=matiere,
                            periode=periode,
                            annee_scolaire=annee_scolaire,
                            defaults={
                                'note': note_decimal,
                                'absent': False,
                                'cree_par': user
                            }
                        )
                    else:
                        # Notes mensuelles
                        note_obj, created = NoteMensuelle.objects.update_or_create(
                            eleve=eleve,
                            matiere=matiere,
                            mois=periode,
                            annee_scolaire=annee_scolaire,
                            defaults={
                                'note': note_decimal,
                                'absent': False,
                                'cree_par': user
                            }
                        )
                    
                    if created:
                        stats['importees'] += 1
                    else:
                        stats['modifiees'] += 1
                
                except (InvalidOperation, ValueError) as e:
                    stats['erreurs'] += 1
                    stats['erreurs_details'].append(
                        f"Ligne {index + 4}, {matiere.nom}: Format invalide '{note_value}'"
                    )
    
    stats['matieres_traitees'] = list(stats['matieres_traitees'])
    return stats
