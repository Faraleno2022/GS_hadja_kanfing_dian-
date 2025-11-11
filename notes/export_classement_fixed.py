"""
Module corrigé pour exporter les classements par classe avec gestion améliorée des élèves sans notes
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from decimal import Decimal
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import ClasseNote, MatiereNote, NoteMensuelle, CompositionNote
from eleves.models import Eleve, Classe as ClasseEleve


def formater_rang(rang, sexe):
    """
    Formate le rang avec l'accord grammatical selon le sexe
    
    Args:
        rang: Le numéro de rang (int ou str)
        sexe: 'M' pour masculin, 'F' pour féminin
    
    Returns:
        str: Le rang formaté (ex: "1er", "1ère", "2ème", "3ème", etc.)
    """
    if rang == '-' or rang is None or rang == 0:
        return '-'
    
    try:
        rang_num = int(rang)
        
        # Cas spécial pour le rang 1
        if rang_num == 1:
            if sexe == 'F':
                return "1ère"
            else:
                return "1er"
        
        # Pour tous les autres rangs, on utilise "ème"
        return f"{rang_num}ème"
    except (ValueError, TypeError):
        return '-'


@login_required
def exporter_classement_classe_fixed(request):
    """
    Exporter le classement d'une classe avec rang, nom complet, matricule et moyenne
    Version corrigée avec meilleure gestion des élèves sans notes
    """
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    type_note = request.GET.get('type_note', 'mensuelle')
    periode = request.GET.get('periode', '')
    
    if not classe_id:
        return HttpResponse("Classe non spécifiée", status=400)
    
    # Récupérer la classe
    classe_note = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Récupérer la classe élève correspondante
    try:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_note.nom,
            annee_scolaire=classe_note.annee_scolaire,
            ecole=classe_note.ecole
        ).first()
        
        if not classe_eleve:
            # Essayer de trouver par nom seulement si année scolaire ne correspond pas
            classe_eleve = ClasseEleve.objects.filter(
                nom=classe_note.nom,
                ecole=classe_note.ecole
            ).first()
            
            if not classe_eleve:
                return HttpResponse("Classe élève non trouvée", status=404)
        
        # Récupérer TOUS les élèves actifs de la classe
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom')
        
        # Si aucun élève actif, essayer avec tous les statuts
        if not eleves.exists():
            eleves = Eleve.objects.filter(classe=classe_eleve).order_by('nom', 'prenom')
            
    except Exception as e:
        return HttpResponse(f"Erreur lors de la récupération des élèves: {str(e)}", status=500)
    
    # Vérifier qu'il y a des élèves
    if not eleves.exists():
        return HttpResponse("Aucun élève trouvé dans cette classe", status=404)
    
    # Préparer les données de classement
    if matiere_id:
        classement_data, titre_export = _generer_classement_matiere_fixed(
            eleves, classe_note, matiere_id, type_note, periode
        )
    else:
        classement_data, titre_export = _generer_classement_general_fixed(
            eleves, classe_note, type_note, periode
        )
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classement"
    
    # Styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre du document
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = titre_export
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Informations supplémentaires
    ws.merge_cells('A2:D2')
    info_cell = ws['A2']
    info_cell.value = f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    info_cell.font = Font(italic=True, size=10)
    info_cell.alignment = Alignment(horizontal='center')
    
    # En-têtes du tableau
    headers = ['Rang', 'Matricule', 'Nom Complet', 'Moyenne /20']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données du classement
    row_num = 5
    for eleve_data in classement_data:
        # Rang avec accord grammatical
        rang_cell = ws.cell(row=row_num, column=1)
        if 'rang' in eleve_data and eleve_data['rang'] and eleve_data['rang'] != '-':
            rang_cell.value = formater_rang(eleve_data['rang'], eleve_data.get('sexe', 'M'))
        else:
            rang_cell.value = '-'
        rang_cell.alignment = Alignment(horizontal='center', vertical='center')
        rang_cell.border = border
        
        # Couleurs pour les premiers rangs
        if eleve_data.get('rang') == 1:
            rang_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            rang_cell.font = Font(bold=True)
        elif eleve_data.get('rang') == 2:
            rang_cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            rang_cell.font = Font(bold=True)
        elif eleve_data.get('rang') == 3:
            rang_cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
            rang_cell.font = Font(bold=True)
        
        # Matricule
        matricule_cell = ws.cell(row=row_num, column=2)
        matricule_cell.value = eleve_data.get('matricule', 'N/A')
        matricule_cell.alignment = Alignment(horizontal='center', vertical='center')
        matricule_cell.border = border
        
        # Nom complet
        nom_cell = ws.cell(row=row_num, column=3)
        nom_cell.value = eleve_data.get('nom_complet', '')
        nom_cell.alignment = Alignment(horizontal='left', vertical='center')
        nom_cell.border = border
        
        # Moyenne avec code couleur
        moyenne_cell = ws.cell(row=row_num, column=4)
        if eleve_data.get('moyenne') is not None:
            moyenne = eleve_data['moyenne']
            moyenne_cell.value = f"{moyenne:.2f}"
            moyenne_cell.font = Font(bold=True)
            
            # Code couleur selon la moyenne
            if moyenne >= 16:
                moyenne_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif moyenne >= 14:
                moyenne_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            elif moyenne >= 12:
                moyenne_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            elif moyenne >= 10:
                moyenne_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            else:
                moyenne_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        else:
            # Afficher clairement pourquoi pas de moyenne
            if eleve_data.get('absent'):
                moyenne_cell.value = "Absent"
            elif eleve_data.get('pas_de_notes'):
                moyenne_cell.value = "Pas de notes"
            else:
                moyenne_cell.value = "Non saisi"
            moyenne_cell.font = Font(italic=True, color="999999")
        
        moyenne_cell.alignment = Alignment(horizontal='center', vertical='center')
        moyenne_cell.border = border
        
        row_num += 1
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    
    # Statistiques
    eleves_avec_moyenne = [e for e in classement_data if e.get('moyenne') is not None]
    if eleves_avec_moyenne:
        stats_row = row_num + 2
        
        ws.merge_cells(f'A{stats_row}:D{stats_row}')
        stats_title = ws[f'A{stats_row}']
        stats_title.value = "STATISTIQUES"
        stats_title.font = Font(bold=True, size=12)
        stats_title.alignment = Alignment(horizontal='center')
        
        moyennes = [e['moyenne'] for e in eleves_avec_moyenne]
        moyenne_classe = sum(moyennes) / len(moyennes) if moyennes else 0
        note_max = max(moyennes) if moyennes else 0
        note_min = min(moyennes) if moyennes else 0
        
        stats_data = [
            ('Nombre total d\'élèves:', len(classement_data)),
            ('Élèves avec notes:', len(eleves_avec_moyenne)),
            ('Élèves sans notes:', len(classement_data) - len(eleves_avec_moyenne)),
            ('Moyenne de classe:', f"{moyenne_classe:.2f}" if moyenne_classe else "N/A"),
            ('Note maximale:', f"{note_max:.2f}" if note_max else "N/A"),
            ('Note minimale:', f"{note_min:.2f}" if note_min else "N/A"),
        ]
        
        for i, (label, value) in enumerate(stats_data, stats_row + 1):
            label_cell = ws.cell(row=i, column=2)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='right')
            
            value_cell = ws.cell(row=i, column=3)
            value_cell.value = value
            value_cell.alignment = Alignment(horizontal='left')
    
    # Message d'avertissement si des élèves n'ont pas de notes
    eleves_sans_notes = [e for e in classement_data if e.get('moyenne') is None]
    if eleves_sans_notes:
        warning_row = ws.max_row + 2
        ws.merge_cells(f'A{warning_row}:D{warning_row}')
        warning_cell = ws[f'A{warning_row}']
        warning_cell.value = f"⚠️ ATTENTION: {len(eleves_sans_notes)} élève(s) n'ont pas de notes saisies"
        warning_cell.font = Font(color="FF0000", bold=True)
        warning_cell.alignment = Alignment(horizontal='center')
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"Classement_{classe_note.nom}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def _generer_classement_matiere_fixed(eleves, classe_note, matiere_id, type_note, periode):
    """Générer le classement pour une matière spécifique avec meilleure gestion"""
    matiere = get_object_or_404(MatiereNote, pk=matiere_id)
    classement_data = []
    
    for eleve in eleves:
        note_value = None
        absent = False
        pas_de_notes = True
        
        # Récupérer la note selon le type
        if type_note == 'mensuelle' and periode:
            try:
                note_obj = NoteMensuelle.objects.get(
                    eleve=eleve,
                    matiere=matiere,
                    mois=periode,
                    annee_scolaire=classe_note.annee_scolaire
                )
                note_value = float(note_obj.note) if note_obj.note else None
                absent = note_obj.absent
                pas_de_notes = False
            except NoteMensuelle.DoesNotExist:
                pas_de_notes = True
        
        elif type_note == 'composition' and periode:
            try:
                note_obj = CompositionNote.objects.get(
                    eleve=eleve,
                    matiere=matiere,
                    periode=periode,
                    annee_scolaire=classe_note.annee_scolaire
                )
                note_value = float(note_obj.note) if note_obj.note else None
                absent = note_obj.absent
                pas_de_notes = False
            except CompositionNote.DoesNotExist:
                pas_de_notes = True
        
        classement_data.append({
            'matricule': eleve.matricule or 'N/A',
            'nom_complet': f"{eleve.nom} {eleve.prenom}",
            'moyenne': note_value,
            'absent': absent,
            'pas_de_notes': pas_de_notes,
            'sexe': eleve.sexe
        })
    
    # Trier et attribuer les rangs
    classement_data = _calculer_rangs_fixed(classement_data)
    
    titre_export = f"Classement - {classe_note.nom} - {matiere.nom}"
    if periode:
        titre_export += f" - {periode}"
    
    return classement_data, titre_export


def _generer_classement_general_fixed(eleves, classe_note, type_note, periode):
    """Générer le classement général avec meilleure gestion des cas sans notes"""
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True)
    classement_data = []
    
    # Debug: afficher le nombre de matières
    print(f"DEBUG: {matieres.count()} matières actives pour la classe {classe_note.nom}")
    
    for eleve in eleves:
        total_notes = Decimal('0')
        total_coefficients = Decimal('0')
        nb_notes_trouvees = 0
        toutes_absentes = True
        
        for matiere in matieres:
            note_value = None
            coefficient = matiere.coefficient or Decimal('1')
            
            # Récupérer la note selon le type
            if type_note == 'mensuelle' and periode:
                try:
                    note_obj = NoteMensuelle.objects.get(
                        eleve=eleve,
                        matiere=matiere,
                        mois=periode,
                        annee_scolaire=classe_note.annee_scolaire
                    )
                    if not note_obj.absent and note_obj.note:
                        note_value = note_obj.note
                        toutes_absentes = False
                        nb_notes_trouvees += 1
                except NoteMensuelle.DoesNotExist:
                    pass
            
            elif type_note == 'composition' and periode:
                try:
                    note_obj = CompositionNote.objects.get(
                        eleve=eleve,
                        matiere=matiere,
                        periode=periode,
                        annee_scolaire=classe_note.annee_scolaire
                    )
                    if not note_obj.absent and note_obj.note:
                        note_value = note_obj.note
                        toutes_absentes = False
                        nb_notes_trouvees += 1
                except CompositionNote.DoesNotExist:
                    pass
            
            if note_value is not None:
                total_notes += note_value * coefficient
                total_coefficients += coefficient
        
        # Calculer la moyenne générale
        if total_coefficients > 0:
            moyenne_generale = float(total_notes / total_coefficients)
            classement_data.append({
                'matricule': eleve.matricule or 'N/A',
                'nom_complet': f"{eleve.nom} {eleve.prenom}",
                'moyenne': round(moyenne_generale, 2),
                'absent': False,
                'pas_de_notes': False,
                'nb_notes': nb_notes_trouvees,
                'sexe': eleve.sexe
            })
        else:
            # Déterminer pourquoi pas de moyenne
            classement_data.append({
                'matricule': eleve.matricule or 'N/A',
                'nom_complet': f"{eleve.nom} {eleve.prenom}",
                'moyenne': None,
                'absent': toutes_absentes and nb_notes_trouvees == 0,
                'pas_de_notes': nb_notes_trouvees == 0,
                'nb_notes': 0,
                'sexe': eleve.sexe
            })
    
    # Debug: afficher les statistiques
    nb_avec_notes = len([e for e in classement_data if e['moyenne'] is not None])
    print(f"DEBUG: {nb_avec_notes}/{len(classement_data)} élèves ont des notes")
    
    # Trier et attribuer les rangs
    classement_data = _calculer_rangs_fixed(classement_data)
    
    titre_export = f"Classement Général - {classe_note.nom}"
    if periode:
        titre_export += f" - {periode}"
    
    return classement_data, titre_export


def _calculer_rangs_fixed(classement_data):
    """Calculer les rangs avec gestion des ex-aequo et meilleur diagnostic"""
    # Séparer élèves avec et sans notes
    eleves_avec_notes = [e for e in classement_data if e['moyenne'] is not None]
    eleves_sans_notes = [e for e in classement_data if e['moyenne'] is None]
    
    # Trier par moyenne décroissante
    eleves_avec_notes.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Attribuer les rangs avec gestion des ex-aequo
    rang_actuel = 1
    for i, eleve_note in enumerate(eleves_avec_notes):
        if i > 0 and abs(eleve_note['moyenne'] - eleves_avec_notes[i-1]['moyenne']) < 0.01:
            # Ex-aequo si la différence est inférieure à 0.01
            eleve_note['rang'] = eleves_avec_notes[i-1]['rang']
        else:
            eleve_note['rang'] = rang_actuel
        rang_actuel += 1
    
    # Marquer les élèves sans notes avec un rang 0 (sera affiché comme '-')
    for eleve_note in eleves_sans_notes:
        eleve_note['rang'] = 0
    
    # Retourner la liste complète : d'abord ceux avec notes (triés), puis ceux sans notes
    return eleves_avec_notes + eleves_sans_notes
