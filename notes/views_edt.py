"""
Emploi du temps par classe + calendrier de présence des professeurs.

Le calendrier des professeurs réutilise le modèle salaires.PresenceEnseignant
(heure d'arrivée / heure de départ) : les données restent cohérentes avec
le module Salaires.
"""
import calendar as _cal
import io
from datetime import date, datetime, time

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404

from utilisateurs.utils import filter_by_user_school
from .models import ClasseNote, MatiereNote, CreneauEmploiDuTemps

JOURS = CreneauEmploiDuTemps.JOUR_CHOICES


def _classes_utilisateur(request):
    profil = getattr(request.user, 'profil', None)
    ecole = profil.ecole if profil else None
    qs = ClasseNote.objects.filter(actif=True)
    if ecole:
        qs = qs.filter(ecole=ecole)
    return qs.order_by('niveau', 'nom')


def _parse_heure(val):
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(val, fmt).time()
        except (ValueError, TypeError):
            continue
    return None


def _enseignants_ecole(ecole):
    from salaires.models import Enseignant
    qs = Enseignant.objects.all()
    if ecole:
        qs = qs.filter(ecole=ecole)
    return qs.order_by('nom', 'prenoms')


# =====================================================================
#  EMPLOI DU TEMPS
# =====================================================================
def _grille_edt(classe):
    """Construit la grille: lignes = plages horaires, colonnes = jours."""
    creneaux = list(CreneauEmploiDuTemps.objects.filter(classe=classe)
                    .select_related('matiere', 'enseignant')
                    .order_by('heure_debut'))
    # Plages horaires uniques (heure_debut, heure_fin)
    plages = sorted({(c.heure_debut, c.heure_fin) for c in creneaux})
    # index rapide (jour, (hd,hf)) -> creneau
    index = {}
    for c in creneaux:
        index[(c.jour, (c.heure_debut, c.heure_fin))] = c
    lignes = []
    for hd, hf in plages:
        cells = []
        for code, _lib in JOURS:
            cells.append(index.get((code, (hd, hf))))
        lignes.append({'hd': hd, 'hf': hf, 'cells': cells})
    return lignes, creneaux


@login_required
def emploi_du_temps(request):
    classes = list(_classes_utilisateur(request))
    classe_id = (request.GET.get('classe_id') or request.POST.get('classe_id') or '').strip()
    classe = None
    matieres = []
    enseignants = []
    if classe_id.isdigit():
        classe = next((c for c in classes if c.id == int(classe_id)), None)
        if classe:
            matieres = list(MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom'))
            enseignants = list(_enseignants_ecole(classe.ecole))

    if request.method == 'POST' and classe and request.POST.get('action') == 'ajouter':
        jour = (request.POST.get('jour') or '').strip()
        hd = _parse_heure(request.POST.get('heure_debut'))
        hf = _parse_heure(request.POST.get('heure_fin'))
        matiere_id = (request.POST.get('matiere_id') or '').strip()
        libelle = (request.POST.get('libelle') or '').strip()
        enseignant_id = (request.POST.get('enseignant_id') or '').strip()
        salle = (request.POST.get('salle') or '').strip()

        if not (jour and hd and hf):
            messages.error(request, "Jour, heure de début et heure de fin sont obligatoires.")
        elif hf <= hd:
            messages.error(request, "L'heure de fin doit être après l'heure de début.")
        else:
            matiere = None
            if matiere_id.isdigit():
                matiere = next((m for m in matieres if m.id == int(matiere_id)), None)
            enseignant = None
            if enseignant_id.isdigit():
                enseignant = next((e for e in enseignants if e.id == int(enseignant_id)), None)
            CreneauEmploiDuTemps.objects.create(
                classe=classe, jour=jour, heure_debut=hd, heure_fin=hf,
                matiere=matiere, libelle=libelle if not matiere else '',
                enseignant=enseignant, salle=salle, cree_par=request.user)
            messages.success(request, "Créneau ajouté.")
        return redirect(f"{request.path}?classe_id={classe.id}")

    lignes, creneaux = ([], [])
    if classe:
        lignes, creneaux = _grille_edt(classe)

    return render(request, 'notes/emploi_du_temps.html', {
        'titre_page': "Emploi du temps",
        'classes': classes,
        'classe': classe,
        'classe_id': classe_id,
        'matieres': matieres,
        'enseignants': enseignants,
        'jours': JOURS,
        'lignes': lignes,
        'creneaux': creneaux,
    })


@login_required
def supprimer_creneau(request, creneau_id):
    creneau = get_object_or_404(CreneauEmploiDuTemps, pk=creneau_id)
    classe_id = creneau.classe_id
    if request.method == 'POST':
        creneau.delete()
        messages.success(request, "Créneau supprimé.")
    return redirect(f"/notes/emploi-du-temps/?classe_id={classe_id}")


@login_required
def emploi_du_temps_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    classe = get_object_or_404(ClasseNote, pk=request.GET.get('classe_id'))
    lignes, _ = _grille_edt(classe)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=1 * cm, bottomMargin=1 * cm,
                            leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    titre = ParagraphStyle('T', parent=styles['Heading1'], fontSize=14,
                           textColor=colors.HexColor('#007bff'), alignment=TA_CENTER)
    cell_style = ParagraphStyle('C', parent=styles['Normal'], fontSize=7.5, alignment=TA_CENTER, leading=9)

    elements = [Paragraph(f"<b>{(classe.ecole.nom if classe.ecole else '').upper()}</b>", titre)]
    elements.append(Paragraph(f"EMPLOI DU TEMPS — {classe.nom} — {classe.annee_scolaire}",
                              ParagraphStyle('S', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=8)))

    header = ['Horaire'] + [lib for _c, lib in JOURS]
    data = [header]
    for lg in lignes:
        row = [f"{lg['hd']:%H:%M}\n{lg['hf']:%H:%M}"]
        for cell in lg['cells']:
            if cell:
                txt = f"<b>{cell.intitule}</b>"
                if cell.enseignant:
                    txt += f"<br/>{cell.enseignant.nom} {cell.enseignant.prenoms or ''}"
                if cell.salle:
                    txt += f"<br/><i>Salle {cell.salle}</i>"
                row.append(Paragraph(txt, cell_style))
            else:
                row.append('')
        data.append(row)

    if len(data) == 1:
        elements.append(Paragraph("Aucun créneau saisi pour cette classe.", styles['Normal']))
    else:
        page_w = landscape(A4)[0] - 2 * cm
        col0 = 2.2 * cm
        colj = (page_w - col0) / len(JOURS)
        table = Table(data, colWidths=[col0] + [colj] * len(JOURS), repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#eef3f8')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{classe.nom}.pdf"'
    return resp


@login_required
def emploi_du_temps_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("Module openpyxl non installé", status=500)

    classe = get_object_or_404(ClasseNote, pk=request.GET.get('classe_id'))
    lignes, _ = _grille_edt(classe)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Emploi du temps'
    entete_font = Font(bold=True, color='FFFFFF')
    entete_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
    hd_fill = PatternFill(start_color='EEF3F8', end_color='EEF3F8', fill_type='solid')
    centre = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BBBBBB')
    bordure = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:G1')
    ws['A1'] = f"EMPLOI DU TEMPS — {classe.nom} — {classe.annee_scolaire}"
    ws['A1'].font = Font(bold=True, size=13, color='007BFF')
    ws['A1'].alignment = centre

    headers = ['Horaire'] + [lib for _c, lib in JOURS]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = entete_font
        cell.fill = entete_fill
        cell.alignment = centre
        cell.border = bordure

    r = 3
    for lg in lignes:
        r += 1
        hcell = ws.cell(row=r, column=1, value=f"{lg['hd']:%H:%M} - {lg['hf']:%H:%M}")
        hcell.fill = hd_fill
        hcell.font = Font(bold=True)
        hcell.alignment = centre
        hcell.border = bordure
        for i, cell in enumerate(lg['cells'], 2):
            if cell:
                parts = [cell.intitule]
                if cell.enseignant:
                    parts.append(f"{cell.enseignant.nom} {cell.enseignant.prenoms or ''}".strip())
                if cell.salle:
                    parts.append(f"Salle {cell.salle}")
                val = "\n".join(parts)
            else:
                val = ''
            cc = ws.cell(row=r, column=i, value=val)
            cc.alignment = centre
            cc.border = bordure

    ws.column_dimensions['A'].width = 16
    for col in 'BCDEFG':
        ws.column_dimensions[col].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{classe.nom}.xlsx"'
    return resp


# =====================================================================
#  CALENDRIER PRÉSENCE PROFESSEURS (réutilise salaires.PresenceEnseignant)
# =====================================================================
@login_required
def calendrier_professeurs(request):
    """Calendrier mensuel: enseignants (lignes) x jours du mois (colonnes)."""
    from salaires.models import PresenceEnseignant
    profil = getattr(request.user, 'profil', None)
    ecole = profil.ecole if profil else None

    today = date.today()
    try:
        annee = int(request.GET.get('annee') or today.year)
        mois = int(request.GET.get('mois') or today.month)
    except ValueError:
        annee, mois = today.year, today.month

    nb_jours = _cal.monthrange(annee, mois)[1]
    jours_mois = list(range(1, nb_jours + 1))

    enseignants = list(_enseignants_ecole(ecole))
    ens_ids = [e.id for e in enseignants]
    presences = PresenceEnseignant.objects.filter(
        enseignant_id__in=ens_ids, date__year=annee, date__month=mois)
    pmap = {(p.enseignant_id, p.date.day): p for p in presences}

    # Abréviation par statut
    abbr = {'PRESENT': 'P', 'ABSENT': 'A', 'RETARD': 'R', 'CONGE': 'C',
            'MALADIE': 'M', 'PERMISSION': 'Pm'}
    lignes = []
    for e in enseignants:
        cells = []
        for j in jours_mois:
            p = pmap.get((e.id, j))
            cells.append({
                'jour': j,
                'statut': p.statut if p else '',
                'abbr': abbr.get(p.statut, '') if p else '',
                'arrivee': p.heure_arrivee.strftime('%H:%M') if p and p.heure_arrivee else '',
                'depart': p.heure_depart.strftime('%H:%M') if p and p.heure_depart else '',
            })
        lignes.append({'enseignant': e, 'cells': cells})

    mois_noms = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    return render(request, 'notes/calendrier_professeurs.html', {
        'titre_page': "Calendrier des professeurs",
        'annee': annee, 'mois': mois, 'mois_nom': mois_noms[mois],
        'jours_mois': jours_mois,
        'lignes': lignes,
        'mois_noms': list(enumerate(mois_noms))[1:],
        'annees': range(today.year - 2, today.year + 2),
        'aujourdhui': today,
    })


@login_required
def pointage_professeurs(request):
    """Pointage entrée/sortie des professeurs pour une date donnée."""
    from salaires.models import PresenceEnseignant
    profil = getattr(request.user, 'profil', None)
    ecole = profil.ecole if profil else None

    jour = _parse_date(request.GET.get('date') or request.POST.get('date')) or date.today()
    enseignants = list(_enseignants_ecole(ecole))

    presences = {p.enseignant_id: p for p in PresenceEnseignant.objects.filter(
        enseignant__in=enseignants, date=jour)}

    if request.method == 'POST':
        maj = 0
        statuts = dict(PresenceEnseignant.STATUT_CHOICES)
        for e in enseignants:
            statut = (request.POST.get(f'statut_{e.id}') or 'PRESENT').strip()
            if statut not in statuts:
                statut = 'PRESENT'
            arr = _parse_heure(request.POST.get(f'arrivee_{e.id}'))
            dep = _parse_heure(request.POST.get(f'depart_{e.id}'))
            PresenceEnseignant.objects.update_or_create(
                enseignant=e, date=jour,
                defaults={'statut': statut, 'heure_arrivee': arr, 'heure_depart': dep})
            maj += 1
        messages.success(request, f"Pointage enregistré pour {maj} professeur(s) — {jour:%d/%m/%Y}.")
        return redirect(f"{request.path}?date={jour.isoformat()}")

    lignes = []
    for e in enseignants:
        p = presences.get(e.id)
        lignes.append({
            'enseignant': e,
            'statut': p.statut if p else 'PRESENT',
            'arrivee': p.heure_arrivee.strftime('%H:%M') if p and p.heure_arrivee else '',
            'depart': p.heure_depart.strftime('%H:%M') if p and p.heure_depart else '',
        })

    return render(request, 'notes/pointage_professeurs.html', {
        'titre_page': "Pointage des professeurs",
        'jour': jour,
        'lignes': lignes,
        'statut_choices': PresenceEnseignant.STATUT_CHOICES,
    })


def _parse_date(val):
    if not val:
        return None
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None
