import secrets
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth import get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods, require_POST

from salaires.models import Enseignant
from utilisateurs.utils import filter_by_user_school, user_is_account_principal
from .acces_enseignants import (SESSION_KEY, empreinte, nouveau_lien, verifier_session,
    classes_autorisees, matieres_autorisees, eleves_autorises, configuration, enregistrer_cellules)
from .forms_acces_enseignants import CreerAccesEnseignantForm, ExpirationForm
from .models import AccesEnseignantTemporaire, NoteMensuelle, CompositionNote, AppreciationMaternelle


@login_required
@never_cache
@require_http_methods(['GET', 'POST'])
def gerer_acces(request):
    if not user_is_account_principal(request.user):
        raise PermissionDenied('Seul le compte principal peut gérer les accès enseignants.')
    enseignants = filter_by_user_school(Enseignant.objects.all(), request.user).filter(
        statut='ACTIF', sync_deleted_at__isnull=True, type_enseignant__in=['MATERNELLE', 'PRIMAIRE', 'SECONDAIRE'],
    ).order_by('nom', 'prenoms')
    acces_qs = filter_by_user_school(AccesEnseignantTemporaire.objects.all(), request.user)
    enseignant_id = request.POST.get('enseignant') if request.method == 'POST' else request.GET.get('enseignant')
    enseignant = None
    if enseignant_id:
        if not str(enseignant_id).isdigit():
            raise PermissionDenied('Enseignant invalide.')
        enseignant = get_object_or_404(enseignants, pk=enseignant_id)
    form = CreerAccesEnseignantForm(
        request.POST if request.method == 'POST' and request.POST.get('action') == 'creer' else None,
        user=request.user, enseignant=enseignant,
    ) if enseignant else None
    renouvellement = ExpirationForm(initial={'expire_le': timezone.now() + timedelta(days=7)})
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'creer' and form and form.is_valid():
            token, digest = nouveau_lien()
            with transaction.atomic():
                user = get_user_model().objects.create_user(
                    username='notes_tmp_' + secrets.token_hex(12), password=None,
                    first_name=enseignant.prenoms[:150], last_name=enseignant.nom[:150],
                )
                profil = user.profil
                for field in profil._meta.fields:
                    if field.name.startswith('peut_'):
                        setattr(profil, field.name, False)
                profil.role, profil.ecole = 'ENSEIGNANT', enseignant.ecole
                profil.est_compte_principal, profil.lecture_seule = False, False
                profil.compte_principal = getattr(request.user, 'profil', None)
                profil.actif, profil.is_validated = True, True
                profil.allowed_menus = ['notes']
                profil.save()
                acces = AccesEnseignantTemporaire.objects.create(
                    utilisateur=user, enseignant=enseignant, ecole=enseignant.ecole, cree_par=request.user,
                    empreinte_lien=digest, expire_le=form.cleaned_data['expire_le'],
                )
                acces.classes.set(form.cleaned_data['classes'])
                acces.matieres.set(form.cleaned_data['matieres'])
            request.session['nouveau_lien_notes'] = request.build_absolute_uri(reverse('notes:enseignant_lien', args=[token]))
            messages.success(request, 'Compte temporaire créé. Copiez le lien personnel ci-dessous.')
            return redirect('notes:gerer_acces_enseignants')
        elif action in ('revoquer', 'renouveler'):
            renouvellement = ExpirationForm(request.POST)
            if not str(request.POST.get('acces_id', '')).isdigit():
                raise PermissionDenied('Accès invalide.')
            with transaction.atomic():
                acces = get_object_or_404(acces_qs.select_for_update(), pk=request.POST.get('acces_id'))
                if action == 'revoquer':
                    acces.revoque_le = timezone.now()
                    acces.save(update_fields=['revoque_le'])
                    messages.success(request, 'Accès révoqué, y compris les sessions déjà ouvertes.')
                    return redirect('notes:gerer_acces_enseignants')
                if renouvellement.is_valid():
                    token, acces.empreinte_lien = nouveau_lien()
                    acces.expire_le, acces.revoque_le = renouvellement.cleaned_data['expire_le'], None
                    acces.save(update_fields=['empreinte_lien', 'expire_le', 'revoque_le'])
                    request.session['nouveau_lien_notes'] = request.build_absolute_uri(reverse('notes:enseignant_lien', args=[token]))
                    messages.success(request, 'Nouveau lien créé. L’ancien lien et les anciennes sessions sont invalidés.')
                    return redirect('notes:gerer_acces_enseignants')
        elif action != 'creer' or not form:
            raise PermissionDenied('Action invalide.')
    return render(request, 'notes/enseignants/gestion.html', {
        'enseignants': enseignants, 'enseignant': enseignant, 'form': form,
        'renouvellement': renouvellement,
        'acces_liste': acces_qs.select_related('enseignant', 'ecole', 'utilisateur__profil').prefetch_related('classes', 'matieres'),
        'nouveau_lien': request.session.pop('nouveau_lien_notes', None),
    })


def acces_par_token(token):
    if len(token) != 64 or any(c not in '0123456789abcdef' for c in token):
        return None
    acces = AccesEnseignantTemporaire.objects.select_related('utilisateur__profil', 'enseignant', 'ecole').filter(empreinte_lien=empreinte(token)).first()
    return acces if acces and acces.est_valide else None


def lien_invalide(request):
    return render(request, 'notes/enseignants/message.html', {
        'erreur': 'Ce lien est invalide, expiré ou révoqué. Demandez un nouveau lien au compte principal de votre école.',
    }, status=403)


@never_cache
@require_http_methods(['GET'])
def ouvrir_lien(request, token):
    acces = acces_par_token(token)
    if not acces:
        return lien_invalide(request)
    return render(request, 'notes/enseignants/lien.html', {'acces': acces, 'token': token})


@never_cache
@require_POST
def connecter(request):
    # URL de POST sans secret : les journaux d'authentification ne reçoivent pas le lien.
    acces = acces_par_token(request.POST.get('token', ''))
    if not acces:
        return lien_invalide(request)
    login(request, acces.utilisateur, backend='django.contrib.auth.backends.ModelBackend')
    request.session[SESSION_KEY] = acces.empreinte_lien
    request.session.set_expiry(acces.expire_le)
    return redirect('notes:enseignant_accueil')


@require_POST
def deconnexion(request):
    logout(request)
    return render(request, 'notes/enseignants/message.html', {'info': 'Vous êtes déconnecté. Utilisez votre lien personnel pour revenir.'})


def accueil(request):
    acces = verifier_session(request)
    classes = list(classes_autorisees(acces))
    for classe in classes:
        classe.matieres_portail = matieres_autorisees(acces, classe)
    return render(request, 'notes/enseignants/accueil.html', {'acces': acces, 'classes': classes})


def contexte_saisie(request, classe_id):
    acces = verifier_session(request)
    classe = get_object_or_404(classes_autorisees(acces), pk=classe_id)
    params = request.POST if request.method == 'POST' else request.GET
    mode = params.get('mode', 'simple')
    if mode not in ('simple', 'intelligent'):
        raise PermissionDenied('Mode invalide.')
    maternelle = classe.niveau_enseignement == 'MATERNELLE'
    genre = params.get('genre', 'maternelle' if maternelle else 'mensuelle')
    periodes = (AppreciationMaternelle.TRIMESTRE_CHOICES if maternelle else
                CompositionNote.PERIODE_CHOICES if genre == 'composition' else NoteMensuelle.MOIS_CHOICES)
    periode = params.get('periode', periodes[0][0])
    configuration(classe, genre, periode)
    matieres = list(matieres_autorisees(acces, classe))
    selected = params.get('matiere', str(matieres[0].pk) if matieres else '')
    colonnes = matieres if mode == 'intelligent' else [m for m in matieres if str(m.pk) == selected]
    if not colonnes:
        raise PermissionDenied('Aucune matière autorisée sélectionnée. Contactez votre école.')
    return {'acces': acces, 'classe': classe, 'mode': mode, 'genre': genre, 'periode': periode,
            'periodes': periodes, 'matieres': matieres, 'matiere_selectionnee': selected,
            'colonnes': colonnes, 'maternelle': maternelle, 'maximum': 10 if classe.niveau_enseignement == 'PRIMAIRE' else 20,
            'appreciations': AppreciationMaternelle.APPRECIATION_CHOICES}


@require_http_methods(['GET', 'POST'])
def saisie(request, classe_id):
    try:
        context = contexte_saisie(request, classe_id)
    except ValidationError as exc:
        return render(request, 'notes/enseignants/message.html', {'erreur': exc.messages[0]}, status=400)
    classe, erreur = context['classe'], None
    from django.conf import settings
    eleves = list(eleves_autorises(classe))
    max_fields = min(4000, (settings.DATA_UPLOAD_MAX_NUMBER_FIELDS or 10000) - 10)
    if len(eleves) * len(context['colonnes']) > max_fields:
        return render(request, 'notes/enseignants/message.html', {
            'erreur': 'Ce tableau contient trop de cellules. Utilisez la saisie par matière ou l’import Excel.',
        }, status=400)
    if request.method == 'POST':
        cellules = []
        allowed = {m.pk for m in context['colonnes']}
        for name, values in request.POST.lists():
            if name.startswith('cell_'):
                parts = name.split('_')
                if len(parts) != 3 or not all(p.isdigit() for p in parts[1:]) or len(values) != 1 or int(parts[2]) not in allowed:
                    raise PermissionDenied('Cellule non autorisée.')
                cellules.append({'eleve': parts[1], 'matiere': parts[2], 'valeur': values[0]})
        try:
            count = enregistrer_cellules(request, classe.pk, context['genre'], context['periode'], cellules)
            messages.success(request, f'{count} note(s) / appréciation(s) enregistrée(s).')
            from urllib.parse import urlencode
            return redirect(request.path + '?' + urlencode({k: context[k] for k in ('mode', 'genre', 'periode')} | {'matiere': context['matiere_selectionnee']}))
        except ValidationError as exc:
            erreur = ' '.join(exc.messages)
    model, field = configuration(classe, context['genre'], context['periode'])
    notes = model.objects.filter(eleve__in=eleves, matiere__in=context['colonnes'], annee_scolaire=classe.annee_scolaire, **{field: context['periode']})
    existing = {(n.eleve_id, n.matiere_id): 'ABS' if n.absent else (n.appreciation if context['maternelle'] else str(n.note) if n.note is not None else '') for n in notes}
    lignes = []
    for eleve in eleves:
        cells = []
        for matiere in context['colonnes']:
            name = f'cell_{eleve.pk}_{matiere.pk}'
            value = request.POST.get(name, '') if request.method == 'POST' else existing.get((eleve.pk, matiere.pk), '')
            cells.append({'name': name, 'valeur': value, 'matiere': matiere})
        lignes.append({'eleve': eleve, 'cellules': cells})
    context.update(lignes=lignes, erreur=erreur)
    return render(request, 'notes/enseignants/saisie.html', context, status=400 if erreur else 200)


@require_http_methods(['GET', 'POST'])
def importer(request, classe_id):
    from .imports_enseignants import analyser_import
    try:
        context = contexte_saisie(request, classe_id)
        if request.method == 'POST':
            if request.POST.get('action') == 'confirmer':
                pending = request.session.get('notes_import_apercu', {})
                if (not pending or pending.get('nonce') != request.POST.get('nonce')
                        or pending.get('classe') != classe_id
                        or pending.get('empreinte') != context['acces'].empreinte_lien
                        or pending.get('date', 0) < timezone.now().timestamp() - 900
                        or any(pending.get(k) != context[k] for k in ('mode', 'genre', 'periode'))):
                    raise ValidationError('Aperçu expiré ou remplacé. Sélectionnez à nouveau votre fichier.')
                count = enregistrer_cellules(request, classe_id, pending['genre'], pending['periode'], pending['cellules'])
                request.session.pop('notes_import_apercu', None)
                messages.success(request, f'Import terminé : {count} note(s) / appréciation(s) enregistrée(s).')
                return redirect('notes:enseignant_accueil')
            request.session.pop('notes_import_apercu', None)
            cellules, preview = analyser_import(request.FILES.get('fichier'), context)
            nonce = secrets.token_hex(16)
            request.session['notes_import_apercu'] = {
                'nonce': nonce, 'classe': classe_id, 'cellules': cellules,
                'empreinte': context['acces'].empreinte_lien, 'date': timezone.now().timestamp(),
                **{k: context[k] for k in ('mode', 'genre', 'periode')},
            }
            context.update(preview=preview, nonce=nonce)
    except ValidationError as exc:
        if 'context' not in locals():
            return render(request, 'notes/enseignants/message.html', {'erreur': ' '.join(exc.messages)}, status=400)
        context['erreur'] = ' '.join(exc.messages)
    return render(request, 'notes/enseignants/import.html', context, status=400 if context.get('erreur') else 200)


def modele_import(request, classe_id):
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    try:
        context = contexte_saisie(request, classe_id)
    except ValidationError as exc:
        return render(request, 'notes/enseignants/message.html', {'erreur': exc.messages[0]}, status=400)
    book = Workbook()
    sheet = book.active
    sheet.title = 'Notes'
    headers = ['Matricule', 'Prénom', 'Nom'] + (
        [f'{m.code} - {m.nom}' for m in context['colonnes']] if context['mode'] == 'intelligent'
        else ['Appréciation' if context['maternelle'] else 'Note']
    )
    sheet.append(headers)
    for eleve in eleves_autorises(context['classe']):
        sheet.append([eleve.matricule, eleve.prenom, eleve.nom] + [''] * len(context['colonnes']))
    for row in sheet:
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = 's'  # Les noms/matricules ne sont jamais des formules.
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
    from openpyxl.utils import get_column_letter
    for idx in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(idx)].width = 26
    sheet.freeze_panes = 'D2'
    buffer = BytesIO()
    book.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modele_notes_enseignant.xlsx"'
    return response
