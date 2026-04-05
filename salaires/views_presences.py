from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import date, datetime, timedelta
from decimal import Decimal
import csv

from .models import Enseignant, PresenceEnseignant
from .forms import PresenceForm
from utilisateurs.utils import user_school, user_is_admin


@login_required
def liste_presences(request):
    """Liste des présences avec filtres"""
    user_school_obj = user_school(request.user)
    
    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    enseignant_id = request.GET.get('enseignant')
    statut = request.GET.get('statut')
    
    # Date par défaut: aujourd'hui
    if not date_debut:
        date_debut = date.today().strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = date.today().strftime('%Y-%m-%d')
    
    # Requête de base
    presences = PresenceEnseignant.objects.filter(
        enseignant__ecole=user_school_obj
    ).select_related('enseignant', 'pointe_par')
    
    # Appliquer les filtres
    if date_debut:
        presences = presences.filter(date__gte=date_debut)
    if date_fin:
        presences = presences.filter(date__lte=date_fin)
    if enseignant_id:
        presences = presences.filter(enseignant_id=enseignant_id)
    if statut:
        presences = presences.filter(statut=statut)
    
    # Statistiques
    stats = presences.aggregate(
        total=Count('id'),
        presents=Count('id', filter=Q(statut='PRESENT')),
        absents=Count('id', filter=Q(statut='ABSENT')),
        retards=Count('id', filter=Q(statut='RETARD')),
        total_heures=Sum('heures_travaillees')
    )
    
    # Liste des enseignants pour le filtre
    enseignants = Enseignant.objects.filter(
        ecole=user_school_obj,
        statut='ACTIF'
    ).order_by('nom', 'prenoms')
    
    context = {
        'presences': presences,
        'stats': stats,
        'enseignants': enseignants,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'enseignant_id': enseignant_id,
        'statut': statut,
        'statuts': PresenceEnseignant.STATUT_CHOICES,
    }
    
    return render(request, 'salaires/presences/liste.html', context)


@login_required
def pointer_presence(request):
    """Pointer la présence d'un ou plusieurs enseignants"""
    user_school_obj = user_school(request.user)
    
    if request.method == 'POST':
        date_pointage_str = request.POST.get('date')
        enseignants_ids = request.POST.getlist('enseignants')
        
        if not date_pointage_str or not enseignants_ids:
            messages.error(request, "Veuillez sélectionner une date et au moins un enseignant.")
            return redirect('salaires:pointer_presence')
        
        # Convertir la date string en objet date
        try:
            date_pointage = datetime.strptime(date_pointage_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Format de date invalide.")
            return redirect('salaires:pointer_presence')
        
        count_created = 0
        count_updated = 0
        
        # Valider que tous les enseignants sélectionnés appartiennent à l'école de l'utilisateur
        enseignants_valides = set(
            Enseignant.objects.filter(
                id__in=enseignants_ids,
                ecole=user_school_obj
            ).values_list('id', flat=True)
        )
        
        for ens_id in enseignants_ids:
            if int(ens_id) not in enseignants_valides:
                continue  # Ignorer les enseignants qui n'appartiennent pas à l'école
            statut = request.POST.get(f'statut_{ens_id}', 'PRESENT')
            heure_arrivee_str = request.POST.get(f'heure_arrivee_{ens_id}') or None
            heure_depart_str = request.POST.get(f'heure_depart_{ens_id}') or None
            heures_travaillees_str = request.POST.get(f'heures_travaillees_{ens_id}') or None
            observations = request.POST.get(f'observations_{ens_id}', '')
            justifie = request.POST.get(f'justifie_{ens_id}') == 'on'
            
            # Convertir les heures en objets time
            heure_arrivee = None
            heure_depart = None
            heures_travaillees = None
            
            if heure_arrivee_str:
                try:
                    heure_arrivee = datetime.strptime(heure_arrivee_str, '%H:%M').time()
                except ValueError:
                    pass
            
            if heure_depart_str:
                try:
                    heure_depart = datetime.strptime(heure_depart_str, '%H:%M').time()
                except ValueError:
                    pass
            
            if heures_travaillees_str:
                try:
                    heures_travaillees = Decimal(heures_travaillees_str)
                except:
                    pass
            
            # Créer ou mettre à jour la présence
            presence, created = PresenceEnseignant.objects.update_or_create(
                enseignant_id=ens_id,
                date=date_pointage,
                defaults={
                    'statut': statut,
                    'heure_arrivee': heure_arrivee,
                    'heure_depart': heure_depart,
                    'heures_travaillees': heures_travaillees,
                    'observations': observations,
                    'justifie': justifie,
                    'pointe_par': request.user,
                }
            )
            
            if created:
                count_created += 1
            else:
                count_updated += 1
        
        messages.success(
            request,
            f"Pointage enregistré: {count_created} nouveau(x), {count_updated} mis à jour."
        )
        return redirect('salaires:liste_presences')
    
    # GET: Afficher le formulaire
    date_pointage_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    
    # Convertir en objet date pour les calculs
    try:
        date_pointage_obj = datetime.strptime(date_pointage_str, '%Y-%m-%d').date()
    except ValueError:
        date_pointage_obj = date.today()
        date_pointage_str = date_pointage_obj.strftime('%Y-%m-%d')
    
    # Récupérer les enseignants actifs
    enseignants = Enseignant.objects.filter(
        ecole=user_school_obj,
        statut='ACTIF'
    ).order_by('nom', 'prenoms')
    
    # Récupérer les présences existantes pour cette date
    presences_existantes = {}
    total_heures_jour = Decimal('0')
    for presence in PresenceEnseignant.objects.filter(date=date_pointage_obj, enseignant__ecole=user_school_obj):
        presences_existantes[presence.enseignant_id] = presence
        if presence.heures_travaillees:
            total_heures_jour += presence.heures_travaillees
    
    # Calculer les heures cumulées du mois pour chaque enseignant
    debut_mois = date_pointage_obj.replace(day=1)
    fin_mois = date_pointage_obj
    
    heures_mois_par_enseignant = {}
    presences_mois = PresenceEnseignant.objects.filter(
        enseignant__ecole=user_school_obj,
        date__gte=debut_mois,
        date__lte=fin_mois
    ).values('enseignant_id').annotate(
        total_heures=Sum('heures_travaillees'),
        jours_presents=Count('id', filter=Q(statut='PRESENT')),
        jours_absents=Count('id', filter=Q(statut='ABSENT')),
        jours_retards=Count('id', filter=Q(statut='RETARD'))
    )
    
    for pm in presences_mois:
        heures_mois_par_enseignant[pm['enseignant_id']] = {
            'heures': pm['total_heures'] or Decimal('0'),
            'presents': pm['jours_presents'],
            'absents': pm['jours_absents'],
            'retards': pm['jours_retards'],
        }
    
    # Statistiques globales du jour
    stats_jour = {
        'total_enseignants': enseignants.count(),
        'pointes': len(presences_existantes),
        'presents': sum(1 for p in presences_existantes.values() if p.statut == 'PRESENT'),
        'absents': sum(1 for p in presences_existantes.values() if p.statut == 'ABSENT'),
        'retards': sum(1 for p in presences_existantes.values() if p.statut == 'RETARD'),
        'total_heures': total_heures_jour,
    }
    
    # Statistiques globales du mois
    total_heures_mois = sum(h['heures'] for h in heures_mois_par_enseignant.values())
    
    context = {
        'enseignants': enseignants,
        'date_pointage': date_pointage_str,
        'date_pointage_obj': date_pointage_obj,
        'presences_existantes': presences_existantes,
        'heures_mois_par_enseignant': heures_mois_par_enseignant,
        'statuts': PresenceEnseignant.STATUT_CHOICES,
        'stats_jour': stats_jour,
        'total_heures_mois': total_heures_mois,
        'mois_courant': date_pointage_obj.strftime('%B %Y'),
    }
    
    return render(request, 'salaires/presences/pointer.html', context)


@login_required
def modifier_presence(request, presence_id):
    """Modifier une présence existante"""
    user_school_obj = user_school(request.user)
    presence = get_object_or_404(
        PresenceEnseignant,
        id=presence_id,
        enseignant__ecole=user_school_obj
    )
    
    if request.method == 'POST':
        form = PresenceForm(request.POST, instance=presence)
        if form.is_valid():
            presence = form.save(commit=False)
            presence.pointe_par = request.user
            presence.save()
            messages.success(request, "Présence modifiée avec succès.")
            return redirect('salaires:liste_presences')
    else:
        form = PresenceForm(instance=presence)
    
    context = {
        'form': form,
        'presence': presence,
    }
    
    return render(request, 'salaires/presences/modifier.html', context)


@login_required
def supprimer_presence(request, presence_id):
    """Supprimer une présence"""
    user_school_obj = user_school(request.user)
    presence = get_object_or_404(
        PresenceEnseignant,
        id=presence_id,
        enseignant__ecole=user_school_obj
    )
    
    if request.method == 'POST':
        enseignant_nom = presence.enseignant.nom_complet
        date_presence = presence.date
        presence.delete()
        messages.success(
            request,
            f"Présence de {enseignant_nom} du {date_presence} supprimée."
        )
        return redirect('salaires:liste_presences')
    
    context = {
        'presence': presence,
    }
    
    return render(request, 'salaires/presences/supprimer.html', context)


@login_required
def rapport_presences(request):
    """Rapport de présences par enseignant et période"""
    user_school_obj = user_school(request.user)
    
    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    enseignant_id = request.GET.get('enseignant')
    
    # Dates par défaut: mois en cours
    if not date_debut:
        date_debut = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = date.today().strftime('%Y-%m-%d')
    
    # Requête de base
    presences = PresenceEnseignant.objects.filter(
        enseignant__ecole=user_school_obj,
        date__gte=date_debut,
        date__lte=date_fin
    ).select_related('enseignant')
    
    if enseignant_id:
        presences = presences.filter(enseignant_id=enseignant_id)
    
    # Regrouper par enseignant
    from collections import defaultdict
    rapport_par_enseignant = defaultdict(lambda: {
        'presents': 0,
        'absents': 0,
        'retards': 0,
        'conges': 0,
        'maladies': 0,
        'permissions': 0,
        'total_heures': Decimal('0'),
        'absences_injustifiees': 0,
    })
    
    for presence in presences:
        ens = presence.enseignant
        rapport_par_enseignant[ens]['enseignant'] = ens
        
        if presence.statut == 'PRESENT':
            rapport_par_enseignant[ens]['presents'] += 1
        elif presence.statut == 'ABSENT':
            rapport_par_enseignant[ens]['absents'] += 1
            if not presence.justifie:
                rapport_par_enseignant[ens]['absences_injustifiees'] += 1
        elif presence.statut == 'RETARD':
            rapport_par_enseignant[ens]['retards'] += 1
        elif presence.statut == 'CONGE':
            rapport_par_enseignant[ens]['conges'] += 1
        elif presence.statut == 'MALADIE':
            rapport_par_enseignant[ens]['maladies'] += 1
        elif presence.statut == 'PERMISSION':
            rapport_par_enseignant[ens]['permissions'] += 1
        
        if presence.heures_travaillees:
            rapport_par_enseignant[ens]['total_heures'] += presence.heures_travaillees
    
    # Convertir en liste
    rapport = list(rapport_par_enseignant.values())
    
    # Liste des enseignants pour le filtre
    enseignants = Enseignant.objects.filter(
        ecole=user_school_obj
    ).order_by('nom', 'prenoms')
    
    context = {
        'rapport': rapport,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'enseignant_id': enseignant_id,
        'enseignants': enseignants,
    }
    
    return render(request, 'salaires/presences/rapport.html', context)


@login_required
def export_presences_csv(request):
    """Exporter les présences en CSV"""
    user_school_obj = user_school(request.user)
    
    # Filtres
    date_debut = request.GET.get('date_debut', date.today().strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', date.today().strftime('%Y-%m-%d'))
    enseignant_id = request.GET.get('enseignant')
    
    # Requête
    presences = PresenceEnseignant.objects.filter(
        enseignant__ecole=user_school_obj,
        date__gte=date_debut,
        date__lte=date_fin
    ).select_related('enseignant')
    
    if enseignant_id:
        presences = presences.filter(enseignant_id=enseignant_id)
    
    # Créer le CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="presences_{date_debut}_{date_fin}.csv"'
    response.write('\ufeff')  # BOM UTF-8
    
    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Enseignant', 'Statut', 'Heure arrivée', 'Heure départ',
        'Heures travaillées', 'Justifié', 'Observations'
    ])
    
    for presence in presences:
        writer.writerow([
            presence.date.strftime('%d/%m/%Y'),
            presence.enseignant.nom_complet,
            presence.get_statut_display(),
            presence.heure_arrivee.strftime('%H:%M') if presence.heure_arrivee else '',
            presence.heure_depart.strftime('%H:%M') if presence.heure_depart else '',
            str(presence.heures_travaillees) if presence.heures_travaillees else '',
            'Oui' if presence.justifie else 'Non',
            presence.observations,
        ])
    
    return response


@login_required
def export_presences_excel(request):
    """Exporter les présences en Excel avec statistiques"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "Le module openpyxl n'est pas installé.")
        return redirect('salaires:liste_presences')
    
    user_school_obj = user_school(request.user)
    
    # Filtres
    date_debut_str = request.GET.get('date_debut', date.today().replace(day=1).strftime('%Y-%m-%d'))
    date_fin_str = request.GET.get('date_fin', date.today().strftime('%Y-%m-%d'))
    enseignant_id = request.GET.get('enseignant')
    
    try:
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
    except ValueError:
        date_debut = date.today().replace(day=1)
        date_fin = date.today()
    
    # Requête
    presences = PresenceEnseignant.objects.filter(
        enseignant__ecole=user_school_obj,
        date__gte=date_debut,
        date__lte=date_fin
    ).select_related('enseignant').order_by('enseignant__nom', 'date')
    
    if enseignant_id:
        presences = presences.filter(enseignant_id=enseignant_id)
    
    # Créer le workbook
    wb = Workbook()
    
    # ===== FEUILLE 1: Détail des présences =====
    ws1 = wb.active
    ws1.title = "Détail Présences"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"Pointage des Présences - {user_school_obj.nom}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal="center")
    
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f"Période: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws1['A2'].alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = ['Date', 'Enseignant', 'Type', 'Statut', 'Arrivée', 'Départ', 'Heures', 'Observations']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Couleurs par statut
    statut_colors = {
        'PRESENT': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'ABSENT': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'RETARD': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'CONGE': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
        'MALADIE': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'PERMISSION': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    }
    
    # Données
    row = 5
    for presence in presences:
        ws1.cell(row=row, column=1, value=presence.date.strftime('%d/%m/%Y')).border = thin_border
        ws1.cell(row=row, column=2, value=presence.enseignant.nom_complet).border = thin_border
        ws1.cell(row=row, column=3, value=presence.enseignant.get_type_enseignant_display()).border = thin_border
        
        statut_cell = ws1.cell(row=row, column=4, value=presence.get_statut_display())
        statut_cell.border = thin_border
        statut_cell.fill = statut_colors.get(presence.statut, PatternFill())
        
        ws1.cell(row=row, column=5, value=presence.heure_arrivee.strftime('%H:%M') if presence.heure_arrivee else '').border = thin_border
        ws1.cell(row=row, column=6, value=presence.heure_depart.strftime('%H:%M') if presence.heure_depart else '').border = thin_border
        ws1.cell(row=row, column=7, value=float(presence.heures_travaillees) if presence.heures_travaillees else 0).border = thin_border
        ws1.cell(row=row, column=8, value=presence.observations or '').border = thin_border
        row += 1
    
    # Ajuster les largeurs
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 30
    
    # ===== FEUILLE 2: Récapitulatif par enseignant =====
    ws2 = wb.create_sheet(title="Récapitulatif")
    
    # Titre
    ws2.merge_cells('A1:I1')
    ws2['A1'] = f"Récapitulatif des Présences - {date_debut.strftime('%B %Y')}"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal="center")
    
    # En-têtes récap
    headers_recap = ['Enseignant', 'Type', 'Présents', 'Absents', 'Retards', 'Congés', 'Maladies', 'Permissions', 'Total Heures']
    for col, header in enumerate(headers_recap, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Calculer les stats par enseignant
    from collections import defaultdict
    stats_enseignant = defaultdict(lambda: {
        'enseignant': None,
        'type': '',
        'presents': 0,
        'absents': 0,
        'retards': 0,
        'conges': 0,
        'maladies': 0,
        'permissions': 0,
        'heures': Decimal('0'),
    })
    
    for presence in presences:
        ens = presence.enseignant
        stats_enseignant[ens.id]['enseignant'] = ens.nom_complet
        stats_enseignant[ens.id]['type'] = ens.get_type_enseignant_display()
        
        if presence.statut == 'PRESENT':
            stats_enseignant[ens.id]['presents'] += 1
        elif presence.statut == 'ABSENT':
            stats_enseignant[ens.id]['absents'] += 1
        elif presence.statut == 'RETARD':
            stats_enseignant[ens.id]['retards'] += 1
        elif presence.statut == 'CONGE':
            stats_enseignant[ens.id]['conges'] += 1
        elif presence.statut == 'MALADIE':
            stats_enseignant[ens.id]['maladies'] += 1
        elif presence.statut == 'PERMISSION':
            stats_enseignant[ens.id]['permissions'] += 1
        
        if presence.heures_travaillees:
            stats_enseignant[ens.id]['heures'] += presence.heures_travaillees
    
    # Écrire les données récap
    row = 4
    total_presents = total_absents = total_retards = total_conges = total_maladies = total_permissions = 0
    total_heures = Decimal('0')
    
    for ens_id, stats in stats_enseignant.items():
        ws2.cell(row=row, column=1, value=stats['enseignant']).border = thin_border
        ws2.cell(row=row, column=2, value=stats['type']).border = thin_border
        ws2.cell(row=row, column=3, value=stats['presents']).border = thin_border
        ws2.cell(row=row, column=4, value=stats['absents']).border = thin_border
        ws2.cell(row=row, column=5, value=stats['retards']).border = thin_border
        ws2.cell(row=row, column=6, value=stats['conges']).border = thin_border
        ws2.cell(row=row, column=7, value=stats['maladies']).border = thin_border
        ws2.cell(row=row, column=8, value=stats['permissions']).border = thin_border
        ws2.cell(row=row, column=9, value=float(stats['heures'])).border = thin_border
        
        total_presents += stats['presents']
        total_absents += stats['absents']
        total_retards += stats['retards']
        total_conges += stats['conges']
        total_maladies += stats['maladies']
        total_permissions += stats['permissions']
        total_heures += stats['heures']
        row += 1
    
    # Ligne de totaux
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws2.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row, column=1).fill = total_fill
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=2, value="").fill = total_fill
    ws2.cell(row=row, column=2).border = thin_border
    ws2.cell(row=row, column=3, value=total_presents).font = Font(bold=True)
    ws2.cell(row=row, column=3).fill = total_fill
    ws2.cell(row=row, column=3).border = thin_border
    ws2.cell(row=row, column=4, value=total_absents).font = Font(bold=True)
    ws2.cell(row=row, column=4).fill = total_fill
    ws2.cell(row=row, column=4).border = thin_border
    ws2.cell(row=row, column=5, value=total_retards).font = Font(bold=True)
    ws2.cell(row=row, column=5).fill = total_fill
    ws2.cell(row=row, column=5).border = thin_border
    ws2.cell(row=row, column=6, value=total_conges).font = Font(bold=True)
    ws2.cell(row=row, column=6).fill = total_fill
    ws2.cell(row=row, column=6).border = thin_border
    ws2.cell(row=row, column=7, value=total_maladies).font = Font(bold=True)
    ws2.cell(row=row, column=7).fill = total_fill
    ws2.cell(row=row, column=7).border = thin_border
    ws2.cell(row=row, column=8, value=total_permissions).font = Font(bold=True)
    ws2.cell(row=row, column=8).fill = total_fill
    ws2.cell(row=row, column=8).border = thin_border
    ws2.cell(row=row, column=9, value=float(total_heures)).font = Font(bold=True)
    ws2.cell(row=row, column=9).fill = total_fill
    ws2.cell(row=row, column=9).border = thin_border
    
    # Ajuster les largeurs
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    for col in range(3, 10):
        ws2.column_dimensions[get_column_letter(col)].width = 12
    
    # Générer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"presences_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
