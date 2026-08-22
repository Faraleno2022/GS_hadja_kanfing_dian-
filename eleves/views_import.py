"""
Vues pour l'importation d'élèves depuis Excel/CSV
"""
import os
import tempfile
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction

from eleves.models import Classe, Eleve


@login_required
@require_http_methods(["GET", "POST"])
def importer_eleves(request):
    """
    Vue principale pour importer des élèves
    """
    # Vérifier les permissions
    peut_importer = (
        request.user.is_staff or 
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrateurs', 'Directeurs', 'Comptables']).exists() or
        (hasattr(request.user, 'profil') and request.user.profil.peut_importer_eleves) or
        (hasattr(request.user, 'profil') and request.user.profil.role == 'COMPTABLE')
    )
    
    if not peut_importer:
        messages.error(request, "Vous n'avez pas la permission d'importer des élèves.")
        return redirect('eleves:liste_eleves')
    
    # Déterminer l'année scolaire à utiliser : on prend la plus récente existante
    annee_courante = Classe.objects.order_by('-annee_scolaire').values_list('annee_scolaire', flat=True).first()

    # Récupérer les classes pour cette année scolaire
    classes = Classe.objects.all()
    if annee_courante:
        classes = classes.filter(annee_scolaire=annee_courante)
    classes = classes.order_by('nom')
    
    # Un utilisateur sans école rattachée ne doit jamais voir les classes des
    # autres établissements.
    from utilisateurs.utils import filter_by_user_school
    classes = filter_by_user_school(classes, request.user)
    
    context = {
        'classes': classes,
        'annee_courante': annee_courante,
    }
    
    if request.method == 'POST':
        return _traiter_import_eleves(request)
    
    return render(request, 'eleves/importer_eleves.html', context)


def _ecole_utilisateur(request):
    """École de l'utilisateur (None pour un superutilisateur multi-écoles)."""
    if request.user.is_superuser:
        return None
    from utilisateurs.utils import user_school
    return user_school(request.user)


def _resoudre_ecole_ligne(nom_ecole, ecole_utilisateur, cache):
    """Trouve l'école d'une ligne d'import à partir de son nom.

    Un utilisateur rattaché à une école importe toujours dans la sienne ;
    seul un superutilisateur peut répartir sur plusieurs écoles.
    """
    from eleves.models import Ecole
    from eleves.import_eleves import normaliser_libelle

    if ecole_utilisateur is not None:
        nom = (nom_ecole or '').strip()
        if nom:
            attendu = normaliser_libelle(ecole_utilisateur.nom)
            recu = normaliser_libelle(nom)
            if recu != attendu:
                from eleves.import_eleves import ImportElevesError
                raise ImportElevesError(
                    f"L'école « {nom} » n'est pas accessible avec ce compte."
                )
        return ecole_utilisateur

    nom = (nom_ecole or '').strip()
    if not nom:
        return Ecole.objects.order_by('id').first()
    if nom in cache:
        return cache[nom]

    nom_normalise = normaliser_libelle(nom)
    ecoles = list(Ecole.objects.all())
    ecole = next(
        (candidate for candidate in ecoles
         if normaliser_libelle(candidate.nom) == nom_normalise),
        None,
    )
    if ecole is None:
        ecole = next(
            (candidate for candidate in ecoles
             if nom_normalise in normaliser_libelle(candidate.nom)),
            None,
        )
    cache[nom] = ecole
    return ecole


def _importer_multi_classes(request, df, generer_matricules):
    """Importe un fichier contenant les colonnes École / Classe / Année scolaire.

    Les lignes sont regroupées par classe, puis chaque groupe est confié à
    l'importateur existant. Les classes absentes sont créées.
    """
    from eleves.import_eleves import (
        ImportElevesError,
        ImportElevesProcessor,
        ImportElevesValidator,
        resoudre_classe,
    )

    ecole_utilisateur = _ecole_utilisateur(request)
    cache_ecoles = {}

    totaux = {
        'total': 0,
        'crees': 0,
        'modifies': 0,
        'erreurs': 0,
        'matricules_generes': 0,
        'doublons_ignores': 0,
    }
    doublons_details = []
    classes_creees = []
    classes_traitees = []
    echecs = []

    colonnes_groupe = ['Classe']
    if 'École' in df.columns:
        colonnes_groupe.insert(0, 'École')
    if 'Année scolaire' in df.columns:
        colonnes_groupe.append('Année scolaire')

    df = df.copy()
    for colonne in colonnes_groupe:
        df[colonne] = df[colonne].fillna('').astype(str).str.strip()

    for cles, groupe in df.groupby(colonnes_groupe, dropna=False, sort=False):
        if not isinstance(cles, tuple):
            cles = (cles,)
        valeurs = dict(zip(colonnes_groupe, cles))
        nom_classe = valeurs.get('Classe', '')
        nom_ecole = valeurs.get('École', '')
        annee = valeurs.get('Année scolaire', '')

        if not nom_classe:
            echecs.append("Des lignes sans nom de classe ont été ignorées.")
            totaux['erreurs'] += len(groupe)
            continue

        try:
            ecole = _resoudre_ecole_ligne(nom_ecole, ecole_utilisateur, cache_ecoles)
            classe, creee = resoudre_classe(nom_classe, annee, ecole)
        except ImportElevesError as exc:
            echecs.append(str(exc))
            totaux['erreurs'] += len(groupe)
            continue

        if creee:
            classes_creees.append(f"{classe.nom} ({classe.annee_scolaire})")

        # Les lignes fautives sont écartées, les autres sont importées :
        # une seule ligne incomplète ne doit pas bloquer toute une classe.
        validator = ImportElevesValidator(groupe, classe.id)
        validator.valider()
        if validator.erreurs:
            for erreur in validator.erreurs[:3]:
                echecs.append(f"{classe.nom} — {erreur}")
            totaux['erreurs'] += len(validator.lignes_invalides)

        groupe_valide = validator.lignes_valides()
        if groupe_valide.empty:
            continue

        processor = ImportElevesProcessor(
            df=groupe_valide, classe_id=classe.id, user=request.user,
            generer_matricules=generer_matricules,
        )
        stats = processor.importer()
        for cle in totaux:
            totaux[cle] += stats.get(cle, 0)
        doublons_details.extend(stats.get('doublons_details') or [])
        classes_traitees.append(classe.nom)

    messages.success(
        request,
        f"✅ Importation terminée : {len(classes_traitees)} classe(s) traitée(s), "
        f"{totaux['crees']} élève(s) créé(s), {totaux['modifies']} mis à jour."
    )
    if classes_creees:
        messages.info(request, "🏫 Classes créées : " + ", ".join(classes_creees))
    if totaux['matricules_generes']:
        messages.info(request, f"🔢 {totaux['matricules_generes']} matricule(s) généré(s) automatiquement")
    if totaux['doublons_ignores']:
        messages.warning(
            request,
            f"🚫 {totaux['doublons_ignores']} ligne(s) rejetée(s) : matricule déjà utilisé. "
            "Aucun élève en double n'a été créé."
        )
        if doublons_details:
            messages.warning(
                request,
                "Matricules en doublon — " + " ; ".join(doublons_details[:10])
                + (" ; …" if len(doublons_details) > 10 else "")
            )
    for echec in echecs[:5]:
        messages.warning(request, echec)
    if len(echecs) > 5:
        messages.warning(request, f"... et {len(echecs) - 5} autre(s) avertissement(s)")

    return redirect('eleves:gestion_classes')


def _traiter_import_eleves(request):
    """
    Traite l'importation d'élèves
    """
    from eleves.import_eleves import (
        ImportElevesError,
        ImportElevesValidator,
        ImportElevesProcessor,
        lire_fichier_eleves,
        normaliser_colonnes_localisation,
    )

    try:
        # Récupérer les paramètres
        classe_id = request.POST.get('classe_id')
        generer_matricules = request.POST.get('generer_matricules') == 'on'
        repartir_auto = request.POST.get('repartition_auto') == 'on'
        fichier = request.FILES.get('fichier')

        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier.")
            return redirect('eleves:importer_eleves')

        if not classe_id and not repartir_auto:
            messages.error(
                request,
                "Veuillez sélectionner une classe, ou cocher « Répartir automatiquement » "
                "si le fichier contient les colonnes École / Classe / Année scolaire."
            )
            return redirect('eleves:importer_eleves')

        classe_cible = None
        if classe_id:
            from utilisateurs.utils import filter_by_user_school
            classe_cible = filter_by_user_school(
                Classe.objects.filter(id=classe_id), request.user
            ).first()
            if classe_cible is None:
                messages.error(
                    request,
                    "Cette classe n'appartient pas à votre école : importation refusée."
                )
                return redirect('eleves:importer_eleves')

        # Sauvegarder temporairement le fichier
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(fichier.name)[1]) as tmp_file:
            for chunk in fichier.chunks():
                tmp_file.write(chunk)
            tmp_path = tmp_file.name

        try:
            # Lire le fichier
            df = lire_fichier_eleves(tmp_path)
            df = normaliser_colonnes_localisation(df)

            # Fichier exporté depuis un autre poste : répartition par classe
            if repartir_auto or (not classe_id and 'Classe' in df.columns):
                if 'Classe' not in df.columns:
                    messages.error(
                        request,
                        "Le fichier ne contient pas de colonne « Classe » : "
                        "sélectionnez une classe de destination."
                    )
                    return redirect('eleves:importer_eleves')
                return _importer_multi_classes(request, df, generer_matricules)

            # Valider les données
            validator = ImportElevesValidator(df, classe_id)
            
            if not validator.valider():
                # Afficher les erreurs
                for erreur in validator.erreurs[:5]:  # Limiter à 5 erreurs
                    messages.error(request, erreur)
                if len(validator.erreurs) > 5:
                    messages.error(request, f"... et {len(validator.erreurs) - 5} autres erreurs")
                return redirect('eleves:importer_eleves')
            
            # Afficher les avertissements
            for avertissement in validator.avertissements[:3]:
                messages.warning(request, avertissement)
            
            # Importer les données
            processor = ImportElevesProcessor(
                df=df,
                classe_id=classe_id,
                user=request.user,
                generer_matricules=generer_matricules
            )
            
            stats = processor.importer()
            
            # Afficher les résultats
            classe = classe_cible
            
            messages.success(
                request,
                f"✅ Importation terminée pour la classe {classe.nom}!"
            )
            
            if stats['crees'] > 0:
                messages.success(
                    request,
                    f"📝 {stats['crees']} élève(s) créé(s)"
                )
            
            if stats['modifies'] > 0:
                messages.info(
                    request,
                    f"✏️ {stats['modifies']} élève(s) mis à jour"
                )
            
            if stats['matricules_generes'] > 0:
                messages.info(
                    request,
                    f"🔢 {stats['matricules_generes']} matricule(s) généré(s) automatiquement"
                )

            if stats.get('doublons_ignores', 0) > 0:
                messages.warning(
                    request,
                    f"🚫 {stats['doublons_ignores']} ligne(s) rejetée(s) : matricule déjà "
                    "utilisé. Aucun élève en double n'a été créé."
                )
                details = stats.get('doublons_details') or []
                if details:
                    messages.warning(
                        request,
                        "Matricules en doublon — " + " ; ".join(details[:10])
                        + (" ; …" if len(details) > 10 else "")
                    )
            
            if stats['erreurs'] > 0:
                messages.warning(
                    request,
                    f"⚠️ {stats['erreurs']} erreur(s) rencontrée(s)"
                )
            
            messages.info(
                request,
                f"📊 Total traité: {stats['total']} élève(s)"
            )
            
            # Rediriger vers la liste des élèves de la classe
            return redirect('eleves:gestion_classes')
            
        finally:
            # Nettoyer le fichier temporaire
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    except ImportElevesError as e:
        messages.error(request, f"Erreur d'importation: {e}")
    except Exception as e:
        messages.error(request, f"Erreur inattendue: {e}")
        import traceback
        print(traceback.format_exc())
    
    return redirect('eleves:importer_eleves')


@login_required
def telecharger_template_eleves(request):
    import pandas as pd
    from eleves.import_eleves import generer_template_eleves

    """
    Télécharge un template Excel pour l'importation d'élèves
    """
    try:
        classe_id = request.GET.get('classe_id')
        
        # Générer le template
        df = generer_template_eleves(classe_id)
        
        # Créer la réponse Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nom du fichier
        if classe_id:
            from utilisateurs.utils import filter_by_user_school
            classe = filter_by_user_school(
                Classe.objects.filter(id=classe_id), request.user
            ).first()
            if classe is None:
                messages.error(request, "Cette classe n'appartient pas à votre école.")
                return redirect('eleves:importer_eleves')
            filename = f"template_eleves_{classe.nom.replace(' ', '_')}.xlsx"
        else:
            filename = f"template_eleves_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Écrire le DataFrame dans la réponse
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Élèves', index=False)
            
            # Obtenir la feuille pour la formater
            worksheet = writer.sheets['Élèves']
            
            # Ajuster la largeur des colonnes
            for idx, col in enumerate(df.columns, 1):
                column_letter = chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)
                if col in ['Prénom', 'Nom', 'Lieu de Naissance', 'Adresse']:
                    worksheet.column_dimensions[column_letter].width = 20
                elif col == 'Matricule':
                    worksheet.column_dimensions[column_letter].width = 15
                elif col in ['Date de Naissance', 'Téléphone Principal', 'Téléphone Secondaire']:
                    worksheet.column_dimensions[column_letter].width = 18
                else:
                    worksheet.column_dimensions[column_letter].width = 15
            
            # Ajouter un style à l'en-tête
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajouter des instructions dans une nouvelle feuille
            instructions_sheet = writer.book.create_sheet('Instructions')
            instructions = [
                ["INSTRUCTIONS POUR L'IMPORTATION DES ÉLÈVES"],
                [""],
                ["1. COLONNES OBLIGATOIRES:"],
                ["   - Prénom: Le prénom de l'élève"],
                ["   - Nom: Le nom de famille de l'élève"],
                ["   - Sexe: M (Masculin) ou F (Féminin)"],
                ["   - Date de Naissance: Format JJ/MM/AAAA (ex: 15/01/2010)"],
                ["   - Lieu de Naissance: Ville ou commune de naissance"],
                ["   - Nom du Père/Tuteur: Nom du responsable principal"],
                ["   - Prénom du Père/Tuteur: Prénom du responsable principal"],
                ["   - Téléphone Principal: Numéro de téléphone (8 chiffres minimum)"],
                ["   - Adresse: Adresse complète de la famille"],
                [""],
                ["2. COLONNES OPTIONNELLES:"],
                ["   - Matricule: Si vide, sera généré automatiquement"],
                ["   - Nom de la Mère: Nom du responsable secondaire"],
                ["   - Prénom de la Mère: Prénom du responsable secondaire"],
                ["   - Téléphone Secondaire: Numéro secondaire"],
                ["   - Email: Adresse email de contact"],
                [""],
                ["3. RÈGLES IMPORTANTES:"],
                ["   - Ne pas modifier les noms des colonnes"],
                ["   - Respecter le format de date JJ/MM/AAAA"],
                ["   - Le sexe doit être uniquement M ou F"],
                ["   - Les téléphones doivent contenir au moins 8 chiffres"],
                ["   - Supprimer les lignes d'exemple avant l'import"],
                [""],
                ["4. GÉNÉRATION AUTOMATIQUE DES MATRICULES:"],
                ["   - Format: [CODE_CLASSE]-[ANNÉE]-[NUMÉRO]"],
                ["   - Exemple: 6A-2024-001"],
                ["   - Cochez l'option lors de l'importation"],
                [""],
                ["5. EN CAS D'ERREUR:"],
                ["   - Vérifier que toutes les colonnes obligatoires sont remplies"],
                ["   - Vérifier le format des dates"],
                ["   - Vérifier que les téléphones sont valides"],
                ["   - S'assurer qu'il n'y a pas de doublons"]
            ]
            
            for row_idx, instruction in enumerate(instructions, 1):
                for col_idx, text in enumerate(instruction, 1):
                    instructions_sheet.cell(row=row_idx, column=col_idx, value=text)
            
            # Formater la feuille d'instructions
            instructions_sheet.column_dimensions['A'].width = 80
            title_cell = instructions_sheet['A1']
            title_cell.font = Font(bold=True, size=14, color="366092")
            
            for row in [3, 14, 21, 27, 32]:
                instructions_sheet.cell(row=row, column=1).font = Font(bold=True, color="366092")
        
        return response
    
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du template: {e}")
        return redirect('eleves:importer_eleves')


# Colonnes à conserver en texte : Excel transformerait sinon un numéro de
# téléphone ou un matricule numérique en nombre, cassant le réimport.
COLONNES_TEXTE = {
    'Matricule', 'Téléphone Principal', 'Téléphone Secondaire', 'Date de Naissance',
}


def _ecrire_feuille_template(writer, df, nom_feuille):
    """Écrit un DataFrame au format template avec un en-tête mis en forme."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    df.to_excel(writer, sheet_name=nom_feuille, index=False)
    worksheet = writer.sheets[nom_feuille]

    for idx, colonne in enumerate(df.columns, start=1):
        lettre = get_column_letter(idx)
        worksheet.column_dimensions[lettre].width = 20
        if colonne in COLONNES_TEXTE:
            for cellule in worksheet[lettre][1:]:
                cellule.number_format = '@'

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1746A2", end_color="1746A2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"


@login_required
def exporter_tous_eleves_template(request):
    """Exporte tous les élèves dans un fichier Excel unique et réimportable.

    Colonnes : École, Classe, Année scolaire, puis les 14 colonnes du modèle
    d'importation. Le fichier peut être importé tel quel sur un autre poste :
    la répartition dans les classes est déduite des trois premières colonnes.
    """
    import io

    import pandas as pd

    from eleves.import_eleves import exporter_tous_les_eleves

    try:
        ecole = None
        if not request.user.is_superuser:
            from utilisateurs.utils import user_school
            ecole = user_school(request.user)

        inclure_inactifs = request.GET.get('inclure_inactifs') == '1'
        df = exporter_tous_les_eleves(ecole=ecole, inclure_inactifs=inclure_inactifs)

        if df.empty:
            messages.warning(request, "Aucun élève à exporter.")
            return redirect('eleves:liste_eleves')

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            _ecrire_feuille_template(writer, df, 'Eleves')

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        nom = f"export_eleves_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nom}"'
        response['X-Eleves-Exportes'] = str(len(df))
        return response

    except Exception as e:
        messages.error(request, f"Erreur lors de l'export: {e}")
        return redirect('eleves:liste_eleves')


@login_required
def exporter_eleves_classe(request, classe_id):
    import pandas as pd
    from eleves.import_eleves import exporter_liste_eleves

    """
    Exporte la liste des élèves d'une classe
    """
    try:
        classe = get_object_or_404(Classe, id=classe_id)
        
        # Vérifier les permissions
        if not request.user.is_superuser:
            from utilisateurs.utils import user_school
            ecole = user_school(request.user)
            if ecole and classe.ecole != ecole:
                messages.error(request, "Vous n'avez pas accès à cette classe.")
                return redirect('eleves:liste_eleves')
        
        # Exporter les élèves
        df = exporter_liste_eleves(classe_id)
        
        # Créer la réponse Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"eleves_{classe.nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Écrire le DataFrame
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Élèves {classe.nom}', index=False)
            
            # Formater
            worksheet = writer.sheets[f'Élèves {classe.nom}']
            
            # Largeur des colonnes
            for idx, col in enumerate(df.columns, 1):
                column_letter = chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)
                worksheet.column_dimensions[column_letter].width = 18
            
            # Style de l'en-tête
            from openpyxl.styles import Font, PatternFill, Alignment
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        return response
    
    except Exception as e:
        messages.error(request, f"Erreur lors de l'export: {e}")
        return redirect('eleves:gestion_classes')
