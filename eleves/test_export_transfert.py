from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from eleves.import_eleves import COLONNES_TRANSFERT, exporter_eleves_modele_import
from eleves.models import Classe, Ecole, Eleve, Responsable


ENTETES_TRANSFERT = [
    'École',
    'Classe',
    'Année scolaire',
    'Matricule',
    'Prénom',
    'Nom',
    'Sexe',
    'Date de Naissance',
    'Lieu de Naissance',
    'Nom du Père/Tuteur',
    'Prénom du Père/Tuteur',
    'Téléphone Principal',
    'Adresse',
    'Nom de la Mère',
    'Prénom de la Mère',
    'Téléphone Secondaire',
    'Email',
]


class ExportElevesTransfertTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.ecole = Ecole.objects.create(
            nom='Groupe Scolaire Exemple',
            adresse='Conakry',
            telephone='+224622000001',
            directeur='Direction',
        )
        cls.classe = Classe.objects.create(
            ecole=cls.ecole,
            nom='6ème A',
            niveau='PRIMAIRE_6',
            annee_scolaire='2025-2026',
        )
        pere = Responsable.objects.create(
            nom='DIALLO',
            prenom='Mamadou',
            relation='PERE',
            telephone='+224622000002',
            adresse='Ratoma',
            email='parent@example.com',
        )
        mere = Responsable.objects.create(
            nom='BAH',
            prenom='Fatoumata',
            relation='MERE',
            telephone='+224622000003',
            adresse='Ratoma',
        )
        cls.eleve = Eleve.objects.create(
            matricule='GS-2025-001',
            prenom='Aïssatou',
            nom='DIALLO',
            sexe='F',
            date_naissance=date(2013, 5, 12),
            lieu_naissance='Conakry',
            classe=cls.classe,
            responsable_principal=pere,
            responsable_secondaire=mere,
            statut='ACTIF',
        )

    def test_dataframe_respecte_le_format_de_transfert_exact(self):
        dataframe = exporter_eleves_modele_import(
            Eleve.objects.filter(pk=self.eleve.pk)
        )

        self.assertEqual(COLONNES_TRANSFERT, ENTETES_TRANSFERT)
        self.assertEqual(list(dataframe.columns), ENTETES_TRANSFERT)
        self.assertEqual(
            list(dataframe.iloc[0]),
            [
                'Groupe Scolaire Exemple',
                '6ème A',
                '2025-2026',
                'GS-2025-001',
                'Aïssatou',
                'DIALLO',
                'F',
                '12/05/2013',
                'Conakry',
                'DIALLO',
                'Mamadou',
                '+224622000002',
                'Ratoma',
                'BAH',
                'Fatoumata',
                '+224622000003',
                'parent@example.com',
            ],
        )

    def test_fichier_excel_telecharge_a_les_memes_entetes(self):
        user = get_user_model().objects.create_superuser(
            username='admin-export',
            password='secret-test',
            email='admin@example.com',
        )
        self.client.force_login(user)

        response = self.client.get(reverse('eleves:exporter_tous_eleves_import'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        feuille = workbook['Élèves']
        entetes = [cell.value for cell in next(feuille.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(entetes, ENTETES_TRANSFERT)
        self.assertEqual(feuille.cell(row=2, column=1).value, 'Groupe Scolaire Exemple')
        self.assertEqual(feuille.cell(row=2, column=2).value, '6ème A')
        self.assertEqual(feuille.cell(row=2, column=3).value, '2025-2026')

