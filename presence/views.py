"""
Pointage journalier des élèves par classe + rapport de présence
et alertes d'absences consécutives.
"""
import io
from datetime import date, datetime, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect

from eleves.models import Classe, Eleve
from utilisateurs.utils import filter_by_user_school
from .models import PresenceJournaliere

SEUIL_ABSENCES_CONSECUTIVES = 3  # alerte à partir de N absences d'affilée


def _classes_utilisateur(request):
    qs = Classe.objects.select_related('ecole').order_by('nom')
    return filter_by_user_school(qs, request.user, 'ecole')


def _parse_date(val, defaut=None):
    if not val:
        return defaut
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return defaut


@login_required
def pointage(request):
    """Sélection classe + date, puis pointage de tous les élèves."""
    classes = list(_classes_utilisateur(request))
    classe_id = (request.GET.get('classe_id') or request.POST.get('classe_id') or '').strip()
    jour = _parse_date(request.GET.get('date') or request.POST.get('date'), date.today())

    classe = None
    eleves = []
    presences_existantes = {}
    if classe_id.isdigit():
        classe = next((c for c in classes if c.id == int(classe_id)), None)

    if classe:
        eleves = list(Eleve.objects.filter(classe=classe, statut='ACTIF')
                      .order_by('prenom', 'nom'))
        presences_existantes = {
            p.eleve_id: p for p in PresenceJournaliere.objects.filter(classe=classe, date=jour)
        }

    if request.method == 'POST' and classe:
        cree = 0
        for eleve in eleves:
            statut = (request.POST.get(f'statut_{eleve.id}') or 'PRESENT').strip()
            motif = (request.POST.get(f'motif_{eleve.id}') or '').strip()
            if statut not in dict(PresenceJournaliere.STATUT_CHOICES):
                statut = 'PRESENT'
            PresenceJournaliere.objects.update_or_create(
                eleve=eleve, date=jour,
                defaults={'classe': classe, 'statut': statut, 'motif': motif,
                          'cree_par': request.user},
            )
            cree += 1
        messages.success(request, f"Pointage enregistré pour {cree} élève(s) — {classe.nom} le {jour.strftime('%d/%m/%Y')}.")
        return redirect(f"{request.path}?classe_id={classe.id}&date={jour.isoformat()}")

    # Pré-remplir le statut courant pour l'affichage
    lignes = []
    for eleve in eleves:
        p = presences_existantes.get(eleve.id)
        lignes.append({
            'eleve': eleve,
            'statut': p.statut if p else 'PRESENT',
            'motif': p.motif if p else '',
        })

    context = {
        'titre_page': "Pointage journalier",
        'classes': classes,
        'classe': classe,
        'classe_id': classe_id,
        'jour': jour,
        'lignes': lignes,
        'statut_choices': PresenceJournaliere.STATUT_CHOICES,
        'deja_pointe': bool(presences_existantes),
    }
    return render(request, 'presence/pointage.html', context)


def _calculer_absences_consecutives(eleve_id, jusqua=None):
    """Nombre d'absences consécutives (statut ABSENT) les plus récentes."""
    jusqua = jusqua or date.today()
    qs = (PresenceJournaliere.objects
          .filter(eleve_id=eleve_id, date__lte=jusqua)
          .order_by('-date')
          .values_list('statut', flat=True))
    streak = 0
    for statut in qs:
        if statut == 'ABSENT':
            streak += 1
        else:
            break
    return streak


def _collecter_rapport(request):
    """Données du rapport de présence, filtrées par classe + période."""
    classes = list(_classes_utilisateur(request))
    classe_id = (request.GET.get('classe_id') or '').strip()
    du = _parse_date(request.GET.get('du'), date.today().replace(day=1))
    au = _parse_date(request.GET.get('au'), date.today())
    seuil = request.GET.get('seuil')
    try:
        seuil = int(seuil) if seuil else SEUIL_ABSENCES_CONSECUTIVES
    except ValueError:
        seuil = SEUIL_ABSENCES_CONSECUTIVES

    classe = None
    if classe_id.isdigit():
        classe = next((c for c in classes if c.id == int(classe_id)), None)

    lignes = []
    if classe:
        eleves = list(Eleve.objects.filter(classe=classe, statut='ACTIF').order_by('prenom', 'nom'))
        # Agrégats par élève sur la période
        agg = (PresenceJournaliere.objects
               .filter(classe=classe, date__gte=du, date__lte=au)
               .values('eleve_id')
               .annotate(
                   present=Count('id', filter=Q(statut='PRESENT')),
                   absent=Count('id', filter=Q(statut='ABSENT')),
                   retard=Count('id', filter=Q(statut='RETARD')),
                   justifie=Count('id', filter=Q(statut='JUSTIFIE')),
                   total=Count('id'),
               ))
        agg_map = {a['eleve_id']: a for a in agg}
        for eleve in eleves:
            a = agg_map.get(eleve.id, {})
            total = a.get('total', 0)
            absent = a.get('absent', 0)
            taux_abs = (absent / total * 100) if total else 0
            consecutives = _calculer_absences_consecutives(eleve.id, au)
            lignes.append({
                'eleve': eleve,
                'present': a.get('present', 0),
                'absent': absent,
                'retard': a.get('retard', 0),
                'justifie': a.get('justifie', 0),
                'total': total,
                'taux_absence': round(taux_abs, 1),
                'consecutives': consecutives,
                'alerte': consecutives >= seuil,
            })

    return {
        'classes': classes,
        'classe': classe,
        'classe_id': classe_id,
        'du': du,
        'au': au,
        'seuil': seuil,
        'lignes': lignes,
        'nb_alertes': sum(1 for l in lignes if l['alerte']),
    }


@login_required
def rapport_presence(request):
    d = _collecter_rapport(request)
    d['titre_page'] = "Rapport de présence"
    return render(request, 'presence/rapport.html', d)


@login_required
def alertes_absences(request):
    """Tableau de toutes les alertes (absences consécutives ≥ seuil), toutes classes."""
    classes = list(_classes_utilisateur(request))
    seuil = request.GET.get('seuil')
    try:
        seuil = int(seuil) if seuil else SEUIL_ABSENCES_CONSECUTIVES
    except ValueError:
        seuil = SEUIL_ABSENCES_CONSECUTIVES
    aujourdhui = date.today()

    # Ne considérer que les élèves absents à leur dernier pointage (optimisation)
    classe_ids = [c.id for c in classes]
    derniers_absents = (PresenceJournaliere.objects
                        .filter(classe_id__in=classe_ids, statut='ABSENT')
                        .values_list('eleve_id', flat=True)
                        .distinct())
    alertes = []
    for eleve in (Eleve.objects.filter(id__in=list(derniers_absents), statut='ACTIF')
                  .select_related('classe')):
        n = _calculer_absences_consecutives(eleve.id, aujourdhui)
        if n >= seuil:
            alertes.append({'eleve': eleve, 'consecutives': n})
    alertes.sort(key=lambda x: x['consecutives'], reverse=True)

    return render(request, 'presence/alertes.html', {
        'titre_page': "Alertes d'absences",
        'alertes': alertes,
        'seuil': seuil,
        'classes': classes,
    })


@login_required
def rapport_presence_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("Module openpyxl non installé", status=500)

    d = _collecter_rapport(request)
    if not d['classe']:
        return HttpResponse("Sélectionnez une classe", status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Présence'
    entete_font = Font(bold=True, color='FFFFFF')
    entete_fill = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')
    centre = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='DDDDDD')
    bordure = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:I1')
    ws['A1'] = f"PRÉSENCE — {d['classe'].nom} — du {d['du'].strftime('%d/%m/%Y')} au {d['au'].strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=12, color='007BFF')
    ws['A1'].alignment = centre

    headers = ['N°', 'Matricule', 'Élève', 'Présent', 'Absent', 'Retard',
               'Justifié', 'Taux abs. %', 'Abs. consécutives']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = entete_font
        c.fill = entete_fill
        c.alignment = centre
        c.border = bordure

    ligne = 3
    for i, l in enumerate(d['lignes'], 1):
        ligne += 1
        e = l['eleve']
        valeurs = [i, e.matricule or '', f"{e.prenom} {e.nom}",
                   l['present'], l['absent'], l['retard'], l['justifie'],
                   l['taux_absence'], l['consecutives']]
        for col, val in enumerate(valeurs, 1):
            c = ws.cell(row=ligne, column=col, value=val)
            c.border = bordure
            if l['alerte'] and col == 9:
                c.font = Font(bold=True, color='C0392B')

    for col, larg in zip('ABCDEFGHI', [5, 14, 28, 9, 9, 9, 10, 12, 16]):
        ws.column_dimensions[col].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="presence_{d["classe"].nom}_{d["au"].isoformat()}.xlsx"'
    return resp


@login_required
def rapport_presence_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    d = _collecter_rapport(request)
    if not d['classe']:
        return HttpResponse("Sélectionnez une classe", status=400)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
                            leftMargin=0.8 * cm, rightMargin=0.8 * cm)
    styles = getSampleStyleSheet()
    titre = ParagraphStyle('T', parent=styles['Heading1'], fontSize=13,
                           textColor=colors.HexColor('#007bff'), alignment=TA_CENTER, spaceAfter=2)
    sous = ParagraphStyle('S', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=6)

    ecole = d['classe'].ecole
    elements = [Paragraph(f"<b>{(ecole.nom if ecole else '').upper()}</b>", titre)]
    elements.append(Paragraph(
        f"RAPPORT DE PRÉSENCE — {d['classe'].nom} — du {d['du'].strftime('%d/%m/%Y')} au {d['au'].strftime('%d/%m/%Y')}", sous))
    if d['nb_alertes']:
        elements.append(Paragraph(
            f"<font color='#C0392B'><b>{d['nb_alertes']} alerte(s) d'absences consécutives (seuil {d['seuil']})</b></font>", sous))

    data = [['N°', 'Matricule', 'Élève', 'Prés.', 'Abs.', 'Ret.', 'Just.', 'Taux abs.', 'Abs. conséc.']]
    for i, l in enumerate(d['lignes'], 1):
        e = l['eleve']
        data.append([str(i), e.matricule or '', f"{e.prenom} {e.nom}",
                     str(l['present']), str(l['absent']), str(l['retard']), str(l['justifie']),
                     f"{l['taux_absence']}%",
                     f"⚠ {l['consecutives']}" if l['alerte'] else str(l['consecutives'])])

    table = Table(data, repeatRows=1,
                  colWidths=[1.0*cm, 2.8*cm, 8.0*cm, 2.0*cm, 2.0*cm, 2.0*cm, 2.0*cm, 2.6*cm, 3.0*cm])
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f7')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]
    for idx, l in enumerate(d['lignes'], 1):
        if l['alerte']:
            style.append(('TEXTCOLOR', (8, idx), (8, idx), colors.HexColor('#C0392B')))
            style.append(('FONTNAME', (8, idx), (8, idx), 'Helvetica-Bold'))
    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="presence_{d["classe"].nom}_{d["au"].isoformat()}.pdf"'
    return resp
